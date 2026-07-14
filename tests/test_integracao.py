"""Testes end-to-end: import + geração de ZIP + validação + render Express.

Estes testes cobrem caminhos que tocam parser + handler + render juntos.
Usam fixture xlsx in-memory pra evitar dependência de arquivos binários.
"""
import io
import json
import zipfile

import openpyxl

from parsers import parse_excel
from ai_rounds import validar_evento
from campo_generator import render_workout


# ── #1: import end-to-end do layout grades-e-dias ─────────────────────────────
def test_parse_excel_grades_e_dias_end_to_end(xlsx_grades_e_dias_bytes):
    result = parse_excel(xlsx_grades_e_dias_bytes)
    assert result["tipo"] == "evento_multidia"
    assert len(result["dias"]) == 1
    dia = result["dias"][0]
    assert dia["label"] == "Sábado"

    # 2 categorias detectadas: Rx Masculino e Scaled Masculino
    nomes_cats = {c["nome"] for c in dia["categorias"]}
    assert nomes_cats == {"Rx Masculino", "Scaled Masculino"}

    # Rx Masculino: 2 baterias (#1 Heat 1, #1 Heat 2 via mista, #2 Final)
    rx = next(c for c in dia["categorias"] if c["nome"] == "Rx Masculino")
    assert len(rx["workouts"]) == 2
    nums_rx = sorted(
        a["numero"]
        for b in rx["baterias"]
        for a in b["alocacoes"]
    )
    # 4 atletas Rx (601, 602, 603, 604)
    assert nums_rx == ["601", "602", "603", "604"]

    # Bateria mista: filtragem por faixa funciona — Scaled NÃO leva atletas RX
    scaled = next(c for c in dia["categorias"] if c["nome"] == "Scaled Masculino")
    nums_scaled = sorted(
        a["numero"]
        for b in scaled["baterias"]
        for a in b["alocacoes"]
    )
    assert nums_scaled == ["901", "902"]   # só os 2 da faixa Scaled

    # Roster
    assert len(result["roster"]) == 6
    nums_roster = {a["numero"] for a in result["roster"]}
    assert nums_roster == {"601", "602", "603", "604", "901", "902"}


# ── #2: geração de ZIP via _handle_generate ───────────────────────────────────
def test_handle_generate_produz_zip_com_html_por_workout(xlsx_grades_e_dias_bytes,
                                                          fonts_empty):
    """Replica internamente o que _handle_generate faz, sem subir HTTP server."""
    from campo_generator import render_workout_combined, sanitize
    from parsers import assign_workout_numbers
    from ai_rounds import enriquecer_workouts

    result = parse_excel(xlsx_grades_e_dias_bytes)
    dia = result["dias"][0]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat in dia["categorias"]:
            workouts = cat["workouts"]
            assign_workout_numbers(workouts)
            enriquecer_workouts(workouts)
            for wkt_pos, wkt in enumerate(workouts, start=1):
                # Junta atletas de baterias que rodam esse workout
                atletas = []
                for b in cat["baterias"]:
                    if (b.get("workouts_que_rodam") or []) and wkt_pos not in b["workouts_que_rodam"]:
                        continue
                    for a in b.get("alocacoes", []):
                        atletas.append({
                            "nome": a["nome"], "box": a["box"], "raia": a["raia"],
                            "bateria": b["numero"], "numero": a["numero"],
                        })
                if not atletas:
                    continue
                html = render_workout_combined(
                    {"nome": "TESTE", "categoria": cat["nome"], "data": "Sábado"},
                    wkt, fonts_empty, "", "", atletas,
                )
                caminho = f"Sábado/{sanitize(cat['nome'])}/{wkt_pos:02d}_{sanitize(wkt['nome'])}.html"
                zf.writestr(caminho, html.encode("utf-8"))

    zip_bytes = buf.getvalue()
    assert len(zip_bytes) > 0
    # Confere conteúdo do zip
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        nomes = zf.namelist()
        # Pelo menos 1 arquivo por categoria × workout que tem atletas
        assert any("Rx_Masculino" in n and "TWENTIES" in n for n in nomes)
        # Lê 1 e confere que tem nome de atleta
        rx_wkt1 = [n for n in nomes if "Rx_Masculino" in n and n.endswith(".html")][0]
        html = zf.read(rx_wkt1).decode("utf-8")
        assert "ATLETA RX 1" in html


# ── Bug fix v1.34.0: workout só aparece no dia que tem bateria rodando ──────
def test_handle_generate_pula_workout_em_dia_sem_bateria_que_o_rode(fonts_empty):
    """Monstar Games tem 3 dias. Workout #5 (Monstar Recap) só roda no Sábado.
    ZIP gerado deve ter o arquivo SÓ em Sábado/, não em Sexta/ ou Domingo/.

    Antes do fix: workouts sem bateria do dia eram gerados em BRANCO mesmo
    assim, poluindo o ZIP com súmulas vazias em dias errados."""
    # Config com 3 dias: cada um tem a mesma categoria, mas baterias rodam
    # workouts diferentes. Mesmos 5 workouts globais em todos os dias.
    workouts = [
        {"numero": 1, "nome": "TWENTIES",      "tipo": "for_time", "time_cap": "9 min",
         "modalidade": "individual", "movimentos": [{"nome": "PULL-UPS", "reps": 20}, {"chegada": True}]},
        {"numero": 2, "nome": "FRAN",          "tipo": "for_time", "time_cap": "5 min",
         "modalidade": "individual", "movimentos": [{"nome": "THRUSTERS", "reps": 21}, {"chegada": True}]},
        {"numero": 3, "nome": "MAX CLEAN",     "tipo": "for_load", "modalidade": "individual"},
        {"numero": 4, "nome": "AMRAP TEST",    "tipo": "amrap",    "time_cap": "10 min",
         "modalidade": "individual", "movimentos": [{"nome": "BURPEES", "reps": 10}]},
        {"numero": 5, "nome": "MONSTAR RECAP", "tipo": "amrap",    "time_cap": "12:30",
         "modalidade": "trio",       "movimentos": [{"nome": "SWIM", "reps": 50}]},
    ]

    def cat(workouts_que_rodam):
        """Cria categoria com 1 bateria que roda os workouts indicados."""
        return {
            "nome": "Trio Rx Misto",
            "workouts": workouts,
            "baterias": [{
                "numero": "1", "codigo_evento": "#1",
                "horario_aquecimento": "08:00", "horario_fila": "08:20",
                "workouts_que_rodam": workouts_que_rodam,
                "alocacoes": [{"raia": "1", "numero": "401", "nome": "Trio A", "box": "CF"}],
            }],
        }

    config = {
        "evento": {"nome": "MONSTAR 2026", "categoria": "", "data": ""},
        "dias": [
            {"label": "Sexta",   "data": "29/05/2026", "categorias": [cat([1, 2])]},
            {"label": "Sábado",  "data": "30/05/2026", "categorias": [cat([3, 5])]},
            {"label": "Domingo", "data": "31/05/2026", "categorias": [cat([4])]},
        ],
        "roster": [],
    }

    # Replica logic de _handle_generate
    from campo_generator import render_workout_combined, render_workout, sanitize
    from parsers import assign_workout_numbers
    from ai_rounds import enriquecer_workouts

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for dia in config["dias"]:
            dia_label = dia["label"]
            dia_data = dia["data"]
            for cat_d in dia["categorias"]:
                workouts_d = cat_d["workouts"]
                assign_workout_numbers(workouts_d)
                enriquecer_workouts(workouts_d)
                baterias = cat_d["baterias"]
                baterias_com_cron = [b for b in baterias if b.get("workouts_que_rodam")]
                algum_sem_cron = len(baterias_com_cron) < len(baterias)
                data_combinada = " ".join(filter(None, [dia_label, dia_data])).strip()
                ev_local = {"nome": "MONSTAR 2026", "categoria": cat_d["nome"], "data": data_combinada}
                for wkt_pos, wkt in enumerate(workouts_d, start=1):
                    roda = algum_sem_cron or any(
                        wkt_pos in b["workouts_que_rodam"] for b in baterias_com_cron)
                    if baterias and not roda:
                        continue   # APLICANDO O FIX
                    atletas = []
                    for b in baterias:
                        wqr = b.get("workouts_que_rodam") or []
                        if wqr and wkt_pos not in wqr: continue
                        for a in b.get("alocacoes", []):
                            atletas.append({**a, "bateria": b["numero"]})
                    if atletas:
                        html = render_workout_combined(ev_local, wkt, fonts_empty, "", "", atletas)
                    else:
                        html = render_workout(ev_local, wkt, fonts_empty, "", "")
                    caminho = f"{dia_label}/{sanitize(cat_d['nome'])}/{wkt_pos:02d}_{sanitize(wkt['nome'])}.html"
                    zf.writestr(caminho, html.encode("utf-8"))

    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as zf:
        nomes = zf.namelist()
    # Sexta: SÓ workouts 1 e 2
    sexta = [n for n in nomes if n.startswith("Sexta/")]
    assert len(sexta) == 2, f"Sexta deve ter 2 wkts, got {len(sexta)}: {sexta}"
    assert any("TWENTIES" in n for n in sexta)
    assert any("FRAN" in n for n in sexta)
    assert not any("MONSTAR" in n for n in sexta), "Monstar NÃO deve aparecer em Sexta"

    # Sábado: SÓ workouts 3 e 5 (Max Clean + Monstar Recap)
    sabado = [n for n in nomes if n.startswith("Sábado/")]
    assert len(sabado) == 2, f"Sábado deve ter 2 wkts, got {len(sabado)}: {sabado}"
    assert any("MAX_CLEAN" in n for n in sabado)
    assert any("MONSTAR" in n for n in sabado)

    # Domingo: SÓ workout 4 (Amrap Test)
    domingo = [n for n in nomes if n.startswith("Domingo/")]
    assert len(domingo) == 1, f"Domingo deve ter 1 wkt, got {len(domingo)}: {domingo}"
    assert "AMRAP_TEST" in domingo[0]

    # Verifica que header tem label do dia + data
    sabado_monstar = [n for n in nomes if "Sábado" in n and "MONSTAR" in n][0]
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as zf:
        html = zf.read(sabado_monstar).decode("utf-8")
    assert "Sábado 30/05/2026" in html or "SÁBADO 30/05/2026" in html, \
        "Header da súmula deve conter 'Sábado 30/05/2026'"


# ── #3: render Express com 2 fórmulas ─────────────────────────────────────────
def test_render_workout_express_emite_amrap_e_for_time(workout_express, evento_basico,
                                                        fonts_empty):
    html = render_workout(evento_basico, workout_express, fonts_empty,
                          logo_src="", logo_evento="")
    # Header tem ambos os números
    assert ">1<" in html and ">2<" in html
    # Tem badge Express + ambas as fórmulas
    assert "Express" in html
    assert "Fórmula 1" in html or "F&#243;rmula 1" in html or "AMRAP" in html
    assert "Fórmula 2" in html or "F&#243;rmula 2" in html or "For Time" in html
    # Movimentos da F1 e F2 aparecem
    assert "HANDSTAND PUSH-UPS" in html
    assert "DEADLIFTS" in html
    assert "DOUBLE UNDERS" in html


# ── #4: validar_evento detecta problemas reais ────────────────────────────────
def test_validar_evento_detecta_workout_sem_movimentos_e_competidor_duplicado():
    config = {
        "dias": [{
            "label": "Sexta",
            "categorias": [{
                "nome": "Rx Masculino",
                "workouts": [
                    {"nome": "VAZIO", "tipo": "for_time", "movimentos": []},
                    {"nome": "OK", "tipo": "for_time", "time_cap": "9 min",
                     "movimentos": [{"nome": "PULL-UPS", "reps": 10}, {"chegada": True}]},
                ],
                "baterias": [
                    {"numero": "1", "codigo_evento": "#1", "horario_aquecimento": "07:00",
                     "horario_fila": "07:20", "workouts_que_rodam": [2],
                     "alocacoes": [
                         {"raia": "1", "numero": "601", "nome": "Foo", "box": "X"},
                     ]},
                    {"numero": "2", "codigo_evento": "#1", "horario_aquecimento": "07:30",
                     "horario_fila": "07:50", "workouts_que_rodam": [2],
                     "alocacoes": [
                         # Mesmo número do atleta acima — duplicado
                         {"raia": "2", "numero": "601", "nome": "Foo", "box": "X"},
                     ]},
                ],
            }],
        }],
    }
    avisos = validar_evento(config)
    # Procura aviso por workout vazio
    assert any("VAZIO" in a["msg"] and "sem movimentos" in a["msg"].lower()
               for a in avisos), f"esperava aviso de workout vazio, got: {avisos}"
    # Procura aviso por competidor duplicado
    assert any("601" in a["msg"] and "2 lugares" in a["msg"] for a in avisos)


# ── #5: validar_evento detecta time cap ausente em For Time ───────────────────
def test_parse_excel_multi_arena_separa_atletas_por_arena(xlsx_multi_arena_bytes):
    """Excel multi-arena (estilo Monstar): cronograma e montagem com 2 blocos
    paralelos. Atletas de cada arena devem ir pra categoria certa.
    """
    result = parse_excel(xlsx_multi_arena_bytes)
    assert result["tipo"] == "evento_multidia"
    # 2 dias: Sexta (com montagem) e Sábado (sem montagem)
    labels = [d["label"] for d in result["dias"]]
    assert "Sexta" in labels
    assert "Sábado" in labels
    # Sexta: 2 categorias (Elite Mas + Elite Fem), atletas separados
    sexta = next(d for d in result["dias"] if d["label"] == "Sexta")
    masculino = next((c for c in sexta["categorias"] if c["nome"] == "Elite Masculino"), None)
    feminino  = next((c for c in sexta["categorias"] if c["nome"] == "Elite Feminino"), None)
    assert masculino is not None and feminino is not None
    nums_m = {a["numero"] for b in masculino["baterias"] for a in b["alocacoes"]}
    nums_f = {a["numero"] for b in feminino["baterias"] for a in b["alocacoes"]}
    # 4 atletas masculinos (101-104), 3 femininos (201-203), sem vazamento
    assert nums_m == {"101", "102", "103", "104"}
    assert nums_f == {"201", "202", "203"}
    # Roster lido da aba 'Atletas' (sem prefixo)
    assert len(result["roster"]) == 7


def test_parse_excel_dia_sem_montagem_inclui_baterias_sem_atletas(xlsx_multi_arena_bytes):
    """Dia que só tem cronograma (sem <Dia> - Montagem) é incluído pra gerar
    súmulas em branco. Útil pra fase de planejamento antes do sorteio."""
    result = parse_excel(xlsx_multi_arena_bytes)
    sabado = next(d for d in result["dias"] if d["label"] == "Sábado")
    masculino = next((c for c in sabado["categorias"] if c["nome"] == "Elite Masculino"), None)
    assert masculino is not None
    # Bateria existe mas sem alocações
    assert len(masculino["baterias"]) == 1
    assert masculino["baterias"][0]["alocacoes"] == []


def test_parse_excel_grade_com_coluna_dia_e_sem_nenhuma_montagem():
    """Regressão (Pwrd by Coffee 2026): arquivo de programação com grades
    `Workouts - <Modalidade>` que têm COLUNA DE DIA na col A (ex: 'Sexta' sem
    data/quebra-de-linha) e abas de cronograma <Dia> SEM nenhuma `- Montagem`
    (roster ainda não fechado). Deve reconhecer e gerar baterias em branco —
    não estourar 'Excel sem dados reconhecíveis'.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Grade com coluna de dia (col A) — dia SEM quebra de linha, workout na col B
    ws = wb.create_sheet("Workouts - Individuais")
    ws.append([None, "Elite Masculino", "Elite Feminino"])
    ws.append(["Sexta",
               '"Muscle Swim"\nFor time:\n50m Swim\n10 Devil Presses\nTime cap: 8 min',
               '"Muscle Swim"\nFor time:\n50m Swim\n10 Devil Presses\nTime cap: 8 min'])

    # Cronograma do dia — SEM aba `Sexta - Montagem`
    ws = wb.create_sheet("Sexta")
    ws.append(["Arena: Campo"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade", "Aquecimento", "", "Fila"])
    ws.append(['"Muscle Swim"', "Elite Masculino (Heat 1)", 1, None, "7 (7)", "14:20", None, "14:45"])
    ws.append([None, "Elite Feminino (Heat 1)", 2, None, "5 (5)", "14:35", None, "15:00"])

    ws = wb.create_sheet("Inscritos")
    ws.append(["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"])
    ws.append(["Elite Masculino", 45, 7, 101, 199, "Sim"])
    ws.append(["Elite Feminino", 45, 5, 201, 299, "Sim"])

    buf = io.BytesIO()
    wb.save(buf)

    result = parse_excel(buf.getvalue())
    assert result["tipo"] == "evento_multidia", result.get("erro")
    sexta = next(d for d in result["dias"] if d["label"] == "Sexta")
    masc = next(c for c in sexta["categorias"] if c["nome"] == "Elite Masculino")
    # Bateria reconhecida, workout anexado, raias em branco (sem roster)
    assert len(masc["baterias"]) == 1
    assert masc["baterias"][0]["alocacoes"] == []
    assert masc["workouts"] and masc["workouts"][0]["nome"] == "MUSCLE SWIM"


def _wb_para_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_linter_categoria_da_grade_sem_bateria_avisa():
    """Fase 2.0: categoria com workouts na grade mas que não casa com nenhuma
    bateria do cronograma → aviso de erro (não geraria súmula silenciosamente)."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Workouts - Individuais")
    ws.append([None, "Elite Masculino", "Categoria Fantasma"])
    ws.append(["Sexta",
               '"W"\nFor time:\n5 Burpees\nTime cap: 3 min',
               '"W"\nFor time:\n5 Burpees\nTime cap: 3 min'])
    ws = wb.create_sheet("Sexta")
    ws.append(["Arena: Campo"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade", "Aquecimento", "", "Fila"])
    ws.append(['"W"', "Elite Masculino (Heat 1)", 1, None, "3 (3)", "08:00", None, "08:20"])
    result = parse_excel(_wb_para_bytes(wb))
    avisos = result.get("avisos_import", [])
    assert any("Categoria Fantasma" in a["msg"] and a.get("nivel") == "erro" for a in avisos)


def test_linter_colisao_de_bateria_mesma_arena_mas_nao_entre_arenas():
    """Fase 2.0: número de bateria repetido na MESMA arena (categorias diferentes)
    é colisão; o MESMO número em arenas diferentes NÃO é (numeração é por arena)."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Workouts - Individuais")
    ws.append([None, "A", "B"])
    ws.append(["Sexta",
               '"W"\nFor time:\n5 Burpees\nTime cap: 3 min',
               '"W"\nFor time:\n5 Burpees\nTime cap: 3 min'])
    # Duas arenas lado a lado (Campo cols 0-7, Quadra cols 9-16)
    ws = wb.create_sheet("Sexta")
    ws.append(["Arena: Campo", None, None, None, None, None, None, None, None,
               "Arena: Quadra"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade", "Aquecimento", "", "Fila", None,
               "Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade", "Aquecimento", "", "Fila"])
    # Campo: bateria 5 com A e B (colisão). Quadra: bateria 5 com A (NÃO colide com Campo).
    ws.append(['"W"', "A (Heat 1)", 5, None, "3 (3)", "08:00", None, "08:20", None,
               '"W"', "A (Heat 1)", 5, None, "3 (3)", "09:00", None, "09:20"])
    ws.append([None, "B (Heat 1)", 5, None, "3 (3)", "08:30", None, "08:50"])
    avisos = parse_excel(_wb_para_bytes(wb)).get("avisos_import", [])
    colisoes = [a for a in avisos if "duplicada" in a["msg"]]
    assert len(colisoes) == 1, f"esperava 1 colisão (Campo), got: {colisoes}"
    assert "Campo" in colisoes[0]["msg"] and "5" in colisoes[0]["msg"]


def _cfg_wkt(mov):
    return {"dias": [{"label": "Dom", "categorias": [
        {"nome": "X", "baterias": [], "workouts": [
            {"nome": "W", "tipo": "for_time", "time_cap": "8 min", "movimentos": mov}]}]}]}


def test_linter_carga_faltando_entre_levantamentos():
    """Fase 2.0: levantamento de barra sem carga onde outro do mesmo workout tem
    carga → aviso (Rocket Master F original: DL 34kg + Snatch/OHS sem carga)."""
    avisos = validar_evento(_cfg_wkt([
        {"nome": "DEADLIFTS", "reps": 12, "carga": "50 KG"},
        {"nome": "HANG POWER SNATCH", "reps": 9},
        {"nome": "OVERHEAD SQUATS", "reps": 6},
        {"chegada": True},
    ]))
    faltando = [a for a in avisos if "carga esquecida" in a["msg"]]
    assert len(faltando) == 1
    assert "HANG POWER SNATCH" in faltando[0]["msg"] and "OVERHEAD SQUATS" in faltando[0]["msg"]
    # Se todos têm carga, nenhum aviso
    ok = validar_evento(_cfg_wkt([
        {"nome": "DEADLIFTS", "reps": 12, "carga": "50 KG"},
        {"nome": "OVERHEAD SQUATS", "reps": 6, "carga": "50 KG"},
        {"chegada": True},
    ]))
    assert not any("carga esquecida" in a["msg"] for a in ok)


def test_linter_typo_anotacao_atlhetes_mas_athletes_ok():
    """Fase 2.0: 'atlhetes' (typo de athletes) avisa; 'athletes' correto não."""
    avisos = validar_evento(_cfg_wkt([
        {"nome": "SYNC. PULL-UPS (2 ATLHETES)", "reps": 40},
        {"nome": "400M RUN (3 ATHLETES)", "reps": 400},
        {"chegada": True},
    ]))
    typos = [a for a in avisos if "typo" in a["msg"]]
    assert len(typos) == 1
    assert "atlhetes" in typos[0]["msg"]


def test_linter_dumbbell_fora_do_rol_e_nao_flagga_barra():
    """Fase 2.0: carga de dumbbell fora do rol de Equipamentos (16kg inexistente).
    Não pode flaggar a carga de BARRA num composto ('Fat Bar 34kg + DB 15kg')."""
    cfg = {
        "equipamento": {"dumbbells": [10, 15, 22.5], "anilhas": [], "unidade": "kg"},
        "dias": [{"label": "Sáb", "categorias": [{"nome": "X", "baterias": [], "workouts": [
            {"nome": "W", "tipo": "for_time", "time_cap": "12 min", "movimentos": [
                {"nome": "DUMBBELL FRONT SQUAT (16KG) (2 ATHLETES)", "reps": 100},
                {"nome": "FAT BAR THRUSTER (34KG) + SINGLE-ARM DUMBBELL THRUSTER (15KG)", "reps": 20},
                {"chegada": True}]}]}]}]}
    avisos = validar_evento(cfg)
    db = [a for a in avisos if "Dumbbell de" in a["msg"]]
    assert len(db) == 1, f"esperava só o 16kg, got: {[a['msg'] for a in db]}"
    assert "16kg" in db[0]["msg"].replace(" ", "")
    # 34kg é barra (não dumbbell) e 15kg está no rol → nenhum outro aviso
    assert not any("34" in a["msg"].split("disponíveis")[0] for a in db)


def test_linter_movimento_typo_no_nome():
    """Fase 4: typo no NOME do movimento ('Thrustres') avisa; custom não."""
    cfg = {"dias": [{"label": "D", "categorias": [{"nome": "X", "baterias": [], "workouts": [
        {"nome": "W", "tipo": "for_time", "time_cap": "5 min", "movimentos": [
            {"nome": "THRUSTRES", "reps": 10},
            {"nome": "HAY BALE BURPEES", "reps": 20},
            {"chegada": True}]}]}]}]}
    avisos = validar_evento(cfg)
    typos = [a for a in avisos if "provável typo de" in a["msg"]]
    assert len(typos) == 1 and "Thruster" in typos[0]["msg"]


def test_validar_evento_sem_roster_nao_spamma_sem_alocacoes():
    """Fase 2.0: evento sem roster (fase de planejamento) não deve gerar um aviso
    'sem alocações' por bateria — seriam 100+. Só vale quando há roster parcial."""
    cfg = {"dias": [{"label": "Sáb", "categorias": [{"nome": "X", "workouts": [
        {"nome": "W", "tipo": "for_time", "time_cap": "5 min",
         "movimentos": [{"nome": "PULL-UPS", "reps": 10}, {"chegada": True}]}],
        "baterias": [
            {"numero": "1", "codigo_evento": '"W"', "workouts_que_rodam": [1], "alocacoes": []},
            {"numero": "2", "codigo_evento": '"W"', "workouts_que_rodam": [1], "alocacoes": []},
        ]}]}]}
    avisos = validar_evento(cfg)
    assert not any("sem alocações" in a["msg"] for a in avisos)


def test_validar_evento_composto_com_movimentos_nao_avisa_vazio():
    """Fase 2.0: composto guarda movimentos em f1/f2 — não pode ser flaggado como
    'sem movimentos' (Muscle Swim + 3k)."""
    cfg = {"dias": [{"label": "Sex", "categorias": [{"nome": "X", "baterias": [], "workouts": [
        {"nome": "MUSCLE SWIM + 3K", "tipo": "composto",
         "f1": {"nome": "MUSCLE SWIM", "movimentos": [{"nome": "50M SWIM", "reps": 50}]},
         "f2": {"nome": "3K", "movimentos": [{"nome": "3K RUN", "reps": 3000}]}}]}]}]}
    avisos = validar_evento(cfg)
    assert not any("sem movimentos" in a["msg"] for a in avisos)


def test_validar_evento_detecta_for_time_sem_time_cap():
    config = {
        "dias": [{
            "label": "Sábado",
            "categorias": [{
                "nome": "Rx",
                "workouts": [
                    {"nome": "SEM TC", "tipo": "for_time", "time_cap": "",
                     "movimentos": [{"nome": "PULL-UPS", "reps": 10}, {"chegada": True}]},
                ],
                "baterias": [],
            }],
        }],
    }
    avisos = validar_evento(config)
    assert any("SEM TC" in a["msg"] and "time cap" in a["msg"].lower()
               for a in avisos)
