"""Parsers heurísticos: texto livre de workout e Excel do organizador."""
import io
import openpyxl
from parsers import (
    parse_workout_text, parse_excel,
    _quebrar_categoria_composta, _bateria_casa_categoria,
    _propagar_codigos_da_montagem, _filtrar_alocacoes_por_faixa,
    _parse_inscritos, _parse_inscritos_full,
    _bateria_tem_atleta_na_faixa, _alocacoes_tem_atleta_na_faixa,
    _normalizar_categoria, _normalizar_categoria_relaxada,
    _chave_categoria_fuzzy, _extrair_carga,
    _workout_numero_de_codigo,
    _roster_de_abas_atletas,
    _workouts_que_rodam_da_bateria,
)


def test_extrair_carga_dupla_unidade_por_numero():
    """Carga misto/scaled com unidade em CADA número ('70kg/50kg') — antes o
    regex pegava só o 2º e deixava '(70kg/' no nome. Não pode quebrar '50/35 lb'
    (unidade só no fim) nem carga simples."""
    assert _extrair_carga("Sync. Deadlifts (70kg/50kg)") == ("Sync. Deadlifts", "70/50 KG")
    assert _extrair_carga("Overhead Squats (22,5kg/15kg)") == ("Overhead Squats", "22,5/15 KG")
    assert _extrair_carga("Wall-Ball Shots (20lbs/14lbs)") == ("Wall-Ball Shots", "20/14 LBS")
    # sem regressão:
    assert _extrair_carga("Deadlifts (84kg)") == ("Deadlifts", "84 KG")
    nome, carga = _extrair_carga("50/35 lb Thrusters")   # unidade só no fim, no início
    assert carga == "50/35 LB" and nome == "Thrusters"


def test_chegada_dirigida_pelo_excel():
    """A linha CHEGADA (rep final) segue o texto do Excel: se disser que a chegada
    não conta como repetição, não aparece; senão (default), aparece."""
    base = '"T"\nFor time:\n10 Pull-Ups\n20 Thrusters\nTime cap: 5 min\n――― NOTAS ―――\nPontuação\n'
    w_sem = parse_workout_text(base + '- A chegada não conta como repetição.', 1)
    assert not any(m.get("chegada") for m in w_sem["movimentos"])
    w_com = parse_workout_text(base + '- Será o tempo de conclusão do workout.', 1)
    assert any(m.get("chegada") for m in w_com["movimentos"])


def test_clausula_interrupcao_vira_descricao():
    """Cláusula 'a cada N min o workout é interrompido para ...' (ex: Full Penance
    do PWRD) não é movimento nem regulamento — deve sobreviver na descrição, só
    ela, sem repetir os movimentos. EMOM comum ('every 2 min: 5 pull-ups') não."""
    txt = ('"Full Penance"\nFor time:\n100 Wall-Ball Shots (20lbs)\n'
           '20 Fat Bar Thruster (60kg)\n'
           'A cada 2 minutos, o workout será interrompido para a execução de '
           '3 complex (5 Toes-to-Ring + 1 Ring Muscle-Up).\nTime cap: 12 minutes')
    w = parse_workout_text(txt, 1)
    assert w["descricao"] == [
        'A cada 2 minutos, o workout será interrompido para a execução de '
        '3 complex (5 Toes-to-Ring + 1 Ring Muscle-Up).'
    ]
    # Sem cláusula de interrupção → descrição segue vazia (não polui a súmula).
    w2 = parse_workout_text('"T"\nFor time:\n50 Wall-Ball Shots\nTime cap: 5 min', 1)
    assert w2["descricao"] == []


def test_distancia_k_km_vira_metros():
    """'3k'/'2k'/'1k'/'3km' = quilômetros → metros (3k=3000m) pra reps/acumulado.
    'Nm' e 'Nkg' não podem ser afetados."""
    def mov(txt):
        w = parse_workout_text('"T"\nFor time:\n' + txt + '\n10 Burpees\nTime cap: 5 min', 1)
        m = w["movimentos"][0]
        return m.get("nome"), m.get("reps"), m.get("carga")
    assert mov("3k Treadmill Run") == ("3000M TREADMILL RUN", 3000, None)
    assert mov("2k Row")[1] == 2000
    assert mov("1k Ski Erg")[1] == 1000
    assert mov("3km Run") == ("3000M RUN", 3000, None)
    assert mov("500m Row") == ("500M ROW", 500, None)          # metros intactos
    assert mov("84kg Deadlift")[2] == "84 KG"                  # kg continua carga


def test_rounds_fixos_rft_e_linha_solta():
    """rounds_fixos deve pegar 'N RFT' e 'N Rounds:' numa linha só (sem precisar
    de 'for time' ao lado). Não pode disparar em for-time simples nem no buy-in
    'then, N rounds of'."""
    def rf(txt):
        return parse_workout_text('"T"\n\n' + txt + '\nTime cap: 10 min', 1).get("rounds_fixos")
    assert rf("5 RFT:\n500m Row\n20 Taps") == 5
    assert rf("5 Rounds:\n500m Row\n20 Taps") == 5
    assert rf("5 Rounds of:\n500m Row") == 5
    assert rf("For time:\n500m Row\n20 Taps") is None            # sem round
    w = parse_workout_text('"T"\nFor time:\n1000m Ski\nthen, 2 rounds of:\n30 HSPU\nTime cap: 5 min', 1)
    assert w.get("rounds_fixos") is None                          # buy-in + bloco, não multiplica


def test_assign_workout_numbers_global_continuo_entre_dias():
    """Numeração deve ser CONTÍNUA por categoria pelo total de workouts (não
    reinicia por dia). Composto/Express ocupam 2 slots."""
    from parsers import assign_workout_numbers_global
    dias = [
        {"categorias": [{"nome": "A", "workouts": [{"tipo": "composto"}, {"tipo": "for_time"}]}]},
        {"categorias": [{"nome": "A", "workouts": [{"tipo": "for_time"}, {"tipo": "for_time"}]}]},
    ]
    assign_workout_numbers_global(dias)
    d0 = dias[0]["categorias"][0]["workouts"]
    d1 = dias[1]["categorias"][0]["workouts"]
    assert d0[0]["numero"] == 1 and d0[0]["numero_f2"] == 2   # composto ocupa 1-2
    assert d0[1]["numero"] == 3
    assert d1[0]["numero"] == 4 and d1[1]["numero"] == 5      # continua no dia seguinte


def test_detectar_blocos_cronograma_multi_arena_bateria_correta():
    """Bug Pwrd: em cronograma multi-arena, o detector pegava a coluna Bateria do
    bloco ANTERIOR (perdia baterias do 2º/3º bloco, ex: Tap Control). Cada bloco
    deve achar Eventos à esquerda e Bateria à direita da SUA categoria."""
    from parsers import _detectar_blocos_cronograma
    # 2 arenas com gap (col 5) entre elas:
    hdr = ['eventos', 'categoria', 'bateria', 'aquecimento', 'fila', '',
           'eventos', 'categoria', 'bateria', 'aquecimento', 'fila']
    blocos = _detectar_blocos_cronograma(hdr)
    assert len(blocos) == 2
    assert blocos[0]['eventos'] == 0 and blocos[0]['categoria'] == 1 and blocos[0]['bateria'] == 2
    # bloco 1: Bateria é a col 8 (a DELE), não a 2 (do bloco 0)
    assert blocos[1]['eventos'] == 6 and blocos[1]['categoria'] == 7 and blocos[1]['bateria'] == 8


def test_checar_movimento_typo_pega_typo_e_ignora_custom():
    """Fase 4: reconhece movimento canônico (mesmo plural/hífen), flagga typo,
    e NÃO acusa movimento custom legítimo (senão viraria ruído)."""
    from movimentos import checar_movimento_typo
    # typo de canônico → flagga com sugestão
    assert checar_movimento_typo("THRUSTRES") == ("THRUSTRES", "Thruster")
    assert checar_movimento_typo("OVERHAED SQUAT")[1] == "Overhead Squat"
    # reconhecido (plural/hífen/modificador) → None
    assert checar_movimento_typo("SYNC. WALL-BALL SHOTS (2 ATHLETES)") is None
    assert checar_movimento_typo("HANG POWER SNATCHES") is None
    assert checar_movimento_typo("SYNC. TOES RAISES (2 ATHLETES)") is None
    # custom legítimo → None (não é typo)
    assert checar_movimento_typo("HAY BALE BURPEES") is None
    assert checar_movimento_typo("LINE-FACING BURPEES") is None
    assert checar_movimento_typo("FAT BAR THRUSTER") is None


def test_parse_findings_json_ia_tolerante():
    """O parser de findings da IA tolera cerca ```json, texto ao redor e lixo."""
    from ai_rounds import _parse_findings_json
    ok = _parse_findings_json('Segue: [{"severidade":"erro","msg":"X","onde":"Y"},{"msg":"sem sev"}]')
    assert len(ok) == 2 and ok[0]["severidade"] == "erro" and ok[1]["severidade"] == "aviso"
    assert all(a["fonte"] == "ia" for a in ok)
    assert _parse_findings_json("não achei problemas") == []
    assert _parse_findings_json('```json\n[{"msg":"z"}]\n```')[0]["msg"] == "z"


def test_chave_categoria_fuzzy_casa_ordem_genero_e_sinal():
    """Casa variações humanas da MESMA categoria (Pwrd by Coffee 2026):
    ordem das palavras, concordância de gênero e posição do '+'."""
    pares = [
        ("Master Masculino 40-44", "Master 40-44 Masculino"),   # ordem
        ("Master 45+ Masculino",   "Master Masculino 45+"),     # ordem + sinal
        ("Master Feminino 40+",    "Master 40+ Feminino"),      # ordem + sinal
        ("Dupla Rx Masculino",     "Dupla Rx Masculina"),       # gênero
        ("Dupla Rx Feminino",      "Dupla Rx Feminina"),        # gênero
        ("Dupla Rx Misto",         "Dupla Rx Mista"),           # gênero
        ("Trio Master Misto 110+", "Trio Master Misto +110"),   # sinal
    ]
    for a, b in pares:
        assert _chave_categoria_fuzzy(a) == _chave_categoria_fuzzy(b), f"{a!r} != {b!r}"


def test_chave_categoria_fuzzy_nao_cruza_generos_nem_categorias():
    """Gêneros e categorias distintas NÃO podem colapsar na mesma chave."""
    assert _chave_categoria_fuzzy("Elite Masculino") != _chave_categoria_fuzzy("Elite Feminino")
    assert _chave_categoria_fuzzy("Dupla Rx Masculino") != _chave_categoria_fuzzy("Trio Rx Masculino")
    assert _chave_categoria_fuzzy("Master 40-44 Masculino") != _chave_categoria_fuzzy("Master 45+ Masculino")
    assert _chave_categoria_fuzzy("Trio Scaled Masculino") != _chave_categoria_fuzzy("Trio Scaled Feminino")


def test_for_load_janelas_por_atleta_e_notas_u2015():
    """Muscle Coffee (Pwrd): janelas de tempo por atleta viram blocos A/B/C, e
    o marcador '――― NOTAS ―――' (U+2015) é cortado do complex."""
    texto = (
        '"Muscle Coffee"\n\nFor load:\n'
        '(00:00 - 03:00) Athlete A\n1 Snatch + 3 Overhead Squat\n'
        '(04:00 - 07:00) Athlete B\n1 Clean + 3 Shoulder-to-Overhead\n'
        '(08:00 - 11:00) Athlete C\n1 Clean + 3 Front Squat\n\n'
        'Time cap: 11 minutes\n\n'
        '――― NOTAS ―――\n\nPontuação\n- Soma das cargas máximas.'
    )
    wkt = parse_workout_text(texto, "MUSCLE COFFEE")
    assert wkt["tipo"] == "for_load"
    janelas = wkt["sequencia_movimentos"]["janelas"]
    assert [j["label"] for j in janelas] == ["A", "B", "C"]
    assert janelas[0]["atleta"] == "Athlete A"
    assert "SNATCH" in janelas[0]["complex"]
    # NOTAS não pode vazar pra nenhum bloco
    assert all("NOTAS" not in j["complex"] and "PONTUAÇÃO" not in j["complex"].upper()
               for j in janelas)


def test_composto_por_dois_titulos_muscle_swim_3k():
    """Muscle Swim + 3k (Pwrd): dois títulos entre aspas em linhas próprias,
    cada um com sua janela e 'For time:' → composto F1/F2 (2 pontuações)."""
    texto = (
        '"Muscle Swim" (00:00-08:00)\n\nFor time:\n50m Swim\n'
        "10 Devil's Presses (22,5kg)\n50m Swim\n\n"
        '"3k" (20:00-35:00)\n\nFor time:\n3k Treadmill Run\n\n'
        'Time cap: 35 minutes\n\n――― NOTAS ―――\nPontuação\n- Duas independentes.'
    )
    wkt = parse_workout_text(texto, 1)
    assert wkt["tipo"] == "composto"
    assert wkt["f1"]["nome"] == "MUSCLE SWIM"
    assert wkt["f2"]["nome"] == "3K"
    assert wkt["f1"]["janela"] == "00:00–08:00"
    assert wkt["f2"]["janela"] == "20:00–35:00"
    # NOTAS não vaza pra F2
    assert not any("NOTAS" in (m.get("nome") or "") for m in wkt["f2"]["movimentos"])


def test_workout_titulo_unico_nao_vira_composto():
    """Um único título entre aspas (nome do workout) NÃO pode virar composto."""
    wkt = parse_workout_text('"Stack Bad"\n\nFor time:\n30 Pull-Ups\nTime cap: 5 min', 1)
    assert wkt["tipo"] != "composto"


def test_composto_dupla_titulo_com_sufixo_atleta():
    """Dupla Muscle Swim & 2k: títulos com sufixo '— Atleta N' (Pwrd) também
    viram composto (auditoria: individual virava, dupla não)."""
    texto = (
        'Os dois eventos são disputados ao mesmo tempo (00:00-10:00).\n\n'
        '"Muscle Swim" — Atleta 1\nFor time:\n50m Swim\n10 Devil Presses\n\n'
        '"2k" — Atleta 2\nFor time:\n2k Treadmill Run\n\nTime cap: 10 minutes\n\n'
        '――― NOTAS ―――\n- Serão duas pontuações: "Muscle Swim" e "2k".'
    )
    wkt = parse_workout_text(texto, 1)
    assert wkt["tipo"] == "composto"
    assert wkt["f1"]["nome"] == "MUSCLE SWIM"
    assert wkt["f2"]["nome"] == "2K"


def test_for_load_multi_janela_3_tentativas_por_complex():
    """Muscle Coffee individual: 2 complexes, cada um com 3 TENTATIVAS (não 1
    linha por complex), somando os melhores. `tentativas` é attempts POR complex;
    o nº de complexes vem de `janelas`."""
    texto = (
        '"Muscle Coffee"\n\nFor load:\n'
        '(00:00 - 03:00)\n1 Snatch + 3 Overhead Squat\n'
        '(04:00 - 07:00)\n1 Clean + 3 Shoulder-to-Overhead\n\n'
        'Time cap: 7 minutes\n\n――― NOTAS ―――\n'
        '- Soma das cargas máximas dos 2 complex.\n- Cada complex validará 1 tentativa única.'
    )
    wkt = parse_workout_text(texto, 1)
    assert wkt["tipo"] == "for_load"
    assert wkt["tentativas"] == 3          # 3 tentativas POR complex
    assert wkt["soma_complexes"] is True
    assert len(wkt["sequencia_movimentos"]["janelas"]) == 2   # 2 complexes


def test_for_time_buyin_distancia_1000m_e_bloco_rounds():
    """Stack Bad (Pwrd): buy-in '1000m Ski Erg' não pode ser dropado (o cap
    ≥1000 'evita anos' matava distâncias), e 'then, 2 rounds of:' vira banner
    de seção — sem multiplicar reps (o buy-in não entra na conta)."""
    texto = (
        '"Stack Bad"\n\nFor time:\n1000m Ski Erg\nthen, 2 rounds of:\n'
        '30 Handstand Push-Ups\n400m Run\n30 Line-Facing Burpees\n\n'
        'Time cap: 16 minutes'
    )
    wkt = parse_workout_text(texto, "STACK BAD")
    movs = wkt["movimentos"]
    nomes = [m.get("nome") for m in movs if m.get("nome")]
    secoes = [m["secao"] for m in movs if m.get("secao")]
    assert any("1000M SKI ERG" in n for n in nomes), "buy-in 1000m foi dropado"
    assert any("2 ROUNDS OF" in s for s in secoes), "faltou banner do bloco de rounds"
    assert not wkt.get("rounds_fixos"), "'then, N rounds of' não pode multiplicar tudo"
    # rounds_bloco = N: o render usa isso pra rodar o buy-in 1x + bloco N rounds.
    assert wkt.get("rounds_bloco") == 2


def test_distancia_1000m_nao_confundida_com_ano():
    """'1000m Row' é distância (mantém); '2026 Games Open' é ano (dropa)."""
    w1 = parse_workout_text("For time:\n1000m Row\n10 Burpees\nTime cap: 5 min", "T")
    assert any("1000M ROW" in (m.get("nome") or "") for m in w1["movimentos"])
    w2 = parse_workout_text("For time:\n2026 Games Open Standard\n10 Burpees\nTime cap: 5 min", "T")
    assert not any("2026" in (m.get("nome") or "") for m in w2["movimentos"])


def test_parse_equipamento_formato_categoria_equipamento_kg():
    """Aba 'Equipamentos' no formato Categoria|Equipamento|Qtd (peso no nome do
    equipamento) → extrai anilhas e unidade kg; med ball em lb não contamina."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Equipamentos")
    ws.append(["Categoria", "Equipamento", "Qtd.", "Observações"])
    ws.append(["Barras", "Barra EVOBLACK 20 kg", 30, ""])
    ws.append(["Anilhas", "Anilha Color 25 kg", 60, ""])
    ws.append(["Anilhas", "Anilha Color 5 kg", 120, ""])
    ws.append(["Anilhas", "Anilha 1 kg", 60, ""])
    ws.append(["Implementos", "Medicine Ball 20 lb", 30, ""])  # não pode virar lb
    from parsers import _parse_equipamento
    equip = _parse_equipamento(wb)
    assert equip is not None
    assert equip["unidade"] == "kg"
    assert 25.0 in equip["anilhas"] and 1.0 in equip["anilhas"]
    assert 20.0 not in equip["anilhas"]  # 20 é da BARRA, não anilha


def test_parse_workout_text_for_time_extrai_movimentos_e_time_cap():
    texto = (
        '"TWENTIES"\n'
        "For Time:\n"
        "20 Chest-to-Bar Pull-Ups\n"
        "20 Devil's Presses\n"
        "Time cap: 9 min"
    )
    wkt = parse_workout_text(texto, numero=1)
    assert wkt["nome"] == "TWENTIES"
    assert wkt["tipo"] == "for_time"
    assert wkt["time_cap"] == "9 min"
    nomes = [m.get("nome") for m in wkt["movimentos"] if m.get("nome")]
    assert "CHEST-TO-BAR PULL-UPS" in nomes
    # For Time fecha com chegada
    assert any(m.get("chegada") for m in wkt["movimentos"])


def test_parse_excel_modelo_retorna_estrutura_valida(modelo_xlsx_bytes):
    result = parse_excel(modelo_xlsx_bytes)
    assert isinstance(result, dict)
    # Parser unificado sempre retorna shape evento_multidia (formato antigo é adaptado)
    assert result["tipo"] == "evento_multidia"
    assert "dias" in result
    assert isinstance(result["dias"], list)
    assert len(result["dias"]) >= 1
    # Pelo menos 1 categoria com pelo menos 1 workout
    primeiro_dia = result["dias"][0]
    assert "categorias" in primeiro_dia
    assert len(primeiro_dia["categorias"]) >= 1
    primeiro_workout = primeiro_dia["categorias"][0]["workouts"][0]
    assert "nome" in primeiro_workout
    assert "tipo" in primeiro_workout
    assert primeiro_workout["tipo"] in {"for_time", "amrap", "express"}


# ── Layout grades-por-modalidade + dias por aba ───────────────────────────────
def test_quebrar_categoria_composta_separa_e_normaliza():
    partes = _quebrar_categoria_composta(
        "Iniciante Feminino (Heat 3) & Iniciante Masculino (Heat 1)"
    )
    assert partes == ["iniciante feminino", "iniciante masculino"]


def test_quebrar_categoria_composta_aceita_virgula_como_separador():
    # Storm 2026: três cats compartilhando bateria usando `,` + `&`
    partes = _quebrar_categoria_composta(
        "Dupla Iniciante Masculino (Single Heat),  Dupla Iniciante Feminina (Single Heat) & Dupla Iniciante Mista (Single Heat)"
    )
    assert partes == [
        "dupla iniciante masculino",
        "dupla iniciante feminina",
        "dupla iniciante mista",
    ]


def test_quebrar_categoria_composta_protege_virgula_dentro_de_parens():
    # Vírgula DENTRO de parens (descritor) não é separador
    partes = _quebrar_categoria_composta(
        "Iniciante (8, 9 anos) Masculino & Iniciante (10, 11 anos) Feminino"
    )
    assert partes == [
        "iniciante (8, 9 anos) masculino",
        "iniciante (10, 11 anos) feminino",
    ]


def test_bateria_casa_categoria_evita_falso_positivo_dupla_vs_individual():
    # 'rx masculino' não deve casar com 'dupla rx masculino' (categorias distintas)
    assert _bateria_casa_categoria("Rx Masculino (Single Heat)", "rx masculino") is True
    assert _bateria_casa_categoria("Dupla Rx Masculino (Heat 1)", "rx masculino") is False
    assert _bateria_casa_categoria("Dupla Rx Masculino (Heat 1)", "dupla rx masculino") is True


def test_normalizar_categoria_remove_so_sufixos_de_heat():
    # Sufixos de bateria são removidos
    assert _normalizar_categoria("Iniciante Feminino (Heat 1)") == "iniciante feminino"
    assert _normalizar_categoria("Rx Masculino (Single Heat)") == "rx masculino"
    assert _normalizar_categoria("Amador Feminino (Final Heat)") == "amador feminino"
    # Descritor livre é preservado (parêntese não-heat)
    assert _normalizar_categoria("Master 35-39 Masculino (identico ao amador)") == "master 35-39 masculino (identico ao amador)"
    assert _normalizar_categoria("Rx Misto (Iniciante)") == "rx misto (iniciante)"


def test_normalizar_categoria_relaxada_remove_tudo():
    assert _normalizar_categoria_relaxada("Master 35-39 Masculino (identico ao amador)") == "master 35-39 masculino"
    assert _normalizar_categoria_relaxada("Rx Misto (Iniciante)") == "rx misto"
    assert _normalizar_categoria_relaxada("Iniciante Feminino (Heat 1)") == "iniciante feminino"


def test_bateria_casa_categoria_com_fallback_relaxado():
    # Grade tem descritor extra, cronograma só tem sufixo de heat — match
    # relaxado precisa funcionar quando habilitado
    bat_cat = "Master 35-39 Masculino (Single Heat)"
    grade_norm   = "master 35-39 masculino (identico ao amador)"
    grade_relax  = "master 35-39 masculino"
    # Sem fallback: não casa (estrita falha)
    assert _bateria_casa_categoria(bat_cat, grade_norm) is False
    # Com fallback: casa
    assert _bateria_casa_categoria(bat_cat, grade_norm, grade_relax, permite_relaxado=True) is True


def test_bateria_casa_categoria_fallback_desligado_evita_colisao():
    # Cenário hipotético: 'Rx Misto (Iniciante)' e 'Rx Misto (Avançado)' no
    # mesmo evento. Relaxada das duas é 'rx misto' — fallback NÃO pode bater
    # uma com a outra. Caller passa permite_relaxado=False nesse caso.
    bat_cat = "Rx Misto (Avançado) (Heat 1)"
    grade_norm  = "rx misto (iniciante)"
    grade_relax = "rx misto"
    # Estrita falha (descritores diferentes) e fallback desligado → não casa
    assert _bateria_casa_categoria(bat_cat, grade_norm, grade_relax, permite_relaxado=False) is False


def test_propagar_codigos_da_montagem_preenche_cronograma_vazio():
    # Cronograma sem códigos + montagem com códigos → cronograma fica preenchido
    cronograma = [
        {"numero": "1", "codigo_evento": "", "categoria": "X"},
        {"numero": "2", "codigo_evento": "", "categoria": "Y"},
    ]
    montagem = {
        ("#1", "X", "1"): [{"raia": "1", "nome": "Foo"}],
        ("#2 & #3", "Y", "2"): [{"raia": "1", "nome": "Bar"}],
    }
    _propagar_codigos_da_montagem(cronograma, montagem)
    assert cronograma[0]["codigo_evento"] == "#1"
    assert cronograma[1]["codigo_evento"] == "#2 & #3"


def test_propagar_codigos_nao_sobrescreve_quando_cronograma_ja_tem():
    # Se o cronograma já tem ao menos 1 código, nada é alterado
    cronograma = [
        {"numero": "1", "codigo_evento": "#1", "categoria": "X"},
        {"numero": "2", "codigo_evento": "",   "categoria": "Y"},
    ]
    montagem = {("#9", "Y", "2"): [{"raia": "1", "nome": "Bar"}]}
    _propagar_codigos_da_montagem(cronograma, montagem)
    # Bateria 2 fica sem código (propagador só ativa quando cronograma inteiro está vazio)
    assert cronograma[1]["codigo_evento"] == ""


def test_filtrar_alocacoes_por_faixa_mantem_so_numeros_da_categoria():
    # Bateria mista: 3 atletas Scaled Feminino (902-904) + 2 Scaled Masculino (1041-1042)
    alocs = [
        {"raia": "1", "numero": "902",  "nome": "Brianna"},
        {"raia": "2", "numero": "903",  "nome": "Monica"},
        {"raia": "3", "numero": "904",  "nome": "Karla"},
        {"raia": "4", "numero": "1041", "nome": "Dhener"},
        {"raia": "5", "numero": "1042", "nome": "Hiago"},
    ]
    # Filtrando pra Scaled Feminino (901-999): mantém as 3 primeiras, descarta as 2 últimas
    fem_keep, fem_drop = _filtrar_alocacoes_por_faixa(alocs, (901, 999))
    assert [a["nome"] for a in fem_keep] == ["Brianna", "Monica", "Karla"]
    assert [a["nome"] for a in fem_drop] == ["Dhener", "Hiago"]
    # Filtrando pra Scaled Masculino (1001-1099): inverso
    masc_keep, masc_drop = _filtrar_alocacoes_por_faixa(alocs, (1001, 1099))
    assert [a["nome"] for a in masc_keep] == ["Dhener", "Hiago"]
    assert [a["nome"] for a in masc_drop] == ["Brianna", "Monica", "Karla"]


def _wb_com_inscritos(linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscritos"
    for linha in linhas:
        ws.append(linha)
    return wb


def test_parse_inscritos_le_faixas_de_numero():
    wb = _wb_com_inscritos([
        ["Categorias cadastradas"],
        ["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"],
        ["RX Feminino",  10, 8,  501, 599, "Sim"],
        ["RX Masculino", 10, 10, 601, 699, "Sim"],
    ])
    faixas = _parse_inscritos(wb)
    assert faixas.get("rx feminino")  == (501, 599)
    assert faixas.get("rx masculino") == (601, 699)


def test_parse_inscritos_suporta_multiplos_blocos():
    # Individuais + Duplas no mesmo arquivo, separados por linha vazia
    wb = _wb_com_inscritos([
        ["Categorias cadastradas"],
        ["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"],
        ["RX Masculino", 10, 10, 601, 699, "Sim"],
        [],
        ["Categorias cadastradas"],
        ["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"],
        ["Dupla RX Misto", 20, 15, 101, 199, "Não"],
    ])
    faixas = _parse_inscritos(wb)
    assert faixas.get("rx masculino")    == (601, 699)
    assert faixas.get("dupla rx misto") == (101, 199)


def test_parse_inscritos_retorna_vazio_sem_aba():
    wb = openpyxl.Workbook()
    wb.active.title = "OutraAba"
    assert _parse_inscritos(wb) == {}


def test_parse_inscritos_full_le_coluna_individual():
    # Storm 2026: coluna `Individual` desambigua Individual vs Dupla
    # quando faixas de número colidem.
    wb = _wb_com_inscritos([
        ["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"],
        ["RX Masculino",       10, 8,  101, 199, "Sim"],
        ["Dupla RX Masculino", 10, 5,  101, 199, "Não"],
    ])
    full = _parse_inscritos_full(wb)
    assert full["rx masculino"]        == (101, 199, True)
    assert full["dupla rx masculino"]  == (101, 199, False)


def test_parse_inscritos_full_sem_coluna_individual_retorna_none():
    wb = _wb_com_inscritos([
        ["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final"],
        ["RX Masculino", 10, 8, 101, 199],
    ])
    full = _parse_inscritos_full(wb)
    assert full["rx masculino"] == (101, 199, None)


def test_alocacoes_tem_atleta_na_faixa_detecta_minimo_um():
    alocs = [{"numero": "104"}, {"numero": "1601"}, {"numero": "abc"}]
    assert _alocacoes_tem_atleta_na_faixa(alocs, (101, 199)) is True
    assert _alocacoes_tem_atleta_na_faixa(alocs, (1601, 1699)) is True
    assert _alocacoes_tem_atleta_na_faixa(alocs, (5000, 5099)) is False
    # Sem faixa → False
    assert _alocacoes_tem_atleta_na_faixa(alocs, None) is False


def test_bateria_tem_atleta_na_faixa_filtra_por_numero_da_bateria():
    montagem = {
        ("#1", "Cat A", "1"): [{"numero": "104"}, {"numero": "105"}],
        ("#1", "Cat B", "2"): [{"numero": "1601"}],
    }
    # Bateria 1: atletas 104,105 → caem em 101-199
    assert _bateria_tem_atleta_na_faixa("1", montagem, (101, 199)) is True
    # Bateria 2: só 1601 → NÃO cai em 101-199
    assert _bateria_tem_atleta_na_faixa("2", montagem, (101, 199)) is False
    # Bateria 2 cai em 1601-1699
    assert _bateria_tem_atleta_na_faixa("2", montagem, (1601, 1699)) is True


def test_parse_workout_text_composto_detecta_dois_subworkouts():
    """Storm 2026: workout composto `"X" + "Y"` no header, com 2 fórmulas
    encadeadas. Parser quebra em sub-workouts (f1, f2), cada um com seu
    próprio tipo, movimentos e time_cap. F1 do Storm é For Time Goal,
    F2 é For Time normal."""
    texto = (
        '"Barbells and Jump" + "Run In The Park"\n\n'
        '"Barbells and Jump" (0:00-5:00)\n\n'
        'For time:\n'
        '15 Deadlifts (115lb)\n'
        '10 Snatches (115lb)\n'
        'Max Box Jump Over (60cm)\n\n'
        'Goal: 45 Box Jump Over + finishing rep (cross the line).\n\n'
        'Descanse um minuto, depois...\n\n'
        '"Run In The Park" (6:00-9:00)\n'
        'For time:\n'
        '10 Snatches (135lb)\n'
        '200m Run\n'
        '10 Snatches (135lb)\n\n'
        'Time cap: 9 minutes\n'
    )
    wkt = parse_workout_text(texto, numero=2)
    assert wkt['tipo'] == 'composto'
    assert wkt['nome'] == 'BARBELLS AND JUMP + RUN IN THE PARK'
    assert wkt['time_cap'] == '9 minutes'
    assert 'minuto' in wkt['descanso'].lower()

    f1 = wkt['f1']
    assert f1['nome'] == 'BARBELLS AND JUMP'
    assert f1['tipo'] == 'for_time_goal'
    assert f1['goal_reps'] == 45
    assert 'BOX JUMP OVER' in f1['goal_movimento']
    assert f1['janela'] == '0:00–5:00'

    f2 = wkt['f2']
    assert f2['nome'] == 'RUN IN THE PARK'
    assert f2['tipo'] == 'for_time'
    assert f2['janela'] == '6:00–9:00'
    # F2 tem 200m Run + Snatches (sem goal)
    nomes_f2 = [m.get('nome') for m in f2['movimentos'] if m.get('nome')]
    assert '200M RUN' in nomes_f2
    assert any('SNATCH' in n for n in nomes_f2)


def test_parse_workout_text_atleta_n_vira_label_em_dupla():
    """Storm 2026 Dupla SETE MINUTOS: linhas `Atleta 1` / `Atleta 2` marcam
    quem faz cada movimento. Devem virar `label='ATLETA N'` no mov pra
    súmula mostrar de quem é a responsabilidade. Movs antes do primeiro
    `Atleta N` (ex: Cal Air Bike dividido pela dupla) ficam sem label.
    """
    texto = (
        '"Sete Minutos" (Final)\n\n'
        'For time:\n'
        '40 Cal Air Bike\n'
        'Depois, cada atleta:\n'
        'Atleta 1\n'
        '16 Bar Muscle-Ups\n'
        '24m Dumbbell Overhead Walking Lunge (22,5kg)\n'
        'Atleta 2\n'
        '16 Bar Muscle-Ups\n'
        '24m Dumbbell Overhead Walking Lunge (22,5kg)\n\n'
        'Time cap: 7 minutes\n'
    )
    wkt = parse_workout_text(texto, numero=3)
    assert wkt['tipo'] == 'for_time'
    movs = [m for m in wkt['movimentos'] if m.get('nome')]
    assert movs[0]['nome'] == 'CAL AIR BIKE'
    assert 'label' not in movs[0]  # antes dos atletas — sem label
    assert movs[1]['label'] == 'ATLETA 1'
    assert movs[2]['label'] == 'ATLETA 1'
    assert movs[3]['label'] == 'ATLETA 2'
    assert movs[4]['label'] == 'ATLETA 2'


def test_parse_workout_text_atleta_n_com_then_reseta_label():
    """Quando `Atleta N` é seguido de `then...`, o label reseta — movs
    pós-then sem outro Atleta antes ficam sem label (não herdam o anterior).
    Padrão Storm BARBELLS Dupla: Atleta 1 + then + Atleta 2 + then + Max
    Box Jump (dupla junta) → Max Box Jump sem label."""
    texto = (
        '"BARBELLS"\n\n'
        'For time:\n'
        'Atleta 1\n'
        '15 Deadlifts (115lb)\n'
        'then...\n'
        'Atleta 2\n'
        '15 Deadlifts (115lb)\n'
        'then...\n'
        '60 Box Jump Over (60cm)\n\n'
        'Time cap: 5 minutes\n'
    )
    wkt = parse_workout_text(texto, numero=1)
    movs = [m for m in wkt['movimentos'] if m.get('nome')]
    # 2 Deadlifts (cada um de um atleta) + 1 Box Jump (dupla)
    assert movs[0]['label'] == 'ATLETA 1'
    assert movs[1]['label'] == 'ATLETA 2'
    assert 'label' not in movs[2]  # após both — sem atribuição


def test_parse_workout_text_then_sozinho_ainda_usa_block_labels():
    """Compat: workouts antigos que usam só `then...` (sem Atleta N) seguem
    com label `1º BLOCO` / `2º BLOCO` etc. — não foi quebrado pela mudança."""
    texto = (
        '"OLD STYLE"\n\n'
        'For time:\n'
        '15 Deadlifts\n'
        'then...\n'
        '15 Snatches\n\n'
        'Time cap: 5 minutes\n'
    )
    wkt = parse_workout_text(texto, numero=1)
    movs = [m for m in wkt['movimentos'] if m.get('nome')]
    assert movs[0]['label'] == '1º BLOCO'
    assert movs[1]['label'] == '2º BLOCO'


def test_workouts_que_rodam_match_composto_pelo_nome_split_por_amp():
    """Storm Domingo: cronograma diz `"Barbells and Jump & Run in the Park"`
    mas o composto se chama `BARBELLS AND JUMP + RUN IN THE PARK` (com `+`).
    Split por `&` gera as partes `BARBELLS AND JUMP` e `RUN IN THE PARK`,
    cada uma deve bater com F1.nome ou F2.nome do composto. Antes do fix
    `workouts_que_rodam` ficava `[]` e o gerador rodava TODOS os workouts
    naquela bateria (bug grave: VINTE SEIS em bateria do composto)."""
    workouts = [
        {'nome': 'VINTE SEIS', 'tipo': 'for_time'},
        {
            'nome': 'BARBELLS AND JUMP + RUN IN THE PARK', 'tipo': 'composto',
            'f1': {'nome': 'BARBELLS AND JUMP'},
            'f2': {'nome': 'RUN IN THE PARK'},
        },
        {'nome': 'SETE MINUTOS', 'tipo': 'for_time'},
    ]
    # Bateria com codigo `"Barbells and Jump & Run in the Park"` → posição 2
    assert _workouts_que_rodam_da_bateria(
        '"Barbells and Jump & Run in the Park"', workouts) == [2]
    # E sanity: nome exato continua funcionando
    assert _workouts_que_rodam_da_bateria('"VINTE SEIS"', workouts) == [1]


def test_workouts_que_rodam_match_exato_de_composto_com_plus_no_codigo():
    """Quando o cronograma escreve o nome do composto com `+` (formato do
    parser), match exato deve funcionar SEM precisar do split por `&`."""
    workouts = [
        {
            'nome': 'BARBELLS AND JUMP + RUN IN THE PARK', 'tipo': 'composto',
            'f1': {'nome': 'BARBELLS AND JUMP'},
            'f2': {'nome': 'RUN IN THE PARK'},
        },
    ]
    assert _workouts_que_rodam_da_bateria(
        '"Barbells and Jump + Run in the Park"', workouts) == [1]


def test_parse_workout_text_simples_nao_eh_composto():
    """Garantia: workout simples (1 nome só) NÃO vira composto por engano."""
    texto = (
        '"Vinte Seis"\n\n'
        'For time:\n'
        '26 Burpees Over-the-Bar\n'
        '26 Hang Cleans (115lb)\n\n'
        'Time cap: 15 minutes\n'
    )
    wkt = parse_workout_text(texto, numero=1)
    assert wkt['tipo'] == 'for_time'
    assert wkt['nome'] == 'VINTE SEIS'
    assert 'f1' not in wkt
    assert 'f2' not in wkt


def test_roster_de_abas_atletas_dedup_linhas_duplicadas():
    """Excel com copia/cola gera linhas duplicadas. Dedup por (numero, nome)
    pra evitar atleta repetido na súmula combinada de pré-evento."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Atletas - Sábado")
    ws.append([101, "Matheus", "Box A"])
    ws.append([102, "Victor",  "Box B"])
    ws.append([101, "Matheus", "Box A"])  # duplicada exata
    ws.append([101, "MATHEUS", "Box A"])  # duplicada case-insensitive
    ws.append([103, "Outro",   "Box C"])
    roster = _roster_de_abas_atletas(wb)
    assert [a["numero"] for a in roster] == ["101", "102", "103"]
    assert [a["nome"] for a in roster] == ["Matheus", "Victor", "Outro"]


def test_roster_de_abas_atletas_preserva_colisao_individual_dupla():
    """Storm reusa #101 entre Individual (`MATHEUS`) e Dupla (`GOKU E KURIRIN`).
    Ambos devem permanecer no roster — colisão legítima, não duplicação."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Atletas - Sábado")
    ws.append([101, "MATHEUS POLACCHINI VIEIRA", "STORM TANK"])
    ws = wb.create_sheet("Atletas - Domingo")
    ws.append([101, "GOKU E KURIRIN", "DUST-3"])
    roster = _roster_de_abas_atletas(wb)
    assert len(roster) == 2
    nomes = {a["nome"] for a in roster}
    assert nomes == {"MATHEUS POLACCHINI VIEIRA", "GOKU E KURIRIN"}


def test_parse_excel_storm_separa_modalidades_por_dia():
    """Regressão Storm Challenge 2026: faixas colidem entre Individuais
    (Sábado) e Duplas (Domingo). Sem desambiguação por modalidade, cats da
    Dupla vazam pro Sábado e vice-versa. Reproduz o cenário mínimo:
    - Sábado: bateria mista Teen (1301-1399) com adultos
    - Domingo: bateria de Dupla Iniciante Mista (1301-1399)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Inscritos com coluna Individual
    ws = wb.create_sheet("Inscritos")
    ws.append(["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"])
    ws.append(["Teen Intermediario 16-17 Masculino", 5, 4, 1301, 1399, "Sim"])
    ws.append(["Intermediario Masculino",            40, 30, 301, 399, "Sim"])
    ws.append(["Dupla Iniciante Mista",              10, 2, 1301, 1399, "Não"])
    ws.append(["Dupla Rx Misto",                     10, 2, 401, 499,   "Não"])
    # Workouts - Individuais (grade)
    ws = wb.create_sheet("Workouts - Individuais")
    ws.append(["Teen Intermediario 16-17 Masculino", "Intermediario Masculino"])
    ws.append([
        '"Vinte Seis"\n\nFor time:\n10 Burpees\n\nTime cap: 10 minutes',
        '"Vinte Seis"\n\nFor time:\n10 Burpees\n\nTime cap: 10 minutes',
    ])
    # Workouts - Duplas (precisa de 2+ cats pro detector de grade pegar)
    ws = wb.create_sheet("Workouts - Duplas")
    ws.append(["Dupla Iniciante Mista", "Dupla Rx Misto"])
    ws.append([
        '"Vinte Seis"\n\nFor time:\n10 Burpees\n\nTime cap: 10 minutes',
        '"Vinte Seis"\n\nFor time:\n10 Burpees\n\nTime cap: 10 minutes',
    ])
    # Sábado (cronograma): bateria 1 mista Teen+Intermediario
    ws = wb.create_sheet("Sábado")
    ws.append(["Storm Challenge"])
    ws.append(["Sábado"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade",
               "Aquecimento", "Duração Aquec.", "Fila", "Duração Fila",
               "Horário", "Cap", "Transição"])
    ws.append(['"Vinte Seis"',
               "16-17 Masculino (Single Heat) & Intermediario Masculino (Heat 1)",
               1, "", "4 (4)", "08:00", "00:30", "08:30", "00:15",
               "08:45", "00:15", "00:10"])
    # Sábado - Montagem: 2 atletas Teen + 2 Intermediario
    ws = wb.create_sheet("Sábado - Montagem")
    ws.append(["08:45", '"Vinte Seis"'])
    ws.append([1, "16-17 Masculino (Single Heat) & Intermediario Masculino (Heat 1)"])
    ws.append(["Raia", "Número", "Nome", "Box"])
    ws.append([1, 1301, "Teen Atleta 1", "Box X"])
    ws.append([2, 1302, "Teen Atleta 2", "Box Y"])
    ws.append([3, 321,  "Inter Atleta 1", "Box Z"])
    ws.append([4, 322,  "Inter Atleta 2", "Box W"])
    # Atletas - Sábado (roster)
    ws = wb.create_sheet("Atletas - Sábado")
    ws.append([1301, "Teen Atleta 1", "Box X"])
    ws.append([1302, "Teen Atleta 2", "Box Y"])
    ws.append([321,  "Inter Atleta 1", "Box Z"])
    ws.append([322,  "Inter Atleta 2", "Box W"])
    # Domingo (cronograma): bateria 1 só Dupla
    ws = wb.create_sheet("Domingo")
    ws.append(["Storm Challenge"])
    ws.append(["Domingo"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade",
               "Aquecimento", "Duração Aquec.", "Fila", "Duração Fila",
               "Horário", "Cap", "Transição"])
    ws.append(['"Vinte Seis"', "Dupla Iniciante Mista (Single Heat)",
               1, "", "2 (2)", "09:00", "00:30", "09:30", "00:15",
               "09:45", "00:15", "00:10"])
    # Domingo - Montagem
    ws = wb.create_sheet("Domingo - Montagem")
    ws.append(["09:45", '"Vinte Seis"'])
    ws.append([1, "Dupla Iniciante Mista (Single Heat)"])
    ws.append(["Raia", "Número", "Nome", "Box"])
    ws.append([1, 1301, "Dupla A", "Box DA"])
    ws.append([2, 1302, "Dupla B", "Box DB"])
    # Atletas - Domingo
    ws = wb.create_sheet("Atletas - Domingo")
    ws.append([1301, "Dupla A", "Box DA"])
    ws.append([1302, "Dupla B", "Box DB"])

    buf = io.BytesIO()
    wb.save(buf)
    r = parse_excel(buf.getvalue())
    assert r["tipo"] == "evento_multidia"

    sabado = next(d for d in r["dias"] if d["label"] == "Sábado")
    domingo = next(d for d in r["dias"] if d["label"] == "Domingo")

    # Sábado: Teen + Intermediario, NÃO pode ter Dupla Iniciante Mista
    sabado_cats = {c["nome"] for c in sabado["categorias"]}
    assert "Teen Intermediario 16-17 Masculino" in sabado_cats
    assert "Intermediario Masculino" in sabado_cats
    assert "Dupla Iniciante Mista" not in sabado_cats

    # Domingo: só Dupla, NÃO pode ter Teen nem Intermediario
    domingo_cats = {c["nome"] for c in domingo["categorias"]}
    assert "Dupla Iniciante Mista" in domingo_cats
    assert "Teen Intermediario 16-17 Masculino" not in domingo_cats
    assert "Intermediario Masculino" not in domingo_cats

    # Teen no Sábado tem os 2 atletas certos (1301, 1302), filtrados da
    # bateria mista
    teen = next(c for c in sabado["categorias"]
                if c["nome"] == "Teen Intermediario 16-17 Masculino")
    nums_teen = [a["numero"] for b in teen["baterias"] for a in b["alocacoes"]]
    assert sorted(nums_teen) == ["1301", "1302"]

    # Roster: todos os atletas têm categoria atribuída
    sem_cat = [a for a in r["roster"] if not (a.get("categoria") or "").strip()]
    assert sem_cat == []


def test_workout_numero_de_codigo_exige_prefixo_explicito():
    # Aceita formatos com prefixo
    assert _workout_numero_de_codigo("#1") == 1
    assert _workout_numero_de_codigo("#02") == 2
    assert _workout_numero_de_codigo("WKT 4") == 4
    assert _workout_numero_de_codigo("Workout 04") == 4
    assert _workout_numero_de_codigo("  #5  ") == 5
    # Rejeita texto qualquer com dígito (evita falso positivo)
    assert _workout_numero_de_codigo("Bat 12") is None
    assert _workout_numero_de_codigo("foo") is None
    assert _workout_numero_de_codigo("") is None
    assert _workout_numero_de_codigo("123") is None


def test_validate_for_load_aceita_workout_bem_formado():
    """Workout For Load válido passa sem exceção."""
    import pytest
    from sumula_app import _validate_workout_tipos, BadRequest
    wkts = [{
        "tipo": "for_load", "nome": "MAX CLEAN", "tentativas": 3,
        "anilhas": [25, 20, 15, 10, 5, 2.5, 1.25],
        "barra_masculina": 20, "barra_feminina": 15, "unidade": "kg",
    }]
    _validate_workout_tipos(wkts)   # não deve levantar


def test_truncar_descricao_em_notas():
    """Descrição corta no primeiro separador NOTAS/Observações/Pontuação."""
    from parsers import _truncar_descricao_em_notas
    # Caso típico Toll Gate
    lines = [
        'For load (relay format)',
        'Athlete 1 (0:00-4:00)',
        '  10-cal Air Bike',
        'Time cap: 14 minutes',
        '─── NOTAS ───',
        'Ponto de partida',
        'A ordem dos 3 atletas é definida',
    ]
    out = _truncar_descricao_em_notas(lines)
    assert len(out) == 4   # corta no NOTAS
    assert 'Time cap' in out[3]
    # Outros separadores
    assert _truncar_descricao_em_notas(['ok', 'Observações', 'cortou']) == ['ok']
    assert _truncar_descricao_em_notas(['ok', 'Pontuação', 'cortou']) == ['ok']
    assert _truncar_descricao_em_notas(['ok', 'Tiebreak', 'cortou']) == ['ok']
    # Sem separador, intacto
    assert _truncar_descricao_em_notas(['a', 'b', 'c']) == ['a', 'b', 'c']


def test_n_atletas_da_modalidade():
    """N atletas inferido pela modalidade."""
    from types_ds import n_atletas_da_modalidade
    assert n_atletas_da_modalidade('individual') == 1
    assert n_atletas_da_modalidade('dupla')      == 2
    assert n_atletas_da_modalidade('trio')       == 3
    assert n_atletas_da_modalidade('quarteto')   == 4
    assert n_atletas_da_modalidade('time')       == 3   # default
    assert n_atletas_da_modalidade('')           == 1


def test_validate_for_load_rejeita_anilhas_acima_do_cap():
    """Mais de 12 anilhas estoura A4 horizontalmente — backend deve rejeitar."""
    import pytest
    from sumula_app import _validate_workout_tipos, BadRequest
    wkt = {"tipo": "for_load", "tentativas": 3, "anilhas": list(range(1, 14))}
    with pytest.raises(BadRequest) as exc:
        _validate_workout_tipos([wkt])
    assert "máximo" in str(exc.value).lower() or "max" in str(exc.value).lower()


def test_validate_for_load_normaliza_unidade_case():
    """Unidade 'KG'/'LB' (caps) é tolerada e normalizada pra lowercase."""
    from sumula_app import _validate_workout_tipos
    for entrada, esperado in [("KG", "kg"), ("kg", "kg"), ("Lb", "lb"), ("LB", "lb")]:
        wkt = {"tipo": "for_load", "tentativas": 3, "anilhas": [25], "unidade": entrada}
        _validate_workout_tipos([wkt])
        assert wkt["unidade"] == esperado, f"esperava {esperado}, got {wkt['unidade']}"


def test_enriquecer_for_load_aplica_todos_defaults():
    """For Load sem campos → enriquecer popula tentativas/unidade/anilhas/barras."""
    from ai_rounds import enriquecer_workouts
    wkt = {"tipo": "for_load", "nome": "MAX"}
    enriquecer_workouts([wkt])
    assert wkt["tentativas"] >= 1
    assert wkt["unidade"] in ("kg", "lb")
    assert isinstance(wkt["anilhas"], list) and len(wkt["anilhas"]) > 0
    assert wkt["barra_masculina"] > 0
    assert wkt["barra_feminina"] > 0


def test_validate_for_load_rejeita_config_invalida():
    """For Load com tentativas/anilhas/barras inválidas → BadRequest."""
    import pytest
    from sumula_app import _validate_workout_tipos, BadRequest

    casos = [
        ({"tipo": "for_load", "tentativas": 0,  "anilhas": [25]}, "tentativas"),
        ({"tipo": "for_load", "tentativas": 99, "anilhas": [25]}, "tentativas"),
        ({"tipo": "for_load", "tentativas": "3","anilhas": [25]}, "tentativas"),
        ({"tipo": "for_load", "tentativas": 3,  "anilhas": []},  "anilhas"),
        ({"tipo": "for_load", "tentativas": 3,  "anilhas": [25, -5]}, "anilhas"),
        ({"tipo": "for_load", "tentativas": 3,  "anilhas": [25, "abc"]}, "anilhas"),
        ({"tipo": "for_load", "tentativas": 3,  "anilhas": [25], "barra_masculina": -1}, "barra_masculina"),
        ({"tipo": "for_load", "tentativas": 3,  "anilhas": [25], "unidade": "g"}, "unidade"),
    ]
    for wkt, contem in casos:
        with pytest.raises(BadRequest) as exc:
            _validate_workout_tipos([wkt])
        assert contem in str(exc.value), f"esperava '{contem}' em '{exc.value}'"


def test_inferir_modalidade():
    """Detecta modalidade a partir do nome da categoria."""
    from parsers import _inferir_modalidade
    assert _inferir_modalidade("Elite Masculino")    == "individual"
    assert _inferir_modalidade("Rx Feminino")         == "individual"
    assert _inferir_modalidade("Dupla Misto")         == "dupla"
    assert _inferir_modalidade("Dupla Amador Masc")   == "dupla"
    assert _inferir_modalidade("Trio Rx Misto")       == "trio"
    assert _inferir_modalidade("Quarteto Amador")     == "quarteto"
    assert _inferir_modalidade("Team Battle")         == "time"
    assert _inferir_modalidade("Equipe Master")       == "time"
    assert _inferir_modalidade("")                    == "individual"


def test_parse_workout_text_detecta_for_load():
    """For Load: detecta tipo, tentativas e não cria movimentos."""
    texto = '"MAX CLEAN"\nFor Load — 5 tentativas\nEncontre a maior carga em 8 minutos.'
    wkt = parse_workout_text(texto, numero=1)
    assert wkt["tipo"] == "for_load"
    assert wkt["nome"] == "MAX CLEAN"
    assert wkt.get("tentativas") == 5
    assert wkt["movimentos"] == []
    # Texto vira descrição (notas)
    assert any("Encontre" in l for l in wkt.get("descricao", []))


def test_parse_excel_vazio_retorna_erro_explicito():
    # Excel sem nenhum dado reconhecível NÃO deve retornar estrutura fantasma
    # com "Único/Geral" vazia — deve sinalizar erro explícito pra UI.
    wb = openpyxl.Workbook()
    buf = io.BytesIO(); wb.save(buf)
    result = parse_excel(buf.getvalue())
    assert result.get('tipo') == 'erro'
    assert 'erro' in result and result['erro']


def test_filtrar_alocacoes_remove_numero_invalido_ou_vazio():
    alocs = [
        {"raia": "1", "numero": "902", "nome": "Foo"},
        {"raia": "2", "numero": "",    "nome": "Sem número"},
        {"raia": "3", "numero": "abc", "nome": "Não numérico"},
        {"raia": "4", "numero": None,  "nome": "None"},
    ]
    keep, drop = _filtrar_alocacoes_por_faixa(alocs, (900, 999))
    assert len(keep) == 1 and keep[0]["nome"] == "Foo"
    assert len(drop) == 3


def test_parse_workout_text_detecta_simultaneous_buyin_marca_paralelos():
    """Simple Dimension: SkiErg + DUs em paralelo no buy-in."""
    texto = (
        "For Time\n"
        "Simultaneous buy-in:\n"
        "900m Ski Erg (1 athlete to completion)\n"
        "150 Double-Unders (1 athlete to completion)\n"
        "After both buy-ins are completed:\n"
        "21 Sync. Pull-Ups (2 athletes)"
    )
    wkt = parse_workout_text(texto, numero=1)
    movs = [m for m in wkt["movimentos"] if m.get("nome")]
    paralelos = [m for m in movs if m.get("paralelo")]
    assert len(paralelos) == 2, f"esperava 2 paralelos, got {len(paralelos)}"
    assert any("SKI ERG" in m["nome"] for m in paralelos)
    assert any("DOUBLE-UNDERS" in m["nome"] for m in paralelos)
    # Mov após "After both" NÃO é paralelo
    pullups = next(m for m in movs if "PULL-UPS" in m.get("nome", ""))
    assert not pullups.get("paralelo")


def test_parse_workout_text_detecta_relay_1_round_per_athlete():
    """Spin: 1 round per athlete vira rounds_per_atleta no wkt."""
    texto = (
        "For Time\n"
        "1 round per athlete:\n"
        "3 Legless Rope Climbs\n"
        "30 Wall-Ball Shots"
    )
    wkt = parse_workout_text(texto, numero=1)
    assert wkt.get("rounds_per_atleta") == 1
    # E a frase NÃO virou movimento
    nomes = [m.get("nome") for m in wkt["movimentos"] if m.get("nome")]
    assert not any("ATHLETE" in (n or "") and "PER" in (n or "") for n in nomes)


def test_parse_workout_text_detecta_emom_e_tiebreak_por_round():
    """Monstar Recap: EMOM 2:30 × 5 + tie-break por round."""
    texto = (
        "Every 2:30 minutes, for 5 rounds:\n"
        "50-metres Swim (2 athletes)\n"
        "10 Dumbbell Thrusters\n"
        "Tiebreak: Tempo no Final de Cada Round"
    )
    wkt = parse_workout_text(texto, numero=1)
    assert wkt.get("tipo") == "amrap"
    assert wkt.get("emom_janela") == "2:30"
    assert wkt.get("emom_rounds") == 5
    assert wkt.get("tiebreak_por_round") is True


def test_parse_equipamento_le_anilhas_e_unidade_kg():
    """Aba Equipamento (Anilha|Peso|Qtd) vira lista de pesos + unidade."""
    import openpyxl
    from parsers import _parse_equipamento
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Equipamento")
    ws.append(["Anilha", "Peso", "Qtd"])
    ws.append(["Vermelho", "25kg", 8])
    ws.append(["Azul", "20 kg", 8])
    ws.append(["Verde", "10kg", 4])
    ws.append(["Mini", "2,5kg", 4])
    r = _parse_equipamento(wb)
    assert r is not None
    assert r["unidade"] == "kg"
    assert r["anilhas"] == [25, 20, 10, 2.5]  # ordenado desc


def test_parse_equipamento_detecta_libras():
    import openpyxl
    from parsers import _parse_equipamento
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Equipamentos")
    ws.append(["Anilha", "Peso", "Qtd"])
    ws.append(["A", "45lb", 8])
    ws.append(["B", "35 lb", 8])
    r = _parse_equipamento(wb)
    assert r["unidade"] == "lb"
    assert r["anilhas"] == [45, 35]


def test_parse_equipamento_aba_inexistente_retorna_none():
    import openpyxl
    from parsers import _parse_equipamento
    wb = openpyxl.Workbook()
    wb.create_sheet("Outra")
    assert _parse_equipamento(wb) is None


def test_parse_mov_line_aceita_reps_gendered():
    """30/24 cal Row — formato gendered (M/F) deve virar movimento."""
    from parsers import _parse_mov_line
    result = _parse_mov_line("30/24 cal Row")
    assert result is not None
    reps, nome = result
    assert reps == 30
    assert "30/24" in nome and "CAL ROW" in nome


# ── _extrair_sequencia_for_load ────────────────────────────────────────────────
def test_extrair_sequencia_for_load_complex_simples():
    """Toll Gate típico: linha com COMPLEX: e buy-in caloric antes."""
    from parsers import _extrair_sequencia_for_load
    lines = [
        "ATHLETE 1 (0:00–4:00) 12-CALORIE AIR BIKE 1-REP-MAX COMPLEX: "
        "1 SQUAT CLEAN 1 FRONT SQUAT 1 SHOULDER-TO-OVERHEAD",
        "ATHLETE 2 (5:00–9:00) 12-CALORIE AIR BIKE 1-REP-MAX COMPLEX (MESMO PADRÃO)",
    ]
    r = _extrair_sequencia_for_load(lines, "TOLL GATE")
    assert r['buy_in'] == '12-CALORIE AIR BIKE'
    assert 'SQUAT CLEAN' in r['complex']
    assert 'FRONT SQUAT' in r['complex']
    assert 'SHOULDER-TO-OVERHEAD' in r['complex']


def test_extrair_sequencia_for_load_trunca_notas():
    """Linhas após NOTAS/OBSERVAÇÕES não entram na sequência."""
    from parsers import _extrair_sequencia_for_load
    lines = [
        "1 Squat Clean + 1 Push Jerk + 1 Split Jerk",
        "—— NOTAS ——",
        "PONTO DE PARTIDA - ATLETA 1 NA AIR BIKE...",
        "OBSERVAÇÕES",
        "CADA ATLETA TEM 4:00",
    ]
    r = _extrair_sequencia_for_load(lines, "MAX")
    assert r['complex'] == '1 SQUAT CLEAN + 1 PUSH JERK + 1 SPLIT JERK'
    # Notas/regulamento não vazaram
    assert 'PONTO DE PARTIDA' not in (r['complex'] or '')
    assert 'CADA ATLETA' not in (r['complex'] or '')


def test_extrair_sequencia_for_load_buyin_marker_explicito():
    """'Buy-in: X' + 'Then: Y' organiza em duas partes."""
    from parsers import _extrair_sequencia_for_load
    lines = ['Buy-in: 30 Wall-Ball Shots', 'Then: 1 Squat Clean + 1 Push Jerk']
    r = _extrair_sequencia_for_load(lines, '')
    assert r['buy_in'] == '30 WALL-BALL SHOTS'
    assert r['complex'] == '1 SQUAT CLEAN + 1 PUSH JERK'


def test_extrair_sequencia_for_load_fallback_nome():
    """Sem texto útil, infere complex do nome do workout."""
    from parsers import _extrair_sequencia_for_load
    r = _extrair_sequencia_for_load(['For Load', '3 tentativas'], 'MAX DEADLIFT')
    assert r['complex'] == 'DEADLIFT'
    assert r['buy_in'] is None


def test_extrair_sequencia_for_load_filtra_atletas_repetidos():
    """ATHLETE 2/3 com '(mesmo padrão)' não duplica complex."""
    from parsers import _extrair_sequencia_for_load
    lines = [
        "ATHLETE 1 (0:00-4:00) 12-cal Air Bike 1-REP-MAX COMPLEX: 1 Squat Clean + 1 Jerk",
        "ATHLETE 2 (5:00-9:00) 12-cal Air Bike 1-REP-MAX COMPLEX (mesmo padrão)",
        "ATHLETE 3 (10:00-14:00) 12-cal Air Bike 1-REP-MAX COMPLEX (mesmo padrão)",
    ]
    r = _extrair_sequencia_for_load(lines, "TOLL GATE")
    # Complex aparece UMA vez (não 3x duplicado)
    assert r['complex'].count('SQUAT CLEAN') == 1


# ── Goal de For Time (Simple Mind/Dimension) ──────────────────────────────────
def test_goal_reps_combo_para_no_plus():
    """Goal: 75 DB Snatches + 50 Burpees → captura SÓ 'DB SNATCHES' (não soma)."""
    txt = "For Time\n21 Pull-Ups\nGoal: 75 DB Snatches + 50 Burpees + finishing rep"
    wkt = parse_workout_text(txt, 1)
    assert wkt['goal_reps'] == 75
    assert wkt['goal_movimento'] == 'DB SNATCHES'


def test_goal_reps_basico_en_pt():
    """Detecta tanto 'Goal:' (EN) quanto 'Objetivo:' (PT)."""
    txt_en = "For Time\nGoal: 75 Snatches + finishing rep"
    txt_pt = "For Time\nObjetivo: 100 Wall-Balls + chegada"
    assert parse_workout_text(txt_en, 1)['goal_reps'] == 75
    assert parse_workout_text(txt_pt, 1)['goal_reps'] == 100
    assert parse_workout_text(txt_pt, 1)['goal_movimento'] == 'WALL-BALLS'


def test_goal_aceita_movimentos_sem_reps_lideres():
    """Quando wkt tem goal_reps, linhas 'Snatches 95/65 lb' (sem reps) viram movs."""
    txt = """For Time
21 Pull-Ups
then...
Snatches 95/65 lb
then...
Snatches 135/95 lb
Goal: 75 Snatches + finishing rep"""
    wkt = parse_workout_text(txt, 1)
    nomes = [m.get('nome') for m in wkt['movimentos'] if m.get('nome')]
    assert 'PULL-UPS' in nomes
    # 2 entradas SNATCHES com cargas diferentes
    snatches = [m for m in wkt['movimentos'] if m.get('nome') == 'SNATCHES']
    assert len(snatches) == 2
    assert snatches[0]['carga'] == '95/65 LB'
    assert snatches[1]['carga'] == '135/95 LB'


def test_goal_filtra_notas_de_movimentos():
    """Linhas de NOTAS / OBSERVAÇÕES não viram movimentos quando goal_reps set."""
    txt = """For Time
21 Pull-Ups
then...
Snatches 95/65 lb
Goal: 75 Snatches + finishing rep
Notes:
Athletes must alternate
Cross the finish line"""
    wkt = parse_workout_text(txt, 1)
    nomes = [m.get('nome') for m in wkt['movimentos'] if m.get('nome')]
    assert not any('NOTES' in n for n in nomes)
    assert not any('ATHLETES MUST' in n for n in nomes)
    assert not any('CROSS THE FINISH' in n for n in nomes)


# ── Progressão de reps por round ──────────────────────────────────────────────
def test_progressao_reps_aplica_apenas_marcados():
    """'*Add N reps' SÓ aplica em movs com '*'; sem markers, ignora directive."""
    txt = """Every 2:30 minutes, for 5 rounds:
50-metres Swim (2 athletes)
10 Dumbbell Thrusters
10 Sync. Pogo Burpees (2 athletes)*
*Add 2 reps each round, last round MAX"""
    wkt = parse_workout_text(txt, 1)
    burpees = next(m for m in wkt['movimentos'] if 'BURPEE' in m['nome'])
    thrusters = next(m for m in wkt['movimentos'] if 'THRUSTER' in m['nome'])
    swim = next(m for m in wkt['movimentos'] if 'SWIM' in m['nome'])
    assert burpees['reps_por_round'] == [10, 12, 14, 16, 'MAX']
    assert thrusters.get('reps_por_round') is None
    assert swim.get('reps_por_round') is None


def test_progressao_sem_marker_nao_aplica():
    """Sem '*' em nenhum mov, directive '*Add' é ignorada (não chuta geral)."""
    txt = """Every 2:30 minutes, for 5 rounds:
10 Burpees
10 Pull-Ups
*Add 2 reps each round"""
    wkt = parse_workout_text(txt, 1)
    for m in wkt['movimentos']:
        if m.get('chegada') or m.get('separador'): continue
        assert m.get('reps_por_round') is None


def test_marker_progressivo_em_varias_posicoes():
    """Aceita '*', '★', '↑' antes/depois de '(N athletes)'."""
    cases = [
        '10 Burpees*',
        '10 Burpees (2 athletes)*',
        '10 Burpees* (2 athletes)',
        '10 Burpees ★',
        '10 Burpees (prog)',
    ]
    for line in cases:
        txt = f"Every 2:30 minutes, for 5 rounds:\n{line}\n*Add 2 reps each round"
        wkt = parse_workout_text(txt, 1)
        mov = next(m for m in wkt['movimentos'] if 'BURPEE' in m['nome'])
        assert mov.get('progressivo'), f"falhou pra: {line!r}"


# ── Carga inline em movimentos ────────────────────────────────────────────────
def test_extrair_carga_unidades_pesos():
    """Captura carga em formatos kg/lb/# com unidade obrigatória."""
    from parsers import _extrair_carga
    casos = [
        ('THRUSTERS 50/35 LB',   '50/35 LB'),
        ('SNATCHES 75#',          '75 #'),
        ('CLEAN 100 KG',          '100 KG'),
        ('POWER SNATCHES @135 LB','135 LB'),
        ('DEADLIFT @200KG',       '200 KG'),
    ]
    for nome, carga_esperada in casos:
        _, c = _extrair_carga(nome)
        assert c == carga_esperada, f"{nome}: esperava {carga_esperada}, got {c}"


def test_extrair_carga_nao_captura_altura_distancia():
    """'@24"' (altura) e '@800m' (distância) NÃO devem virar carga."""
    from parsers import _extrair_carga
    casos = [
        'BOX JUMP-OVERS @ 24"',
        'WALL BALL @ 14',         # sem unidade — também não
        '30/24 CAL ROW',          # cal, não peso
        '900M SKI ERG',           # m, não peso
        'DOUBLE-UNDERS',          # sem nada
    ]
    for nome in casos:
        _, c = _extrair_carga(nome)
        assert c is None, f"{nome}: NÃO devia virar carga, mas got {c!r}"


# ── _enriquecer_roster_com_categoria ──────────────────────────────────────────
def test_roster_categoria_match_unico_da_match():
    """Atleta na faixa única → categoria atribuída pelo nome real dos dias."""
    from parsers import _enriquecer_roster_com_categoria
    roster = [{'numero': '301', 'nome': 'A', 'box': 'X'}]
    inscritos = {'rx masculino': (301, 350)}
    dias = [{'categorias': [{'nome': 'Rx Masculino', 'workouts': [], 'baterias': []}]}]
    _enriquecer_roster_com_categoria(roster, inscritos, dias)
    assert roster[0]['categoria'] == 'Rx Masculino'


def test_roster_categoria_ambiguo_fica_vazio():
    """Múltiplas categorias com mesmo normalizado → categoria='' (não chuta)."""
    from parsers import _enriquecer_roster_com_categoria
    roster = [{'numero': '225', 'nome': 'B', 'box': 'X'}]
    inscritos = {'rx misto': (201, 300)}   # forma relaxada
    dias = [{'categorias': [
        {'nome': 'Rx Misto (Iniciante)', 'workouts': [], 'baterias': []},
        {'nome': 'Rx Misto (Avançado)',  'workouts': [], 'baterias': []},
    ]}]
    _enriquecer_roster_com_categoria(roster, inscritos, dias)
    assert roster[0]['categoria'] == ''


def test_roster_categoria_numero_invalido_vazio():
    """Atleta sem número numérico válido → categoria=''."""
    from parsers import _enriquecer_roster_com_categoria
    roster = [{'numero': 'abc', 'nome': 'C', 'box': 'X'}]
    inscritos = {'rx masculino': (301, 350)}
    dias = [{'categorias': [{'nome': 'Rx Masculino', 'workouts': [], 'baterias': []}]}]
    _enriquecer_roster_com_categoria(roster, inscritos, dias)
    assert roster[0]['categoria'] == ''


# ── EMOM detection (janela + rounds) ──────────────────────────────────────────
def test_emom_detecta_janela_e_rounds():
    """'Every 2:30 minutes, for 5 rounds' → emom_janela='2:30' + emom_rounds=5."""
    txt = "Every 2:30 minutes, for 5 rounds:\n10 Burpees"
    wkt = parse_workout_text(txt, 1)
    assert wkt['emom_janela'] == '2:30'
    assert wkt['emom_rounds'] == 5
    assert wkt['tipo'] == 'amrap'


# ── _parse_equipamento heurística libra ───────────────────────────────────────
def test_parse_equipamento_heuristica_lb_sem_unidade():
    """Pesos 45/35/25/15/10/5 sem unidade explícita → assume lb."""
    import openpyxl
    from parsers import _parse_equipamento
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Equipamento")
    ws.append(["Anilha","Peso","Qtd"])
    for p, q in [(45,8),(35,8),(25,4),(15,4),(10,4),(5,4)]:
        ws.append(["X", p, q])
    r = _parse_equipamento(wb)
    assert r['unidade'] == 'lb'


def test_parse_equipamento_heuristica_kg_quando_tem_25():
    """Pesos com 2.5 ou 1.25 (fracionários) → kg."""
    import openpyxl
    from parsers import _parse_equipamento
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Equipamento")
    ws.append(["Anilha","Peso","Qtd"])
    for p, q in [(25,8),(20,8),(15,4),(10,4),(2.5,4),(1.25,4)]:
        ws.append(["X", p, q])
    r = _parse_equipamento(wb)
    assert r['unidade'] == 'kg'


def test_chegada_negada_varias_formas_omite_rep():
    """Excel dizendo que NÃO tem chegada (várias formas PT/EN) → for_time não
    ganha a rep de chegada. Regressão: só pegávamos 'chegada não conta'."""
    def tem_chegada(nota):
        w = parse_workout_text('"T"\n\nFor time:\n21 Thrusters\n21 Pull-Ups\n\n'
                               + nota + '\nTime cap: 8 min', 1)
        return any(m.get("chegada") for m in w["movimentos"])
    for nota in ("A chegada não conta como repetição", "Sem rep de chegada",
                 "Não há chegada", "Não tem chegada", "A chegada não é contabilizada",
                 "chegada não pontua", "Não conta a chegada final", "No finish rep",
                 "The finish doesn't count"):
        assert not tem_chegada(nota), f"deveria omitir chegada: {nota!r}"
    # Sem menção → default mantém a chegada; menção positiva também.
    assert tem_chegada("A chegada conta normalmente")
    w = parse_workout_text('"T"\n\nFor time:\n21 Thrusters\n21 Pull-Ups\nTime cap: 8 min', 1)
    assert any(m.get("chegada") for m in w["movimentos"])


def test_chegada_negada_nas_notas_vale_para_composto():
    """Composto ("Muscle Swim" + "3k"): o split manda as NOTAS pra fora dos
    blocos F1/F2, então a cláusula "a chegada não conta" não chegava nos
    sub-workouts e cada um re-adicionava a chegada. A regra vale pros dois."""
    txt = (
        'Arena: (Piscina)\n\n'
        '"Muscle Swim" (00:00-08:00)\n\n'
        'For time:\n50m Swim\n10 Devil\'s Presses (22,5kg)\n50m Swim\n\n'
        'Rest 12 minutes, then...\n\n'
        '"3k" (20:00-35:00)\n\n'
        'For time:\n3k Treadmill Run\n\n'
        'Time cap: 35 minutes\n\n'
        '――― NOTAS ―――\n\n'
        'Observações\n'
        '- A chegada não conta como repetição: o workout termina na última rep '
        'do último movimento prescrito.'
    )
    w = parse_workout_text(txt, 1)
    assert w['tipo'] == 'composto'
    for f in ('f1', 'f2'):
        assert w[f]['movimentos'], f'{f} sem movimentos'
        assert not any(m.get('chegada') for m in w[f]['movimentos']), \
            f'{f} não deveria ter chegada'
    # Sem a cláusula, o mesmo composto mantém a chegada nas duas partes.
    w2 = parse_workout_text(txt.split('――― NOTAS ―――')[0], 1)
    assert all(any(m.get('chegada') for m in w2[f]['movimentos'])
               for f in ('f1', 'f2'))


PWRD_LOOP = (
    '"PWRD Loop"\n\n'
    'AMRAP 4 minutes:\n'
    '30 Sync. Toes Raises (2 athletes)\n'
    '30 Sync. Fat Bar Thruster (34kg) + Single-Arm Dumbbell Thruster (15kg) (2 athletes)\n'
    'Max. Wall-Ball Shots (14lbs) + Dumbbell Front Squat (15kg) (2 athletes)\n\n'
    'Rest 1 minute\n\n'
    'AMRAP 4 minutes:\n'
    '30 Sync. Toes Raises (2 athletes)\n'
    '30 Sync. Fat Bar Thruster (34kg) + Single-Arm Dumbbell Thruster (22,5kg) (2 athletes)\n'
    'Max. Wall-Ball Shots (14lbs) + Dumbbell Front Squat (22,5kg) (2 athletes)\n\n'
    '――― NOTAS ―――\n'
    'Pontuação\n'
    '- Será o total de reps de Wall-Ball Shots + DB Front Squat somadas das 2 rounds.\n'
    '- Toes Raises e Thrusters são reps prescritas e não contarão para a pontuação.'
)


def test_amrap_multijanela_pwrd_loop_estrutura():
    """PWRD Loop (Pwrd): 2 janelas AMRAP com reps prescritas (não pontuam) +
    linha MAX (pontua). Antes: virava 1 amrap achatado e as linhas MAX (o que
    conta!) eram DROPADAS por não começarem com número."""
    w = parse_workout_text(PWRD_LOOP, 1)
    assert w["tipo"] == "amrap"
    jans = w.get("janelas") or []
    assert len(jans) == 2, "deveria separar as 2 janelas AMRAP"
    for jan in jans:
        nomes = [m["nome"] for m in jan["movimentos"]]
        maxes = [m for m in jan["movimentos"] if m.get("max")]
        prescr = [m for m in jan["movimentos"] if not m.get("max")]
        assert len(maxes) == 1, "cada janela tem 1 linha MAX (a que pontua)"
        assert "WALL-BALL SHOTS" in maxes[0]["nome"], "MAX = Wall-Ball + DB Front Squat"
        assert maxes[0]["pontua"] is True
        assert all(m["pontua"] is False for m in prescr), "prescritos não pontuam"
        assert any("TOES RAISES" in n for n in nomes)
    # progressão de carga entre janelas (15kg → 22,5kg)
    assert "15KG" in jans[0]["movimentos"][2]["nome"]
    assert "22,5KG" in jans[1]["movimentos"][2]["nome"]
    assert w.get("rest_entre") and "1 minute" in w["rest_entre"].lower()
    assert "wall-ball" in (w.get("score_regra") or "").lower()


def test_amrap_uma_janela_nao_vira_multijanela():
    """AMRAP simples (1 bloco) NÃO deve ativar o path multi-janela."""
    w = parse_workout_text('"X"\n\nAMRAP 12 minutes:\n10 Pull-Ups\n20 Push-Ups\n30 Squats', 1)
    assert w["tipo"] == "amrap"
    assert not w.get("janelas"), "1 janela não é multi-janela"
    assert any("PULL-UPS" in (m.get("nome") or "") for m in w["movimentos"])


def test_time_cap_mmss_e_posicao_no_fim():
    """Corpus achou 2 formatos de time cap que sumiam: mm:ss ('12:30 minutes')
    e time cap no FIM, depois de Note/Score (fora da região de movimentos)."""
    w1 = parse_workout_text('"T"\n\nEvery 2:30, for 5 rounds:\n50m Swim\n10 Burpees\n\nTime cap: 12:30 minutes', 1)
    assert w1.get("time_cap") == "12:30 min"
    # time cap no fim, depois de Note/Score
    w2 = parse_workout_text(
        '"T"\n\nFor time:\n25m Shuttle Run\n40 Wall Ball Shots\n\n'
        'Note\nrestarts each part\n\nScore\nTotal reps\n\nTime cap: 14 minutes', 1)
    assert w2.get("time_cap") == "14 min", "time cap depois de Note/Score foi perdido"
    # negação continua sem cap
    w3 = parse_workout_text('"T"\n\nFor time:\n30 Pull-Ups\n\n- O workout não terá time cap.', 1)
    assert not w3.get("time_cap")


def test_validar_workout_schema_pega_falhas_estruturais():
    """O validador do schema canônico detecta: pontuação Max/Goal dropada e
    time cap perdido (as classes de bug que mordem em produção)."""
    from parsers import validar_workout_schema
    # workout ok não gera problema
    ok = parse_workout_text('"T"\n\nFor time:\n21 Thrusters\n21 Pull-Ups\nTime cap: 8 min', 1)
    assert validar_workout_schema(ok, "For time: 21 Thrusters. Time cap: 8 min") == []
    # simula parse que perdeu o time cap presente no texto
    ruim = {"tipo": "for_time", "nome": "X", "movimentos": [{"nome": "PULL-UPS", "reps": 10}]}
    codigos = [c for c, _ in validar_workout_schema(ruim, "For time:\n10 Pull-Ups\nTime cap: 10 minutes")]
    assert "timecap_perdido" in codigos


def test_rounds_of_no_cabecalho_nao_vira_secao_repetida():
    """'Four rounds of:' como PRIMEIRA linha é o round count (rounds_fixos), não
    um bloco aninhado — não pode virar {secao} (senão o render repete 'FOUR
    ROUNDS OF' em cada round). Regressão do Rocket Master."""
    w = parse_workout_text('"Rocket"\n\nFour rounds of:\n12 Deadlifts\n9 Snatches\nTime cap: 12 min', 1)
    assert w.get("rounds_fixos") == 4
    assert not w.get("rounds_bloco")
    assert not [m for m in w["movimentos"] if m.get("secao")], "não pode criar seção do cabeçalho"
    # Mas 'then, N rounds of' DEPOIS de um buy-in continua sendo bloco aninhado
    sb = parse_workout_text('"SB"\n\nFor time:\n1000m Ski Erg\nthen, 2 rounds of:\n30 HSPU\nTime cap: 16 min', 1)
    assert sb.get("rounds_bloco") == 2 and not sb.get("rounds_fixos")
    assert any("2 ROUNDS OF" in m.get("secao", "") for m in sb["movimentos"])
