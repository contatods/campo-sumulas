"""Render de súmulas: HTML único e HTML combinado por workout."""
from campo_generator import render_workout, render_workout_combined, render_for_load_team_summary


def test_render_workout_for_time_emite_doc_completo(evento_basico, workout_for_time, fonts_empty):
    html = render_workout(evento_basico, workout_for_time, fonts_empty, logo_src="", logo_evento="")
    # Estrutura básica: 1 doc, 1 page
    assert html.count("<html") == 1
    assert html.count("<body>") == 1
    assert html.count('<div class="page">') == 1
    # Conteúdo: nome do workout, evento, time cap, movimento
    assert "TWENTIES" in html
    assert "SUN2026" in html
    assert "9 min" in html
    assert "CHEST-TO-BAR PULL-UPS" in html


def test_render_workout_combined_n_paginas_na_ordem_dos_atletas(
    evento_basico, workout_for_time, fonts_empty, atletas_desordenados
):
    # Não ordena de propósito — o renderer aceita a ordem que receber.
    # Quem ordena por bateria/raia/nome é o handler em _handle_generate.
    html = render_workout_combined(
        evento_basico, workout_for_time, fonts_empty,
        logo_src="", logo_evento="", atletas=atletas_desordenados,
    )
    # 1 doc, 4 pages (uma por atleta)
    assert html.count("<html") == 1
    assert html.count("<body>") == 1
    assert html.count('<div class="page">') == len(atletas_desordenados)
    # Nomes dos atletas aparecem no HTML, na ordem que foi passada
    pos_nomes = [html.find(a["nome"]) for a in atletas_desordenados]
    assert all(p > 0 for p in pos_nomes)
    assert pos_nomes == sorted(pos_nomes)


def test_render_for_load_individual_emite_tentativas_e_barra_correta(fonts_empty):
    """Súmula For Load: barra deduzida do gênero, tentativas, anilhas."""
    ev = {"nome": "EVT", "categoria": "Rx Feminino", "data": "2026", "unidade_default": "kg"}
    wkt = {
        "numero": 1, "nome": "MAX CLEAN", "tipo": "for_load", "modalidade": "individual",
        "tentativas": 3,
        "descricao": [],
    }
    html = render_workout(ev, wkt, fonts_empty, logo_src="", logo_evento="")
    # Tem as 3 tentativas
    for t in ("T1", "T2", "T3"):
        assert t in html, f"esperava {t} na súmula"
    # Anilhas default kg
    for p in (25, 20, 15, 10, 5, 2.5, 1.25):
        assert str(p) in html
    # Barra feminina (categoria Rx Feminino) — 15kg
    assert "15 kg" in html or ">15<" in html
    # NÃO usa barra masculina
    assert "20 kg" not in html  # confere que não vazou a M
    # Melhor Carga aparece
    assert "Melhor Carga" in html or "MELHOR CARGA" in html.upper()


def test_render_for_load_categoria_masculina_usa_barra_de_20kg(fonts_empty):
    ev = {"nome": "EVT", "categoria": "Rx Masculino", "data": "2026", "unidade_default": "kg"}
    wkt = {
        "numero": 1, "nome": "MAX CLEAN", "tipo": "for_load", "modalidade": "individual",
        "tentativas": 3,
    }
    html = render_workout(ev, wkt, fonts_empty, logo_src="", logo_evento="")
    assert "20 kg" in html or ">20<" in html


def test_render_for_load_categoria_mista_individual_usa_barra_masculina(fonts_empty):
    """Categoria MISTO individual (sem time): default conservador = barra M."""
    ev = {"nome": "EVT", "categoria": "Rx Misto", "data": "2026", "unidade_default": "kg"}
    wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
           "modalidade": "individual", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, logo_src="", logo_evento="")
    assert "20 kg" in html, "MISTO individual deve usar barra masculina (20kg)"
    assert "15 kg" not in html, "MISTO individual não deve renderizar barra feminina"


def test_render_for_load_team_dupla_trio_quarteto(fonts_empty):
    """For Load em modalidade dupla/trio/quarteto gera sub-blocos por atleta
    com soma do time no fim. Quarteto entra em layout super-compacto."""
    import re
    ev = {"nome": "EVT", "categoria": "Trio Rx Misto", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "TIME X", "box": "CF", "raia": "1", "numero": "1", "bateria": "1"}
    for modalidade, n_atletas_esperado, super_compact_esperado in [
        ("individual", 0, False),  # sem sub-blocos
        ("dupla", 2, False),
        ("trio", 3, False),
        ("quarteto", 4, True),     # super-compact pra caber em A4
    ]:
        wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
               "modalidade": modalidade, "tentativas": 3}
        html = render_workout(ev, wkt, fonts_empty, "", "", atl)
        n_blocos = len(re.findall(r'class="fl-atleta-bloco"', html))
        n_tents = len(re.findall(r'>T\d+<', html))
        assert n_blocos == n_atletas_esperado, (
            f"{modalidade}: esperava {n_atletas_esperado} sub-blocos, got {n_blocos}"
        )
        # Tentativas totais = N atletas × 3 (em team) ou 3 (em individual)
        esperado_tents = (n_atletas_esperado or 1) * 3
        assert n_tents == esperado_tents, (
            f"{modalidade}: esperava {esperado_tents} tentativas, got {n_tents}"
        )
        m = re.search(r'<div class="fl-zone[^"]*"', html)
        super_c = "fl-zone-super-compact" in m.group()
        assert super_c == super_compact_esperado, (
            f"{modalidade}: super-compact esperado {super_compact_esperado}, got {super_c}"
        )
        # Team tem 'Soma do Time' no fim
        if n_atletas_esperado > 0:
            assert "Soma do Time" in html


def test_render_for_load_team_pre_workout_modalidade(fonts_empty):
    """Sub-bloco por atleta exibe 'Atleta 1', 'Atleta 2', etc — e label
    'Melhor Carga' por atleta (não 'Melhor Atleta N' — esse termo confunde
    semântica: o campo registra o peso lifted, não escolhe qual atleta)."""
    import re
    ev = {"nome": "EVT", "categoria": "Trio", "data": "2026"}
    atl = {"nome": "T", "box": "C", "raia": "1", "numero": "1", "bateria": "1"}
    wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
           "modalidade": "trio", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, "", "", atl)
    for pos in (1, 2, 3):
        assert f"Atleta {pos}" in html
    # 'Melhor Carga' rotula cada sub-bloco (3 atletas)
    import re
    n_label = len(re.findall(r'class="fl-atleta-melhor-lbl">Melhor Carga<', html))
    assert n_label == 3


def test_render_for_load_compact_para_tentativas_altas(fonts_empty):
    """Pra 5+ tentativas, layout compacto é aplicado (cabe em A4).
    Pra 4 ou menos, layout expandido (2 linhas por tentativa)."""
    import re
    ev = {"nome": "EVT", "categoria": "Rx Masculino", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "X", "box": "CF", "raia": "1", "numero": "101", "bateria": "1"}
    for n, esperado_compact in [(3, False), (4, False), (5, True), (8, True)]:
        wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
               "modalidade": "individual", "tentativas": n}
        html = render_workout(ev, wkt, fonts_empty, "", "", atl)
        m = re.search(r'<div class="fl-zone( fl-zone-compact)?"', html)
        assert m, f"div fl-zone não encontrado pra {n} tentativas"
        is_compact = bool(m.group(1))
        assert is_compact == esperado_compact, (
            f"{n} tentativas: esperado compact={esperado_compact}, got={is_compact}"
        )


def test_render_modalidades_aplica_label_correto(fonts_empty):
    """Modalidade muda o label de 'Nome do X' no pré-workout."""
    ev = {"nome": "EVT", "categoria": "Rx", "data": "2026"}
    atl = {"nome": "X", "box": "CF", "raia": "1", "numero": "1", "bateria": "1"}
    wkt_base = {"numero": 1, "nome": "WKT", "tipo": "for_time", "time_cap": "5 min",
                "movimentos": [{"nome": "PULL-UPS", "reps": 10}, {"chegada": True}]}
    for modalidade, label_esperado in [
        ("individual", "Nome do Atleta"),
        ("dupla", "Nome da Dupla"),
        ("trio", "Nome do Trio"),
        ("quarteto", "Nome do Quarteto"),
        ("time", "Nome do Time"),
    ]:
        wkt = {**wkt_base, "modalidade": modalidade}
        html = render_workout(ev, wkt, fonts_empty, "", "", atl)
        assert label_esperado in html, f"modalidade {modalidade!r}: esperava {label_esperado!r}"


def test_render_for_load_libras(fonts_empty):
    ev = {"nome": "EVT", "categoria": "Rx Masculino", "data": "2026", "unidade_default": "lb"}
    wkt = {
        "numero": 1, "nome": "MAX CLEAN", "tipo": "for_load", "modalidade": "individual",
        "tentativas": 3,
    }
    html = render_workout(ev, wkt, fonts_empty, logo_src="", logo_evento="")
    # Barra M default em lb = 45
    assert "45 lb" in html
    # Anilha default lb inclui 55
    assert ">55<" in html or "55" in html


def test_render_for_load_team_summary_lista_atletas_e_soma(fonts_empty):
    """Resumo de time For Load: lista cada atleta com campo de melhor carga + soma."""
    ev = {"nome": "EVT", "categoria": "Dupla Misto", "data": "2026", "unidade_default": "kg"}
    wkt = {"numero": 1, "nome": "MAX CLEAN", "tipo": "for_load",
           "modalidade": "dupla", "tentativas": 3, "unidade": "kg"}
    atletas = [
        {"nome": "João Silva", "box": "CF ALFA", "numero": "401", "raia": "1", "bateria": "1"},
        {"nome": "Maria Souza", "box": "CF DELTA", "numero": "402", "raia": "1", "bateria": "1"},
    ]
    html = render_for_load_team_summary(ev, wkt, fonts_empty, "", "", atletas)
    # Tem todos os atletas pelo nome
    for a in atletas:
        assert a["nome"].upper() in html
    # Tem o campo "Soma do Time"
    assert "Soma do Time" in html or "SOMA DO TIME" in html.upper()
    # Header tem "Resumo do Time"
    assert "RESUMO DO TIME" in html.upper()


def test_render_for_time_relay_renderiza_blocos_atleta_com_cum_continuo(fonts_empty):
    """For Time com `rounds_per_atleta` + modalidade team renderiza UMA tabela
    de movimentos, mas com separador 'Atleta N' entre blocos. Reps cumulativos
    seguem entre atletas — workout contínuo, score do time, UM tempo total."""
    import re
    ev = {"nome": "EVT", "categoria": "Trio Rx", "data": "2026"}
    wkt = {
        "numero": 1, "nome": "SPIN", "tipo": "for_time", "modalidade": "trio",
        "time_cap": "12 min", "rounds_per_atleta": 1,
        "movimentos": [
            {"nome": "ROPE CLIMBS", "reps": 3},
            {"nome": "30/24 CAL ROW", "reps": 30},
            {"chegada": True},
        ],
    }
    html = render_workout(ev, wkt, fonts_empty, "", "")
    # Nota do formato no topo
    assert "Formato Relay" in html
    assert "1 Round por Atleta" in html
    # 3 separadores 'Atleta N' com linha de nome
    n_atletas = len(re.findall(r'class="atleta-sep-row"', html))
    assert n_atletas == 3
    # Movimentos repetem por atleta (mas em UMA tabela só, reps cum acumulam)
    assert html.count("ROPE CLIMBS") == 3
    assert html.count("CAL ROW") == 3
    # Chegada aparece uma vez no fim
    assert html.count("chegada-inline") >= 1


def test_render_for_time_paralelo_marca_movimentos(fonts_empty):
    """Movimentos com `paralelo: True` ganham classe mov-row-paralelo."""
    ev = {"nome": "EVT", "categoria": "Trio Rx", "data": "2026"}
    wkt = {
        "numero": 1, "nome": "SIMPLE", "tipo": "for_time", "modalidade": "trio",
        "time_cap": "12 min",
        "movimentos": [
            {"nome": "900M SKI ERG", "reps": 900, "paralelo": True},
            {"nome": "DOUBLE-UNDERS", "reps": 150, "paralelo": True},
            {"nome": "PULL-UPS", "reps": 21},  # sem paralelo
            {"chegada": True},
        ],
    }
    html = render_workout(ev, wkt, fonts_empty, "", "")
    # Classe presente nos paralelos
    assert 'class="mov-row mov-row-paralelo"' in html
    # Mark visual presente
    assert "mr-paralelo-mark" in html


def test_render_amrap_emom_mostra_header_correto(fonts_empty):
    """AMRAP com emom_janela + emom_rounds mostra 'EMOM X × Y rounds'."""
    ev = {"nome": "EVT", "categoria": "Trio Rx", "data": "2026"}
    wkt = {
        "numero": 1, "nome": "RECAP", "tipo": "amrap", "modalidade": "trio",
        "emom_janela": "2:30", "emom_rounds": 5,
        "movimentos": [
            {"nome": "SWIM", "reps": 50},
            {"nome": "THRUSTERS", "reps": 10},
        ],
    }
    html = render_workout(ev, wkt, fonts_empty, "", "")
    assert "EMOM 2:30" in html
    assert "5 rounds" in html
    # EMOM não emite linha R+ (apenas N rounds fixos)
    assert "amrap-row rplus-row" not in html
    assert ">R+<" not in html


def test_render_amrap_tiebreak_por_round_adiciona_coluna(fonts_empty):
    """AMRAP com tiebreak_por_round adiciona coluna de tiebreak no scorecard."""
    import re
    ev = {"nome": "EVT", "categoria": "Trio Rx", "data": "2026"}
    wkt = {
        "numero": 1, "nome": "RECAP", "tipo": "amrap", "modalidade": "trio",
        "emom_janela": "2:30", "emom_rounds": 5, "tiebreak_por_round": True,
        "movimentos": [{"nome": "SWIM", "reps": 50}],
    }
    html = render_workout(ev, wkt, fonts_empty, "", "")
    n_tb = len(re.findall(r'class="ar-tb-cell"', html))
    assert n_tb == 5, f"esperava 5 células tiebreak (1/round), got {n_tb}"
    assert "Tie-break" in html


def test_render_for_load_trio_misto_atleta_1_usa_barra_feminina(fonts_empty):
    """Trio Rx Misto: atleta 1 = F (15kg), atletas 2 e 3 = M (20kg)."""
    import re
    ev = {"nome": "EVT", "categoria": "Trio Rx Misto", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "TRIO X", "box": "CF", "raia": "1", "numero": "401", "bateria": "1"}
    wkt = {"numero": 1, "nome": "MAX CLEAN", "tipo": "for_load",
           "modalidade": "trio", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, "", "", atl)
    # 3 sub-blocos
    n_blocos = len(re.findall(r'class="fl-atleta-bloco"', html))
    assert n_blocos == 3
    # Header da zone indica misto
    assert "Misto · Barras conforme atleta" in html
    # Atleta 1 = F (com label 'Barra 15 kg' no header do bloco)
    blocos = html.split('class="fl-atleta-bloco"')
    # blocos[0] é antes do 1º bloco; blocos[1..3] são os 3 sub-blocos
    assert "(F)" in blocos[1] and "15 kg" in blocos[1]
    assert "(M)" in blocos[2] and "20 kg" in blocos[2]
    assert "(M)" in blocos[3] and "20 kg" in blocos[3]


def test_render_for_load_dupla_misto_atleta_1_F_atleta_2_M(fonts_empty):
    ev = {"nome": "EVT", "categoria": "Dupla Misto", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "DUPLA", "box": "CF", "raia": "1", "numero": "1", "bateria": "1"}
    wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
           "modalidade": "dupla", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, "", "", atl)
    blocos = html.split('class="fl-atleta-bloco"')
    assert "(F)" in blocos[1] and "15 kg" in blocos[1]
    assert "(M)" in blocos[2] and "20 kg" in blocos[2]


def test_render_for_load_quarteto_misto_2F_2M(fonts_empty):
    ev = {"nome": "EVT", "categoria": "Quarteto Misto", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "QUARTETO", "box": "CF", "raia": "1", "numero": "1", "bateria": "1"}
    wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
           "modalidade": "quarteto", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, "", "", atl)
    blocos = html.split('class="fl-atleta-bloco"')
    # 4 atletas: 1=F, 2=F, 3=M, 4=M
    assert "(F)" in blocos[1] and "(F)" in blocos[2]
    assert "(M)" in blocos[3] and "(M)" in blocos[4]


def test_render_for_load_trio_rx_nao_misto_nao_aplica_genero_por_atleta(fonts_empty):
    """Trio Rx Masculino: todos os sub-blocos com mesma barra M, sem marca de gênero."""
    ev = {"nome": "EVT", "categoria": "Trio Rx Masculino", "data": "2026", "unidade_default": "kg"}
    atl = {"nome": "TRIO", "box": "CF", "raia": "1", "numero": "1", "bateria": "1"}
    wkt = {"numero": 1, "nome": "MAX", "tipo": "for_load",
           "modalidade": "trio", "tentativas": 3}
    html = render_workout(ev, wkt, fonts_empty, "", "", atl)
    assert "Misto · Barras conforme atleta" not in html
    assert "Barra Masculina 20 kg" in html
    # Sem marca de gênero por atleta (a classe é definida no CSS mas não usada)
    assert 'class="fl-atleta-genero"' not in html


def test_parse_excel_aplica_equipamento_aos_for_load(fonts_empty):
    """parse_excel injeta anilhas + unidade nos workouts For Load do evento."""
    import openpyxl, io
    from parsers import parse_excel
    wb = openpyxl.Workbook()
    # Workouts (com pelo menos um For Load)
    ws_w = wb.create_sheet("Workouts")
    ws_w.append(["Categoria", "WKT 1"])
    ws_w.append(["Rx Masculino", "MAX CLEAN\nFor Load"])
    # Equipamento
    ws_e = wb.create_sheet("Equipamento")
    ws_e.append(["Anilha", "Peso", "Qtd"])
    ws_e.append(["A", "45lb", 8])
    ws_e.append(["B", "35lb", 4])
    ws_e.append(["C", "25lb", 4])
    # Remove a Sheet default
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_excel(buf.getvalue())
    # Equipamento detectado em top-level
    if result.get("tipo") != "erro":
        assert result.get("unidade_default") == "lb"
        assert result.get("equipamento", {}).get("anilhas") == [45, 35, 25]


def test_render_escapa_html_de_input_do_usuario(fonts_empty):
    """Garante que dados externos (nome, box, etc) são escapados — sem XSS."""
    ev = {"nome": "<script>alert(1)</script>", "categoria": "A & B", "data": "2026"}
    wkt = {
        "numero": 1, "nome": '"Hack"', "tipo": "for_time", "modalidade": "individual",
        "time_cap": "9 min",
        "movimentos": [{"nome": "<img src=x>", "reps": 20}, {"chegada": True}],
    }
    atleta = {"nome": 'João <b>X</b>', "box": 'CF "Aspas"',
              "raia": "1", "numero": "001", "bateria": "1"}
    html = render_workout(ev, wkt, fonts_empty, logo_src="", logo_evento="", atleta=atleta)
    # Strings cruas NÃO podem aparecer
    assert "<script>alert" not in html
    assert "<img src=x>" not in html
    assert "<b>X</b>" not in html
    # Versão escapada SIM — template usa |upper no nome do evento
    assert "&lt;SCRIPT&gt;" in html
    assert "A &amp; B" in html
