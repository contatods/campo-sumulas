import pathlib
import sys

import pytest

ROOT = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))


@pytest.fixture
def modelo_xlsx_bytes():
    return (ROOT / "modelo_importacao.xlsx").read_bytes()


@pytest.fixture
def fonts_empty():
    return {"black": "", "bold": "", "reg": "", "light": ""}


@pytest.fixture
def evento_basico():
    return {"nome": "SUN2026", "categoria": "RX MASCULINO", "data": "2026"}


@pytest.fixture
def workout_for_time():
    return {
        "numero": 1,
        "nome": "TWENTIES",
        "tipo": "for_time",
        "modalidade": "individual",
        "time_cap": "9 min",
        "movimentos": [
            {"nome": "CHEST-TO-BAR PULL-UPS", "reps": 20},
            {"nome": "THRUSTERS", "reps": 20, "label": "@ 43kg"},
            {"chegada": True},
        ],
    }


@pytest.fixture
def atletas_desordenados():
    return [
        {"nome": "Carlos", "box": "CF Alfa",  "raia": "10", "bateria": "B", "numero": "003"},
        {"nome": "Ana",    "box": "CF Beta",  "raia": "2",  "bateria": "A", "numero": "001"},
        {"nome": "Bruno",  "box": "CF Gama",  "raia": "1",  "bateria": "B", "numero": "002"},
        {"nome": "Diana",  "box": "CF Delta", "raia": "3",  "bateria": "A", "numero": "004"},
    ]


@pytest.fixture
def workout_express():
    """Workout Express completo (formula1 AMRAP + formula2 For Time)."""
    return {
        "numero": 1,
        "numero_f2": 2,
        "nome": "EXPRESS FORMULA",
        "tipo": "express",
        "modalidade": "individual",
        "time_cap": "14 min",
        "formula1": {
            "janela": "00:00-05:00",
            "n_rounds": 3,
            "descricao": ["AMRAP em 5 minutos:"],
            "movimentos": [
                {"nome": "HANDSTAND PUSH-UPS", "reps": 21},
                {"nome": "DEADLIFTS", "reps": 15, "label": "@ 140kg"},
                {"nome": "RING MUSCLE-UPS", "reps": 9},
            ],
        },
        "formula2": {
            "janela": "06:00-14:00",
            "descricao": ["For time:"],
            "movimentos": [
                {"nome": "DOUBLE UNDERS", "reps": 60},
                {"separador": "then..."},
                {"nome": "HANDSTAND PUSH-UPS", "reps": 40},
                {"chegada": True},
            ],
        },
    }


def _build_xlsx_grades_e_dias():
    """Constrói um xlsx in-memory que segue o layout grades-por-modalidade.

    Mimimiza o Sun Challenge: 1 grade `Individuais` × 2 workouts × 2 categorias,
    1 cronograma `Sábado` + `Sábado - Montagem` + `Inscritos` + `Atleta - Individuais`.
    """
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    # Remove sheet default
    wb.remove(wb.active)

    # Grade Individuais: 2 categorias × 2 workouts
    ws = wb.create_sheet("Individuais")
    ws.append(["Rx Masculino", "Scaled Masculino"])
    ws.append([
        '"Twenties"\nFor time:\n20 Pull-Ups\n20 Thrusters\nTime cap: 9 min',
        '"Twenties Sc"\nFor time:\n20 Jumping Pull-Ups\n20 Thrusters\nTime cap: 9 min',
    ])
    ws.append([
        '"Downhill"\nFor time:\n40 Wall Balls\nTime cap: 8 min',
        '"Downhill Sc"\nFor time:\n40 Wall Balls\nTime cap: 8 min',
    ])

    # Cronograma Sábado: 2 baterias × #1, 1 mista
    ws = wb.create_sheet("Sábado")
    ws.append(["Evento Teste"])
    ws.append(["Eventos", "Categoria", "Bateria", "Arbitragem", "Quantidade", "Aquecimento", "", "Fila"])
    ws.append(["#1", "Rx Masculino (Heat 1)", 1, None, "3 (3)", "07:00", None, "07:20"])
    ws.append([None, "Scaled Masculino (Heat 1) & Rx Masculino (Heat 2)", 2, None, "2/1 (3)", "07:30", None, "07:50"])
    ws.append(["#2", "Rx Masculino (Final Heat)", 3, None, "3 (3)", "10:00", None, "10:20"])

    # Sábado - Montagem
    ws = wb.create_sheet("Sábado - Montagem")
    # Bat 1: Rx Masculino (3 atletas)
    ws.append(["07:00", 1, None, None])
    ws.append(["#1", "Rx Masculino (Heat 1)", None, None])
    ws.append(["Raia", "Número", "Nome", "Box"])
    ws.append([1, 601, "ATLETA RX 1", "BOX A"])
    ws.append([2, 602, "ATLETA RX 2", "BOX B"])
    ws.append([3, 603, "ATLETA RX 3", "BOX C"])
    ws.append([None, None, None, None])
    # Bat 2 mista: 2 Scaled + 1 Rx
    ws.append(["07:30", 2, None, None])
    ws.append(["#1", "Scaled Masculino (Heat 1) & Rx Masculino (Heat 2)", None, None])
    ws.append(["Raia", "Número", "Nome", "Box"])
    ws.append([1, 901, "ATLETA SCALED 1", "BOX D"])
    ws.append([2, 902, "ATLETA SCALED 2", "BOX E"])
    ws.append([3, 604, "ATLETA RX 4", "BOX F"])
    ws.append([None, None, None, None])

    # Inscritos: faixas de número
    ws = wb.create_sheet("Inscritos")
    ws.append(["Categorias cadastradas"])
    ws.append(["Nome", "Max", "Pago", "Nº. Inicial", "Nº. Final", "Individual"])
    ws.append(["Rx Masculino", 10, 4, 601, 699, "Sim"])
    ws.append(["Scaled Masculino", 10, 2, 901, 999, "Sim"])

    # Atleta - Individuais
    ws = wb.create_sheet("Atleta - Individuais")
    for num, nome, box in [
        (601, "ATLETA RX 1", "BOX A"),
        (602, "ATLETA RX 2", "BOX B"),
        (603, "ATLETA RX 3", "BOX C"),
        (604, "ATLETA RX 4", "BOX F"),
        (901, "ATLETA SCALED 1", "BOX D"),
        (902, "ATLETA SCALED 2", "BOX E"),
    ]:
        ws.append([num, nome, box])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def xlsx_grades_e_dias_bytes():
    return _build_xlsx_grades_e_dias()
