#!/usr/bin/env python3
"""
sumula_app.py — Súmulas Digital Score  v1.0.0
Servidor web local. Sem dependências além de Jinja2 + fontes Lato.
Uso: python3 sumula_app.py   →  abre http://localhost:8765
"""

import json, os, io, zipfile, threading, webbrowser, sys, base64, re
from http.server import HTTPServer, BaseHTTPRequestHandler

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from campo_generator import render_workout, load_fonts, img_b64, sanitize

PORT = int(os.environ.get('PORT', 8765))
# Render sempre define PORT via env — usa isso para detectar ambiente cloud
HOST = '0.0.0.0' if 'PORT' in os.environ else 'localhost'
IS_CLOUD = HOST == '0.0.0.0'

def _resolve_logo(value):
    """Retorna uma data-URL de logo.
    Se 'value' já é data-URL (upload do front), usa direto.
    Se é caminho de arquivo, converte com img_b64.
    """
    if not value:
        return ""
    if value.startswith("data:"):
        return value          # já é data-URL — upload via interface
    return img_b64(value)    # caminho local

# ── Parsers opcionais ───────────────────────────────────────────────────────────
try:
    import openpyxl; HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    import pdfplumber; HAS_PDF = True
except ImportError:
    HAS_PDF = False

# ── Carregar fontes na inicialização ────────────────────────────────────────────
print("╔══════════════════════════════════════════════╗")
print("║  Súmulas Digital Score  —  v1.0.0            ║")
print("╚══════════════════════════════════════════════╝\n")
print("⏳ Carregando fontes...")
FONTS = load_fonts()
print()


# ── Parsers de texto de workout ─────────────────────────────────────────────────
BLOCK_LABELS = {1: "1º BLOCO", 2: "2º BLOCO", 3: "3º BLOCO", 4: "4º BLOCO", 5: "5º BLOCO"}

def _parse_mov_line(line):
    """Extrai (reps, nome_upper) de uma linha de movimento.
    O número inicial é SEMPRE tratado como reps — metros, calorias, etc.
    Ex: '20-metres Dumbbell Lunges' → reps=20, nome='20-METRES DUMBBELL LUNGES'
    Ex: '20 Pull-Ups'              → reps=20, nome='PULL-UPS'
    """
    m = re.match(r'^(\d{1,4})([-\s])(.+)$', line.strip())
    if not m: return None
    num_s, sep, rest = m.group(1), m.group(2), m.group(3).strip()
    try: num = int(num_s)
    except: return None
    if num >= 1000: return None  # evita anos
    # Se o separador é hífen (20-metres, 40-ft...), mantém o número no nome
    if sep == '-':
        nome = f"{num_s}-{rest}".upper()
    else:
        nome = rest.upper()
    return (num, nome)

def parse_workout_text(text, numero):
    """Converte o texto livre de uma célula/seção num dict de workout."""
    lines = [l.strip() for l in str(text).split('\n') if l.strip()]
    wkt = {"numero": numero, "nome": f"WKT {numero}", "tipo": "for_time",
           "modalidade": "individual", "time_cap": "", "movimentos": [], "descricao": []}

    # Nome: primeira linha entre aspas (simples, duplas, curvas)
    if lines:
        m = re.match(r'^["“‘](.+?)["”’]', lines[0])
        if m:
            wkt["nome"] = m.group(1).strip().upper()
        elif not re.match(r'^\d', lines[0]):
            wkt["nome"] = lines[0].strip('"“”').upper()[:40]

    # Detecta Express antes de qualquer outra coisa
    if any(re.search(r'express formula', l, re.I) for l in lines):
        return _parse_express(lines, wkt)

    # Tipo
    full = '\n'.join(lines).lower()
    if 'for time' in full or 'por tempo' in full:
        wkt["tipo"] = "for_time"
    elif 'amrap' in full or 'as many reps' in full:
        wkt["tipo"] = "amrap"

    # Movimentos, separadores, time cap
    movs = []
    block = 1
    has_seps = any(re.match(r'^then\.+$', l, re.I) for l in lines)
    skip_prefixes = ('for time', 'por tempo', 'amrap', 'as many reps', 'rest',
                     'atenção', 'atencao', 'obs', 'note', '"', '“')

    for line in lines:
        ll = line.lower()
        # Time cap
        tc = re.search(r'time\s*cap[:\s]+(\d+)\s*min', line, re.I)
        if tc: wkt["time_cap"] = f"{tc.group(1)} min"; continue
        # Separador
        if re.match(r'^then[\.\s]*$', line, re.I):
            if movs: movs.append({"separador": "then..."})
            block += 1; continue
        # Linhas de instrução a ignorar como movimentos
        if any(ll.startswith(p) for p in skip_prefixes): continue
        # Movimento
        parsed = _parse_mov_line(line)
        if parsed:
            reps, nome = parsed
            mov = {"nome": nome}
            if reps is not None: mov["reps"] = reps
            if has_seps and block in BLOCK_LABELS: mov["label"] = BLOCK_LABELS[block]
            movs.append(mov)

    if wkt["tipo"] == "for_time" and movs:
        movs.append({"chegada": True})
    wkt["movimentos"] = movs
    return wkt


def _parse_express(lines, wkt):
    """Extrai fórmulas 1 e 2 de um workout Express."""
    wkt["tipo"] = "express"; wkt["estilo"] = "express"
    f1_lines, f2_lines, current = [], [], None
    f1_janela = f2_janela = ""

    for line in lines:
        m1 = re.search(r'Express Formula 1.{0,5}[([]?([0-9]{2}:[0-9]{2}[^)\]]*)', line, re.I)
        m2 = re.search(r'Express Formula 2.{0,5}[([]?([0-9]{2}:[0-9]{2}[^)\]]*)', line, re.I)
        if re.search(r'Express Formula 1', line, re.I):
            current = 'f1'
            if m1:
                j = m1.group(1).strip().strip(')').replace('-', ' -> ')
                f1_janela = j + '  .  AMRAP'
            continue
        if re.search(r'Express Formula 2', line, re.I):
            current = 'f2'
            if m2:
                j = m2.group(1).strip().strip(')').replace('-', ' -> ')
                f2_janela = j + '  .  FOR TIME'
            continue
        if current == 'f1': f1_lines.append(line)
        elif current == 'f2': f2_lines.append(line)

    def extract_movs(flines, add_chegada=False):
        movs, tc_val = [], ""
        for line in flines:
            tc = re.search(r'time\s*cap[:\s]+(\d+)\s*min', line, re.I)
            if tc: tc_val = f"{tc.group(1)} min"; continue
            if re.match(r'^then[\.\s]*$', line, re.I):
                if movs: movs.append({"separador": "then..."}); continue
            p = _parse_mov_line(line)
            if p:
                reps, nome = p
                mov = {"nome": nome}
                if reps is not None: mov["reps"] = reps
                movs.append(mov)
        if add_chegada and movs: movs.append({"chegada": True})
        return movs, tc_val

    f1_movs, _     = extract_movs(f1_lines, False)
    f2_movs, tc    = extract_movs(f2_lines, True)
    if tc: wkt["time_cap"] = tc

    # Corrige nome: "EXPRESS FORMULA 1" → "EXPRESS FORMULA"
    if re.search(r'\s+[12]$', wkt["nome"]):
        wkt["nome"] = re.sub(r'\s+[12]$', '', wkt["nome"]).strip()

    wkt["formula1"] = {"janela": f1_janela or "00:00 → 05:00  ·  AMRAP 5 MIN",
                       "descricao": [], "movimentos": f1_movs}
    wkt["formula2"] = {"janela": f2_janela or "06:00 → 12:00  ·  FOR TIME",
                       "descricao": [], "movimentos": f2_movs}
    return wkt


def _is_categoria_grid(ws):
    """Detecta se a aba tem formato grade (colunas=categorias, linhas=workouts)."""
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if len(rows) < 2: return False
    r1 = [c for c in rows[0] if c is not None]
    r2 = [c for c in rows[1] if c is not None]
    return (len(r1) >= 2
            and all(isinstance(v, str) for v in r1[:4])
            and r2 and isinstance(r2[0], str) and '\n' in r2[0])


# ── Excel import ────────────────────────────────────────────────────────────────
def parse_excel(data):
    if not HAS_EXCEL:
        raise RuntimeError("openpyxl não disponível — instale com: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    # Detecta formato grade em TODAS as abas e combina tudo
    todas_categorias = {}
    evento_nome = ""
    for sname in wb.sheetnames:
        ws = wb[sname]
        if _is_categoria_grid(ws):
            resultado = _parse_excel_grade(wb, sname)
            todas_categorias.update(resultado.get("por_categoria", {}))
            if not evento_nome:
                evento_nome = resultado.get("evento_nome", "")

    # Lê atletas de qualquer aba com a coluna "Nome" (ou "Atleta")
    atletas_por_categoria = _parse_atletas(wb)

    if todas_categorias:
        return {
            "tipo": "categoria_grid",
            "evento_nome": evento_nome,
            "categorias": list(todas_categorias.keys()),
            "por_categoria": todas_categorias,
            "atletas_por_categoria": atletas_por_categoria
        }

    # Fallback: formato template
    return _parse_excel_template(wb)


def _parse_atletas(wb):
    """Lê atletas de qualquer aba que tenha as colunas: Nome/Atleta, Box, Raia, Bateria, Nº.
    Retorna dict { categoria: [ {nome, box, raia, bateria, numero}, ... ] }
    """
    CAMPOS = {
        "nome":    ["nome", "atleta", "name", "athlete"],
        "box":     ["box", "afiliacao", "afiliação", "affiliate", "team"],
        "raia":    ["raia", "lane"],
        "bateria": ["bateria", "heat", "bat"],
        "numero":  ["numero", "número", "nº", "no", "number", "id", "inscricao", "inscrição"],
        "categoria": ["categoria", "category", "cat"],
    }

    def encontrar_col(header_row, opcoes):
        for i, v in enumerate(header_row):
            if v and str(v).strip().lower() in opcoes:
                return i
        return None

    resultado = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue

        # Procura linha de cabeçalho (primeiras 3 linhas)
        header_row_idx = None
        header_row = None
        for ri in range(min(3, len(rows))):
            row = [str(v).strip().lower() if v else "" for v in rows[ri]]
            if any(op in row for op in CAMPOS["nome"]):
                header_row_idx = ri
                header_row = rows[ri]
                break
        if header_row is None: continue

        col = {k: encontrar_col(header_row, v) for k, v in CAMPOS.items()}
        if col["nome"] is None: continue  # sem coluna de nome, ignora

        for row in rows[header_row_idx + 1:]:
            if not row or all(v is None for v in row): continue
            def cell(idx):
                if idx is None: return ""
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ""

            nome = cell(col["nome"])
            if not nome: continue

            atleta = {
                "nome":    nome,
                "box":     cell(col["box"]),
                "raia":    cell(col["raia"]),
                "bateria": cell(col["bateria"]),
                "numero":  cell(col["numero"]),
            }

            cat = cell(col["categoria"]) if col["categoria"] is not None else sname
            if not cat or cat.lower() in ("atletas","inscritos","participants","athletes"):
                cat = sname

            if cat not in resultado:
                resultado[cat] = []
            resultado[cat].append(atleta)

    return resultado


def _parse_excel_grade(wb, sname):
    """Parseia formato grade: col=categoria, linha=workout."""
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"erro": "Planilha vazia"}

    # Linha 0 = cabeçalhos de categoria
    categorias = []
    for col_idx, val in enumerate(rows[0]):
        if val is not None:
            categorias.append((col_idx, str(val).strip()))

    # Linhas 1+ = workouts
    por_categoria = {}
    for cat_idx, cat_nome in categorias:
        workouts = []
        for row_num, row in enumerate(rows[1:], 1):
            if cat_idx >= len(row) or row[cat_idx] is None: continue
            cell_text = str(row[cat_idx]).strip()
            if not cell_text: continue
            wkt = parse_workout_text(cell_text, row_num)
            workouts.append(wkt)
        if workouts:
            por_categoria[cat_nome] = workouts

    # Tentar extrair nome do evento do nome do arquivo ou da planilha
    evento_nome = sname if sname.lower() not in ('individuais','duplas','equipamento') else ""

    return {
        "tipo": "categoria_grid",
        "evento_nome": evento_nome,
        "categorias": [c for _, c in categorias if c in por_categoria],
        "por_categoria": por_categoria
    }


def _parse_excel_template(wb):
    """Parseia formato template (Evento + Workouts + WKT1, WKT2...)."""
    config = {"evento": {"nome": "", "categoria": "", "data": ""}, "workouts": []}
    wkt_map = {}
    for sname in wb.sheetnames:
        sl = sname.strip().lower()
        if sl == "evento":
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]: continue
                k = str(row[0]).strip().lower()
                v = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if k in ("nome","name","evento"): config["evento"]["nome"] = v
                elif k in ("categoria","category"): config["evento"]["categoria"] = v
                elif k in ("data","date"): config["evento"]["data"] = v
        m = re.match(r'^(?:wkt|workout)\s*[-_]?\s*(\d+)$', sl)
        if not m: continue
        num = int(m.group(1))
        ws = wb[sname]; hdrs = None; movs = []
        for row in ws.iter_rows(values_only=True):
            if not any(row): continue
            if hdrs is None: hdrs = [str(c or "").strip().lower() for c in row]; continue
            first = str(row[0] or "").strip().lower()
            if first in ("then...","then","então","---"): movs.append({"separador":"then..."}); continue
            if first in ("chegada","finish","arrival"): movs.append({"chegada":True}); continue
            mov = {}
            for i, h in enumerate(hdrs):
                if i >= len(row) or row[i] is None: continue
                v = str(row[i]).strip()
                if not v: continue
                if h in ("movimento","exercise","movement","nome","name"): mov["nome"] = v.upper()
                elif h in ("reps","rep","repetições"):
                    try: mov["reps"] = int(float(v))
                    except: mov["reps"] = v
                elif h in ("label","bloco","grupo","block"): mov["label"] = v
            if "nome" in mov: movs.append(mov)
        wkt = {"numero": num, "nome": f"WKT {num}", "tipo": "for_time",
               "modalidade": "individual", "time_cap": "", "movimentos": movs}
        config["workouts"].append(wkt); wkt_map[num] = wkt
    config["workouts"].sort(key=lambda w: w.get("numero", 0))
    return config


# ── PDF import ──────────────────────────────────────────────────────────────────
def parse_pdf(data):
    if not HAS_PDF:
        raise RuntimeError("pdfplumber não disponível — instale com: pip install pdfplumber")
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # Tenta dividir por seções de workout
    # Padrão: "Workout N" ou linha com nome entre aspas seguida de tipo
    sections = re.split(r'\n(?=(?:Workout|WKT)\s+\d+)', full_text, flags=re.I)

    config = {"evento": {"nome": "", "categoria": "", "data": ""}, "workouts": []}

    # Extrai nome do evento das primeiras linhas
    header_lines = [l.strip() for l in full_text.split('\n')[:8] if l.strip()]
    for line in header_lines:
        if len(line) > 4 and not re.match(r'^(workout|wkt|\d)', line, re.I):
            config["evento"]["nome"] = line
            break

    wkt_num = 0
    for sec in sections:
        sec = sec.strip()
        if not sec: continue
        # Detecta seção de workout
        has_wkt_hdr = re.match(r'^(?:Workout|WKT)\s+(\d+)', sec, re.I)
        has_quoted  = re.search(r'["“].+["”]', sec)
        has_movs    = re.search(r'^\d{1,3}\s+\w', sec, re.M)
        if not (has_wkt_hdr or (has_quoted and has_movs)): continue
        wkt_num += 1
        wkt = parse_workout_text(sec, wkt_num)
        config["workouts"].append(wkt)

    # Se não achou seções estruturadas, tenta parse linha a linha
    if not config["workouts"]:
        lines = [l.strip() for l in full_text.split('\n') if l.strip()]
        current = None
        for line in lines:
            tc = re.search(r'time\s*cap[:\s]+(\d+)\s*min', line, re.I)
            p = _parse_mov_line(line)
            m_name = re.match(r'^["“](.+?)["”]', line)
            if m_name and current is None:
                wkt_num += 1
                current = {"numero": wkt_num, "nome": m_name.group(1).upper(),
                           "tipo": "for_time", "modalidade": "individual",
                           "time_cap": "", "movimentos": [], "descricao": []}
                config["workouts"].append(current)
            elif current:
                if tc: current["time_cap"] = f"{tc.group(1)} min"
                elif p:
                    reps, nome = p
                    mov = {"nome": nome}
                    if reps is not None: mov["reps"] = reps
                    current["movimentos"].append(mov)
                elif re.match(r'^then[\.\s]*$', line, re.I):
                    current["movimentos"].append({"separador": "then..."})
        for wkt in config["workouts"]:
            if wkt.get("tipo") == "for_time" and wkt.get("movimentos"):
                wkt["movimentos"].append({"chegada": True})

    for wkt in config["workouts"]:
        if wkt.get("tipo") == "for_time" and wkt.get("movimentos"):
            if not any(m.get("chegada") for m in wkt["movimentos"]):
                wkt["movimentos"].append({"chegada": True})
    return config


# ── HTML Interface ──────────────────────────────────────────────────────────────
HTML_INTERFACE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Súmulas — Digital Score</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0D0D0D;--surface:#111;--surface2:#161616;--surface3:#1C1C1C;
  --border:#1E1E1E;--border2:#252525;--border3:#2E2E2E;
  --text:#D8D8D8;--text2:#888;--text3:#555;--text4:#333;
  --accent:#FFF;--green:#5A9;--red:#A55;--blue:#68A;
  --radius:3px;
}
body{background:var(--bg);color:var(--text);font-family:system-ui,-apple-system,'Segoe UI',sans-serif;
  font-size:13px;height:100vh;display:flex;flex-direction:column;overflow:hidden}

/* ── Header ── */
.hdr{background:#0A0A0A;height:46px;display:flex;align-items:center;padding:0 18px;
  border-bottom:1px solid var(--border);flex-shrink:0;gap:12px}
.hdr-logo{font-size:11px;font-weight:900;letter-spacing:.25em;color:#FFF;text-transform:uppercase}
.hdr-divider{width:1px;height:16px;background:var(--border3)}
.hdr-title{font-size:11px;font-weight:400;color:var(--text3);letter-spacing:.08em}
.hdr-right{margin-left:auto;display:flex;align-items:center;gap:10px}
.hdr-version{font-size:9px;font-weight:700;color:var(--accent);letter-spacing:.1em;
  background:rgba(229,93,0,.12);border:1px solid rgba(229,93,0,.25);
  padding:2px 7px;border-radius:10px;text-transform:uppercase}
.hdr-brand{font-size:10px;color:var(--text4);letter-spacing:.05em}

/* ── Layout ── */
.layout{flex:1;display:flex;overflow:hidden}

/* ── Sidebar ── */
.sidebar{width:280px;flex-shrink:0;background:var(--surface);border-right:1px solid var(--border);
  display:flex;flex-direction:column;overflow:hidden}
.sidebar-scroll{flex:1;overflow-y:auto;padding-bottom:4px}
.sidebar-scroll::-webkit-scrollbar{width:3px}
.sidebar-scroll::-webkit-scrollbar-thumb{background:#222;border-radius:2px}
.sidebar-footer{padding:12px;border-top:1px solid var(--border);flex-shrink:0}

/* ── Sections ── */
.sec{border-bottom:1px solid var(--border)}
.sec-hdr{display:flex;align-items:center;padding:10px 14px 8px;gap:8px}
.sec-title{font-size:9px;font-weight:700;letter-spacing:.22em;text-transform:uppercase;color:var(--text3);flex:1}
.sec-action{background:none;border:1px solid var(--border3);color:var(--text3);cursor:pointer;
  font-size:9px;font-weight:700;letter-spacing:.1em;padding:3px 7px;border-radius:var(--radius);
  transition:border-color .15s,color .15s}
.sec-action:hover{border-color:#444;color:var(--text2)}
.sec-body{padding:0 14px 12px}

/* ── Event display ── */
.ev-display{padding:8px 10px;background:var(--surface2);border-radius:var(--radius);cursor:pointer;
  transition:background .15s}
.ev-display:hover{background:var(--surface3)}
.ev-nome{font-size:12px;font-weight:700;color:var(--text);letter-spacing:.04em;text-transform:uppercase}
.ev-meta{font-size:9px;color:var(--text3);letter-spacing:.1em;text-transform:uppercase;margin-top:2px}
.ev-empty{font-size:11px;color:var(--text3);font-style:italic}

/* ── Forms ── */
.field{display:flex;flex-direction:column;gap:4px;margin-bottom:10px}
.field:last-child{margin-bottom:0}
.field label{font-size:9px;font-weight:700;letter-spacing:.16em;text-transform:uppercase;color:var(--text3)}
.field input,.field select,.field textarea{
  background:var(--surface2);border:1px solid var(--border3);color:var(--text);
  padding:7px 9px;border-radius:var(--radius);font-size:12px;font-family:inherit;
  transition:border-color .15s,background .15s;outline:none;width:100%}
.field input:focus,.field select:focus,.field textarea:focus{
  border-color:#3A3A3A;background:var(--surface3)}
.field select{cursor:pointer}
.field textarea{resize:vertical;min-height:72px;line-height:1.5}
.field input::placeholder,.field textarea::placeholder{color:var(--text4)}
.field-row{display:flex;gap:8px}
.field-row .field{flex:1}
.field-row .field.w80{flex:0 0 80px}
.field-row .field.w100{flex:0 0 100px}
.logo-upload-wrap{
  border:1.5px dashed var(--border);border-radius:6px;
  padding:8px 10px;cursor:pointer;min-height:44px;
  display:flex;align-items:center;justify-content:center;gap:8px;
  transition:border-color .2s,background .2s;
}
.logo-upload-wrap:hover{border-color:var(--accent);background:rgba(229,93,0,.06)}
.logo-upload-wrap span{font-size:10px;color:var(--text3);pointer-events:none}

/* ── Workout list ── */
.wkt-empty{padding:14px;font-size:11px;color:var(--text4);text-align:center;font-style:italic}
.wkt-card{display:flex;align-items:center;padding:9px 14px;gap:10px;cursor:pointer;
  border-bottom:1px solid var(--border);transition:background .1s}
.wkt-card:hover{background:var(--surface2)}
.wkt-card.active{background:var(--surface2);border-left:3px solid var(--text2)}
.wkt-num{font-size:17px;font-weight:900;color:var(--border3);width:22px;
  flex-shrink:0;text-align:center;line-height:1}
.wkt-card.active .wkt-num{color:var(--text2)}
.wkt-info{flex:1;min-width:0}
.wkt-name{font-size:11.5px;font-weight:700;color:var(--text2);
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;line-height:1}
.wkt-card.active .wkt-name{color:var(--text)}
.wkt-tags{display:flex;gap:5px;margin-top:3px;align-items:center}
.tag{font-size:8px;font-weight:700;letter-spacing:.1em;padding:1px 5px;
  border-radius:2px;text-transform:uppercase;background:var(--surface3);color:var(--text3)}
.tag.for_time{color:#7AB}
.tag.amrap{color:#9A7}
.tag.express{color:#BA7}
.wkt-actions{display:flex;gap:4px;flex-shrink:0}
.icon-btn{background:none;border:none;color:var(--text4);cursor:pointer;padding:3px 5px;
  font-size:11px;border-radius:var(--radius);transition:color .12s,background .12s}
.icon-btn:hover{color:var(--text2);background:var(--surface3)}
.icon-btn.danger:hover{color:var(--red)}

/* ── Import buttons ── */
.import-btns{display:flex;gap:8px}
.btn-import{flex:1;padding:8px 6px;background:var(--surface2);border:1px solid var(--border3);
  color:var(--text2);cursor:pointer;font-size:10px;font-weight:700;letter-spacing:.1em;
  border-radius:var(--radius);transition:background .15s,color .15s,border-color .15s;text-align:center}
.btn-import:hover{background:var(--surface3);border-color:#3A3A3A;color:var(--text)}
.import-hint{font-size:9.5px;color:var(--text4);margin-top:7px;line-height:1.5;text-align:center}

/* ── Generate button ── */
.btn-generate{width:100%;padding:11px;background:#FFF;color:#000;font-size:10px;
  font-weight:700;letter-spacing:.18em;text-transform:uppercase;border:none;cursor:pointer;
  border-radius:var(--radius);transition:background .15s;display:flex;align-items:center;
  justify-content:center;gap:8px}
.btn-generate:hover{background:#E8E8E8}
.btn-generate:disabled{background:var(--surface3);color:var(--text4);cursor:default}
.spinner{display:none;width:11px;height:11px;border:2px solid #333;
  border-top-color:#000;border-radius:50%;animation:spin .7s linear infinite}
.generating .spinner{display:block}
@keyframes spin{to{transform:rotate(360deg)}}

/* ── Editor panel ── */
.editor{width:400px;flex-shrink:0;display:none;flex-direction:column;
  background:var(--surface);border-right:1px solid var(--border);overflow:hidden}
.editor.open{display:flex}
.ed-hdr{display:flex;align-items:center;padding:12px 16px;border-bottom:1px solid var(--border);
  gap:10px;flex-shrink:0}
.ed-hdr-title{font-size:10px;font-weight:700;letter-spacing:.18em;text-transform:uppercase;
  color:var(--text2);flex:1}
.ed-close{background:none;border:none;color:var(--text3);cursor:pointer;font-size:16px;
  line-height:1;padding:2px 4px;border-radius:var(--radius);transition:color .12s}
.ed-close:hover{color:var(--text)}
.ed-body{flex:1;overflow-y:auto;padding:16px}
.ed-body::-webkit-scrollbar{width:3px}
.ed-body::-webkit-scrollbar-thumb{background:#222;border-radius:2px}
.ed-footer{padding:12px 16px;border-top:1px solid var(--border);
  display:flex;gap:8px;flex-shrink:0}

/* ── Movements editor ── */
.mov-section-hdr{font-size:9px;font-weight:700;letter-spacing:.18em;text-transform:uppercase;
  color:var(--text3);margin-bottom:8px}
.mov-table{border:1px solid var(--border3);border-radius:var(--radius);overflow:hidden;margin-bottom:8px}
.mov-table-hdr{display:flex;background:var(--surface3);padding:4px 8px;gap:6px}
.mov-table-hdr span{font-size:8px;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:var(--text4)}
.mth-nome{flex:1}
.mth-reps{width:52px;text-align:center}
.mth-label{width:72px}
.mth-ctrl{width:52px;text-align:center}
.mov-row{display:flex;align-items:center;gap:6px;padding:5px 8px;
  border-top:1px solid var(--border);background:var(--surface)}
.mov-row:first-child{border-top:none}
.mov-row.sep-row{background:var(--surface2)}
.mov-row.chegada-row{background:#0A1A0A;border-top:2px solid #1A3A1A}
.mov-row input{background:var(--surface3);border:1px solid var(--border3);color:var(--text);
  padding:4px 6px;border-radius:var(--radius);font-size:11.5px;font-family:inherit;outline:none;
  transition:border-color .12s;width:100%}
.mov-row input:focus{border-color:#3A3A3A}
.mov-row .mi-nome{flex:1}
.mov-row .mi-reps{width:52px;text-align:center}
.mov-row .mi-label{width:72px;font-size:10.5px}
.mov-row .mi-sep{flex:1;color:var(--text3);font-style:italic;font-size:11px}
.mov-row .mi-chegada{flex:1;font-size:10px;font-weight:700;letter-spacing:.12em;
  text-transform:uppercase;color:#3A7;padding:2px 0}
.mov-row .mi-ctrl{width:52px;display:flex;justify-content:center;gap:2px}
.empty-table{padding:12px;text-align:center;font-size:11px;color:var(--text4);font-style:italic}
.mov-actions{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px}
.btn-mov{padding:5px 10px;background:var(--surface3);border:1px solid var(--border3);
  color:var(--text2);cursor:pointer;font-size:10px;font-weight:700;letter-spacing:.08em;
  border-radius:var(--radius);transition:background .12s,color .12s}
.btn-mov:hover{background:#1E2A1E;border-color:#2A402A;color:#8C8}

/* ── Express sections ── */
.express-section{border:1px solid var(--border3);border-radius:var(--radius);
  margin-bottom:14px;overflow:hidden}
.express-hdr{background:var(--surface3);padding:8px 12px;display:flex;align-items:center;gap:10px}
.express-hdr-title{font-size:9px;font-weight:700;letter-spacing:.18em;text-transform:uppercase;
  color:var(--text2);flex:1}
.express-body{padding:12px}

/* ── Divider ── */
.divider{border:none;border-top:1px solid var(--border);margin:14px 0}

/* ── Preview panel ── */
.preview{flex:1;display:flex;flex-direction:column;background:#333;overflow:hidden}
.preview-bar{background:#0A0A0A;height:38px;display:flex;align-items:center;
  padding:0 16px;border-bottom:1px solid var(--border);flex-shrink:0;gap:10px}
.pb-label{font-size:8.5px;font-weight:700;letter-spacing:.2em;text-transform:uppercase;color:var(--text4)}
.pb-name{font-size:11px;font-weight:700;color:var(--text2);flex:1;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.pb-hint{font-size:9px;color:var(--text4);letter-spacing:.04em;flex-shrink:0}
.preview-empty{flex:1;display:flex;flex-direction:column;align-items:center;
  justify-content:center;gap:12px;background:var(--bg)}
.pe-icon{font-size:28px;color:var(--border3)}
.pe-text{font-size:10px;letter-spacing:.2em;text-transform:uppercase;color:var(--text4)}
.pe-sub{font-size:10px;color:var(--text4);max-width:260px;text-align:center;line-height:1.6}
.preview-frame{flex:1;border:none;width:100%;display:none;background:#FFF}

/* ── Buttons ── */
.btn-primary{flex:1;padding:9px 14px;background:#FFF;color:#000;font-size:10px;
  font-weight:700;letter-spacing:.14em;text-transform:uppercase;border:none;cursor:pointer;
  border-radius:var(--radius);transition:background .15s}
.btn-primary:hover{background:#E8E8E8}
.btn-secondary{padding:9px 14px;background:none;color:var(--text3);font-size:10px;
  font-weight:700;letter-spacing:.14em;text-transform:uppercase;
  border:1px solid var(--border3);cursor:pointer;border-radius:var(--radius);transition:all .15s}
.btn-secondary:hover{border-color:#3A3A3A;color:var(--text2)}

/* ── Status ── */
.status-bar{padding:6px 14px;font-size:9.5px;min-height:28px;letter-spacing:.04em}
.status-bar.ok{color:var(--green)}
.status-bar.err{color:var(--red)}
.status-bar.info{color:var(--text2)}

/* ── Toast ── */
.toast{position:fixed;bottom:20px;right:20px;padding:10px 16px;border-radius:var(--radius);
  font-size:11px;font-weight:700;letter-spacing:.08em;opacity:0;
  transition:opacity .25s;pointer-events:none;z-index:999}
.toast.show{opacity:1}
.toast.ok{background:#1A3A1A;color:#6C6;border:1px solid #2A4A2A}
.toast.err{background:#3A1A1A;color:#C66;border:1px solid #4A2A2A}

/* ── Modal de categoria ── */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.75);display:flex;
  align-items:center;justify-content:center;z-index:100}
.modal{background:#141414;border:1px solid var(--border3);border-radius:4px;
  padding:24px;width:440px;max-width:90vw;max-height:80vh;overflow-y:auto}
.modal-title{font-size:11px;font-weight:700;letter-spacing:.2em;text-transform:uppercase;
  color:var(--text2);margin-bottom:6px}
.modal-sub{font-size:11px;color:var(--text3);margin-bottom:18px;line-height:1.5}
.cat-grid{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:16px}
.cat-btn{padding:9px 10px;background:var(--surface2);border:1px solid var(--border3);
  color:var(--text2);cursor:pointer;font-size:10.5px;font-weight:600;text-align:left;
  border-radius:var(--radius);transition:all .12s;line-height:1.3}
.cat-btn:hover{background:var(--surface3);border-color:#3A3A3A;color:var(--text)}
.modal-cancel{background:none;border:none;color:var(--text4);cursor:pointer;
  font-size:10px;letter-spacing:.1em;text-transform:uppercase;padding:4px}
</style>
</head>
<body>

<!-- ══ Header ═══════════════════════════════════════════════════════════════ -->
<header class="hdr">
  <span class="hdr-logo">Súmulas</span>
  <div class="hdr-divider"></div>
  <span class="hdr-title">Digital Score</span>
  <div class="hdr-right">
    <span class="hdr-version">v1.0.0</span>
    <span class="hdr-brand">© Digital Score</span>
  </div>
</header>

<!-- ══ Main layout ══════════════════════════════════════════════════════════ -->
<div class="layout">

  <!-- ── Sidebar ──────────────────────────────────────────────────────────── -->
  <aside class="sidebar">
    <div class="sidebar-scroll">

      <!-- Evento -->
      <div class="sec">
        <div class="sec-hdr">
          <span class="sec-title">Evento</span>
          <button class="sec-action" id="btnToggleEvento" onclick="toggleEventoForm()">Editar</button>
        </div>
        <div class="sec-body">
          <div id="eventoDisplay" class="ev-display" onclick="toggleEventoForm()">
            <div class="ev-empty">Clique para configurar o evento</div>
          </div>
          <div id="eventoForm" style="display:none;padding-top:10px">
            <div class="field">
              <label>Nome do Evento</label>
              <input id="evNome" type="text" placeholder="Ex: SUN 2026" oninput="onEventoChange()">
            </div>
            <div class="field-row">
              <div class="field">
                <label>Categoria</label>
                <input id="evCat" type="text" placeholder="Ex: RX MASCULINO" oninput="onEventoChange()">
              </div>
              <div class="field w80">
                <label>Data</label>
                <input id="evData" type="text" placeholder="2026" oninput="onEventoChange()">
              </div>
            </div>
            <div class="field-row" style="margin-top:6px">
              <div class="field">
                <label>Logo do Evento <span style="font-size:9px;opacity:.6">(JPG/PNG — aparece no header)</span></label>
                <div class="logo-upload-wrap" id="logoEventoWrap" onclick="document.getElementById('inputLogoEvento').click()">
                  <img id="logoEventoPreview" style="display:none;height:28px;object-fit:contain">
                  <span id="logoEventoPlaceholder">Clique para selecionar</span>
                </div>
                <input id="inputLogoEvento" type="file" accept="image/*" style="display:none" onchange="onLogoEvento(this)">
              </div>
              <div class="field">
                <label>Logo Digital Score <span style="font-size:9px;opacity:.6">(aparece no header)</span></label>
                <div class="logo-upload-wrap" id="logoEmpresaWrap" onclick="document.getElementById('inputLogoEmpresa').click()">
                  <img id="logoEmpresaPreview" style="display:none;height:28px;object-fit:contain">
                  <span id="logoEmpresaPlaceholder">Clique para selecionar</span>
                </div>
                <input id="inputLogoEmpresa" type="file" accept="image/*" style="display:none" onchange="onLogoEmpresa(this)">
              </div>
            </div>
          </div>
        </div>
      </div>

      <!-- Workouts -->
      <div class="sec">
        <div class="sec-hdr">
          <span class="sec-title">Workouts</span>
          <button class="sec-action" onclick="novoWorkout()">+ Novo</button>
        </div>
        <div id="workoutList">
          <div class="wkt-empty">Nenhum workout ainda.<br>Clique em "+ Novo" para começar.</div>
        </div>
        <div class="status-bar" id="wktStatus"></div>
      </div>

      <!-- Importar -->
      <div class="sec">
        <div class="sec-hdr">
          <span class="sec-title">Importar</span>
        </div>
        <div class="sec-body">
          <div class="import-btns">
            <button class="btn-import" onclick="triggerImport('excel')">📊 Importar Excel (.xlsx)</button>
          </div>
          <div class="import-hint">Importe todas as categorias e workouts<br>a partir do seu arquivo Excel</div>
        </div>
      </div>

    </div>

    <!-- Generate -->
    <div class="sidebar-footer">
      <button class="btn-generate" id="btnGerar" onclick="gerarZIP()" disabled>
        <div class="spinner"></div>
        <span id="btnGerarLabel">⬇&nbsp;&nbsp;Gerar Súmulas (ZIP)</span>
      </button>
    </div>
  </aside>

  <!-- ── Editor panel (slides in) ─────────────────────────────────────────── -->
  <div class="editor" id="editor">
    <div class="ed-hdr">
      <span class="ed-hdr-title" id="edTitle">Novo Workout</span>
      <button class="ed-close" onclick="fecharEditor()" title="Fechar">×</button>
    </div>
    <div class="ed-body">

      <!-- Campos básicos -->
      <div class="field">
        <label>Nome do Workout</label>
        <input id="edNome" type="text" placeholder="Ex: TWENTIES">
      </div>
      <div class="field-row">
        <div class="field">
          <label>Tipo</label>
          <select id="edTipo" onchange="onTipoChange()">
            <option value="for_time">For Time</option>
            <option value="amrap">AMRAP</option>
            <option value="express">Express (2 fases)</option>
          </select>
        </div>
        <div class="field w100">
          <label>Time Cap</label>
          <input id="edTimeCap" type="text" placeholder="Ex: 10 min">
        </div>
      </div>

      <hr class="divider">

      <!-- Movimentos (for_time / amrap) -->
      <div id="secMovimentos">
        <div class="mov-section-hdr">Movimentos</div>
        <div class="mov-table" id="movTable">
          <div class="mov-table-hdr">
            <span class="mth-nome">Movimento</span>
            <span class="mth-reps">Reps</span>
            <span class="mth-label">Label</span>
            <span class="mth-ctrl">Ações</span>
          </div>
          <div id="movTableBody">
            <div class="empty-table">Adicione movimentos abaixo</div>
          </div>
        </div>
        <div class="mov-actions">
          <button class="btn-mov" onclick="addMov('main')">+ Movimento</button>
          <button class="btn-mov" onclick="addSep('main')">⁝ Separador</button>
          <button class="btn-mov" onclick="addChegada('main')" id="btnChegadaMain">✓ Chegada</button>
        </div>
      </div>

      <!-- Express (2 fórmulas) -->
      <div id="secExpress" style="display:none">
        <!-- Fórmula 1 -->
        <div class="express-section">
          <div class="express-hdr">
            <span class="express-hdr-title">Fórmula 1</span>
          </div>
          <div class="express-body">
            <div class="field">
              <label>Janela de Tempo</label>
              <input id="edF1Janela" type="text" placeholder="Ex: 00:00 → 05:00  ·  AMRAP 5 MIN">
            </div>
            <div class="mov-table" id="movTableF1">
              <div class="mov-table-hdr">
                <span class="mth-nome">Movimento</span>
                <span class="mth-reps">Reps</span>
                <span class="mth-label">Label</span>
                <span class="mth-ctrl">Ações</span>
              </div>
              <div id="movTableF1Body">
                <div class="empty-table">Adicione movimentos abaixo</div>
              </div>
            </div>
            <div class="mov-actions">
              <button class="btn-mov" onclick="addMov('f1')">+ Movimento</button>
              <button class="btn-mov" onclick="addSep('f1')">⁝ Separador</button>
            </div>
          </div>
        </div>
        <!-- Fórmula 2 -->
        <div class="express-section">
          <div class="express-hdr">
            <span class="express-hdr-title">Fórmula 2</span>
          </div>
          <div class="express-body">
            <div class="field">
              <label>Janela de Tempo</label>
              <input id="edF2Janela" type="text" placeholder="Ex: 06:00 → 12:00  ·  FOR TIME">
            </div>
            <div class="mov-table" id="movTableF2">
              <div class="mov-table-hdr">
                <span class="mth-nome">Movimento</span>
                <span class="mth-reps">Reps</span>
                <span class="mth-label">Label</span>
                <span class="mth-ctrl">Ações</span>
              </div>
              <div id="movTableF2Body">
                <div class="empty-table">Adicione movimentos abaixo</div>
              </div>
            </div>
            <div class="mov-actions">
              <button class="btn-mov" onclick="addMov('f2')">+ Movimento</button>
              <button class="btn-mov" onclick="addSep('f2')">⁝ Separador</button>
              <button class="btn-mov" onclick="addChegada('f2')">✓ Chegada</button>
            </div>
          </div>
        </div>
      </div>

      <hr class="divider">

      <!-- Descrição -->
      <div class="field">
        <label>Descrição (texto da súmula — uma linha por item)</label>
        <textarea id="edDescricao" placeholder="For time:&#10;20 Chest-to-Bar Pull-Ups&#10;20 Devil's Presses (22,5kg)&#10;Time cap: 10 minutos"></textarea>
      </div>

    </div>
    <div class="ed-footer">
      <button class="btn-primary" onclick="salvarWorkout()">Salvar e Visualizar</button>
      <button class="btn-secondary" onclick="fecharEditor()">Cancelar</button>
    </div>
  </div>

  <!-- ── Preview panel ─────────────────────────────────────────────────────── -->
  <div class="preview" id="preview">
    <div class="preview-bar">
      <span class="pb-label">Visualização</span>
      <span class="pb-name" id="pbName">—</span>
      <span class="pb-hint">Ctrl+P para imprimir / salvar como PDF</span>
    </div>
    <div class="preview-empty" id="previewEmpty">
      <div class="pe-icon">◻</div>
      <div class="pe-text">Nenhum workout selecionado</div>
      <div class="pe-sub">Crie ou selecione um workout para ver a súmula aqui</div>
    </div>
    <iframe id="previewFrame" class="preview-frame"></iframe>
  </div>

</div><!-- /layout -->

<!-- Hidden file inputs -->
<input type="file" id="fileExcel" accept=".xlsx,.xls" style="display:none" onchange="handleImport(this,'excel')">

<!-- Modal: seletor de categoria -->
<div id="catModal" style="display:none">
  <div class="modal-overlay">
    <div class="modal">
      <div class="modal-title">Selecionar Categoria</div>
      <div class="modal-sub" id="catModalSub">Escolha a categoria para gerar as súmulas:</div>
      <div class="cat-grid" id="catGrid"></div>
      <button class="modal-cancel" onclick="document.getElementById('catModal').style.display='none'">Cancelar</button>
    </div>
  </div>
</div>

<!-- Toast -->
<div class="toast" id="toast"></div>

<script>
// ═══════════════════════════════════════════════════════════════════
//  STATE
// ═══════════════════════════════════════════════════════════════════
const DS_LOGO_DEFAULT = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAyAAAADICAYAAAAQj4UaAACUN0lEQVR4nOy9d5xeRb34/55Tnr41lZAECKGFFnpTioKgqKAIiOhVuVa8Kuq13vv7YrnXfhUbWK6KgkrTSxGkBAFFuvRQU0glbevTn3POzO+PmafsZnezu9lNniTzfrFk93nOnDMzp30+82lgsVgsFovFYrFYLNsIMdE7VCBQE71Xi8VisVgsFovFsj0QzSbdq0tx7rn0JE9dh6suxdne/bFYLBaLxWKxWCwThwKhzsVVl57kXXcurtpKI8a4G6tzcbkOOVgjWvr5eW2eCJ3WuFD9ZTXhFhaLxWKxWCwWi8Uy+bTGhWoHxFdW9A7+Tl2Hy3lIwditI2NWEK47F/fcBSjxFSRA/xfnL1CuPDqS4kSFmheEaoFwhKfkWPdssVgsFovFYrFYmgnhgCN42VFitUL93XPUA3e+vPyf511PBFoREefp30e9z7Fs3HiA3i/OP00K9VmlOCUVEx5AKKESqiZzErNYLBaLxWKxWCzjJeaC5wgcAcVAoRTPSCEvW+cnrz7oK89V1KU4fBk12liRUSkgSiH4MkJ8Bbnus/NeH/fF511HnOYIyFUkSqlIIJQSCJSNA7FYLBaLxWKxWHYKBAiQSgEopRBuyhci7gmKgXo2jOT3p39r+a9Ae0pVLSNb2OXIqEtxqu5W6z+/15cTnnup50B/WUoESoAz/H7EJOTZslgsFovFYrFYLJOKGt6YoRRSoVTKd9yU75CrRNet6ytdvOAna7tG45I1onpQVT6e/dTszlmJ2O/Tcff0nkIkFSghcIfcnePUOy1DsMEgFovFYrFYLBbLDoQA1wdhVAUlh1RIlEKCkh1JzysF8uX+srxw7v8sf1SdiytGsIQMq4Aopb9b88VZnQmR/Etr3DmqqxAGQgh/870IEC5EAaqcQymJcDyczBTwYmMfs8VisVgsFovFYtk+KIXMboSoglIKJ54GL66VkCGMC1KqsCXhelGkegpSnj7rm8sfbfSiGszwCojRXNZ/bt5dnSn31I2FMHCGUj4crXjIUj9OegrxfU8kNv8EvJn74k7dCyeeQWfnsr5YFovFYrFYLBZLs6NkSLj+ZaLulVSWPkT5xXuJNq1AxJKIWEp7OQ1CKhWlfdcNpNoQIg+Z8Y3lG7hUx5AP3nZIrUBdd64rzrs+WvPZvb48Le1dOqzlw/GQhR7clmmkXvuvJBaehTd974kYt8VisVgsFovFYmkCZHYj5efuIn/fzwlWP4OT7tRfDLKGSKXC9oTr9ZXkopnJZafzHGIoV6zNFJCq5WPtF/c8Ke1695QCKRWD4j2MP5jM95A4+I20vu1ruFP3qnek6iMmbEIsi8VisVgsFotlh6NRuXC0KqBKWXK3f4f8336B8GLaLUsO1C+kUuG0lOd1FcP/mPmt5V8fKih9cwXE+Gut/dxe97TF3ZP7yjJyGgPOhQApUWGZzBmfJXPaJeZooVY4hlU61DjqJI4TMQp3r6Ei+0fTbqhxiNr/JueYE9nOYrFYLBaLxWIZzEjybDX2wygipadvpf/6zyGLfcYlK2rcVMVcoSKpeuMitm/rN17oBmisEeIN2Pe5uOIrRK9+ce9T0q44OVsZpHxUtwtLtJ3/PZJHv1MfUAhwvMGbDQxUcdzmCgMZlbIxZMPxj2O8x9zW7SwWi8VisVgsux5K6p/BRoVqwimlQEYkDjkTt3MuPT+/AFXKghuryfxCICqRjDpTXmd3sfzxNsGX1XW4ULeCDJBQ1aW6iOD64l6LWuLuKf0VGYlG9yvHQ+Y20vrWL5M+9RMQVfQBN+v8QC0JABkhS1nEJLllKRRCuKhyDpHuQHjxETaWRP0bcOIpXU5FuKigiPDiiGQrIwXNy+xGcD1E1RQlI4hCnJZpw/YMBKrYjwrLCD+JUhFCCGS5gNs6fURXNRWWUfkeRDxj2o1njM2l+1ksFovFYrFYmgelJMJPIPxE44eAGlpOjQJwfSpLHqD7indoK0iD541SKN8VSKm6Q8c5YNY3lmxUClG1gtTMFlXXq9Vf2GvfhCNem61INVD5cFGFHhIHn6mVDxnq/MCDkZGxdrhEfa9SfvpWKssfJdq4lKh/o1FKJt41SAgXWeojseANtJ7zDdM3sbkVQElQUHr8T+Tu+j5Osh1VKeB2zKbt3ZfjJVrMDtm8nXAI1yym95pPQq3gu6T9nT8gtv/JdY1xQDv9vyi7kb6rLybqWY2IpZDFXjKnfYr0SR8BhmpnTnpQJnvL1yg9dydOqhNZ6B77GBOtKFuPxWKxWCwWi8WyGQJUhBPP4E6fjz/nUBIHvwlvt/31d1XZvhHXhyggNv94Wt74Bfpv/jJOZlotO1bNCpL0pvSUojOBK/nySS7cF8IAF6yTHLhP+ojj0nHX6y2GkRDCre6FKMBpmUrr2//b9HWIAuimgzK7kdxfvkXp6VuJshu11cP1Ea7PpASCCIEMA5zWabSc9RVtxVByaBck81n6lIupvPx3ys/dDV6M9vdcoTN4DTXJoMcrI2L7n0zquH8hd9s3Aci86Qta+Ri2nT5x3vS9aXnzf9D983dBdiPxBa8nfcrFDGueEEIXY0m20nLWV6gsfYCo6xXc9lljG+NLf6f83F2IRCuoEYtSWiwWi8VisVh2SQRhrptwwxJKT91C/u4fEZt3DJk3fgF/7sKhF9kdD1RE+pSLKT2/iGD5o4hEBmTNFUspUEh1MnAlB04fIgbEfCiler35pK4pCBdZ7Kb1tE/hTpmrtZvBMR9Su1xVXv47fdd8inDjMpxkO056it5VdUV/MnBcqBTx5xyG0zJVC9piCGVADwZUCMIjvt9JlJ66DW/aPGJ7Hm2UgS24iClFfP9TyN31fQDi+5+y5WBv4YBSxPY8GrdjNuHa54nvd5LZXwRiiPiZWrsIp2Uq3tzDCde9iDf38LGPcfEdOlPBEDmbLRaLxWKxWCwW4XoQSyKM1aP8/D1Ulj5Iy1suJfXai0yMR4PnjRAgAdej9cz/oOvHZw8U9ZVwyqEUQojjll+6R4Lzri8jANWogCy+XimFWPc59o4iBcLsvWr9yEwlcdjZQwvpZvW/+Oi19P3+E+DFdUxEFG0bobeqAEQV8/voIh5UUNa/yEjHZ8QSo1AmBCoo1bZTQclYK7ZwMCFQYbmWJaB27C2iLSFEFf37eMeo1JbHZrFYLBaLxWLZdWnwlhGpNpCSvms/RbhxifaCknLgArjjgpL4ex6BP3chwaqnEPF01QoiAq2z7Bav+C0CSkrqOBAHtKeP+ArylUv2aBOCfcqRAmWCHISDKmeJ73eytn4MDkYxykdlyT/o+/0nEPEMwo9DFLLt8u5WGSIeYsTNxdC/N2W76vZbMUaLxWKxWCwWi2U0yAgEOG27kb/7R+Tv/rFWOAbV/dCJpzySR7xDL8xXVQiBiCIlk56TRooDADhPfznAlOEklQDhD1YblFLE93kt2pWqIZjZ+IPJ7CZ6f/9JnZXJcWq+XxaLxWKxWCwWi2UHRSmd7bV1JtnbvkFlyT+GUEK0OhGbdyxOonWA95MChMARRAPiDYYKeBiofyiFcD3cGfPRq/BiwHcIQe72bxNtWgaxlFU+LBaLxWKxWCyWnQZV86jpv/FS7d5v4puB2ndu51xEqt0UJx/kgeOIAfrFyBHXQoAMEekpeFP2GHAQlNIZr/o3UHzqzzjJtu0f5DzWGIfG7cfSdlu3a9x+a8ZosVgsFovFYrGMFRkh4hmC1U9TWXK/iX82Rodq5tZYEm/qXqiwssUQgC1XBVQS4cd1Gld9FPO5Nr2Unr4VmV1vKiBuT2FXNdQlGUU/lNSZoapB9a43uv4rCV5D/RPPH+iWNmw7pY9hNEZ97NFYi0yfXN/sYxxjtFgsFovFYrFYthIBFB+7vv5HDS3nikRLPVvWCAyT/3UQSm1eQ8JoNsErj5nq5ttR+VAK4fpEvavNoB1doXE47UspcBzCjcsQfhxZ7EVmN+BO3QvCio5jGQoZgRcn6lpZ832LulbCnkdBVGbYOZASvBgyuwFZ7EX4ccKNy2q1RYa1HCllcixLZM9qRDyF7BnHGCep+KPFYrFYLBaLZRfBGCWC1c9CUAY/XgvHqG8zuppzW7aA1Bgk6AoHwjLB+hfrloTthZIQSxOueloXCBRCWwocb+gf16f09K0UHrkWke5AFfvpv+GLqFIWvNjw7bw40fol5G79eq1cfe7WrxOtXwJefIR2MVQpq49R7EekOyg8ci2lp2/dYj8Rgtxt3yRY9SSiZRrBqifHNsZHr9XWq8EZCywWi8VisVgsltFiPHFk/6uEPavqnw1gdNlXR2cBGaoDQqAqJWT/BnAmqcI5YuRxNA5aRohkK7lFPyB49Xliex61edVGUzk86l5F4dHrtGVASkQ8Q/nFe+n68dtIHPImbdEZkPpWH0eWshQfvQ6Z70b4SQCi/vV0Xf52kkedh5Noqfe72k4plJKUnr6NcM2z2jQlJcJx6b3qo6SOOg+3c87m9VVM3yuvPEp58Z268nlYGd8YjW9efVpHmFRVH6/FYrFYLBaLxVJDuKhSFtm3DqbPZ7wy4/gUkFonhKmIPgkCq+Nq16RohJV710cX6jOxFEohEq2UF9+lrQvVcouDu+24JqbFfK8UItFCuO4FsiufMFsNpdEJnERGKx/mmMJPoor95O68zLQZWhMUsWSDXxwgXITrUHjgN6ghrRO6b8KN6b5uzRgHZCnQhSWHPWfC0fNqLSYWi8VisVgslsHU5P/xs3WtgUlRPoSjXZUSLYh0+zDHEMjcJuOPlmwQ0CUi2YIY0btMGQG7MSuV3o+IpUdup+TmtVBcDyfTyYjmmsHtzLFFqkOXvB8WuXlq4/GMUQhd8T0o4rTOHN4KElaQ2Y3a4mIzaFksFovFYrFYNmPrZMQJUEAmGMdBFbPE9j+Z1rO+gts6Y5A7FMbNCIKVT9F3w+eJelYPVEKkBMZRj2QzJWG07YYI0h8t47U0jGWMRvlAOLS+41skDn2LcTNrVGC0JUgGRQr3XEH+b79AxNNWCbFYLBaLxWKxTCjNpYAIB1Up4u1+IB0X/boWZzEcsf1OouP9v6Lrx2fVBGwrMA+FQFWKtJ33XZLHvXvELV3aaTn7q8hiP8VH/oBItll3LIvFYrFYLBbLhDGGLFjbAOGgKnmSh52tlY+oQnVlfsgfGeLNWkBsr2NQ5fygFX0LoK0fUQWnbSaJw99u0v7K4ec0CkBJkse+iwHxIxaLxWKxWCwWywTQXBYQoC70VrNCjZQJy1o8Ro0wwefx1OY5mwdsZ1yzZLjFKpYWi8VisVgsFstYaWKTwSiFXyskj54xzZWdV4vFYrFYLBbLxNPECojFYrFYLBaLxWLZ2bAKiMVisVgsFovFYtlmNKkCMta4DhsHMjrGMk/be06Fda+zWCwWi8Vi2QlpwiB0QLijDy5XSm9vGRkhGL2+2ZAAYHshw3o/bHYzi8VisVgslp2GplRAVDlnsjZFaKF5OGVEgeOZ7a2QOizCQQVlLdQrUyF9SOuCKVjouKhKySgB20MJUYhEC7geBGVUULLWEIvFYrFYLJadhOaS2lWEk2yl8NDvCDcsBS8OjgOOO8yPR+mJm6gsewgRz9iCeUOhFLg+KruR7F3f04K86w0znw54MQgr5O76vt5uW7tiCQdVKdF2/v8w7XP3kTz+X5ClPnCaUle2WCwWi8VisYyR5pLqlALHRxZ66PnZ+WRO/yzulLmb160wf1eWPkT+7h8ivDjbP2ahiZERItlK4W//iyr2kzzyXOPaNMiqoBSqnCN/z+UEyx/RSp2S26HD2gIikq26IKWt9WKxWCwWi8Wy09BcCgiAkgg/icxuou+aS0Z2rVISEUthK3aPAqUQ8TTFR6+n+NgNI26H42xH5cNQdRfbnn2wWCwWi8VisUw4zeWCZbHUsFmwLBaLxWKxWHZGms8CIhxUUMRtmzlqFyxAxzBYK8jwCIEq50kedd6oXLAqyx7e/lYQi8VisVgsFstOR3MpIEKADHBSHXR8+Fq86XuPuHls7+Pwps2j9+qPImJpbBzIMDguqtBL6uQP03r217a4eXy/k+m64lyCVx7V82qVEIvFYrFYLBbLBNFcLljCRRb7SR17oVY+wjJIqbNbDfkTkjjsLGLzjtWpeB1bD2QzhIAoQLRMo+W0T2srURQOM58Swgp4MTKnfUpvtz1rgVgsFovFYrFYdjqaywJi0K4/qp4adjiUNMHV1lVoRJRE+HGdylYIPa/Dxle4gELEEib1rbUqWSwWi8VisVgmjuaygFRRwxXKGwIh9PaWkVEKGK2SJoxCZ5UPi8VisVgsFsvE0pwKyJjdfqyb0OgYyzzZObVYLBaLxWKxTDxNqoBYLBaLxWKxWCyWnRGrgFgsFovFYrFYLJZtRhMrIKOMP7C1P0bPmObKzqvFYrFYLBaLZeJpwixY1aKDDYHQw8nCSppMTZYtohS4vk63O1J8h5KghJ5Xq9xZLBaLxWKxWCaY5rKAKImIpSk+cSMqKIIbA4RWSIb6cTzCtc9RWf4wIm4L5g2JUuDGkH3rKD3+p3pq4+Hm1PVBOBQf+j11ZdBisVgsFovFYpkYmst8oCQiliRcs5ieX72f1rO+gts6o24RadgOAcHKp+i74fOoShHhJ60CMiwKEUvSf9OlKBWROPQtCOGAcAZsg1LIoEjhnisoPnotItFiLCaWCUUM1vuVtTZZdh7s9W2xWCyWLdBcCgiAlIhEC5UX7qXrlTdrIXhIHyyBzG0yRfYGKR+Ow8jGHTW0YL2ZUD5EO1P8cGA7YdptybVpCAXJcUduh9QVyjdrN4YxVos6yoj+Gz5P7s7vD2/ZCCvI7EZEstUKDZOEquTN3ApAIbwYOD5bH3cjJmAfOwt2LrYLSunrm8m4vi0Wi8Wys9B8CghopSLZCjJE5buG3Ux4xkWrUbAXDqqYRUUVhhNChOMiEq0DvxcOKiiiKsVqJwa3AgROIqPdlKrHFA5EAbKUM22GagcilhykKOljq0IPakgrg3l5uzGthG3tGI0SIuItqELvkG2q+xapdmv5mBQUCBd/7uEI16udk6h7FTK7cZyV54VWRmUEyHrszq5qDawuIMhI/16bG8ukoxTC8/F2PwjhOPXru2ulXiwa1/VtsVgslp2R5lRAwAgNDrgjrPKrQQK/EKhSP/ED30Bsz6OMq1ZDeyVBCKLuVRQevQ7huCBc0y6Lt/tBJA55k3FPanT70seQpSzFR69D5ru1MgGooIiT7iRz4gdwEi3VjtTbKYVSktLTtxGuedYoEwpUhJIRqePfi9s5R3+2WV8dKq88SnnxnfV24xmj45pjmvlyRzjtCiuwTQZCQBQhki10fOhqnGR77Xz23/j/kb/nCpx059jmXgiQEpnvw4mnwfVRxT5wPEQsDWoXO4/CQZVzWhBOtkBQQRULejFDCGvRm0yEQEUBTmoGnRdfj/DiVK0g/Td8gcLff4lId9hni8VisViAZlZAgBEzYA3GcVGFXjKnXULmzC9tcfPY/qfQe9VHEb6PKmWJ73cy7e/7hXH5Gp7UUe+k+2fnI3PaMuO2zqDzw9fizpg/Yrv0iR+k98oPUn7pPkSiBRVUaH/PFSQOOXPkdkDu1q+Tu+syRGYKKtc19jEOdiuzgtj2RUbUXORcd3znwwh8wvVpeePniB90OsJPEL76Avl7f0qw4vHNLWc7M8JBlbLE5h9P6qQP4U3bG1XOUnryzxTu/5VJWjHOubaMjaqSISM75xaLxWIZkubKgjVehAOVPN6cQ8i86Qv6hRcFIMOhf6KAxCFnkjr6fFS+B5FspfUd39ACW1gZvl1Yxp0xn8yZX0IFJVRQInPml7TyEZZHaFdBJFr0MZKtqHwPqaPP18rHFvqJUmTe9AX8OQtR2Y34cxaObYxHnY8q9ZtYE0vzIAb9O8a2UiIcj46LriRzxmfxZx+CN2NfEgvfSufH/kh8v5O1NcDZOW7xEXFcVKmfxMK30PnR60kcdAbejH3w5x5Oy1v/H23vuUJbgyZKEB5NzJfFYOfIYrFYLJuzc0gnZjXYbZ9tBAOp4zQcb+gfoeNGvGnzUEEZJ9mO0zK9XitjuHaOB0riTplr0tm6+vdqPZLhflwflMJpmY6TbEcFZbxp82ruUiO2Q7tYOR2zUeUCTsc4xril2h+WHQvHQZb6Sb32ImL7vEYrzdUkB1GA8JO0nvttrVBHO7vLi4AwwGmZRuvbv67vyygYMB+Jg99I8pgLkcWJUcRVUEYFpV3Pxc1isVgslgli51BAABBa8Kj+vsXNHVRYqQnqROHoal4IB8Kg/ncYbCFzVrWd0McwSoc+9mim3/QpCsw+xjFGy86FkggvTuLA043y69YzuJkECe6UucT2PBJVKezc1i/HQVUKxPY8CqdlmlYKTC2b2o+SJA59cz3wf7wYdzZv6h540+Yh4pldx8XNYrFYLJYJZCdSQBh70bzG7cfSdlu3a9x+a8Zo2TlQCuH64CepFeocgl2nOKfSYx0qZswkkxCx1NZlYRIOKijR8tZLmfKZRUz5zF0kDjodVcrt3AqexWKxWCyTwM6lgFgsuwKOiyznCdc8azJhNbgCGbcgFZYJ1ixGeImdWwlRErwYwdrnQKr6Z1WMZTNY8QQExVFaHQej51jEM8T3PwXhJ0xtiybP4WGxWCwWS5NiFRCLZUdDKYSfIHf3j5CFHhMrhEnrq92x8nddRrRhCcQSO3cWIqUQsRThmsXk//bzujtaNeW0FyPqXUv+np9ALL0Vc6Hd3nQKa9XwY7FYLBaLZazsZAqIGpuA0bht07erbr8VY7TsHCiJ8BNEG5fS87N3UVn+qCm8J5D9G8j++b/ILfqhKea5CwRKywiRyJh01d9H5rtqdVcqL/+dnp9dQNS3TlstxmMNEmjrip/AiWWoFiW1WCwWi8UyPnYiHwIFjl8rzjayfKBM1d5YvQBgTTjZgmBhAoCrCC8+OqHGuIpUV2drx97SKqpSOpWqY6qvj2eMdqV250NJRDxDsPppeq54B97M/RHxNOHGZcjetVr52KWUTwGuT+7Wb1B88He4U/ZAlfoJXn0B0FaS8buiCZSKcONp8KvPDKuAWCwWi8UyXnYOC4iSCD9JuPZZVKFXu2HIsJ6Ks/FHRrU0teUlDyASGWT/OoKVT2jlQEZDt1MS0MpK5eX79XYy0r8LR383XDsZgXAIVj6B7F+HSGQoL3mgnoJ3uGPKsFZgMVzzDCLdSbjmmbGNcekDDVWJLTsVSprgap9gzbNUlj6IKvQiUu27mPIB1etbpNqRuY1Ulj5AsPZ5hB9H+Mmtj4NROtB93FYUi8VisVgsNXYSBUSBFyfqXkXv7z6G7N9gBHBn8x/HhSgke+vXKS++E5FsQcmIvms/TbDqSe3jPVQ74YDjUXriRnKLLtPCSDxNbtFllJ64cfjjCQdcj2DVk/Rd+2mUjBDJFsqL7yR769d1kGxjGtVBx5P9G+j93ceIulfiJFuJuleOfYzxll3DFWdXxCjGIpbSaWFdb9c+11Kn4RXxDCKW1M+GrVUYBCYIPa3/2OWUO4vFYrFYJpbmdsESzsiuDkrWhQEZIZJtlJ9bRNcP34K32/5DCAradUL2rSdY+YR2U4lChJ8k6llN9xXnE9vrqGEz5ahKnsrSh/QqaLVfCnqv/hixB69CxNLD9rOy/FFUJa9XY6MQkWghf9dlVF68D6dthunroLEKQfjqC0RdryCSbaigPL4xDhDAxMjVsSdCYLNse+w5qzMa18YxoWsFOYkWvXuUjQCxWCwWi2UraF4FxHFR5QLIYNhNhJcAP15f8ZUhTqqDqGe1SVHKQDlEmNXLWBI3Mw0VmSJ9KtLKQ1TR1ozhVjhdD6d1xkAh3fEQcZ/yc4u0NWPIjgqclmn6GDKguorqpKdQWfFPqBTrfau10X0XqXacVIfuqxDjHyPUXMxUMcewAppw9Er6WIPdLZadGoVIttV+t1gsFovFMn6aUwERAlXoxZ9zaF3gb7SEmL/DdS8Rda/QgkEUgushCz24U/bEW3DqKK0DOnWpquQRsTSJw84enQWkWnxMhqiwQnzBqaO3gJhq6DLfRWyPI0ZtAdm6MTqooIiIpYjvf0qDFaRqyVFaLyoXqCx/pD5Gq4RYLCilcKoKiL0lLBaLxWLZKppPAREOqtRP+rRLaDnjs/UaB0Mg+zfQd+2nKD+3CCfdicx3E19wKm3nfx+ndfrwx4gCsrd/h/yiHyBS7ahyHrdjNu3/8lP8OQtH7F7piRvpu+ZTAwTz9nf/RCsuIxCsepLe336EqGc1Ip5GFXqbf4xSakXFKiGWXRptVRTJVvO3vR8sFovFYtkamisI3XFR5SzxA99Ay5lfqgfUDpMhymmdTvuFP8HtnIss9uN2zqX9wp9owXykDFGuR8uZXyJ+4BtQxSzCcWk7/3taMI+GaWeOmTjsbDKnXoIq51HlPJlTL9HKx3DHUxKiEH/OQtrO/x7CcVHFHWWM2XFWjrZYdia0dbJmAbFYLBaLxbJVNJl0KVBhmfjex+tVdzlyhihkhEi14+1+MCrfjbf7wToFaS0N7TAZomQIShGffzyqlMNpnYk/9zAtgA93POFQDUaN7fMavZ3j6t+r9UOGzZ7lgpL4cw/DaZ2JKuWIz98RxuhZ64fFAgjHabCA2BB0i8VisVi2huZzwUKgQhNwrbbwoq8GXMvABFgHm8eLDHMMRONxJIQVqKbtHLGpgwrLtT9VWNbHVltIfSocfQwT/7EjjXGXp5aNzUT8D8iyJDb/bltnpBp8LWwLpXFMcyK3rSI7GfMhmtUCYuZ6cIxcs5yLnYHB2RgHJAupft4s8zvc9dC4SZNcD0LUF72q1+uQc2vYns+R0R63NqbGdia+s5Y5cwLfEdVzucXrs3rcic4QaLGMnyZUQBiFcD1424aXwOBsUqM9zpiPua3bbccx7opUXyRKooIihGWUlAjHM5Xpza1jClIqGSIcB7wEwk/o9tuqHkcUDHynOM7Al+BEUZ0TaeYkKKNUdU7cAYkZ9JxECMcFL94wJ6ag52ShJEQNL3dB/VyNe58KHBeRaK3vc3tTFYqjEBWUIAx0emDHBTHcufB0bRQvtn0U5R2J2rUeoSoFvZCDQgh3YBpzJfX9oCTC9XTRVy++ba71RqrnOwpQlRIqCsww3IEKiVKgIvMsq96b8dqzbtIF/MbnaliGoKKfncLZ/LmlJMo8Q4VTndtYfUFtMvuqVD2rpRDaVXqk41Xvx7CMqhTNc9HVzx7HgSjSCWuURHgxnYxma6ie77CCCssDz7fTIC80zqEb08/h6ljs/W/ZzjSnAmKxbE8cF8IKspxD+Em86fPxZx+Mt9sC3M45OJlOnQIabQGT+W6i7lWEaxcTrHqKcMNSVFTBSbQad7jJVEQETmaqeXHrlTZVyaMqpYlVIh0XwjKyrLPFeTP3w9/9YLxZB+B2zMFJd2pBRklUUEbmu4i6VhG+uphg1dOEG5ehwgpOIqOTLozl5adq/9vCdqYgYyxFbdVRRchC3+iOM9RqIkL/5ycRiUz9M6Dm7lgVBsbCeK8JI6CpSl7PZ6oTf+5h+LMOxJ2xL27H7jipdi2sgT4X2Q1Em16hsuopglVPEHWv1kJQPL1theQdAlMnyVzrTjyDv/tBeLMPxpu5H277LJxkKzi+Vk7KOaLsRmT3KoJ1LxCufZ6oZxXIUBeAnezCoI6rFaBiPyiJ0zINf85CvN32x502D7dtJiLZhnBj+t6sFJC5TUQ9awjXvUiwdjHRphWosIQTz2gBfzKuiariEZSQlQLCT+B2zsXf7QC8mfvidMzGbZmmr0mhXYhlKYvsX6+frete1Bkhe9foOlrx9OT1VSlwfZzMFP1nWEHmu3VmyKFwXK10BCXczjnEDz4af+7heFP3RKTaEa6v78O+VwnXvUj55ft1Cv3xPDfM+0QV+0ApnNYZ+DOPwNvtALype+G0zsRJtoAb04pmMUvU9yrhhiWEa54lWPcCMteF8OJ6Dq0iYtmOWAXEYqliVuZkvhu3bTfSx15I4rC34s85bPiXzyBUUCJ45TGKj11P6ak/o4p9OoXyhK/W6ZowIpGh86PX47RM1+55rk/2tm+Sv/dnOOmOrRd+Guekcw6ZE95HYuFZ+LMP0cLVKFCVIsGKf1J84kbKT99GlN2IE08zesFBbPll7XjIfBfpE95Ly+mfQUUBwvGJetfS/eOztcVGuCMeU4XlhgQNCqWk3lxJhB83ik3dAKIqeWS+WwfSjWWeHddUVR8DRoBT5RxIib/HESQOO5v4Aa/Dm773qHaRBmR2I+XFd1H4x5UmTXeLnhcrhAy81qfsQebED5JY+Fb8WQeNXLy1AVnoI1jxKKUnb6b07B2oXDci1QYmtm5C+wqoQi/EksQXnErisLOJ7X0cbsfuo96NqhQIVj5J6ak/U3r6FqLeV3XBTdefOMXJcSEoIyt5vGnzSB3yJuIHnoE/55DaPTUaZKGHYMXjlJ75C+Vn7yDqm4S+CgdVyRGbs5COD/zWWA8Dsjd9heJj1+lnefVYZsFC5bvxZu5P6sQPkDjkTJyWaSMeIgNkb/4Kub/+BKcazzmKfoFCFXoR8TTxg99IcuFZ+Hsfi9u226iHF21aTvm5uyn+848EKx8H1zf1yYapYWaxTCLNqYCMVVirbj/edmNtu63bNW6/Pca4K+B4qHIO4XqkT/4I6ZM/2vAiV6iwol2stuDOI/wEsX1eQ2yf15A+8UNkb/825WduQyTaJsctChDJVr06ryQIxyhLE3D+TDFQ4flkTvsU6RM/WE/9PAZhqnFOotd/nPw9l1N65najVIyin1JqwXtLKGVW9vRcCNEYOL6lthJvyh44mamIeAaRbMVJtuOk2hCJVpyWqbjt5kVvlKHkMe/Cn3u4tjaM6n6RIFyintXk77l89NeD40IUIEu9xOYdQ+Z1/0b8oNPr7asCzAhKmgorCFMQNXnsu0gc+Q4Kf/sFuTu+i4rKZgy7sBLiuKhKAeH6tLzhM6RO/CBOy1T9XTVL4ZbOl4xwUm3EDziV+AGnkt64lMI9V1B45FpAW+cmRFB2XO12FwUkFr6F9MkX4+95REM/Qn1bbWmBQEYIP0ls/vHE5h9P+vUfp/DAbyjc/2tUsXegsD0ejIAu8z14U/cgc+J/kDzqXJxUh9lAafdRxxvZWmvc3JxUB/EDXk/8gNcTveHTFB74LYV//AaVN0rehFmalC4wnMjoZ4oQxA95E8VHr2kYW9U9r0j6lIvJnPFZRKJFfxeFqFpMaKMLNAgpwY9rC7qMBn4/HI6HquRBKRJHnkv6pA/jzzmkobujuDbNdu7UvUid+AFSr3k/pSdvJnfX9wnXPo9It4O08SGWbUtTKiDCi43yZWhuFtevmU0HfD5iU2mOo4v0bdHHs6EdXkPdDm+U7iRK6WMIxwhKO8IYm8HZfRvgeMhCN/7uB9P2jm/h73UUgF5FF0K/jIwFRPavJ+pdi8x3oyoFqpXjncwU3I7ZekXL4O1+IB3/+hsKf/8l2Vu+Vvd1nmhMxjNdqNKfGOXR8VDFXryZ+9P2zu/h76EFHBVVtDuHWY0PX32ecP3LRH3rUOU8ACKewW2bgTdjX7zdDqiv9kcB7pQ9aH3Ht2g58z9ovIZrvw1QmCMQDrJrJV1XnKeFlZqr2TBUAy2rczGalT3hoIISred+m9jex49qe4DYvGOJzTt2y9sPItywhPxffzK628vxUMU+nFQHbe/4T1InvK+WVY9IW7yqiofMdyH7NyLL2dr4nVQ7bvus+kqzSdMtvBjp130Mf4/D6f3NB5HF/l1XCXFcVLEfb/p82i74QU2YV2GloSCrJNywhGjTK8jsRq0AOC5OshWndSbu1D0GrESrKMCbtjet532XxMK30vfHLxKue0k/H7Zmjs196U7Zg5azvkzi4Dfpz2Vk4lD82iKJKvYR9a1D5ntQQQHQroROsg2nfbcGRUD3123fjZY3fYHkEeeQvfkrlJ+9XSsh40E4OuahUiT9mveReePntasoDfOKANfXbqx965DZTchyTlt2XR+RaMFpmY7bNqOuTCmFkgFu+yzd1yPP1X195i/G0sQELZ6Z54gM9fO/GltlCv3q2L+Qtnf9gORR55lxGUXe9RDDiVbV0I3uFXqfW3qPOy4y340/awEtZ3+F+H4n68+r9z7oZ2R2I+Gm5cie1UT5bgjK+vpMteN2zsGbPl8XdYba9Zc4/G3EDzyN7J//m/zff6XdY8cSX2qxbCVNpoDoIMpw47LaCsOwAoRSJk2sRPasRsRTyJ7V5uZyjLAyzBteKXAcwo3LEH4cWexFZjfgTt1LZ4oaztwuI/DiRF0ra6stUddK2PMoiMoM+zCRErwYMrsBWexF+PEdZIxm/zvz88jxUPkukgvfStsFl+lAY3M+hHnAB6ufofTULVSWPEDU9Qqy2GcymmGESIHwE4jMFPyZ+xE74HUkDjmzJpCkXvuveNPn0/PriyCqTIIlpCHzzUTEfTguqthLbN6xtF/0a5x0Z+2FJ9wYwaqnKT50NeWX7iPqWYMKylSzuwE1hVf4cdyO2cT3O5nksRfizz4YUFoAHmSZGLHXYWX0q4UwaC5G2UZJ7YqwLRjtC95xkbku4vucQOt538WbsY9Rrozw4TqEa56l9NSfKS99kKhrBaqU1cKx0tYWEUvitEzFn3sYycPfTnzBqeCa545SxPY+jo5//S3dPz1Xf7YlBW9nw3FRpSz+nEPo+MDV2sIXBdpNzosR9aym+PAfKD93l45jKuW0sG9ufiFETdHzdtuf+ILTSBz6lpr1VIUVYvueyJR/u5G+6/6dyvN/BX+0FrPBfdXPqviCU2m94Ae4rTMGKB4C/e4sP30b5ZfvJ9qwBFnoQVWKdaVHOAg/gZNqw52xD/H9TiZ+8Bvxpu5V6683Yx86Png1uTu/T37RZWNf1DBZFIXr0/bun5A84hy97yjQ/fRiyEIv5ecWUX7+bsI1zxL1b9Cr/FHDu9CL4cQzOG0z8ecsJH7gacQPeF1dUZYR3vS96fjAb8kt+gG5276pFe2JEqIbnyN+YsBXKizT/p4rSCx8K4RlHczvxUFGhOteMIsyr+pYPMfBSXXouLlZB6AQhOteqieDGA6jfCSPOIfWc7+ts/BFJm2/6yPzPZSeuoXyM7cRrH0emeuCqGJcR5U5FQ44Op7Fn30QicPeRmLhW3UwehQg4hlaz/kG3sz96f/TlxoSEuxCzwDLdqO5FBAZIRKtFB69ltj+J5M45MwtNsnd+g2CVU8iWqYRrHqS3G3fJHPml0asLg5QevpWCo9ci0h3oIr99N/wRdrf94u6GXUoHI9o/RJyt35d38BA7tavE5t9KO6M+SO0A1XK0n/DF1HFfkS6g8Ij1xLbr9nHmNy5V0RN3EDqmAtou+CH9ZUtUxMlfPV5cnf8D+XnFumAdC+mM4l4cfNCGphWUeU2UX5+DaXFd5K/8zKSx1xA5tRPIBItxPY7ifZ/vZK+qz6mBZxmxdGWDW/WgbR/4LfmpacFXlUpkP3zf1N86GoTjJ7SilcsBY2pNBtSfEa9a8n//X8pPnINyePeQ8ubv1S/rmrKrar536ughArKqHIOVepHFvspPnINKijpDE6TdT0Kh/xff0LisLOoBp43ZpIRfpL4fieZ1XC9ChqsfoZo0/ItCxJVTGaccNOKUSmKMt9N+jXvp+Wcr+tA1urKsetTWfYQ+Xt/SuXF+5ClrLku9Wp93a9eZzyKetYQblhK6Z9/JL7gNFrf8U3cjtk1K4q/x+G0vu2/6fv9JxDJdraYUnxnwVi+3I7ZtP/rb+vKhxG484t+SP7eK4iyGxAmu51ItgBioFqrFKpSoLLkAcov3kd+0Q9JHvkO0q//Nx2bFZa169tR5+mV+lhy7HNsnlXJI8+l/cIfmRpRoVYoHJ9ww1Lyi35A6Zm/IAs99axRrrd5vJGJc4le+hvl5/+Kc9dlxBecSvqUi/F3P6im5Gbe8CmQIbm/fBsx2niyqvIRT9Nx0a+1hVBGgNLXcKVI/u+/oPjAVYSbXtEKnBfXiomfAL9xZhUqKBKuf4lgzTMUH/4d3m4HkD7xQySPfZdWpKMQHIfMqZ/EbZtF3zWX1DPuTaAQLTwjmAsXle+i5eyvDlA+ZKGHwgO/pfTkzUQblurMaY3zJQTC0wsyuB6yZ412wxrueWaUj/TJH6H1bf+lPwsr+lkTBeT/9r8U/v5Lwq5X9Ln2E/q6EilzbdYi1fT1Wc5Rfv4eyosXkb/ncjKn/zuJQ99cs/KkTngvIpak7w+fHHt8msUyTraRAiIYcSFyUN5q4bj0XvVRUkedjztlTu2FP2B74VBZ/gjlxXfq1dSwgki2klv0A4JXnye219EDV2Vr7QRR1yoKj16rTaBSIuIZyi/eS9eP30bikDP1qkFjn033ZClL8ZFrkYXuWhq9qH89XT95O8mjz9cBcWa49XYKJSWlp28lXPOsFv5NCsQdYowDHpBjOI/NjuOhCj0kDn4jbRf8QH+mIqrBvvm//S+5276BKmV1DEB6CrXUj7V86oNwfR2DYDJR5e78H8rP3k7bu36IP/cw4vu8lrYLvt8QxLjthjsqhIAoQsRStL/nigHKh8xuoOdX76ey5AGc9BQzHxHVtJ7D7rKaclJG5P/6Y4JVT9Fx0a+0O4aMwHGJulfTd80lqHIeWcppV66whAorZjU/QsQyk6d8KImIJXUg7hM3DTwvQqCCCm7H7kz7z4e0lURFIDwKf/8lhXt/ish0js3/XDgjB98KAWGFljd9gczp/w4oiLTyIfPdZP/8XxQfvQ7CAJHI1K/NIevQCHMO9IJJ6dnbCde9RMdH/oA3bW9dilZGJI9+J6WnbzXPmon0p29yZEjb+f+D2zazFo+gKgX6rr6Y4hM34aTacdJTzb0vTdalIRCujh0SDiookvvrTyg+dQstb/wCyaPPR+Y2kb35q+OzflSfVYecSfu7f1K3nJsUu4X7f032tm8g8z04yZYtXA8GY80UprZU8bHrKT/zF1InfojM6Z9GeHGi/vWUnv7z6GpHgVnACRGOR8dFVxKbd0zdbdLxCFY+Sd91/06w8nFEvEUnyQAztzBkjQrhInyvdr+EG5bS+4dPUHryJlrP+y5u55yaF0HyqHNBhloJSbRuvq9xYcR5o+CrQi/xg88gfcpHzYJAnNIzt5G98cuEG5ciYkmtrKbaB6mp+lzIvle17cwdIU7PWLrSJ1yklQ9j7cGLEW5YSt81l1BZ8g8zh1MAWT/Pww1ZuDWrc7hhCb2/fj+pEz9E69lf1cpsFJA86jxk/3qyN38FkR7jM81iGQeTr4CYAMphH9zGD7T+sFTmoeOT/8evIQwGLq6aJigQ8ZResau6MCmFSLZTfm4RpSduHrYdnq/dSqop/JREJFuJ1r9M9uYv17dvXNAFEI5JN1oXzIWf1MLmHd+tP+gHtwOcZLt+AMhQf7EDjVEPYIznsZkRDqpSwJ0+n7Z3/ajhha6Vj/4//Qf5e67ASbcjUh2mlsIoYgkahXHHxWmZTrhxKd2Xv4P29/0v8f1PIX7A6xs7MinDGzfCQRV7ybz1/+HN3K/miqIqBXp//a9Ulj2s/YijYPRZU5QCpbd1WmdQWfoAvb/+AB0f+YNxWQhxO2bjdsym8PdfIjLTEMiaFWpA7MJkopRJATronAgBfsVYDQd+J2JJLWiMVWAfKfWl46KKfSQWnqWVj9C47LkxKkseoO+6z+hYgnQHVIOat3QuGq5LJzOVqHsFfb/5EJ0fv6khZTFk3vAZyi/eu3NbPas4LqrQS+KIc4jte6Lx9dcO+n1/+CTFx2/EaZs5+nu/UdgXLk5mKjLXpff16LXIQi9R72qEnxrb/IoGi+SFP26w0urFoP4/flFnvEu1mef9GJ5V1UKECB0PIiNyt3+HYNlDJI46j9JjN+gA5VTbCM/9AZ1FVQq0vftyo3yEVFNVl569g76rLkYFRZ0lytRPGkVHzbyaI/gJRDxF+YV76P7RWbT/65U6I5+MIApJHnMB4cZl5O76vhbOJyq7k69jAEWyhZazvkI1vjJ/z+Vkb/oy+AmdtEBWFdVhxubF9FNkuHekuf9j+59C6zu+UVd8XZ9gxeP0/PK9yOwGk/VwtNcmYFxfQcssxFLk//oTZG4T7e/WCTFUFJJ+/cepvPIo5Wfv2LUWIizbhclVQIRAFXr1A2e4NKZKIfvX6ZuiGlwpHFSpHzczdUR3IZndiKrkjbCsV4RUJV83dQ6DKmV1LEa8pXY8ogAlA7zp+wzrHqHCMrJ/vW7X0H9VzuF2zqnl3h9qjFHfWkTk7JhjHNd5bHIlREnazvvOwDSIwiF706Xk//oTLWirUQh3w+5f1XxsCcv0/vr9dHz4Wv1iNuexqRACghLu9L1Jv+b9+roWAiEccnd8l/LL9+O0ztQxLOMlCnAy0yi/dB+5O/6Hljf/p/ZfF4LMm75IefEiVFgC4Q0SkLYRQ60gVgW+ofpRFTRGLUyNrg8ilqKy9AHCDUvwpmvXzsKDV9H/p/9A13qYatIFj+OYUYBIdVBZ8Ti5RT+k5cwv1QRnf+5C4vu8lvLzf9WuRqMSOndQlAQvTvqkDwKqVlCz+NDvKD52g1Y+xu0qqd1ahBcDL05l6YPaNW6syofpp3B9E5/WMuBZ1X/DF8jf91Oclq18VkEtxsppmUrllceoLHlAZ2tKjlL5cDxUvpvkCe/VMR+yqnw4lF+8j94rP4BwXJ1ZKtqKfppCoyLdQdS/np6fX0jnx/6k46PMfZh54+cJlj5EZcVjA+dsK3D8JKqcI/3aS/CmzQOgcP+v6f+//08rfjC6cY1YzFBbPp3MFNre+X1T80MHwYfrXqTnf9+DLPYhUu1b58ZrnnNO20yKj1yL0zKd1rd9rbbP1rO/xqalD5sFKBsPYpk8Jk8BEQJVzpM++SOkTvkojp9ks6BQpSvIlp66RZunZaR9zYv9xPY/mdazvqID7YZoh4Bg5VP03fB5op7ViFgaVcnjdszWmYzmHqqFiQEBv3o1JepfT/9Nl1J54V5tlTBm4tZzvkHi0Ldos/QQ7WRQpHDPFeT/9ouan6Qq50mf9OGdf4xjPY/NqoQY39rUCe8jNv+E+svJcSk+cg25u3+klQ/jt7zVmKB+VSnS+9sPM+WTt+oA1WZbZRYuspwnfcTbEYmWWsBouHEZhX/8xgSib4XyUSWq4KQ7Kdz/a5LHXIA3bW9UFOK2zSS+4PUUH/6D9jdvtvnZVpjEEzLXRe9VHyF1/HsJVjxO8eHfa/cvx906AQ4gCnFS7RQevJrUa96P27abOd8OiUPfQnnxnTSddW4iMRZQf/bB+LP1M1Q4OhtT/r6fmwKNE6BQGiVaxE0By7Fe046HynWROf0z+HMOrQv1wiF/38/J3/czY5GcqBoOJkFELAVjKVJXXbyYMqeu0CLAEUTdq+j73cdMrMcE1uuIQkQ8g8xuou+qjxprXtIobB4tb/saXT9664Q9RxTgdMwm9dr3A1BZ9hD9//efxo1sgqz+xtrVcvZX9eJi1QJdztF31cW6GGKideLOdxTgtE6ncO9Pie9zAvGDzkBFFdwpe5A64X3k7viudpVVtkaIZXKYHAXEmBGTR19Ay9lfHXFTAaSOfy9CuPRd/zmQUqcvvejXtTiL4YjtdxId7/8VXT8+W/tIx1J0vP9XeLMWjNjOS7XTcdGv6frBmwnXL4EooO3cb5M87t0jtnNpp+Xsr+rA2EevAyB51Pl2jIPHeN2/G6tOkykgQphV+KlkTr2kvsouXKLuVWRv+rKOe1ATpHxUkZHOYNa7lv4bPk/7RVeaGJwmQkU48RTxA99A4wu19Oh1yEKvidmYoBeR4yFzmyg+cq0RWCLAJX7wGRQfuXZijrEjo3TMVvjqi/Rf82lwnLpP+4QIVApcnZWv/OQtpE76UO2b2PwTjP+3cRXdGRECFZbx5x6u31VhgPB8wlVPE25YogXwiVSAx7OvqlA/bU9Sr7u4vg/hEKxdTO7WrxuXq0lwkRkplmDIvjrISp6W111adwMTOsaw///+E9m3vuGamkBkiEi1aWvend+j5S3/n35uyQh/7kKSh7+d4kO/G30A/UgEJZKHvw23fXdUpaCtkdXkGRNxDoSDKufx9zyS5DEXaKuTUTZzd/wPlVVPaLeriU5gohT4cbK3foPYvifpLFhKkXrN+yk+dLVJNT/Kek0WyxiZHCnIrIAkj30XtXz1jcG7jT9SuzEkDn87bvssVKGX5OFna8E8qlD3AR2qbYg3awGxecci+zcQm3esFsyrdRGG+qkGdPpJkoefjSr04rbPInH4240JVw7fNgpAST0ugx3j5mPU7guVYd28thvCQZayJI84B7fTZAESDghB7o7vEOU2gTvKjEZjRUaIeIZgxeOoUi9NlepQCAgquO2zdewHQqcgVpLyi/fpAOYJFshELEH5hXv1vLi6JoA/+1CcjEn5u7MKv6NFSe3vnu7QbjANqTUnbP+uR/mFewBqNQncztnaKhWUB1lIdyZ0hiR36p7mb31th+tfRoVN8twSLrKSI3X8e82iiKy5beb+/HWdWrcZrMxGUfKm70PyqPPr/REO5cV36Kxf6Y6JVz6qRAFOuoPC/b8iXPdiw5wo0id/RLt8bXUxRXAyU0geo9+JhQevJlj5xIS5d9WOE1ZIn/QhnS1MaS+C8NUXKNz/a5xU58QrH0A1/XiwdjGlJ2/SipAMtEX6wNORpdzwKfstlq1k8q4skxGjqsUPyKk9+Kca4NxYR6C6ujC4xsGAnyGEOJM9atg2mHaNZlOzMo7jjtzP6lhkWP/MjnH4MTYbMsKJp41ypWrzGL76AqUnbsZpDPafDJQ0MTRNNj/CQUUVnClzTbVmnVkt6ltH1LWiXsxyolAK4caJulcQ9awx14vCbZuBY9yBmvYa2pY0xphM/M4RXpxw/Yu6rk0tEYODN2N+cy4gTBgKYQqIDvi0UqA5VnoFRBXc1pk6sxP1Z3jl5fspv/DXCa78vRUIB1nOkzzyHO26VlWUlCR/zxW1wP5JxfGQpSz5+35u+qQVTG+3/Ynte6Ku3TLufhgFpHUG3ox9UKUshft/hRPfSsVmwCEEVIp4M/cjfvAbqV6fAPl7f4os5ydZCdBxRsVHr9XPZiM7JBa+FeE2gZJr2WmZZNV2DC+wzV52o2w7VNaa0TUcZ7vBbe0Yx3+MbYhj/L73OBJ/1oH6nW76Wnj497oC77Z4WTbrw1xFuC3T9K8muFn2rUOWsmZeJrLfSp+PUo6o71X9URSB4+HWXL2a9DraWVAKHB+Z6ybqXq0/MsHG7tQ9ze877zlQxkrciGiZUhP8tivm3ojvd7JO/FBdcAIKD/y2ORSPKjLCSbbpmhhg+ip0IPvyR7WSN9n9lRFOooXys7cT9a83wdvajTZ52NuoWri27hh6Yar45M2EG5aAP8rUxKNBOMhKgfjBb9RZNU2xwah3LeXFd+j0/pOZEELqxBfBqqcINy7RaXkRxPY8ArdjjsnEt/M+CyzbjyZ42los2wKBCivEDzxV/2kC5VUlT3nxXTixFLtMAbahUNQLUJkXqyrnJs8dymSXUqWsObx5mceSINXOLPs2D0KgghKyf/2Aj52W6bVV0J0TvUKux11fbY7tcYRJld4EzwEhiB90OlTTpwqHqPdVKi/fb4LkmyBJQ0Mwvzdjn5ryAVB68uZt6ManwPWJ+tdTfu6uxg4S2/e1OG2ztBC9NQ8VoetpFR+5RruMTqhLqkLEksQPPK32N0B58Z1E/Rvr6e0nE8dFFvsJlj+m/44CRKIFf9YCk5nQioqWicdeVZZdAylxEhli818D1AXeYMXjRF0rwBtHgbCdjVqgq/nb1BqYnGNN0n4to0cIUBJZ6B3wsZNs28kFDh3/Eqx5FjCuoyrC7ZxL8ugLkLlNw6cbn3S0q6zTMp3YvKNpvAErL99PlN1GAuloEAIVBjqbIKK+qBMFVJY8gIhNcOzYFrvjUHnhXv2HiQVxMlPw5xyCCorgjPNhZuquBKue1LEf8QlMUlDLILYH/u4H649cnRuo/OK92gVqG57rYO1iAJQZnzdrgbbIWCyTwCS/ZcZy46hhfh9L2/G2G2vbbd2ucfsdaYxNgnBQYRm3Y45eqYNaldrK0oeaJ/B0eyIEsqitEdW5EfGMDsqfrHPq1KvzVlehVTXocVdXBrcVSkFYGviZH9+5LVBSaZeTVx4j6l2LtogAStLy5i+ROORMZO+rJl5uGytixirlzthnM/erytIHmuy0KITn4e95pPlL37PRhiVEm5bruljbSgExSRuCtc/pAP1qHCPg73GEEaK37lyWnrwJFUywNcC8m/xZB+o0wsbaJUtZwjWLJz4ByIhdcYm6Vtb6BeBO3as53BItOyWTeGU1BEKPtitVIVCMIfBJKWpp4sbVDnPc0U6FGvgAsmMcRMMYmwWhgzq9GfvUH+gm3iNY9VRtxWmXRSltgu9fp/82c+O27YaTbNXxGRMq+ujVUpFowW3brXZMFVaQ2Y3GB9myrVCD3HmEcNHne2dVArXLjsxuIn/vT2tBywiB8JO0v/9XpF57EarQiyrn9f3gVOdkkhECFQX41TTrKqpZaIK1i7VlpimUcwFRhEi2NyzqaMI1i7cy8HscKHNOe9cSbVqmPzJf+bMPNln9xjlv5l0YrFmMcLwJnn+BkpGOSwSUcf+LNr2iLXEmQ2A9Ccwk/SBMfbJe3auqAtK2G7gTPWaLRTNJb3qdOUlVStRMs8O90IzwgwxrqwuqnDNCY4QWmoe7+JUu1lTO1f81vuXD3zDme6+6vaOPW01ra8ytw7ZzXD0usyphxzjUGJsvhaeKQtwZ++o/pATXQwVlok2vaJeGXbXwHQAK4caIulbqmh+pdu2+0DoDd+peBK/802S4maigS4EKyvgz9sFtn1UT/mTPaqK+V2spgC2WSUNGiFQbhft/RWLB64nte5Iu8Oa4CD9B2/nfI3Hg6eTu/iHBK/+kWqG+5po10WmRB1EV6qvHkNlNyN61Jv6gCYRBASoM8Dpn62KI1Q+BYP1LKKW2vbXGcVGlLOHG5XhGoAdwp+xRD4YXW6FYT8rCmkI4Lu70vQd8Gq5/CZndiEh1bJuYJBPPE+V7alYYQFuot0Z5s1hGYJIUEAWuR+6u7xOff/yo/Gmzd30P2b8eJ9NJ4cHfkTjiXLxBN+VQlJ64icrSB3HaZlJZ+iClJ24icdhZIzdy4oQbllJ48Hc4mU5k/3qyd32P1rO/prX9YdvFIKyQu+v7Jm8+doxDjFFlNyJS7c0RzGkQwsGbsgegXQUEILMbkPkubQHZlZ+vSoEXI+p7lXD108T2ea2ujO3FiB/wOu3PnWiZQL9nrRDHD3idsXyUEV6MyiuPIQs9k1dgbZemMZ6naoV16mm5d1GE49J71cV0fPB3+HMXNiyyQPyg04kf9AbKixdR/OcNVJY8QNS/HgG6hpMXoxpHM3ECmkI4Hm7nXPOX7k3Utw5Z7Dc1c5rhYeWYStoz62m6zXUUda3YToVWtTUh6lph/jI9bZmGk2rTFgXHb47pq2JSs1ctwdV3bmzv45jyqdu3nfAvqGXDqilpCoTrGw+BZpo0y87C5Cgg1eI2yx+h+6fnkT7lYr0CMfhFpxQoSfGx6yk+ep2pni2QhR56fnY+mdM/iztl7oCHW62dEFSWPkT+7h/WHoDCi9F3zSWEG5cR2/vYYdtFXSvJ3fEdZKEH4ScRiRYKf/tfVLGf5JHnUqt3Maivqpwjf8/lBMsfqeWQt2McYozNkkmmhrbOOG0zB3wq813Icn7i61zsiAiBikJKT99KbN8TaytgySPPI3/vz3QWmQmJzRAQhTjpDl24jOpLV1B66s+1F7BlnFRr+TSipCkUKnVwqZTGDVGgCn06BmpXRCldEb6UpefnF9B67ndIHPpm/Z2MQAbg+sQPPI34gach+9ZRWfIApefvJlj2kK5hIyOIJXW8w0QoI1Ii/DhOy9SBH2c3QlgGP9Ecz1YBKIlb7aeStVoV2o1ye9SPMJbU7EbTR1PDI55GpDpQfeubSIEDbUaS+v2cajMf6T677bO0dXh74ye2dw8sOzGT52ytJCKe0fnAf/m+kVfZlGpw8VAIP4nMbqLvmktGduWpmsXNjVyNk8jd/u0tt/PiehXLrOqKeJrio9dTfOyGEfuJYwpY1drZMQ49xibC+Ac7qfYBH8tCrxasvTjN81LaTiidJaz45C2kT70Et22mdlvr2J3MKRfTf9P/00GxW1uN1/WQfetoOfNLut5EFCBcn2DVUzrry0RWF94VaFxIkBEqLEMYoEwArhAueD4ilkL4SZxYChFPI2JJ/VwJK3hT5pp97YKWEPOcVJUivb/5IMmjLyBz2iV6UQgtRKsoQDgOTttMEke8ncQRb0fme6gsf4TK83dTfvnvRJteQUUhTjxdj9MYs8XQCKRewixU1ZGFnqZ7riqltKVb/4Xuf6SrZ28nF1whHKJ8d/Uv/Y/j4SQyNGWadaUQfsy84xs+Dsv1+hvb8rQL9LlTEoRDsPRBVKU4QB6wWCYKG+1p2flRSpuSY8mBH5fzdQtSc73btz1GSVPZjeTv/B6t530HTA2C9OsuprLiMUpP3qL9vauxRGNCaOWjfwPxg95A+rRLtB++ERJyt39Hv3D9RHMKCs1EVemQEapS0C5swtFB/Z1z8absgTttHu6Uubgdc3BbpyNS7TjJNp2EwYsbt4qBCscum+1G6Zgw4foUHrqK8uI7SBxxDskjz9UBzA1xHyoKEcLBSXeQOOh0EgedjirnqCx9iNLTf6b8wr1EPWsQXqxeGXws94rS7svOoJVnVc6jlGyuLFjG00H/bv4JAwhMFqrt9VCtFOu/G0FaVAsHNtOzXmBqHjk67g1q/c3d9k2Kj92wfSveK6XjUBoWMS2WiWTyFBATaB2bd8zoXXfiaUCggiJu28xRuycBOnOOWfXLnPG5UbknRX3r9M2FQpXzJI86b1TuSZVlD9fck+wYhxlj06zWCcBkvao+5KuEFZrnbdQEyAiRaqfw0NXEF5yqC6FFFXB82t99Ob0ISk/erC1J1WJcW3oxVbOsyADZv574gafR/i8/10JdWAEvRuGB31J+9g5EehsFXO6oGLeWqtLhJFrwdz8Qf48j8Pc8En/3A3E75262mmoZBcYy7aQ6UJUC+Xsup/jgVfhzDydx6JnE938d7tQ9G5QRbRlBCEQ8o++XBacisxsoPf0Xig9dTbDySf0s9GKju64FVN1Fq5ngqk/a5qvFYFJ1D44nVFJb37aXNU0IVLXCfWMfnG1bT2PcmC7KQi9R10pEefvGwwnX3zUto5ZtwuQoIMJBVfL4ex1N50eu22Lwcmyf1yCSrRTu+zkinsFJddDx4Wu3GKAd2/s4vGnz6L36YkTcR4UV2t99+ZYDtPc+Dn/PI+n+8dn6ZV7OkTrpQzpAewvE9zuZrivOJVzxTwA7xqHGeO/Pmi4I3TJaFMKL0/uHT9L5kevw5xwCYQURz9Bx0ZXk//pj8vdcTtS/EeEnEH68Id1mTVzS/8gIFRRRQQkn3UnLm75A5g2f1opgpJWP8ov3kb3x/zNxQ3aVbUjMarIq9IHr4s8+lMTBZxA/4FS83fYbVdpiFZT1uShl9Wq6UWJksQ9/1gId9Nw0iwbbERmBcGuJECpLH6D80n046Sn4exxG4oBTie13It6MfevKCAoVGjetlumkTngvyWMuoPjw78nd+T1k/4axxcUZZaiRppUBN7tmBIMta9sUxTCT1YTXdrWvSuoMbA2IWAr8uL7Gtud71D4TLJPI5KXhjUIyp31KC61hxbwkh0/f2nLapyk9cRNR1ypazvygFszDss5aMUKK2sRhZxF78CrKi+8kfuAbtGAuQ4Z/CAqQAd70vUkddyHZW7+FO2UOLad9mi2nqA3Bi5E57VP0/PwCADvGIcZY/OefoFLYToGIm3WMWvzM4Af5rl7/YyiUAs9HVfL0/OJC2t/7c2J7H6e/E5B+/cdJHP42ig/9nvJzdxJuWKYDmVVDWmghEMJFxNN4M/YlfsDrSB7zLrxp8+rpo90Y5WfvoPfqj+kdV/2OLQNx3FpxxvghZ5I64b3E933tsD72MruRcNMKwg0vIze9Qti9Etm/HpnvRhX7tRISVkzguULlumh99+WkX/P+5nPx2W6o2rNCxDMIISAKqLxwL+XnFuEk2/BmHUh8/5OJH/B6/DmH1pWRKNSXsxcjdcL7iO9/Cn3XfIrKy/eblKojWDJqAmlUs3hUs2DhxZvs3KhB1gbzseM01NzYHrVklAk0Z6B3QBRsp/5sCV33RQZFk4xfZ2gUsZR5VtIE71CLZXKYvDS8joeIJaiblIfxL64+JByvViROBzxtoR3U/Gt1ju+w3q5WXGc43Ho7U0EVxzP9GCktpTbjiliitn87xqHGGNf1R2iWjEZasVJheeCnfpKtygu/syIlwksii330/O97SL/u46Re8z6cRAsqCnA7ZpN54+fInP4Zwg1LCDcuQ/a+iizn9MsznsZp2w1v2jytZDf6NzsuKiiRX/RDcnf/UAsrtu7H5pj7U+a7ie19HC1v+gKx+Sfo7xoEElnoJVj+KJUlDxCseoJw43JkvkvX4lFSZxUzP6Ja2Vs4xnLlQRQ2rORbNkNJowXoGBthYm+CVx6jsuQB8ot+gDfrQBIL30Ly8HNwWqc3tJO4U/ag48PX0Hvlv1JefCciuQWffiFQYYCqFAZ87MQz2y2wezgEQivHjZ+5MR1rt93uZ1VzHdYdMpm5qtXRm+pRb97jQdm8L6nd2066Yzv2y2LZNkziEnA1C8ho120agvVUNPpgMbNiVM3AUatqO+p2mO1H+8AUAx+udoyDmNwCXePCrNSpxuBE0JlmHLfJXkpNgop0prbcJspP/5nkYWdBonWgsOq4eDP3w5u53yh3GVB6+jYK91xOZeUT9axkVvkYiHHLUJUiLWd8lszpnzHKgk4LixBUlj1M8bEbqLx4D1H3apQMdaIFL65d46qZ81AN96My17r5rGoVbLb7tSmpzpn+S8TTiEQGpCRY9SSV5Y+Qv/enpE54P+mTP6IXfMAo8zHa33MFXT94C+H6l/S5GfKa1wKpCkqoUnbANyI9ZYLSYE8gQiBrGaec2iJUTcna5ms7AiUlTmaK+VsfXIUVVLFv21ZmHy2OQJXLyFzXwI9bZ5haKk10vi2WCWaSfVDGYjQWw/w+lrbjbTfWttu6XeP2O9IYmwSBjkco9Az42Em21ixSlkaE1kELvWRO/zQtb/pibSUxf9/PUZU8iQVvwJ02b7PMYoNR5RzhuhcpP38PpWduJVz7nE6LaYsNDo1ZXFBhhbZ3/VAnjFBKu/a4PsHKJ8nd8T+UX7wHFZR1Wt3qynw1dkApbCaxSaZqGQFELI2It6AKfWRv+Srlp2+l7cIfG8VcK3kinqHtHd+k6/JzGFGoFA5EFWT/hgEfu63TmisbkbHeR9WaG47QMVzCxW2bWV8w29YI9PGrfQRUqR9Z6DU1hppNoBcoGRJ1rx7wqds5B9x485xvi2USsE7wll0AUyG3f735S+NkpiBiaVQ5q1eYm2l1cXsiBKrYT8vbvkr6pA8DoIr99P7hk5SevBGER/6uH+B2zMadsgdu+yy96lhd9a0UiXKbkL1rCDetIOpdg6oU9Mp8so1aHJJlCASqlKX1/O9q5SMKtFDqeuTv/jG5O76DCkqIRGt9Jb1BGLZsB6rnwPVwWqYTrH6G7svPofOj1+PtdgBVJcSfdwzxBW+g/MytI7piKRkRdq0gTsOzqm03nFQHMt+17apjj4hObS77XkWVcw11IlzcqXuhpKyl2N52SITj407dy/RQz1/U+yqq2N8kMYmD0XMUbVyi/zKul96UuTipdlQlT7X2l8Wys2EVEMsugF4djnpW1/8GnFQnTmYaYaEHMWIigF0Ix0Xlu0m//uNa+YgClIzo/e2HKD17h64DYlx3wq4VhBuWaOvSoJU6IRz9wvdiuiaCn7CKx5ZwXFShl8QR55A67j0moFnPY/9Nl5Jf9AOc9JS68Gov1+ZCKYgCRKoNmdtE79UXM+UTt+haGSaWL3HImyg/dcsWdxWtf0n/IrRrk5Nqx50yl6jv1eao5m3qBsnsJqKe1Xgz968J/N5uB5j0vNu4j1GESLbiTdtr4Mcbl6AqeZMEoNmePxLheoSvvqD/NElenJYZuFP3IFj5hLl+7M1u2fmY5Ki2Mdw0m91go2w7uN2ob9Txthvc1o5x/MfYViiE4xJtXK7/dJzaiqU7bU+dErZp81xuQ4SDKufx5hxK5k1f1C9r1yd35/coPfMXnDZTCb2aIciPI5KtiHQHTmbqgB+R7kAkW3TMSFXxsO4EIyNDRLKNljd+DlAok/ih+PAfyC/6oVb+TBC0pYmJQkSynWDlUxQf+6O2KJqsUf6sBSZ2ZJhzqBTCixGsfU7fL049O5w/Z6GpPdIkweiOiypldV8b8Hc/ECfVYVLLbqPnqnBQURm3c65OKQ01C0yw8ilULStXk6EUwksQrH9JV7uv3t9CENvzSFMNvUnOt8UywUzela2UcWupugio4X+MoNOYxrMeCD1COyU3FxyrebWHa9MYHF9ta1ZzagGZIx1Pybq7jh3jyGNsFpQCN0a4cVl9VdnUnPBnH4yKtpO/crMhBCqskD7pw1pxEA5R3zoKD16Fk54CYTBw++o5l5FOLzrgJ9Jz3GzXQrNi0u3G9z9Fu5BIqV1civ3k7vweTqLVzueOhJIIz6f8wj0AtSrzTmaKLk4oh3nmKIXwE4QblhB1r6JqvQWI7ftahOMx+mQik49CESx9CDBjVAq3Y3dtEQmKOjZkWyAEKqgQ2+MwbXmVEbguKEnllcfMQkjzzFuNqiWpdx3ByqcAVbMmx/d/nckgaO95y87J5CggJrVp8aHfG/9lU01zqB9HuxiUHv8TUe9aRKqd4uM36oeXG6MmRA/Z1iNc+xyVZQ/htE6nsuyhWpDrsG0Q4MZQQZHi4zciUu1EvWspPf6neprZ4dq6PghHj8tgx7j5GGXfOlMpu0kenEoi/BhR90qi7pUDvorNO0anJN1WL6dmtbSYWgdu6wzi+58MZsU2XLsYle8xNVOa5HzupCglie13MqB0NWmg8tLfiLpWwLa8Ri1bj1m4kf3rqAZsA+YZan4f8lFg2uW7qSx9kMZ7LjbvGNwpe0JQbo7niJIIP0ll6YOooFQX/BHEF7ze1JrZdqv3wnGILzit1jcQhBuXEb76fHMF8A9GOKgooPz8IkDXUALw5x2DN30fnb2xGc63xTLBTE4MiIwQiRaKj16Lk2wldcpHcWo1FxpuJCVRSlJ66hayN39VC4KuT7hmMT2/ej+tZ30Ft+p2MKgdQptW+274vA5wjaV18bRfX0TbO76FP/dQqvnbGxqCUkT96+m/6VLCNYt1hVrHof+mS1EqInHoW/RKzhDtZFCkcM8VFB+9Vq9igR3jUGOMJWk6YdXxkPkeKiv+SXLaXrXu+3MW4k3dS682epOsNBkLQ3Oa1HXfvOnztbWjmlJzxJoxlglDSYQXx5sxHxqqSYdrnpn8gF57fiePqkUa0PdYuV5EdoRHjXBcSs/8heTR79SLOyaTVuKQM8gt+rF+DqsRihpuCxqsNZXlj+gCmWZQiUPfQm7RD0cuvDhRCAGVIt70+cT2eU31QwDKz96BLPTiZKZum76MBxXhxFOUn7sb9aYv6vduFCL8BIkj3kb2lv9CxFObVUu3WHZ0Ji8IXSlELE3+vp9RfOJGhBcfdruob63+3nG1H3SihcqL99L9o7N0rYZhkNmNOv+9nwQZIPwkUc9qen75XpyWacN3rZRFFnv1vmVYW7np/+MXyS/60bAvZBWWkf3rEfGWmqBqxzjMGJvF+tGAEILyC/fo7ELCvNRjKeILTiN39yS/1M2qZub49+r6F9V6Lk2H6VM1G8vsQ3BapiPz3ebFGIzQ1jJujO9/rYhatRBhKWdOyWTcT0bRsed0cpBSL/6YZ41O8NCjiwyOVONBRohYmsrL/yDcuBxv2p611fvkce+h8MBVzSNMCwFRROmx64nve6J59kvcqXuSOPgMCg9fY1JuT2J/HRdZKZA5+p2mgrh530UBpSdubG7rB+h3pVHkyovvIHH429HXhiJ17IUU7v+1zuLl2kyNlp2Lyc2CJQQintGry9UHQH0xqIaTbDfxCdWVIYVItBLlNkHvq2xW0Mj8LeIpRKK1oZ3UVoJSP+G6l4Zth+ebh6KstcP1EVISbni5vv3gvgpHt2ssBGjHOPwYmwkpEbEUlZfvR+Y26RUxUyshedyFFB747eS9JF0Ple8mvt/JZM78orEuOE32UlQI10PmNiJLWZxUG8gIJ91J69lfpff3/4bMd+MkMiZTS7VZwwqvAprN8rUjUS0MqP8A2GKdlXFTiyMLcVpnTs4xdmWEru/gTd8HACUjhOMSblqOKucRqfaRkwm4PirfRfGh39Hylv8EU4TWm7Y3yWMuIH/PFTgtUyd3VXw0BW9lhEhmKD1zO5lNy7WLmFlcSb/u45Se+nPD+2ASng3CQVWKeNP3Jnncewa8s0rPLSJY/fSWq883A0ohPJ/8/b8isfCs2oKh0zKd9CkX0//HL+G0Tt82iwWjLXRssWwlk6eACAcVFHFSnWRecxFOdZW/KpcaYUVJSenpWwnXPKtX65UCFaGCCukT3o87ZU7NHaSGEeAqyx+hvPjOejshUMVe4ge+gdheR7NZALfZJupaReHRa417iav7WuzH2/0g0oecaSqQikF9BVnKUnzkWmShW6+qgB3jsGNsJuEaQIEXI+pZS+nJW0i95v36Yynxpu9D8uh3kr/vZ9qqNJEPecfT8z59Pu3vvnx4C9L2xgTqR72vEqx4nPgBr6OaXCBx2FlMmboHubsuo/LyP5D5HjCZxXA8c41V44qq19VAS8pmF1pjwoQmR8DQWby8GAo1Maq2cHTF5lK//ttMi9ux+0TsfbNjgUIWemg9+2skDj5dL2w0Y6XoSUXUq4tPwvNKOA7x/U8a8Fll+SOjc6lTxo35kT+QPvED9fTXKDKnfYrys3cQ9a9DeEkmvOhkdXEkDIxb6hbmxri35v76E9rO+y4ooZ+rM/cl/bqPkf3zf+v+T4bwLBxUpUDLm76kLcsyqlmc8n/9sX737QjCtJKIeIbKskcoPfVnEoedRTXxS+q1/0r5mdupLHtQK66TpnSa+yEs12NMd4S5s+ywTI4CIgSEZdzWGXR++FrcGfNH3Dx94gfovfKDlF+6D5FoQQUV2t9zBYlDzhy5HZC79evk7roMkZmCynWROe0SMmd+aYtdjO1/Mr1XfRTh+6hSlvh+J9P+vl+M6A4FkDrqfLp/dj4y1wVgxzjkGO81VpsmW3VSChFLUHjgtySPfRfC9WufZ874d8ov3E3Us0bnXZ8Ia4jroQp9uJ2zaf/Ab3HaZiL7N6BkiNs+i6YUvh2H3F2XEd/vRG3ylyEoB3/OQjouupJo43IqrzxCsGYx0cZlyOwGZKFXB0pGgU4TGlVQUVh7gaLqmV2E42ihwPV0Rp9qYoNJEgInBOGgogoqKJuCa/q8uS3Tze9jUEEas9IN+lxFAdGmV2D+8bWP/T2OMK6BEzQ3jgtRiKoUaHvHN0m99gP1fu1K1dNNulNV6Ne1amKpiXteGaHY221/YvuepJ87ro8KK1RevA8RS2x5rs2CgOxfT+6u79P6jm/VFnuczBRaz/suPT+7ANxoYq2pjouqFBGuh9MylahvvenvCM8qGeEk2yg+ci2po9+Jv+eR5rkhSZ/6SSrLHqb8wj3a6jyRSojrI/vXkz7xgyQOP9ucPwXCpfjgVQTLHkakO5vvPTQcSiG8ONm/fIv4Aa/TLq/m2mm74DK6fvRW7Qo7Ue+nRozFRRV6cTvnIAt9WtGxbl+WSWSSsmBp60fmzC9pobUaeDfUT1hBJFpofcc3EMlWVL6H1NHna8E8CoZvFwVacHzTF3R+9OxG/DkLybzpC/qG2ULbxCFnkjr6fFS+B5Fs1cdPtOi828P2tYw7Yz6ZM7+ECkqooGTHONQY4+YB2WwxDsZ9LVzzDMWHrm54cSucdCftF/5ECwpBcaCb0Vgx2ctkrgtvtwN0ReTp80Ep+n7/ccJXnzfdaTKB26zCBcsfpvc3H9ar8Y5Xr0WgJO60vUgedT6tZ3+Vjg9ezZRP3c7Uz/6VqZ/9K1M+fQdTLrmNzk/cQue//YmOD/2e9vf9krZ3/YDWt/8XmdM/Q/KYC4nvfwrezH21Il4pIPPdqHJO96HZVuEVWigr55H57gFfebMPQXhjSJPpuDoJQTT8vRGseBwwwf9K4c86EH+Pw1Dl/NbPjevp/SBpf8/lNeWjutCwyyAEhAEiniZ59DvxZu6vz23Vkrd1O6+5BWXe8BmEnzC1OwTlxXcSrn3OVLAfxTUjQ0SqncKDV1N5+f56bJ2MiO93Em3nfQdZytbiS7au2zoDoir24Wam0HHRlUz597uJH3S6Fka3tH/j0td/wxf089Ok5BWuT/t7rsCfsxCZ22TSym7le0EIcD1k/3oSC99Ky9v+q66AOR5R1wqyt33dxDE22TN2JJR2E47Wv0T2lq/VLJXICHfqnnT865U4iVb9XK4unm0twtFW+lIWooDM6Z9hymfuou3CH+lMfFb5sEwik2MBUQocH3fKXGo1JYZ7gAntiuG0TMdJthP2vIo3bV7dtWgkQVBFIFycjtmol+/H6ZhtHnzRyDeoWZ3xps1DBWXcZDtOdTWzuiI7dGe1EDZlbm08doybj1Ek21G5jSbFcJNhXBtyt/8P8QWn6aJVxvfe3/NIOi66kt7ffAhZ7NO+w9VV/NFgMoupoIQq95I8/Gxaz/22jqkBsjddSvGJm0id8tFJHOBWIiNEso3SM7cSblpO+qQPET/g9QNiZkD7twsA4SDi6Xrw9ChRUYDKdxNuXEbwyj8pv3gfwYp/1pTl5omRUaZCeR/RpuU6S5VJMx6bdzTe9H0IN72iYzWGW2k1rmmq0Is7fW9UOY8q9IHXUDNHSZxYivLLf0eVc6b6cQSOR/oNn6Fyxbm6L2N2ixC1LEoyuwl/1gJaz/8esb2OAqDw4FVUXv4H7f/yU5SSCJpMAZwkVBTQ/s7LiC84FVXK0X/TpRQe+A0iljJBy9HYhS9z/8v+daRf/wkSC9+q7ycTEJ2/+0fjqJOkr52+az7NlEtu1S6ipvZO8tgLwfXpv/5zqFJWW53H8rxq6DNhBZnvJr7Pa2m74DJdiwZo/5cr6Nq4jHDDEoSfGH7fVReiVU/R/8cv0fbO7+s5lBFOZiqdH/4DvVd9lPLzd+OkO8Dxxz7HQoBwUWFZL+Id925az/sfXeNDRuDoZ2/v7/4NWeg1CVh2EOtHFRki0h0U/nEl/pyFJI99l16wkBH+3MPpvPgGeq++mGD107rYo+ONcx61oq0qBVRQIrb3sbS85f8Zt25F4uA30nr2V+m/4QtbjleyWMbJJOYDVdqHdDQrSkLom8wI5KNPVWqEaLPCVDfvjmKFxfhc14IxR1iVHNxuQEE2O8aGdg1jbLYg9CrVwk/5bvr+cIlenazGLMiI2H4n0fmJm/HnHYPMd5kVY0z9FLe+Siqchs+0v6wKijrAvWUqbRdcRvv7/reufPz5v8nf+zOcdAdq0Ep6U+LGCVY8Tu7uHxGs+Kc+nY5HNWe9cH0tTDkuozrX5vpTxk1LOC5O6wxiex9H+vX/RufF1zPlEzeTes1F2hUgKDaRNUSgZGTqMpjMUVLXQMic8VlUJa9f0K5XvzZq14eHCivIXBeJhW9hyqdup/09V+jrrlFoMJlwoo3LKD56XYOrliS+30lk3vQFZP/G+kLASAsIQtSvSxXpOi4yIn3KR+n8xC115ePvv6Lvd/9G1P+qbtlsFsvJwCiTqddepJWPoIxIZGg7/39of88VuC3TkblNqDBouN/raZEH0jjXLiooovJdpE/9BK1v+6pOAKIkOC65RT8kWPFP41YzBgVBSUQsSdS9kt7ffEi7OlbvCxmRPOo8Ov/tRvw9jmh4Xon6YlHj9Tjouqwulsh8FyKRpvWsr9D5sT9q5SMoAxCueU67/YzGFUeGOOkOig9eRe4v36rPnQxxWqbR+ZFraDnzP3TWqnxXvar7qPta1s/XVDtt7/w+be/6kVE+TNYrBX2//wSVpQ81pwvwaFEKkWih/4bPU3rmL/UaTDLCm7WAKZ+4hfQpF4NSeh6rskR1Hp2h5rA+j0QhqtCLKvbhzdiHtgsuY8q/3URsr6Pr8oIMCVY+2ZyFhS07DZOeBWtc2471RbhZwOskH3Nbt2vcfkcaY7NiVvnLL/2d/ms+RduFPwaTZUYHT+7HlH/7P4qPXU/xwd8RrH4aWewD0EWiqsUXldRuVCa2xJuxL4nDziZ17IU4mSm1Y/X/6T/J//0XuC3TiLKbkEUTaNxsMSDV1K/5bmJzDiV96iUkDnmjEWQVpSdvovDAb5C5LtzOOTitM3FbpyGSbYhEC048g0hktEXETyHiKf1ZPA1+AuH6m4txSqKiEOF6eLsfROu53yJ55Dvou/bTetW1GVYxTZ7+0rN3kDnjc7qWj7EUJha+ldZzvkX2z19DFfq0S1bVVSYKUSrCm7In6bM/TuqE9wKK2PwTSL/+4+TvugyRasjQoyQinia36IfED3kzbut0850k84ZP6+/+8m1krgsRS9aq1dcTDCmQej4JyigV4bZMI3HoW0i99gP4cw4xwwnI3fI1nXQh2Yoq9KOiCsKZILeOZkZJiCUJlj1CuHYx3qwDa7EDySPPJX7A6yn8/ZcUH7uecONyrZT5CR2vVE3coXek5zoMtGsq4M1aQOYNn9aWj6olwvUpPXkzuTv/Z/wryeZ5VVlyPz2/voiO9/68nl5dRvhzFzLl4zdRfOwGCg9dTbj6aZ2CvWpdryaIqMZZyUhn5fLieDP2IbHwLSSPvRC3bTfzfQh+nMrSB+j99QeQxX5t4RuN4iQjRLqD3O3fRUUBLW/+TyP0BuD6ZM74dxKHn03h/ispL76dqHt1bUGiroQM0Vc/gTdtHomFbyV1wnuNNd9kjXN9VFCi7/cfp/j4/+k6Rs2Spng8VLMkuh59v/0I6oLvkzz87fq7KNCuzm/7L1LHvIvCQ1dTfm6RnsewXI+xqz0XVO1aVFIiPB+ndSbxg04nsfAs4gtOrVeJjwKEF0Pmuui75hJKT9+mF9CawhJt2RmZXAXEYmlWzGpd4ZFrUDKi7Z3fM7VWoppPdfKo80kedb6uRP/KY4TrXiDqfVXHK5giXE7rDLyZ++LvcQSxOQvNapUm6l1D/7X/Tvm5u3DSnUZZkdrfttkwmWNUUCLzhk+TOe1TtRSwlWUPkbv1G1SWPmBW2TzCdS9oAawaXN6YBUs4WhB3Y7q2hZ9AJFtxUh046Sm4nbPxZuyLP3ch3m4H6BcgUE3X6e91FJ0X/5Hun55PuP5F4460HV+CVevE+pco/ONK0qd8tCZQoSTpkz5IfP7xFB+9lmDNM8hiFuF6uO2ziO13MomFbzUZekLAAUfobHLOoNSkSoEXR/ato/8Pn6Tjg1fX6gYBpE/6MPH9X0fhgd9QeenvRD1ratcioDOSxVO4bTPxdzuA2L4nEj/gFO1maAg3LKX/hs9TefEeRHpqzfVGFbOIqtK8M2MCfYM1z9D1k3NoPevLutgfoMIKTrqTzBmfJXXihyg/v4jy4jsJ1jyLzG7SVrmqBdrMtdcxC2/3Q0gcfAbxg07XWe6q14brUPznn+i/5lMIL7F1/ZYhItVJ+fm76b78XNou/BHeDJPi11gkk0efT/Lo8wnWPEOw9CGC1c8S9a5BlbLaZdLxELEUTut0vJn7EdvrKPy9jq5n5qv2WzgU7v8V/bd8TSsUo1U+an2ViFQ7ubsuI9r0Cq3nfL3m/quiAG/6fFrf/l/IM/6d4JXHCFY+SbjhJWT/RlQlb1IWe4hEBrd1Jt5u++PvcQT+HodrNzBzroQXA9chXPcSfdd+msqyh3d85aNK1dJJRN/VHyNc9xItZ/y7SXMfoZTEm7WA1rd/HfWmLxKsfIJg1VOE619G9m9AVvImBsdDxNI4LdPwpu2Ft/vB+LMP0Smcq8eRoT6W61B69nayN3+VaMMSvYi2vRd/LDs1VgGx7LqYOhfFx24g6nqF1nO+iT/n0Np3SkZ6ZX7WArxZC8a06+I//0juz/+ts2pVX4rGdaIWOCrNCt7WmrirK4HVFJRj9l0XJmZI6cxs1RSQQG7RD8jd/h0tiKQ6qKXONcqJaFwRrv6roJZqulJAlXKo/vW1PirjouckMngz9yNxxDmkjn13XdCRIU7rdNrf9wu6f/BmXT16tMUtTZBufS4mSHGREpFoJXfHd/H3OJzYvGNMogWtuHm7H0jL7l/V20YhuINc06oZZWRE9pb/Iv/XH5lV7EH9kxEi2Ur5+bvp/c0HabvgB7W01ioK8GbsQ+vb/gsVVYg2rUD2r68lTRDxDE5mCm7bTB3o3DgtQVm7xtz5fe1yk55qFG2BqugAeyfVbjbeyjmrzr+M9Bw0mwuHkiYZR5m+33+C8nOLyJzxWbyZ++mvowAn1UbyiHNIHnGOdq3sX4/M96CCEgDCi+NkOnFaZ9aE4mrGN1wfVS6Qu/O75O+5Qi9smAJ9W4UMcVIdBKufpvtHbyVzxudIHf8vtWx+KqwgHBd/94Pxdz94wHiVlGZ1fJArbXWFXDjg+trt8o7vUn7uLh3E7cXH128lcdJTKD15M8Gqp8ic8VmSR55bW2xQYQUn1U58wanEF5zaMMaq29rQfVVRBWEWNlSlQOEfV5Jb9ENUsW+CCx6qgc9mMWixYFtg5kHE0+Tu+C6VJQ/QcuYXie19nI7VMgqdSLQQ2/dEYvueOPp9V5UOs2BUWfYIhft+ql2+HM/GfVi2CZOrgIzlxTPYH3pM7RqFn211zG3dbgcc446AUUKClU/R/ZO3kTz2Ql2bZdq8gXURogCFTrkqqj7hDcF8GkXlpb+Tv/cKys/9Ffy4cbFpeCkKR9fRcNx6kTkvxta83ESipe7nC+YlP5ZzpgWC9vf+nMQhZ2pTvhen/8b/R/7uH+oA9Go151qb0bycjBuA5yIwVg7RUAFBRgSrn6Gy/FGKj1xD+7t+ZBQ9rRB50+eTPP5fyN35PeMKsOXUpcKL6bmtzsUYg+NH2Ln2848q9P7qfbSe9916Cm2BcbeS2hLUWFuieg25HpUlD5D7yzcpL3lAC/vD3SsyQqTaKT71Z8JNr9Dy1kuJ73cSwjXXmgwRjq9XwM0q+LC9LmUpPXsHhb/9gsqKx3ESaR3kX70mhYsqZZH962r70srLOK9HIeppvqtj38rre1IwVk6Raqf05M2UX/obyaPOI3XC+3TGuipKIfwk7pQ9dZG9kTDZr0pP3UL+nst1TaRUB/oGmyhFOEQkMtrl6IbPUXzsBlInfoDEgW/Qro61fkstzAvdr9qzrKps1PosUGFAZckDFB/+PeXFd6KCou53VRnYmr6m2on619P3u3+j+OBVJE94H4kFp+lrsNbXhvkxgeY1BQDq15EQCJOauPTUnyk8cBXhWl13quaSNiGooROuOCYWY1tinhFOegrBK4/SfcV5JA4+g+Rx7yG+zwl16zEMPOeDnV2rFuoqjofsW0f5pfso/vNPOr4tKNfPi1U+LNuASVVAxGhXT5TUL6lq6j4vNkjoHq6d0i97xzcve5PdqXYTDttw4HGE01BwaUsFouSAYnJ2jMOMcUdChohEGqSkcM8VlB69jti+JxE/8DT8uYfhdszWbkRDNFWlLOGGJVSW/IPSs3foNKoy0orHgMrW6FXBeJrK84soPv4n3LaZgED2rNb+9+OyXEgqL/3NCAxaqAq7V41+f46LyveQOf0zA5SP/H0/08pH64yG1eyxohr+UQM+qg0hlkIkWgjXLKb7ZxfQ+fEb8absWbOiJA59M4V7f7bl689UEg67VlBZ+iDVytMyu2l01/yohiMRbgxVKdL7mw+SWHgWqePeg7/nEQgvvvkRBMjsRipLH6L42HWUn7/HrAyPojaBjHBS7YTrXqTn5+8ivv/rSB71DmLzjtXnZPhOIvs3EqxdTOXFeyk9t4ho/cvg+jr70OBrEsD1yN76dTKnfhKRbCNY9WQt7mdMCIEKy5RfvNc8M3QGP9m7xjy/mkwJMVa6anG3wr0/pfjo9cT3O4n4Qafj73E4XsfsLaY8VaUswboXqTz/V12M9dXndW2RyXIHMve5k+ogWPUEfb/9CPkZ+xI74BTi+56IN2uBrlHjDvV6N8ky+l4lWPs8laUPUVlyP+G6FyEKtTA/kZXDZaTfQX6Cyiv/pLLsYXLT5hHb97XE5p+Av/tBuO2zNrPY0VAkV5VzRF2rCFY/Rfmlv1NZ8g9tWfbjps7HENf0uNHKhyr0aqG8eskKUIXe7aOEgH5HxTOAovTETZSevg1/94OI7XcSsb2Pw5u5H07LtHptq0GoKEDlugg3LSdY+QSVpQ8RrHwS2feqVsQTGfATVvGwbFMmrxChDKm8fD/+nkeACocXIKQE1yNY+QSyfx0ikaG85AFSJ38UhGdcKYYQHoyPpCr0Eq55BpHuJFzzDKrQa8yH4dCCsDKrGwjKSx5AJDLI/nUEK58gNv947SrhDCdAKxCuzslublQ7xqHGuEEH6jabwDESxhVGpDtRYUDpiRspPfF/iGQbbttuOG0zcVLttRz+qpxD5rqJ+tYi+9YhKwXtGhBPU82otRnmvMh8D32//WjtHAgvZgrcjWW1USuVKizTc+UHBsy17sco9icEBCXc6fNJv/7jNWEh3LiM3O3frgvKk3kelYRIItIdRN2ryN/xPdou/FEt0N+buhdO6wy9Ql9T2ofaT4SIt1B+8iZK/7xhwBiFn5w4Fwqlr3PhepT++UdKT92CN30+3ox9cTt2r2U5ktlNhJteIVz/ErL3VWMZyKC1klG+5GVUs0aUn7uL8uI7cNpm4k3fW6/It0zXcyJDZCmHzG4g6llD1L0Kmd2g3TP8pH5WVF3ThhiP8BMEq5+h55fvpZo1R6dcHZvlUzgeKt9Dz8/eReNc6+t7O8fxjIRx1xPpToiq9/6NiFQHbsfuuJ1zcVtnIFJtZmFGoSolZG4TUe9aou4VRD1rteVgwHxPYixCVXmK6edN2PUKwT2XU/jbz3HSU3DaZuo+J9tqyqAq55HFPmTfeqLsBlSxrxbgXd3P+BcbRtFX82yMetdSuP9Kiv/4DSLZhtMyzcxvO048Y+KeAmQphyr0EPWtR2Y3IktZhBA6jiXdMfw1vZV9FV6CcP2LdP/k7QO+En5i7PfFhPbNvKPM9RWseZbKin8iXB8n3YnTOgO3ZToi2WrcWZXOzFbsJ8pu1EVj8z0mWN3V12p1HpUcpVXbYpk4JqkOiETEW8gtugx36h4kDjt7+G1dh2DVk/Rd+2n9MEy2UF58J9lbv07LGZ8dfvVJOMj+DfRd+ymi7pU46U6dqvB3H6Pt/O/jtE4fph0QBWRv/w7lxXciUu2ocp6+az9N+7/8FH/OwuH7KhxKT9xIbtFlNXO3HePmY8S4KjWd28VoqAojVX94GRJuegXWv2xeANUxaVcb4frgxXCqLyYltzBspQXYVFvDi2yMbnWDELH0oE9GuT/hIMsF0oefjYinzYspTvGB3yLzvdr1alsFdEYBTrKVyrKHdB0M4zol4mmcdAdR7xoTbzLSuJReeW6w3umPJ1jwNXOrBQFJuGEJ4drndGxL9bsGQV6k2nS7aqXmMR3LCB3GNUIV+6ksfQhevp8BhSyFMMf0agqtqKbf3pKQZtyMqBXI24rrUYiBrkBm/83/LFCb3/tRQLj+JcK1iwckXADMfBt3IS+mV+ON0LdNV5HN86bx+KpSIFz3IuGaxQ3PLOM26phrxPW1cgKjfG5NYF9NYgoAZEjUs4po03KzCFS9Vqr9dY3Cbyx41f1M9hwLZ3P3zWZRoM3YRSxVq5auKgVzrT436D1l3IRdD+H49TawbebRYhmBSSxEqH3Ge6/+GLEHrxpCSKpuK6ksfxRVyeuXoDED5++6jMqL9+G0zTAvsME+jYLw1ReIul5BJNt0PvdkG+XnFtH1w7fg7bb/EC9RBUIg+9YTrHxCv9SjEOEniXpW033F+TpH/jAuRKqS1znGvVjdYqGwYxxqjM3ysB4XjUKEo605fmJzK1WjsDaW1aOxbr/F/Y1zro3bUmz+CYDSblsyorzkH4jYCEXHJgOFtuiUsjrNbzyj56laZ2G00tG2FHargoAR3vXV0RCUX+3LRLzkq/swwpiOpWm8HqsubtVrcowC5UQKoDvLvS+cumIGW5jvCb6nx0rj8YWL8D2IDWVVr/aX7Sd8DpgrRy8YeEM8X6vbVud3W/e32a/jxnu2ds5hc3dTc763laJpsYySyYsBMe4mIu5Tfm6RdvsZCiG072IsDdIUhVMKJz2Fyop/QqXIZtV/zUKoSLXjpDpQUb14jpPqIOpZTbjm2c0X4av7iSVxM9N0O6Bmyo4qlJ64cfjVP9fT/teNQXN2jMOPcafAPLxr/+4saHcLJ9mK0zlb/+0IZL4bmd1oakJs6wEragW1oBZToEo5vdLcrGzLl3qjAGmZZKqKxfbux1jZkfq9sz5ftzU70jm3WDSTp4AYYVmFFV3sZiwr50Ig813E9jhi1NaBappLWejBnbIn3oJTR2cdUAqEq48fS2s3o9FYB2rVaO0Yhx2jpblREtwYjp8c+Nn2OHcCVBRqX+Zqjnp0ELfMbhhdJWaLxWKxWCw7BJMYhK6FmPZ3/2Tk2AEgWPUkvb/9CFHPau2LXuglfdolI8dHQC0+ovzcIpx0JzLfTXzBqSPHR0AtPiK/6Ae1+Ai3Y/aW4yOA0hM30nfNpwYIQ3aMQ4zRKiHNT8MigUbhJNtw0p2E+W4t9G+zvrhQKRHb80iEFzf5/n2dqSXfbfPSWywWi8WyEzE5+VKFgypnyZx6iRZaZVhfWR38E4X4cxbSdv73EI6LKmaJH/gGWs78Uq1w15DtqsXKLvwJbudcZLEft3Mu7Rf+RAvmwx1TRuB6tJz5JeIHvkFXAHZc2s7/nhbMoxH6KkMSh51N5tRLUOU8qpy3YxxijDtsAPouhXaRlMV+ot61+hNTCTm255G64Nq2dHsyKZxTx/+L+UAHoZaevHnb9cFisVgsFss2YXIUEBP/EdvnNdRy8JuKm5v9mAqx/tzDcFpnoko54vOPp5bG0HGHaefVCnZ5ux+Mynfj7X5wfaW0ocrnZsczVZ/j849HlXI4rTPx5x6m+zrc8YSjx6GkHpcp/GbHONQYp0MUDB1UaGkehIMKSjqrUgPJ499rLFimkvVk48aQ2Y2kT/wQ/h6Ho6IQ4fqErz5PefFdpsiYtX5YLBaLxbKzMKkV41RYHjbWYADCgbBSi41QoQm43pLwUw24loHehwyoZc4ZueHA4yipjz/KvqqwXPvTjnFgu9oYLc2PkohYktKTN+maEUbh9WcfTOa0S5DZjeC6k6dICgGuj+xfR+LQN5M584sDFI3sn/8bFRQ3r0hssVgsFotlh2ZyS1aPRXBp3HbM7arbi214zG3dbgcco6W5URIRSxGufpriw3/QCqSSICWZ0z9D+pSLkf0b6ta2ibKGCKH3F4XI7AaSR51H+3t/gfB8kBLheuTv+znlZ+/QSRSs9cNisVgslp2KyVVALBZLc6MkItFC7tavE254WSdEMAX1Ws/5Bm3nfRe8ODLfZRQR17jwjUUZEXU3PcdFhQEyvwmRyNB67ndo/5ef6SJqUQieT+mpW8je/BWrfFgsFovFspOyDdPcWCyWpkMpcHxkKUvvL99Px0euwe2YbWJ4XFKv/VfiB7ye/H0/o/TUn4n6XtUVt/04wo2ZGKDhXAnrBcRUWIGgDELgTplL4rAPknrNRbjts+pxUK5P8dFr6bvus7o42eDaOBaLxWKxWHYKJlcBGYvw0LjtWIWO6vbjbTfWttu6XeP2O9IYLTsGKkLE04Qbl9H9k3Nov/DH+Hsdpb8KK7hT96T1nG+QPu2TVJ6/h/KL9xKseRbZtw5VymoFY4i4HyEcXbk7kcGbMh9/zkLi+59MbL+TcFIdeiOTeUsFJXJ/+Rb5e65AxFM6A5eNJbJYLBaLZadkEhUQAZ4/OiFCKZ2OVjiglC6CNyrhwwi7rm/24Q/8fMSm0hzHVF8ebaEzJfW4qtgxNrRrGKNNw7tjISNEIkPUu5bun55H+nUfI33Sh3UGKnSKXrd1JsljLiB5zAWooEjUtZKoZw2ybx2y0KMDxqMQHA8RS+qigu2zcDvn4HbO1haTKkqaa9Kn/NzdZP/yLYKVjxvFRFnlw2KxWCyWnZhJLEQYEHWthD2PgqjMsAKp1Pn/ZXYDstiL8OOEG5eZjE+Rds8YCpPqFyWRPasR8RSyZ7URXJyR08AqBY5DuHEZwo8ji73I7AbcqXvpLE7OMKExMgIvrsdlfNPtGDcfoyr2mqBlyw6FjBB+ApQke9s3KT1+I8kT3kty4VtxWmfUt1MS4SfxZu6HN3O/cR1KRSGVF++l8I8rKb9wLwiBk54y/L1gsVgsFotlp2GS6oBoASV369eJ1i8BL64F0qF+vBiqlKX/hi+iiv2IdAeFR66l9PSterV/uHauD0KQu+2bBKueRLRMI1j1JLnbvllL7zlS29LTt1J45FpEugNV7NfHL2XBi43Q1zjR+iXkbv06wk8g/IQd41BjLOeN4mStIDscpt6Lk+4k7FpB/x+/yKbvnU7f7z9J6cmbiXpWjy4l81C7LuepLHuY7G3foOv7Z9Dzv/9C+fm/IuJpRCxllQ+LxWKxWHYRJmeZWiktyPavp+vyt5M86jwc48pRD1bVAapKSUpP30a45llTcEwiHJfeqz5K6qjzcafM2bzuhXEpqix/hPLiO3W2nLCCSLaSW/QDglefJ7bX0bWaGwPbCaKuVRQevRbhuPp48QzlF++l68dvI3HImQjHFORr6CqALGUpPnItstCtC7WBHuNP3k7y6PPrYxzQTqGkpPT0rTv2GMd8Hm32oh0XZawhcUQsiSr0Unzk9xQf+QNOy1TcqfPwps/HmzYPp303nMwURDyj3f2EA1KiwjKq2EfUt46oeyXh+pcJ171E1LMKFZS0cptoMdZSe61YLBaLxbIrMXl+MsYKoor95O68DC3hDl4R10KsiCW1MFL1+xYuwvfJ/+PXEAZ6MzWomQIRTyGS7fWVU6UQyXbKzy2i9MTNw7bD83HSndptyPibi2Qr0fqXyd785YYxDOgmCAcn3WmqROu+Cj+JquTJ3fHdev8HtwOcZLtJKxrqL3a0MY7nPFp2bJQCFelAchM0ripFgpWPEyx/GGUU6WqKXVFNz6sUSoY6OF1G+pp1tXVN+Alt7VAmzsMaySwWi8Vi2eWY5CxYElwPJ9PJiEXMlBwotAqBKvXjZqbWgmCHQmY3oip5IyxHIFz9txfXqUSHO1wpq2Mx4i31YNgoQMkAb/o+w8ZVqLCM7F+v29U+VKhyDrdzjk4dOmRDRdS3FhGZWgjmmDvOGMd5Hi07B1VFBLTiHEuDEIiqtltztVNaoRAgiJttoKYVm/oitX1ZLBaLxWLZJZn8SOGxChyOgypmie1/Mq1nfQW3dcbmdQaUBAHByqfou+HzRD2rEbE0qpLH7ZhN2zu+hT/3UCMMNfqra2Ep6l9P/02XUnnhXm2ViAIQDq3nfIPEoW/R6UOHaCeDIoV7riD/t18g4mn9TTlP+qQPkzrlozh+csi+KiVNcbWvancT10cV+3eMMTam/7WCo6WqcIxkuVDV7bZRlywWi8VisexQNFeqIuGgKkW83Q+k46Jf12IQhiO230l0vP9XdP34bIgqiFiKjvf/Cm/WghHbeal2Oi76NV0/eDPh+iUQBbSd+22Sx717xHYu7bSc/VVksZ/io9cBkDzqfFrO/urIwwJSx78XIVz6rv8cSLljjPGRPyCSbdZH32KxWCwWi8UyYUxOFqzxIhxUJU/ysLO1YB5VqK+4DvEjQ7xZC4jNOxbZv4HYvGO1YC7D4dugtCDvJ0kefjaq0IvbPovE4W83KXHl8G2jAJQkeey7al1OHvsuba2IghH6KUFGJA5/O277LFShl+ThO8IYbSVqi8VisVgsFsvE0lwWEKAu9JrieY2ZmjbD2VxArhbdG64+BtQL5VXbCqEF77gJjh2ubdVtSYb1bWRoPh+hXbVtY92OHWmMFovFYrFYLBbLBNFcFpABjFL4HSwkj1poHm+7wW3H0G6zY+woY7RYLBaLxWKxWCaGJlZALBaLxWKxWCwWy86GVUAsFovFYrFYLBbLNqNJFZCxBj6rQf+O5zhjabut2zVuvz3GaLFYLBaLxWKxTAzNqYAId/TZl5TS26PG2Q4TGzHaqagGjlf7aoK9R0VD4PgOMUYbB2KxWCwWi8VimViaMAsWqHLOZG2K0ELzcAK3AsfT21f/FUKnmh1WSDffe9XtHVRQqqe1ldEwwdqmneOiKiWqFb9VpVT/rlrxebNuKl0BXYb6WMLZMcYoQ6wSYrFYLBaLxWKZSJrLAqIinGQrhYd+R7hhKXhxcBwtvA/541F64iYqSx/EaZtJZemDlJ64CRxvhDYOeHHCDUspPPg7nEwnsn892bu+p4Vyd7i2DngxCCvk7vo+wnERjkvuru9DWNHfDddX1wMhyN71PWT/epxMJ4UHm3+MuB7WFctisVgsFovFMpE0lwVEKXB8ZKGHnp+dT+b0z+JOmbt53Qrzd2XpQ+Tv/iHCi4FSCC9G3zWXEG5cRmzvY4dtF3WtJHfHd5CFHoSfRCRaKPztf1HFfpJHnjt0jQ2lUOUc+XsuJ1j+CCKeASBY/gjdPz2P9CkX68+GaIeSFB+7nuKj1yESLYDYccZoLD0Wi8VisVgsFstE0FwKCICSCD+J7N9A3+8/rlfmh0NGWkh2HC0omziJ3O3fHhinMdQxvLiuRG4EbBFPU3z0eoqP3TBCOwWOM0AwF/EMlVceo/LL941cZ0MpRDxdK0Ao/CQyu4m+ay7Zcl9jKXSBxgkY4yPXUHz0uhHb4bhW+bBYLBaLxWKxTArNp4CAFny9GMJPjBxwLQRIOfqg7F0dpRDJti3PqbHaWCwWi8VisVgsE01zKiBghOBoC9s0/F4NzAYyZ3xuVO5JUd86bSFAocp5kkedNyr3pMqyh2suWKqcIzbvmNG7YMXTgEAFRdy2maN2wQJ03IcMxzfGqkIhxzCn2xMlzU+zdMhisVgsFovFMhE0rwIyZgQqLNP+7itIHHbWyJvufRz+nkfS/eOzUZUCqpwjddKHaD37a1s8Sny/k+m64lzCFf8EwN/raDo/cp0O3h6B2D6vQSRbKdz3c0Q8g5PqoOPD1+JN33vkdnsfhzdtHr1XX4yI+6iwQvu7Lx/bGMtZrbzsQMK88JNaEfRiNI9WZLFYLBaLxWLZWporC9Z4cVxjiThWC+Yy1Cv9Q/5ICMt40/cmddyFyFw3TusMWk77tBbQo+Haylq2q8xpn0LJCCUjMqd9qpY5CimHbhvp9Lctp30ap3UGMtdN6rgLtfIRlodvJyOQIYnDziK293HIvnXE9j5ubGM89kJksb9eD6TZUQrhxem/+Sv0/OLdlB67HhFv3bI1zGKxWCwWi8WyQ7DzWECUNIHTpojeSAHauCYoPGOC3hPaQiCEDnofNphcB4CLWKK2fxFLoGt1mDS2Q/bNuEk5nolraejrSO3MuGp9leH4xrhDoeueBCsf18qUG0P48R3KemOxWCwWi8ViGZ6dRwEBvUpeDaLeEkLUV9WVAkYbdC0GBmgryeiL9TXENNT6OppDVvsqxj/GHQqFiKVN/20ciMVisVgsFsvOxM6lgIy5arcY5vdma9e4/daMcQdCSRv6YbFYLBaLxbITsnPEgFgsFovFYrFYLJYdAquAWCwWi8VisVgslm3GzqeAjDpeYNB2Y4ozUMP8vqVmg7cdZdvB7cY7RovFYrFYLBaLZTuzk8WA0FCQcJjYB4EJHPfq2a6UAtcfuR2mnRIDa2o4XkPRvBGyYCmpj1FtJwT1gHY1vK6gpD7GVo3RKiIWi8VisVgsluZg57CAKImIp6ksf5hw7XP1lLpD/SDAjaGCIsXHb0Sk2ol611J6/E/1lLjDtXV9EA7Fh35fO3Txod/rdLiuP3w7xwHHpfT4n4h61yJS7RQfvxEVFMGN6T4N29YjXPsclWUP4bROp7LsobGN8YkbdUYpNdosXxaLxWKxWCwWy+Sxc1hAlALhoipFen59EW3v+Bb+3EP1wv+AWhkKlCLqX0//TZcSrlmMSLaC49B/06UoFZE49C2IzWps6HYyKFK45wqKj16LiKcBKD56LU6yldQpH8Xxk3UFoNZUopSk9NQtZG/+qq5p4fqEaxbT86v303rWV3BbZwzZDgHByqfou+HzqEoBEUujKvlxjLFFFye0WCwWi8VisVi2MzuHAgKmoGCSqGc13f/7HpzMVIZ2PRKoUhZVyiISLbqiuOOCjOj/4xfJL/rRsIUIVVhG9q9HxFtqrlQiliZ/38+0pcGLD9M3RdS3Vn/vuLqgYKKFyov30v2js3Q/hkFmN6JkiPCTIINxjtEqHxaLxWKxWCyW5mACFJAmqjNhlBBQqHzX8Ns5rrZ8VOMoTHyGkJJww8sN+zP/1kpwODjpzoGFAIVAxDNE3avqbk6D2wFOst3EmYTmmAqRaCXKbYLeV03cRkMfzd8inkIkWhvajWeMFovFYrFYLBbLRLF18v9WKiCq+SptV5UA1x9hGwYK5sJFFfvxdj+I9CFnIhwHHZfRsD0gS1mKj1yLLHQbJQBUUMRJdZJ5zUU4VUvGgHYKJSWlp28lXPOstkgoPW8qqJA+4f24U+YYN7JGFywFwqGy/BHKi++stxvvGC0Wi8VisVgslq1Fbb38v3UKiOMj4hlUvnurdjMpjDZVreOiSv3E9zuZ9vf9YkR3KIDUUefT/bPzkTltfXBbZ9D54WtxZ8wfsV36xA/Qe+UHKb90HyLRggoqtL/nChKHnDlyOyB369fJ3XUZItU+UKkYU+pgi8VisVgsFotlK1AK/PgW5eUtMb4sWMK4AsVT+DP2RYWVQYHQOwhC6HiMeJrWd3xDT2ZY0e5OQ/2EZdwZ88mc+SVUUEIFJTJnfkkrH2F5hHYVRKJFHyPZisr3kDr6fK18RMHw7aIAlCLzpi/gzTkEKvkdc54tFovFYrFYLDs2QoCs4LRMx5u2t/lsfHLpGFoN8vUybkD+7INRUbD59zsKMkQk23FaptfrgTje8D9K4k6Za1L2uvr3aq2O4X5M/Q+nZTpOsh0VlPGmzTOZrsTI7ZAgHNz22XqehwmQt1gsFovFYrFYJg3hoIIy/vT5iFiqLscO3GhUuxqdAqIkyGDIpvFDzsRJtpkg6R0RE4QehaMT7oUDYcNchMHotD8h9DHMyRq91cj0aUdW8iwWi8VisVgsOzgCJSMSC99a84YagJL1pElbYGQJ2FgEZHYT4abl9Z2DLq6nJN7M/YjtfTyqnN28YveOxFgsC43bbot249neYrFYLBaLxWKZCISAsITbsTvxg89Ap2p19XcmcZKqFAk2LkF4sS0WwN7yErxwUFGFqGulPlhjrljza8uZXwQvoQOkraBssVgsFovFYrHsPDgestRPy+mfxUl16BpzYmC6WJndgMr3mpCFkRMlDVBARFkKxGC7iTaxVJY+qH8fUNzCARnh7X4QrWd/FZnbaKwgVgmxWCwWi8VisVh2eLw4sn8dqaPfRfK4d2uDg+PWvzfWjmDFE8hC92ZlIrT2oAboHA7okASlEHNa9+5HqWUxV0BVnVERTjxD6fm7kfluHbfQqNWYKuKpE95H+nUfQ/auNYHVLjsUY0lp27jtNmtnU+5aLBaLxWKxWLYRQuhQjL61xPZ5La3nfMPEMg9yoDK6QfGJGxGOR02FUCjXEaIUynwk0ZW+F+gv63s4D0d85b5QwApf1+GrtcZLEG1aTvn5u03QyaDiIyYepPXtX6fl7K+iKgVUOVfLFNX0blnCgVH4qwGmEnm83tSPj7odXqx2krR/3CgUi2qBQsff8rYWi8VisVgsFst4qRoRHA+iAJndQPKYd/H/t3dvMXaVVRzA/+vb+1znPrQ0ZXrJDKUxE6AJVSshFRNj2pAYQnT6YAxgjAgaCaQ2UmMynaQJDTbqC8bIg5akXqZEI/JQNSmhaCPYIhUUMi2dmTqhHZiZdi7nti/f8mHvMz3nzJnOlCI5o/9fMi9z9vWcl732WutbnV87DMm0XtmmLC7FCieGURo6Dkk1V82scw0kVL2EdGECALAvepi90jXee7cALyEwclxE7otyIvEJ1EKSWeSOPY30bfdAkumayd3x1HC1aP7cY0huvAMzzw8gGHsj+jiRhjiJxpthIYjSSnMT8MdOI9nzqZqattrtJZpOfvYEysGAd/YEEhu3ArhKEKIKGAN/7DTs7HuQTCtK7/wV2c88Akg5sKhzTrXRoMTCDIIL/4Ikm+LraLDvkYiIiIhWNlUg8KFBCWoDOB1daLt3ANm7How/r5P9gAXgYu5PPwK8PFA5NFtg066YUoCTXQMX8toHRwQhUDUJ/SULAGLw57xvK1rbEQcgTfDH/oG5owfRcm9/tCxsVY2XxANKQiQ3fxqrHv8DSm8fQ+HUc/DH3oSdeQ9anG28bIgYqF/A5UMPof3+nyDZva0muIqphYYBiqeew+wL+1FOHs2+sB+m+Qakt34R4rgLf5g4y+G9cwKXn30YtjANSTWh+NpvMdO5Hi0790QDEAWoCkLiEq1wYgTTv34cwYW3IOk2wC/+t74JIiIiIvp/pApJpGBaV8O58RZktnweqVt3wDTfEFf6yMJnXBsAxkXpjaPIv/rLuDk9rDymOkZEBMcBzCc7gMon3ni4+ZlHNyVbs+HfM67zsbxvVaSiTEsMtDiLjq8eQurWHXWCkPIJqyMk9YsIp/4NOzMOOEt3xn/k4iBEkk1IbtwKmJpm+zhDoaUcvOFXo/Kpco+LDaGBh2T3JyGpJizMZihgFd7oKaiXgyQy89+PFqaRWL8FpnVN/f0gCMaHEE6ej9Jetqb0jYiIiIjoesQJBEm3wF3VHb0YL6ttOJ//fxR8hJfGMPnDndDiXO0zvgoA18ArqN6+7sDwkPbDyEBUMlT1ml8H4cguhO8+0fNAR8r5+aVCEBqpyISIAUIPkmpGx9d/hcS62+IgpN7KVxqVM0GiHhEiIiIiImpsaudne9StXAoDwHGhhRlM/fRL8EdPQtLVL8qtatiZcZ3JQvBs11PDD2gfHDmC+Q2qAxBA0A85M7Up0Zq1p9OubC74Wp0FMQ7Uy8NkO9D+5R8juXk75oONRVe+0uhGGi3zUUvqpJdqWYuFzeDLCLLKP+aCcy7y487vp8trciciIiIi+qAkrgBa7Lm0/CxrHISTo7h8+Jvwz70CyXbGE9CvbGkE1jUI8sZuWbd/ZEj7IeXsB1Cn83k+C/Kdm+9fnTWH3s8FgRGpHnFuHCAoQcMAzTu/jebPPlpVkrTkDRARERERUeMqr9Za7o2OX9IXX/89Zn7zXdi59yGZtigjUsGqBquzrjuRCw6t/f7wg7XZD2CRiYHa1+eg94heLPQMdmScL0zlg8CYmiBEDACFzU0hdct2NO/YjeSmu2qyICsg60FERERERNVqqoKCC28hd+xpFP42CCSzUU90TX+yVdhsQiSwOlwqudtuahmawj6oSHX5UP0ARKN56Oef2NDeBPdUyjXds54NjWBhjZVx49WtgMS625G54z4kbr4TTucGmGz79d46ERERERF9xLQ4g3B6HP7511B8/Xl4Z09AS7ko61EnyaAK6xggYUTmbHjn+gMjr1Q2nldatEaqvMPonp6PtyXljwLpyPnhwnIsYD5CUi8P9Ysw6RZItg3uqp5oZah6y9oSEREREVEDiVdltSGC8TOwhWnY/CWIcaNneuPUXZVVFaExMM0JI1OF4KH1B0eeKbd11DvLVaOC8o7nd3d/oiVtfpdyZO100QYAHJE6+4qJ/mwQL09bYgM1EREREdGKIhA3BTjxVHRgsXEQaq2G2aRxBQjm/PAbXU+NPKP9cGUAQb0doqMvodw4MvLYhrXNGfdnrSlnx0zJwg9tAEj9QARgIzoRERER0UqlNl74tW4/t1pVa0SczoyDOc+eKXj6la6D5/6yVPABLCMAAYDBPji74u71ib093zMi32pKmBtnPQsvtKGJ2kYMJOqRv9b7IyIiIiKixqSAikbDIQBVI+K2pg1ynvVE5PC4LezpPfDu5NXKriotO1hQjbYVgV7c270mBfOICh7OJswahaIUKHyrCMKFjSZERERERLQyOQbGNYKUK3CNYM4LS4D8wlP9wdonz70JXOkfX87xrjlbURnZXNzbvSZjnHtC1e0K3WYtbko40u4Y4eq7REREREQrnAhQ8G1eRCaM2pOukZc9DY+ufnL0bSCKDbALVhap1ap7zA9yIaoQHIGpTLHoIBz8c3PHhaLXmzCOEwQAFq6XRUREREREK0EApF0gZXA201makN1jhfJH2gcHvdDlZj0+NKqQF/vh6mCd+SBERERERPQ/Q/vgaP/drvbDLL314j60hnEFohmHu67vgoiIiIiIqHHs64XuG4BeS5kVERERERFRQ/gPTER+G/NkbG0AAAAASUVORK5CYII=";

let config = {
  evento: { nome: "", categoria: "", data: "", logo_empresa: DS_LOGO_DEFAULT, logo_evento: "" },
  workouts: [],
  atletas: []   // atletas da categoria atual (pode estar vazio)
};
let editingIdx = -1;   // -1 = new workout
let previewIdx  = -1;

// ═══════════════════════════════════════════════════════════════════
//  EVENTO
// ═══════════════════════════════════════════════════════════════════
function toggleEventoForm() {
  const form = document.getElementById('eventoForm');
  const disp = document.getElementById('eventoDisplay');
  const btn  = document.getElementById('btnToggleEvento');
  const open = form.style.display === 'none';
  form.style.display = open ? '' : 'none';
  btn.textContent    = open ? 'Fechar' : 'Editar';
  if (open) {
    document.getElementById('evNome').value = config.evento.nome || '';
    document.getElementById('evCat').value  = config.evento.categoria || '';
    document.getElementById('evData').value = config.evento.data || '';
    // Mostra preview da logo empresa se já estiver carregada (padrão DS)
    const empImg = document.getElementById('logoEmpresaPreview');
    const empPh  = document.getElementById('logoEmpresaPlaceholder');
    if (config.evento.logo_empresa) {
      empImg.src = config.evento.logo_empresa;
      empImg.style.display = '';
      empPh.style.display  = 'none';
    }
  }
}

function onEventoChange() {
  config.evento.nome      = document.getElementById('evNome').value.trim();
  config.evento.categoria = document.getElementById('evCat').value.trim();
  config.evento.data      = document.getElementById('evData').value.trim();
  renderEventoDisplay();
  atualizarBotaoGerar();
  refreshPreview();
}

function onLogoEvento(input) {
  if (!input.files || !input.files[0]) return;
  const reader = new FileReader();
  reader.onload = e => {
    config.evento.logo_evento = e.target.result;
    const img = document.getElementById('logoEventoPreview');
    const ph  = document.getElementById('logoEventoPlaceholder');
    img.src = e.target.result; img.style.display = '';
    ph.style.display = 'none';
    refreshPreview();
  };
  reader.readAsDataURL(input.files[0]);
}

function onLogoEmpresa(input) {
  if (!input.files || !input.files[0]) return;
  const reader = new FileReader();
  reader.onload = e => {
    config.evento.logo_empresa = e.target.result;
    const img = document.getElementById('logoEmpresaPreview');
    const ph  = document.getElementById('logoEmpresaPlaceholder');
    img.src = e.target.result; img.style.display = '';
    ph.style.display = 'none';
    refreshPreview();
  };
  reader.readAsDataURL(input.files[0]);
}

function refreshPreview() {
  if (previewIdx >= 0 && previewIdx < config.workouts.length) {
    previewWorkout(previewIdx);
  }
}

function renderEventoDisplay() {
  const d = document.getElementById('eventoDisplay');
  if (config.evento.nome) {
    d.innerHTML = `<div class="ev-nome">${esc(config.evento.nome)}</div>
      <div class="ev-meta">${esc(config.evento.categoria)}${config.evento.data ? ' · ' + esc(config.evento.data) : ''}</div>`;
  } else {
    d.innerHTML = '<div class="ev-empty">Clique para configurar o evento</div>';
  }
}

// ═══════════════════════════════════════════════════════════════════
//  WORKOUT LIST
// ═══════════════════════════════════════════════════════════════════
const TIPO_LABEL = { for_time: 'For Time', amrap: 'AMRAP', express: 'Express' };

function renderWorkoutList() {
  const el = document.getElementById('workoutList');
  if (!config.workouts.length) {
    el.innerHTML = '<div class="wkt-empty">Nenhum workout ainda.<br>Clique em "+ Novo" para começar.</div>';
    return;
  }
  el.innerHTML = config.workouts.map((w, i) => `
    <div class="wkt-card${previewIdx === i ? ' active' : ''}" id="wcard${i}" onclick="selectWorkout(${i})">
      <div class="wkt-num">${w.numero}</div>
      <div class="wkt-info">
        <div class="wkt-name">${esc(w.nome)}</div>
        <div class="wkt-tags">
          <span class="tag ${w.tipo}">${TIPO_LABEL[w.tipo] || w.tipo}</span>
          ${w.time_cap ? `<span class="tag">${esc(w.time_cap)}</span>` : ''}
        </div>
      </div>
      <div class="wkt-actions">
        <button class="icon-btn" onclick="event.stopPropagation();editarWorkout(${i})" title="Editar">✎</button>
        <button class="icon-btn danger" onclick="event.stopPropagation();deletarWorkout(${i})" title="Excluir">×</button>
      </div>
    </div>`).join('');
}

function selectWorkout(idx) {
  previewIdx = idx;
  renderWorkoutList();
  previewWorkout(idx);
}

function atualizarBotaoGerar() {
  const btn = document.getElementById('btnGerar');
  const lbl = document.getElementById('btnGerarLabel');
  const nWkt = config.workouts.length;
  const nAtl = config.atletas.length;
  btn.disabled = nWkt === 0;
  if (nWkt === 0) {
    lbl.innerHTML = '&#x2B07;&nbsp;&nbsp;Gerar Súmulas (ZIP)';
  } else {
    const cat = config.evento.categoria || config.evento.nome || '';
    const catLabel = cat ? esc(cat) + ' — ' : '';
    if (nAtl > 0) {
      const total = nWkt * nAtl;
      lbl.innerHTML = `&#x2B07;&nbsp;&nbsp;Gerar ${catLabel}${total} súmulas (${nAtl} atletas × ${nWkt} WKTs)`;
    } else {
      lbl.innerHTML = `&#x2B07;&nbsp;&nbsp;Gerar ${catLabel}${nWkt} súmula${nWkt !== 1 ? 's' : ''} (sem atletas)`;
    }
  }
}

// ═══════════════════════════════════════════════════════════════════
//  EDITOR
// ═══════════════════════════════════════════════════════════════════
function novoWorkout() {
  editingIdx = -1;
  const n = config.workouts.length + 1;
  document.getElementById('edTitle').textContent = 'Novo Workout';
  document.getElementById('edNome').value = '';
  document.getElementById('edTipo').value = 'for_time';
  document.getElementById('edTimeCap').value = '';
  document.getElementById('edDescricao').value = '';
  document.getElementById('edF1Janela').value = '';
  document.getElementById('edF2Janela').value = '';
  setMovTableFromArray('main', []);
  setMovTableFromArray('f1', []);
  setMovTableFromArray('f2', []);
  onTipoChange();
  abrirEditor();
}

function editarWorkout(idx) {
  editingIdx = idx;
  const w = config.workouts[idx];
  document.getElementById('edTitle').textContent = `Workout ${w.numero} — ${w.nome}`;
  document.getElementById('edNome').value = w.nome || '';
  document.getElementById('edTipo').value = w.tipo || 'for_time';
  document.getElementById('edTimeCap').value = w.time_cap || '';
  document.getElementById('edDescricao').value = (w.descricao || []).join('\n');
  if (w.tipo === 'express') {
    document.getElementById('edF1Janela').value = (w.formula1 || {}).janela || '';
    document.getElementById('edF2Janela').value = (w.formula2 || {}).janela || '';
    setMovTableFromArray('f1', (w.formula1 || {}).movimentos || []);
    setMovTableFromArray('f2', (w.formula2 || {}).movimentos || []);
  } else {
    setMovTableFromArray('main', w.movimentos || []);
  }
  onTipoChange();
  abrirEditor();
}

function abrirEditor() {
  document.getElementById('editor').classList.add('open');
}

function fecharEditor() {
  document.getElementById('editor').classList.remove('open');
  editingIdx = -1;
}

function onTipoChange() {
  const t = document.getElementById('edTipo').value;
  document.getElementById('secMovimentos').style.display = t !== 'express' ? '' : 'none';
  document.getElementById('secExpress').style.display    = t === 'express' ? '' : 'none';
  document.getElementById('btnChegadaMain').style.display = t === 'amrap' ? 'none' : '';
}

function salvarWorkout() {
  const nome = document.getElementById('edNome').value.trim().toUpperCase();
  if (!nome) { toast('Digite o nome do workout', 'err'); return; }
  const tipo = document.getElementById('edTipo').value;
  const timeCap = document.getElementById('edTimeCap').value.trim();
  const desc = document.getElementById('edDescricao').value.split('\n').map(s=>s.trim()).filter(Boolean);

  let wkt;
  if (editingIdx >= 0) {
    wkt = config.workouts[editingIdx];
  } else {
    wkt = { numero: config.workouts.length + 1, modalidade: 'individual' };
    config.workouts.push(wkt);
    editingIdx = config.workouts.length - 1;
  }

  wkt.nome     = nome;
  wkt.tipo     = tipo;
  wkt.estilo   = tipo;
  wkt.time_cap = timeCap;
  wkt.descricao = desc;

  if (tipo === 'express') {
    wkt.formula1 = {
      janela: document.getElementById('edF1Janela').value.trim(),
      descricao: [],
      movimentos: getMovTableArray('f1')
    };
    wkt.formula2 = {
      janela: document.getElementById('edF2Janela').value.trim(),
      descricao: [],
      movimentos: getMovTableArray('f2')
    };
    delete wkt.movimentos;
  } else {
    wkt.movimentos = getMovTableArray('main');
    delete wkt.formula1;
    delete wkt.formula2;
  }

  previewIdx = editingIdx;
  fecharEditor();
  renderWorkoutList();
  atualizarBotaoGerar();
  previewWorkout(previewIdx);
  toast('Workout salvo!', 'ok');
}

function deletarWorkout(idx) {
  if (!confirm(`Excluir workout "${config.workouts[idx].nome}"?`)) return;
  config.workouts.splice(idx, 1);
  // Renumber
  config.workouts.forEach((w, i) => w.numero = i + 1);
  if (previewIdx >= config.workouts.length) previewIdx = config.workouts.length - 1;
  renderWorkoutList();
  atualizarBotaoGerar();
  if (previewIdx >= 0) previewWorkout(previewIdx);
  else {
    document.getElementById('previewEmpty').style.display = '';
    document.getElementById('previewFrame').style.display = 'none';
    document.getElementById('pbName').textContent = '—';
  }
}

// ═══════════════════════════════════════════════════════════════════
//  MOVEMENTS TABLE
// ═══════════════════════════════════════════════════════════════════
// section: 'main' | 'f1' | 'f2'
const bodyId = s => s === 'main' ? 'movTableBody' : `movTable${s.toUpperCase()}Body`;

function setMovTableFromArray(section, movs) {
  const body = document.getElementById(bodyId(section));
  body.innerHTML = '';
  if (!movs.length) {
    body.innerHTML = '<div class="empty-table">Adicione movimentos abaixo</div>';
    return;
  }
  movs.forEach(m => appendMovRow(section, m));
}

function getMovTableArray(section) {
  const body = document.getElementById(bodyId(section));
  const rows = body.querySelectorAll('.mov-row');
  const arr = [];
  rows.forEach(row => {
    const t = row.dataset.type;
    if (t === 'sep') {
      arr.push({ separador: row.querySelector('.mi-sep-input').value.trim() || 'then...' });
    } else if (t === 'chegada') {
      arr.push({ chegada: true });
    } else {
      const nome = row.querySelector('.mi-nome').value.trim().toUpperCase();
      if (!nome) return;
      const mov = { nome };
      const repsEl = row.querySelector('.mi-reps');
      const reps = parseInt(repsEl.value);
      if (!isNaN(reps) && reps > 0) mov.reps = reps;
      else if (repsEl.value.trim()) mov.reps = repsEl.value.trim();
      const label = row.querySelector('.mi-label').value.trim();
      if (label) mov.label = label;
      arr.push(mov);
    }
  });
  return arr;
}

function appendMovRow(section, mov) {
  const body = document.getElementById(bodyId(section));
  // Remove empty placeholder
  const empty = body.querySelector('.empty-table');
  if (empty) empty.remove();

  const row = document.createElement('div');

  if (mov.chegada) {
    row.className = 'mov-row chegada-row';
    row.dataset.type = 'chegada';
    row.innerHTML = `<div class="mi-chegada">✓ Chegada / Finish</div>
      <div class="mi-ctrl">${ctrlBtns(section)}</div>`;
  } else if (mov.separador !== undefined) {
    row.className = 'mov-row sep-row';
    row.dataset.type = 'sep';
    row.innerHTML = `<input class="mi-sep-input" value="${esc(mov.separador || 'then...')}"
        placeholder="then..." style="flex:1;font-style:italic;color:var(--text3)">
      <div class="mi-ctrl">${ctrlBtns(section)}</div>`;
  } else {
    row.className = 'mov-row';
    row.dataset.type = 'normal';
    row.innerHTML = `
      <input class="mi-nome" value="${esc(mov.nome || '')}" placeholder="Nome do movimento">
      <input class="mi-reps" type="number" min="1" value="${mov.reps || ''}" placeholder="—" style="width:52px;text-align:center">
      <input class="mi-label" value="${esc(mov.label || '')}" placeholder="Label" style="width:72px;font-size:10.5px">
      <div class="mi-ctrl">${ctrlBtns(section)}</div>`;
  }

  body.appendChild(row);
}

function ctrlBtns(section) {
  return `<button class="icon-btn" onclick="movUp(this)" title="Subir">↑</button>
    <button class="icon-btn danger" onclick="removeRow(this)" title="Remover">×</button>`;
}

function addMov(section) {
  appendMovRow(section, { nome: '', reps: '' });
  // Focus first input
  const body = document.getElementById(bodyId(section));
  const last = body.lastElementChild;
  if (last) { const inp = last.querySelector('input'); if (inp) inp.focus(); }
}

function addSep(section) { appendMovRow(section, { separador: 'then...' }); }

function addChegada(section) {
  // Only one chegada allowed
  const body = document.getElementById(bodyId(section));
  if (body.querySelector('.chegada-row')) { toast('Chegada já adicionada', 'err'); return; }
  appendMovRow(section, { chegada: true });
}

function removeRow(btn) {
  const row = btn.closest('.mov-row');
  const body = row.parentElement;
  row.remove();
  if (!body.querySelector('.mov-row')) {
    body.innerHTML = '<div class="empty-table">Adicione movimentos abaixo</div>';
  }
}

function movUp(btn) {
  const row = btn.closest('.mov-row');
  const prev = row.previousElementSibling;
  if (prev && prev.classList.contains('mov-row')) {
    row.parentElement.insertBefore(row, prev);
  }
}

// ═══════════════════════════════════════════════════════════════════
//  PREVIEW
// ═══════════════════════════════════════════════════════════════════
function previewWorkout(idx) {
  const wkt = config.workouts[idx];
  if (!wkt) return;
  document.getElementById('pbName').textContent = `${wkt.numero} — ${wkt.nome}`;
  document.getElementById('previewEmpty').style.display = 'none';
  // Show loading state
  const frame = document.getElementById('previewFrame');
  frame.style.display = 'block';

  fetch('/api/preview', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ config, workout_index: idx })
  })
  .then(r => { if (!r.ok) throw new Error('Erro ' + r.status); return r.text(); })
  .then(html => {
    const blob = new Blob([html], { type: 'text/html' });
    const old = frame.src;
    frame.src = URL.createObjectURL(blob);
    if (old && old.startsWith('blob:')) URL.revokeObjectURL(old);
  })
  .catch(e => { toast('Erro no preview: ' + e.message, 'err'); });
}

// ═══════════════════════════════════════════════════════════════════
//  GENERATE ZIP
// ═══════════════════════════════════════════════════════════════════
function gerarZIP() {
  if (!config.workouts.length) return;
  const btn = document.getElementById('btnGerar');
  const lbl = document.getElementById('btnGerarLabel');
  btn.disabled = true;
  btn.classList.add('generating');
  lbl.textContent = 'Gerando…';

  fetch('/api/generate', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ config })
  })
  .then(r => { if (!r.ok) throw new Error('Falha na geração'); return r.blob(); })
  .then(blob => {
    const url = URL.createObjectURL(blob);
    const a   = document.createElement('a');
    a.href = url;
    const cat = (config.evento.categoria || config.evento.nome || 'sumulas').replace(/\s+/g, '_');
    a.download = `${cat}.zip`;
    a.click();
    URL.revokeObjectURL(url);
    const n = config.atletas.length
      ? config.workouts.length * config.atletas.length
      : config.workouts.length;
    toast(`${n} súmula(s) gerada(s) com sucesso!`, 'ok');
  })
  .catch(e => toast('Erro: ' + e.message, 'err'))
  .finally(() => {
    btn.disabled = false;
    btn.classList.remove('generating');
    atualizarBotaoGerar();
  });
}

// ═══════════════════════════════════════════════════════════════════
//  IMPORT
// ═══════════════════════════════════════════════════════════════════
function triggerImport(type) {
  document.getElementById(type === 'excel' ? 'fileExcel' : 'filePDF').click();
}

function handleImport(input, type) {
  const file = input.files[0];
  if (!file) return;
  input.value = '';
  toast('Importando…', 'info');

  const reader = new FileReader();
  reader.onload = e => {
    const b64 = e.target.result.split(',')[1];
    fetch('/api/import/' + type, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ data: b64, filename: file.name })
    })
    .then(r => r.json())
    .then(result => {
      if (result.error) throw new Error(result.error);

      // ── Formato grade de categorias (Excel do evento real) ──
      if (result.tipo === 'categoria_grid') {
        mostrarSeletorCategoria(result); return;
      }

      // ── Formato simples ──
      aplicarImport(result);
    })
    .catch(e => toast('Erro ao importar: ' + e.message, 'err'));
  };
  reader.readAsDataURL(file);
}

function mostrarSeletorCategoria(data) {
  const cats = data.categorias || Object.keys(data.por_categoria || {});
  if (!cats.length) { toast('Nenhuma categoria encontrada no arquivo', 'err'); return; }

  const sub = document.getElementById('catModalSub');
  sub.textContent = `${cats.length} categorias encontradas. Escolha qual gerar as súmulas:`;

  const grid = document.getElementById('catGrid');
  grid.innerHTML = '';
  cats.forEach(cat => {
    const btn = document.createElement('button');
    btn.className = 'cat-btn';
    const wkts = (data.por_categoria[cat] || []).length;
    btn.innerHTML = `<strong>${esc(cat)}</strong><br><span style="font-size:9px;opacity:.6">${wkts} workout(s)</span>`;
    btn.onclick = () => {
      document.getElementById('catModal').style.display = 'none';
      const workouts = data.por_categoria[cat] || [];
      const atletasCat = (data.atletas_por_categoria || {})[cat] || [];
      config.evento.nome      = data.evento_nome || config.evento.nome || 'Sun2026';
      config.evento.categoria = cat;
      document.getElementById('evNome').value = config.evento.nome;
      document.getElementById('evCat').value  = cat;
      renderEventoDisplay();
      config.workouts = workouts;
      config.atletas  = atletasCat;
      previewIdx = 0;
      renderWorkoutList();
      atualizarBotaoGerar();
      if (workouts.length) previewWorkout(0);
      const msgAtletas = atletasCat.length ? ` · ${atletasCat.length} atleta(s)` : '';
      toast(`${cat} — ${workouts.length} workout(s)${msgAtletas} importado(s)`, 'ok');
    };
    grid.appendChild(btn);
  });
  document.getElementById('catModal').style.display = '';
}

function aplicarImport(result) {
  if (result.evento && result.evento.nome) {
    config.evento = { ...config.evento, ...result.evento };
    document.getElementById('evNome').value = config.evento.nome || '';
    document.getElementById('evCat').value  = config.evento.categoria || '';
    document.getElementById('evData').value = config.evento.data || '';
    renderEventoDisplay();
  }
  if (result.workouts && result.workouts.length) {
    config.workouts = result.workouts;
    previewIdx = 0;
    renderWorkoutList();
    atualizarBotaoGerar();
    previewWorkout(0);
    toast(`${result.workouts.length} workout(s) importado(s)`, 'ok');
  } else {
    toast('Nenhum workout encontrado no arquivo', 'err');
  }
}

// ═══════════════════════════════════════════════════════════════════
//  HELPERS
// ═══════════════════════════════════════════════════════════════════
function esc(s) {
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function toast(msg, type = 'ok') {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className = `toast show ${type}`;
  clearTimeout(el._t);
  el._t = setTimeout(() => el.classList.remove('show'), 3000);
}
</script>
</body>
</html>"""


# ── HTTP Handler ────────────────────────────────────────────────────────────────
class SumulaHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass  # silencia log

    def do_GET(self):
        if self.path in ('/', '/index.html'):
            self._send(200, 'text/html; charset=utf-8', HTML_INTERFACE.encode('utf-8'))
        else:
            self._send(404, 'text/plain', b'Not found')

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))
            routes = {
                '/api/preview':        self._handle_preview,
                '/api/generate':       self._handle_generate,
                '/api/import/excel':   self._handle_import_excel,
                '/api/import/pdf':     self._handle_import_pdf,
            }
            handler = routes.get(self.path)
            if handler: handler(body)
            else: self._send(404, 'text/plain', b'Rota nao encontrada')
        except Exception as e:
            import traceback; traceback.print_exc()
            self._send(500, 'application/json',
                       json.dumps({"error": str(e)}).encode('utf-8'))

    def _handle_preview(self, body):
        cfg      = body['config']
        idx      = int(body['workout_index'])
        ev       = cfg.get('evento', {})
        wkt      = cfg['workouts'][idx]
        logo     = _resolve_logo(ev.get('logo_empresa', ''))
        logo_evt = ev.get('logo_evento', '')   # data-URL vinda do front
        html = render_workout(ev, wkt, FONTS, logo, logo_evt)
        self._send(200, 'text/html; charset=utf-8', html.encode('utf-8'))

    def _handle_generate(self, body):
        cfg      = body['config']
        ev       = cfg.get('evento', {})
        logo     = _resolve_logo(ev.get('logo_empresa', ''))
        logo_evt = ev.get('logo_evento', '')
        atletas  = cfg.get('atletas', [])   # lista de atletas da categoria atual

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            if atletas:
                # Uma súmula por atleta × por workout
                # Organização: WKT01_TWENTIES / 001_Joao_Silva.html
                for wkt in cfg['workouts']:
                    num_w = wkt.get('numero', 1)
                    nome_w = wkt.get('nome', 'wkt')
                    pasta = f"WKT{num_w:02d}_{sanitize(nome_w)}"
                    for atleta in atletas:
                        nome_a = atleta.get('nome', 'atleta')
                        num_a  = atleta.get('numero', '')
                        prefixo = f"{sanitize(num_a)}_" if num_a else ""
                        html = render_workout(ev, wkt, FONTS, logo, logo_evt, atleta)
                        zf.writestr(f"{pasta}/{prefixo}{sanitize(nome_a)}.html",
                                    html.encode('utf-8'))
            else:
                # Sem atletas: um modelo em branco por workout
                for wkt in cfg['workouts']:
                    num  = wkt.get('numero', 1)
                    nome = wkt.get('nome', 'wkt')
                    html = render_workout(ev, wkt, FONTS, logo, logo_evt)
                    zf.writestr(f"{num:02d}_{sanitize(nome)}.html", html.encode('utf-8'))

        cat = sanitize(ev.get('categoria', '') or ev.get('nome', 'sumulas'))
        self._send(200, 'application/zip', buf.getvalue(),
                   {'Content-Disposition': f'attachment; filename="{cat}.zip"'})

    def _handle_import_excel(self, body):
        try:
            data   = base64.b64decode(body['data'])
            result = parse_excel(data)
            self._send(200, 'application/json; charset=utf-8',
                       json.dumps(result, ensure_ascii=False).encode('utf-8'))
        except Exception as e:
            self._send(200, 'application/json',
                       json.dumps({"error": str(e)}).encode('utf-8'))

    def _handle_import_pdf(self, body):
        try:
            data   = base64.b64decode(body['data'])
            result = parse_pdf(data)
            self._send(200, 'application/json; charset=utf-8',
                       json.dumps(result, ensure_ascii=False).encode('utf-8'))
        except Exception as e:
            self._send(200, 'application/json',
                       json.dumps({"error": str(e)}).encode('utf-8'))

    def _send(self, code, content_type, data, extra=None):
        self.send_response(code)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Access-Control-Allow-Origin', '*')
        if extra:
            for k, v in extra.items(): self.send_header(k, v)
        self.end_headers()
        self.wfile.write(data)


# ── Startup ──────────────────────────────────────────────────────────────────────
def main():
    try:
        server = HTTPServer((HOST, PORT), SumulaHandler)
    except OSError:
        print(f"⚠  Porta {PORT} em uso.")
        sys.exit(1)

    if IS_CLOUD:
        print(f"✓ Servidor em: http://0.0.0.0:{PORT}")
        print("  Sumulas Digital Score v1.0.0 online \u2014 pronto para receber conexoes\n")
    else:
        url = f'http://localhost:{PORT}'
        print(f"✓ Servidor em: {url}")
        print("  Pressione Ctrl+C para encerrar\n")
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n✓ Encerrado.")


if __name__ == '__main__':
    main()
