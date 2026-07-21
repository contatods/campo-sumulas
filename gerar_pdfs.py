#!/usr/bin/env python3
"""Conversor de súmulas HTML → PDF, organizado por bateria.

Camada PÓS-geração: consome os HTMLs que o app de súmulas já produz e NÃO
altera nada no gerador. A conversão usa o Chrome em modo headless — o
mesmo motor de renderização do Ctrl+P — então o PDF sai idêntico ao que
o navegador mostra (fontes, logo e quebras de página preservados).

Uso CLI:
    python3 gerar_pdfs.py CAMINHO_DO_ZIP [--excel prog.xlsx | --json backup.json]
                          [--saida PASTA]

Uso como módulo (sumula_app.py):
    from gerar_pdfs import achar_chrome, converter, horarios_do_config
    converter(pasta_com_htmls, pasta_saida, horarios, chrome)

Entradas:
    CAMINHO_DO_ZIP   ZIP gerado pelo app (ou pasta já descompactada)
    --excel          Excel de programação do organizador (mesmos formatos que
                     o app importa) — fornece horários das baterias.
    --json           backup JSON exportado pelo app ("Exportar JSON") — idem.
                     Sem horários, a ordem do dia completo cai pra
                     Categoria → Workout → Bateria.
    --saida          pasta de destino (default: <nome-do-zip>_PDFs ao lado)

Saída:
    <saida>/<Dia>/<Categoria>/<NN_Workout>/Bateria_03.pdf   (uma por bateria)
    <saida>/<Dia>/00_DIA_COMPLETO.pdf  (tudo, ordem horário → bateria → raia)

Requisitos: Google Chrome instalado. Sem dependências Python externas
(--excel usa o openpyxl que o app já tem, via parsers.py).
"""

import argparse
import concurrent.futures
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

# Cópia de campo_generator.sanitize() — mantida igual de propósito pra que
# os nomes de pasta do config/Excel batam com os do ZIP sem importar o gerador.
def sanitize(n):
    if not n:
        return ""
    s = re.sub(r'[\/\\:*?"<>|\r\n\t]+', "_", str(n))
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s)
    return s.strip("_") or "_"


# Navegadores aceitos, por plataforma. Qualquer Chromium serve — no Windows
# o Edge (nativo do sistema) imprime PDF idêntico ao Chrome.
if sys.platform == "win32":
    CHROME_CANDIDATOS = [os.path.expandvars(p) for p in (
        r"%ProgramFiles%\Google\Chrome\Application\chrome.exe",
        r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe",
        r"%LocalAppData%\Google\Chrome\Application\chrome.exe",
        r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe",
        r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe",
    )] + ["chrome", "msedge"]
else:
    CHROME_CANDIDATOS = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
        "google-chrome", "chromium", "chromium-browser",
        "microsoft-edge",
    ]

# Início de cada página-súmula. O sufixo [\s"] evita casar "page-footer".
PAGE_START = re.compile(r'<div class="page[\s"]')
BATERIA_RE = re.compile(
    r'Bateria / Heat</div>\s*<div class="fline fline-filled">([^<]*)</div>'
)
RAIA_RE = re.compile(
    r'>Raia</div>\s*<div class="fline fline-filled">([^<]*)</div>'
)


def achar_chrome():
    """Caminho do Chrome/Chromium, ou None se não houver (ex: Render)."""
    env = os.environ.get("CHROME")
    if env and os.path.exists(env):
        return env
    for c in CHROME_CANDIDATOS:
        if os.path.sep in c:
            if os.path.exists(c):
                return c
        elif shutil.which(c):
            return c
    return None


def dividir_documento(html):
    """Separa o HTML em (cabeca, [paginas], cauda).

    cabeca = tudo até o <body> inclusive (fontes + CSS, aparece 1x).
    paginas = cada <div class="page">...</div> de um atleta.
    cauda = </body></html>.
    """
    i_body = html.find("<body>")
    if i_body < 0:
        return None
    cabeca = html[: i_body + len("<body>")]
    i_fim = html.rfind("</body>")
    corpo = html[i_body + len("<body>"): i_fim]
    cauda = html[i_fim:]
    starts = [m.start() for m in PAGE_START.finditer(corpo)]
    if not starts:
        return None
    paginas = [
        corpo[s: starts[k + 1] if k + 1 < len(starts) else len(corpo)]
        for k, s in enumerate(starts)
    ]
    return cabeca, paginas, cauda


def bateria_da_pagina(pagina):
    m = BATERIA_RE.search(pagina)
    return m.group(1).strip() if m else ""


def raia_da_pagina(pagina):
    m = RAIA_RE.search(pagina)
    return m.group(1).strip() if m else ""


def rotulo_bateria(b):
    """'3' → 'Bateria_03'; 'A' → 'Bateria_A'; '' → 'Sem_Bateria'."""
    if not b:
        return "Sem_Bateria"
    return f"Bateria_{int(b):02d}" if b.isdigit() else f"Bateria_{sanitize(b)}"


def chave_num(b):
    return (0, int(b)) if str(b).isdigit() else (1, str(b))


def horarios_do_config(cfg):
    """Config do app (dias→categorias→baterias) →
    {(dia_pasta, cat_pasta, bateria_str): 'HH:MM'}."""
    horarios = {}
    for dia in cfg.get("dias", []) or []:
        dia_pasta = sanitize(dia.get("label", "Dia"))
        for cat in dia.get("categorias", []) or []:
            cat_pasta = sanitize(cat.get("nome", "Categoria"))
            for b in cat.get("baterias", []) or []:
                h = (b.get("horario_aquecimento") or b.get("horario_fila")
                     or b.get("horario") or "")
                h = str(h).strip()
                if h and len(h) == 4:           # '9:30' → '09:30' pra ordenar
                    h = "0" + h
                num = str(b.get("numero", "")).strip()
                if num:
                    horarios[(dia_pasta, cat_pasta, num)] = h
    return horarios


def carregar_horarios_excel(xlsx_path):
    """Excel de programação do organizador → mapa de horários.

    Reusa o parse_excel do app (parsers.py, mesmo diretório deste script),
    então aceita qualquer formato de Excel que o app aceite."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    try:
        from parsers import parse_excel
    except ImportError:
        raise RuntimeError(
            "ler Excel requer os arquivos do app (parsers.py + openpyxl). "
            "Alternativa sem instalação: exporte o backup JSON no app "
            "(Exportar JSON) e use esse arquivo no lugar do Excel.")
    return horarios_do_config(parse_excel(Path(xlsx_path).read_bytes()))


def carregar_horarios(json_path):
    """Backup JSON do app → mapa de horários."""
    snap = json.loads(Path(json_path).read_text(encoding="utf-8"))
    return horarios_do_config(snap.get("config", snap))


def finais_do_excel(xlsx_path):
    """Lê o Excel de programação e identifica as baterias de FINAL.

    O parser do app remove o marcador "(Final Heat)" dos nomes de categoria,
    mas ele está presente nas abas de cronograma (Sábado/Domingo, coluna
    Categoria). Aqui leio essas abas direto e devolvo, por dia:

        {dia_pasta: {'bats': {'23','24',...},          # nºs das baterias-final
                     'cat_bat': {cat_pasta: (hora, bat)}}}  # final de cada cat

    Uma bateria-final costuma ser compartilhada por 2+ categorias
    ("Scaled Fem (Final Heat) & Intermediario Fem (Final Heat)") — por isso
    o split em '&' e ','. Usado pra (a) detectar páginas-final e (b) ordenar
    o 00_FINAIS.pdf por horário mesmo quando a súmula ainda está em branco
    (aguardando balizamento — sem nº de bateria na página).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pular = ('montagem', 'workout', 'equipamento', 'inscrito', 'atleta')

    # PASSO 1 — números das baterias-final, do Excel cru (coluna Categoria
    # contém "(Final Heat)"). Confiável porque o nº de bateria não é abreviado.
    # MULTI-ARENA (Pwrd By Coffee): a aba tem BLOCOS lado a lado ("Arena: X"
    # na linha 1, cada bloco com suas colunas Categoria/Bateria) — varre TODOS
    # os pares categoria+bateria da linha de cabeçalho, não só o primeiro.
    fin_bats = {}                       # dia_pasta -> {'23','24',...}
    for nome in wb.sheetnames:
        if any(p in nome.lower() for p in pular):
            continue
        ws = wb[nome]
        blocos, hdr = [], None          # blocos = [(ci_cat, ci_bat), ...]
        for r in ws.iter_rows(min_row=1, max_row=8):
            cats = [c.column - 1 for c in r
                    if c.value and str(c.value).strip().lower() == 'categoria']
            bats = [c.column - 1 for c in r
                    if c.value and str(c.value).strip().lower() == 'bateria']
            if cats and bats:
                hdr = r[0].row
                # Pareia cada 'Categoria' com a 'Bateria' seguinte mais próxima
                for ci in cats:
                    cand = [bi for bi in bats if bi > ci]
                    if cand:
                        blocos.append((ci, min(cand)))
                break
        if hdr is None:
            continue
        dia_pasta = sanitize(nome)
        for r in ws.iter_rows(min_row=hdr + 1):
            for ci_cat, ci_bat in blocos:
                cat = r[ci_cat].value if ci_cat < len(r) else None
                if not cat or 'final heat' not in str(cat).lower():
                    continue
                if ci_bat < len(r) and r[ci_bat].value is not None:
                    bat = str(r[ci_bat].value).strip()
                    if bat.endswith('.0'):
                        bat = bat[:-2]
                    if bat:
                        fin_bats.setdefault(dia_pasta, set()).add(bat)
    if not fin_bats:
        return {}

    # PASSO 2 — mapeia cada bateria-final pra sua categoria USANDO A CONFIG
    # parseada (nomes idênticos às pastas do ZIP — ex.: 'Teen_Scaled_14-15_
    # Feminino', não a abreviação '14-15 Feminino' do cronograma). Assim a
    # detecção de finais em branco funciona mesmo com nomes abreviados.
    try:
        from parsers import parse_excel
        cfg = parse_excel(Path(xlsx_path).read_bytes())
    except Exception:
        cfg = {'dias': []}
    finais = {}
    for dia in cfg.get('dias', []) or []:
        dia_pasta = sanitize(dia.get('label', 'Dia'))
        fb = fin_bats.get(dia_pasta, set())
        if not fb:
            continue
        ent = finais.setdefault(dia_pasta, {'bats': set(fb), 'cat_bat': {},
                                            'cat_wkts': {}})
        for cat in dia.get('categorias', []) or []:
            cat_pasta = sanitize(cat.get('nome', 'Categoria'))
            workouts = cat.get('workouts', []) or []
            for b in cat.get('baterias', []) or []:
                num = str(b.get('numero', '')).strip()
                if num in fb:
                    hor = (b.get('horario_aquecimento')
                           or b.get('horario_fila') or '')
                    hor = str(hor).strip()
                    if hor and len(hor) == 4:
                        hor = '0' + hor
                    ent['cat_bat'][cat_pasta] = (hor, num)
                    # Nomes (sanitizados, como no filename do ZIP) dos
                    # workouts que a bateria-final roda. Usado pra NÃO
                    # arrastar pro 00_FINAIS páginas em branco de workouts
                    # comuns quando o dia inteiro aguarda balizamento.
                    nomes = set()
                    for p in b.get('workouts_que_rodam') or []:
                        if isinstance(p, int) and 1 <= p <= len(workouts):
                            n = sanitize(workouts[p - 1].get('nome', ''))
                            if n:
                                nomes.add(n.upper())   # filename sai maiúsculo
                    ent['cat_wkts'][cat_pasta] = nomes
    # Dias com bateria-final achada no PASSO 1 mas sem mapa por categoria
    # (parse_excel falhou / formato inesperado): mantém os números — o
    # converter cai no fallback por nº de bateria em vez de perder as finais.
    for dia_pasta, fb in fin_bats.items():
        finais.setdefault(dia_pasta, {'bats': set(fb), 'cat_bat': {},
                                      'cat_wkts': {}})
    return finais


def imprimir_pdf(chrome, html_path, pdf_path):
    """Imprime 1 HTML em PDF via Chrome headless.

    IMPORTANTE: sem --user-data-dir. Com perfil custom o Chrome (testado na
    v149/macOS) imprime o PDF mas trava no shutdown e nunca sai; sem a flag
    ele usa perfil efêmero próprio e encerra limpo em ~3s, inclusive com
    várias instâncias em paralelo. Timeout + 1 retry cobrem flakiness."""
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    cmd = [chrome, "--headless", "--disable-gpu",
           "--no-pdf-header-footer",
           f"--print-to-pdf={pdf_path.resolve()}",
           Path(html_path).resolve().as_uri()]
    for tentativa in (1, 2):
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            if r.returncode == 0 and pdf_path.exists():
                return
            erro = r.stderr[-300:]
        except subprocess.TimeoutExpired:
            erro = "timeout (90s) — Chrome travado, processo morto"
    raise RuntimeError(f"Chrome falhou em {pdf_path.name}: {erro}")


def converter(raiz, saida, horarios=None, chrome=None, log=print, finais=None):
    """Converte a árvore de HTMLs de `raiz` em PDFs organizados em `saida`.

    raiz: pasta com <Dia>/<Categoria>/NN_workout.html (shape do ZIP do app).
    horarios: mapa de horarios_do_config() — ordena o 00_DIA_COMPLETO.pdf.
    finais: mapa de finais_do_excel() — se houver, gera também um
            <Dia>/00_FINAIS.pdf só com as súmulas das baterias-final (elas
            CONTINUAM no 00_DIA_COMPLETO.pdf também — saída adicional, não
            exclusiva). Funciona com finais preenchidas ou ainda em branco.
    Retorna (n_pdfs_ok, lista_de_erros). Levanta RuntimeError sem Chrome.
    """
    raiz, saida = Path(raiz), Path(saida)
    horarios = horarios or {}
    finais = finais or {}
    chrome = chrome or achar_chrome()
    if not chrome:
        raise RuntimeError("Google Chrome não encontrado nesta máquina")

    htmls = sorted(p for p in raiz.rglob("*.html") if not p.name.startswith("."))
    if not htmls:
        raise RuntimeError("nenhum .html encontrado na entrada")

    tmp = Path(tempfile.mkdtemp(prefix="sumulas_pdf_"))
    try:
        # ── Monta os trabalhos de impressão ───────────────────────────────
        trabalhos = []          # (html_temporário, pdf_destino)
        dias = {}               # dia → páginas do mestre: (sort_key, cabeca, pg)
        finais_dias = {}        # dia → páginas-final: (sort_key, cabeca, pg)
        html_dir = tmp / "html"
        html_dir.mkdir()
        n_tmp = 0

        def agendar(cabeca, paginas, cauda, pdf_destino):
            nonlocal n_tmp
            n_tmp += 1
            f = html_dir / f"{n_tmp:04d}.html"
            f.write_text(cabeca + "".join(paginas) + cauda, encoding="utf-8")
            trabalhos.append((f, pdf_destino))

        for h in htmls:
            rel = h.relative_to(raiz)
            partes = rel.parts
            if len(partes) >= 3:                      # Dia/Categoria/arquivo
                dia_pasta, cat_pasta = partes[0], partes[1]
            elif len(partes) == 2:                    # Categoria/arquivo (legacy)
                dia_pasta, cat_pasta = "", partes[0]
            else:
                dia_pasta, cat_pasta = "", ""
            stem = h.stem
            m = re.match(r"(\d+)", stem)
            prefixo = int(m.group(1)) if m else 999

            doc = dividir_documento(h.read_text(encoding="utf-8"))
            if doc is None:
                log(f"⚠  Pulei (estrutura inesperada): {rel}")
                continue
            cabeca, paginas, cauda = doc

            # Agrupa páginas por bateria preservando a ordem de geração
            grupos, ordem = {}, []
            for p in paginas:
                b = bateria_da_pagina(p)
                if b not in grupos:
                    grupos[b] = []
                    ordem.append(b)
                grupos[b].append(p)

            pasta_pdf = saida / dia_pasta / cat_pasta
            so_sem_bateria = ordem == [""]
            if so_sem_bateria:
                # Súmula em branco ou "aguardando balizamento": sem bateria
                # pra fatiar — vira um PDF único do workout.
                agendar(cabeca, paginas, cauda, pasta_pdf / f"{stem}.pdf")
            else:
                for b in sorted(ordem, key=chave_num):
                    agendar(cabeca, grupos[b], cauda,
                            pasta_pdf / stem / f"{rotulo_bateria(b)}.pdf")

            # Páginas pro PDF mestre do dia, ordenadas página a página:
            # horário → bateria → raia. Em bateria mista (2 categorias na
            # mesma bateria) as raias saem intercaladas como no piso, não
            # em blocos por categoria. Sem horário, agrupa por categoria
            # pra não embaralhar baterias homônimas de arenas diferentes.
            fin = finais.get(dia_pasta, {})
            fin_bats = fin.get('bats', set())
            fin_catbat = fin.get('cat_bat', {})
            # Nome do workout como aparece no filename ('03_SETE_MINUTOS' →
            # 'SETE_MINUTOS') pra casar com cat_wkts na regra das finais.
            stem_wkt = stem.split('_', 1)[1] if re.match(r'\d+_', stem) else stem
            for b in ordem:
                hor = horarios.get((dia_pasta, cat_pasta, b), "")
                for pg in grupos[b]:
                    raia = chave_num(raia_da_pagina(pg))
                    if hor:
                        sort_key = ((0, hor), chave_num(b), prefixo,
                                    raia, cat_pasta)
                    else:
                        sort_key = ((1, ""), cat_pasta, prefixo,
                                    chave_num(b), raia)
                    dias.setdefault(dia_pasta, []).append(
                        (sort_key, cabeca, pg))

                    # Detecta página-final: bateria marcada (Final Heat) no
                    # cronograma, OU página em branco (aguardando balizamento)
                    # de categoria com final E do workout que a final roda —
                    # sem a checagem de workout, um dia inteiro pré-balizamento
                    # (todas as páginas sem bateria) despejaria os workouts
                    # comuns dentro do 00_FINAIS. Ordena pelo horário do
                    # heat-final da categoria — assim o 00_FINAIS.pdf sai em
                    # ordem mesmo com as súmulas ainda em branco.
                    if b:
                        # Preenchida: casa pela bateria-final DA CATEGORIA.
                        # Imune a nº de bateria repetido entre arenas
                        # (multi-arena); `b in fin_bats` fica de fallback
                        # pra quando o mapa por categoria não existir
                        # (parse do Excel falhou → só PASSO 1 disponível).
                        if cat_pasta in fin_catbat:
                            eh_final = fin_catbat[cat_pasta][1] == b
                        else:
                            eh_final = not fin_catbat and b in fin_bats
                    elif cat_pasta in fin_catbat:
                        nomes_final = fin.get('cat_wkts', {}).get(cat_pasta)
                        # Sem info de workout (bateria roda tudo / config
                        # antiga): mantém comportamento permissivo.
                        eh_final = (not nomes_final) or \
                                   (stem_wkt.upper() in nomes_final)
                    else:
                        eh_final = False
                    if eh_final:
                        fhor, fbat = fin_catbat.get(cat_pasta, (hor, b))
                        fkey = ((0, fhor) if fhor else (1, ""),
                                chave_num(fbat or b), raia, cat_pasta)
                        finais_dias.setdefault(dia_pasta, []).append(
                            (fkey, cabeca, pg))

        for dia_pasta, chunks in dias.items():
            chunks.sort(key=lambda c: c[0])
            cabeca = chunks[0][1]
            todas = [pg for _, _, pg in chunks]
            agendar(cabeca, todas, "</body></html>",
                    saida / dia_pasta / "00_DIA_COMPLETO.pdf")

        # PDF separado só das finais (além de seguirem no dia-completo)
        for dia_pasta, chunks in finais_dias.items():
            chunks.sort(key=lambda c: c[0])
            cabeca = chunks[0][1]
            todas = [pg for _, _, pg in chunks]
            log(f"  ★ {len(todas)} súmula(s) de final em {dia_pasta}/00_FINAIS.pdf")
            agendar(cabeca, todas, "</body></html>",
                    saida / dia_pasta / "00_FINAIS.pdf")

        # ── Imprime tudo (3 Chromes em paralelo) ──────────────────────────
        total = len(trabalhos)
        log(f"→ {len(htmls)} HTML(s) → {total} PDF(s) em {saida}")
        feitos, erros = 0, []

        def worker(par):
            f, destino = par
            imprimir_pdf(chrome, f, destino)
            return destino

        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as ex:
            for fut in concurrent.futures.as_completed(
                    ex.submit(worker, t) for t in trabalhos):
                try:
                    destino = fut.result()
                    feitos += 1
                    log(f"  ✓ [{feitos}/{total}] {destino.relative_to(saida)}")
                except Exception as e:
                    erros.append(str(e))
                    log(f"  ✗ {e}")
        return feitos, erros
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("entrada", help="ZIP de súmulas (ou pasta descompactada)")
    ap.add_argument("--json", help="backup JSON do app (horários das baterias)")
    ap.add_argument("--excel", help="Excel de programação (mesmos formatos do app)")
    ap.add_argument("--saida", help="pasta de destino dos PDFs")
    args = ap.parse_args()

    entrada = Path(args.entrada).expanduser()
    if not entrada.exists():
        sys.exit(f"✗ Não encontrei: {entrada}")

    chrome = achar_chrome()
    if not chrome:
        sys.exit("✗ Google Chrome não encontrado. Instale o Chrome ou defina "
                 "a variável de ambiente CHROME com o caminho do executável.")

    horarios, finais = {}, {}
    try:
        if args.json:
            horarios = carregar_horarios(args.json)
        elif args.excel:
            horarios = carregar_horarios_excel(args.excel)
            finais = finais_do_excel(args.excel)
    except RuntimeError as e:
        sys.exit(f"✗ {e}")
    if (args.json or args.excel) and not horarios:
        print("⚠  Arquivo lido mas sem horários de bateria — ordem do dia "
              "completo cai pra Categoria → Workout → Bateria.")

    tmp_zip = None
    try:
        if entrada.is_file():
            tmp_zip = Path(tempfile.mkdtemp(prefix="sumulas_zip_"))
            with zipfile.ZipFile(entrada) as zf:
                zf.extractall(tmp_zip)
            raiz, nome_base = tmp_zip, entrada.stem
        else:
            raiz, nome_base = entrada, entrada.name
        saida = Path(args.saida) if args.saida else entrada.parent / f"{nome_base}_PDFs"

        def log_flush(msg):
            print(msg, flush=True)

        feitos, erros = converter(raiz, saida, horarios, chrome, log_flush, finais)
        if erros:
            sys.exit(f"\n✗ {len(erros)} PDF(s) falharam de {feitos + len(erros)}.")
        print(f"\n✓ Pronto: {feitos} PDF(s) em {saida}", flush=True)
    except RuntimeError as e:
        sys.exit(f"✗ {e}")
    finally:
        if tmp_zip:
            shutil.rmtree(tmp_zip, ignore_errors=True)


if __name__ == "__main__":
    main()
