"""parsers.py — extração de workouts a partir de texto livre, Excel e PDF.

Funções públicas:
    parse_workout_text(text, numero) -> Workout
    parse_excel(data: bytes) -> dict   # categoria_grid OU template
    parse_pdf(data: bytes) -> dict
    assign_workout_numbers(workouts) -> list[Workout]   # mutates in-place
    _atleta_sort_key(a) -> tuple        # bateria → raia (numérica) → nome

Nada aqui depende do servidor HTTP ou da geração de HTML.
"""
from __future__ import annotations

import io
import re
from typing import Any, Optional

from types_ds import Atleta, Movimento, Workout
from movimentos import padronizar_workouts

# Excel e PDF são opcionais (parsers respectivos só ativam se a lib estiver instalada)
try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


BLOCK_LABELS = {1: "1º BLOCO", 2: "2º BLOCO", 3: "3º BLOCO", 4: "4º BLOCO", 5: "5º BLOCO"}

# Header de bloco por atleta em duplas/trios: `Atleta 1`, `Athlete 2`.
# Quando presente, os movimentos seguintes ganham label `ATLETA N` no lugar
# do label numérico de bloco — juiz vê na súmula quem faz cada movimento.
_ATLETA_HEADER_RE = re.compile(r'^\s*(?:atleta|athlete)\s+(\d+)\s*[:\.]?\s*$', re.I)


# Separadores que marcam fim da prescrição "core" e início de
# regras/observações/regulamento que NÃO devem aparecer na súmula impressa.
# A súmula deve conter só o essencial pro atleta executar; o resto é regulamento
# e atleta/árbitro consultam à parte.
# Classe de traços "decorativos" usados como moldura de seção. Inclui U+2015
# (―, HORIZONTAL BAR) além de U+2500/2014/2013 — o Pwrd usa `――― NOTAS ―――`
# com U+2015, que antes escapava do corte.
_TRACO = r'─—–―\-'
_DESC_CUT_RE = re.compile(
    r'^\s*(?:[' + _TRACO + r']+\s*)?'
    r'(?:notas?|notes?|observa[çc][õo]es?|observations?|pontua[çc][ãa]o|tiebreak|'
    r'regras?|rules?|regulamento|crit[ée]rios?|criteria|score|scoring)'
    r'\s*(?:[' + _TRACO + r']+\s*)?\s*:?\s*$',
    re.IGNORECASE,
)


def _truncar_descricao_em_notas(lines: list[str]) -> list[str]:
    """Corta a lista de linhas no primeiro separador tipo `NOTAS`, `Observações`,
    `Pontuação`, etc. Tudo abaixo é regulamento e fica fora da súmula impressa.

    Mantém comportamento original quando não há separadores (lista intacta).
    """
    out: list[str] = []
    for line in lines:
        if _DESC_CUT_RE.match(line):
            break
        out.append(line)
    return out


# Palavras que indicam FRASE EXPLICATIVA (não-movimento). Linhas que começam
# com número seguido dessas palavras NÃO devem virar movimento. Cobre português
# e inglês — listas de "regras" e "divisão de funções" típicas em prescrições
# de eventos sofisticados (trios, duplas, etc).
_FRASE_NAO_MOVIMENTO_RE = re.compile(
    r'\b(?:atletas?|nadar(?:ão|emos|á|em)|executar(?:á|emos|ão|em)|'
    r'iniciar(?:á|emos|ão|em)|completar(?:á|emos|ão|em)|trocar(?:á|emos|ão|em)|'
    r'ser(?:á|ão|emos|ia)?|times?|equipes?|teams?|round\s+seguinte|'
    r'definid[ao]s?|alterad[ao]s?|escolher[áa]?|cada\s+\w+\s+executar|'
    r'minute[s]?\s*,?\s*for\s+\d+\s+rounds?|'
    r'rounds?\s+per\s+athletes?|rounds?\s+por\s+atleta|'
    r'sets?\s+per\s+athletes?|sets?\s+por\s+atleta)\b',
    re.IGNORECASE,
)
# Filtro EXTRA usado SÓ no fallback de flex-mov (For Time com Goal). Mais
# agressivo — pega notas/regulamento em EN que escapam do filtro genérico
# porque movimentos normais nunca têm 'must', 'should', 'cross the', etc.
_NOTA_LIKELY_RE = re.compile(
    r'\b(?:athletes?\s+(?:must|should|will|need|alternate|cross)|'
    r'must\b|should\b|will\s+(?:start|finish|complete)|need\s+to|needs\s+to|'
    r'deve\b|deverá|precisa|tem\s+que|'
    r'cross\s+the\b|finish\s+line|alternat(?:ing|e)\b|'
    r'before\s+(?:start|finishing)|after\s+(?:complet|finish))\b',
    re.IGNORECASE,
)
# Headers de seção/parte que NÃO são movimento. Aparecem em workouts longos
# como Simple Dimension/Mind: 'Part 1 (0:00-6:00)', 'Stage 2', 'Bloco 1'.
# Detecção: começa com palavra-header + número, opcionalmente seguido de
# janela de tempo entre parênteses.
_SECTION_HEADER_RE = re.compile(
    r'^\s*(?:part|parte|stage|phase|fase|bloco|block|round|rodada|chapter|cap[íi]tulo)'
    r'\s+\d+\b',
    re.IGNORECASE,
)
# Linha que contém janela de tempo entre parênteses '(0:00-6:00)' ou '(0:00→6:00)'
# — geralmente é header informativo, não movimento.
_TIME_WINDOW_RE = re.compile(
    r'\(\s*\d{1,2}:\d{2}\s*[-→–—]\s*\d{1,2}:\d{2}\s*\)',
)


# Extrai carga no fim do nome do movimento. Dois formatos:
#   A) NUM UNIT (unidade obrigatória):  '50/35 lb', '20kg', '225/155 LB', '75#'
#   B) @NUM (unit opcional):             '@135/95', '@40lb'
# NÃO captura distâncias/calorias (cal, m, km) — esses ficam no nome
# (ex: '900M SKI ERG' permanece intacto).
_CARGA_END_RE = re.compile(
    r'\s*\(?\s*'
    r'(\d+(?:[\.,]\d+)?(?:/\d+(?:[\.,]\d+)?)?)\s*'   # número (ou par M/F)
    r'(kg|lb|lbs|#|pood)'                            # unidade DE PESO obrigatória
    r'\s*\)?\s*$',
    re.IGNORECASE,
)
_CARGA_AT_END_RE = re.compile(
    r'\s*\(?\s*@\s*'
    r'(\d+(?:[\.,]\d+)?(?:/\d+(?:[\.,]\d+)?)?)\s*'   # número após @
    r'(kg|lb|lbs|#|pood)'                            # unidade DE PESO obrigatória
    r'\s*\)?\s*$',                                   # antes era opcional — capturava
    re.IGNORECASE,                                   # altura (24") e distância (800m)
)
# Carga no INÍCIO ('20kg Sandbag Carry', '95lb Thrusters'). Útil quando o
# organizador formata diferente. SÓ pega se a unidade vier imediatamente
# após o número (sem espaço pra distância). Senão confunde com reps líderes.
_CARGA_AT_START_RE = re.compile(
    r'^\s*'
    r'(\d+(?:[\.,]\d+)?(?:/\d+(?:[\.,]\d+)?)?)\s*'   # número
    r'(kg|lb|lbs|#)'                                 # unidade de peso ATACHADA
    r'\s+(\S.+)$',                                   # resto = nome
    re.IGNORECASE,
)
# Carga dupla com unidade em CADA número: '70kg/50kg', '22,5kg/15kg', '20lbs/14lbs'
# (Rx/scaled ou M/F). Diferente de '50/35 lb' (unidade só no fim, tratada pelo
# _CARGA_END_RE). Normaliza pra 'N1/N2 UNIT'.
_CARGA_DUAL_UNIT_RE = re.compile(
    r'\s*\(?\s*'
    r'(\d+(?:[\.,]\d+)?)\s*(kg|lb|lbs|#|pood)\s*'    # 1º número + unidade
    r'/\s*'
    r'(\d+(?:[\.,]\d+)?)\s*(kg|lb|lbs|#|pood)?'      # 2º número + unidade (opcional)
    r'\s*\)?\s*$',
    re.IGNORECASE,
)


def _extrair_carga(nome: str) -> tuple[str, Optional[str]]:
    """Separa o nome do movimento da carga ao final, se houver.

    Retorna (nome_sem_carga, carga|None). Carga normalizada em uppercase
    ('50/35 LB', '20 KG', '@135/95'). Sem unidade quando o input usa só `@`.
    Genérico — usado em parser de qualquer tipo de workout.
    """
    # Tenta carga no INÍCIO primeiro ('20kg Sandbag Carry', '95lb Thrusters')
    m_start = _CARGA_AT_START_RE.match(nome)
    if m_start:
        num, unit, resto = m_start.group(1), m_start.group(2).upper(), m_start.group(3).strip()
        if resto and len(resto) >= 3:
            return (resto, f"{num} {unit}")
    # Carga dupla com unidade em cada número ('70kg/50kg') ANTES do fallback —
    # senão o _CARGA_END_RE pega só o 2º ('50kg') e deixa '(70kg/' no nome.
    m_dual = _CARGA_DUAL_UNIT_RE.search(nome)
    if m_dual:
        nome_limpo = nome[:m_dual.start()].rstrip(' ,-()@').strip()
        if nome_limpo:
            n1, u1, n2, u2 = m_dual.groups()
            unit = (u2 or u1 or '').upper()
            return (nome_limpo, f"{n1}/{n2} {unit}".strip())
    # Senão tenta no FIM (com unidade obrigatória) ou @-prefixed
    m = _CARGA_END_RE.search(nome) or _CARGA_AT_END_RE.search(nome)
    if not m: return (nome, None)
    nome_limpo = nome[:m.start()].rstrip(' ,-()@').strip()
    if not nome_limpo: return (nome, None)   # não destrói nomes só-carga
    num = m.group(1)
    unit = (m.group(2) or '').upper()
    carga = f"{num} {unit}".strip() if unit else num
    return (nome_limpo, carga)


# ── Helpers genéricos ─────────────────────────────────────────────────────────
def _safe_int(s, default: Optional[int] = None) -> Optional[int]:
    """Tenta int(s) silenciosamente. Retorna default se falhar.

    Substitui o padrão repetido `try: int(...) except ValueError: pass`.
    Mais legível e padroniza tratamento de input não numérico.
    """
    if s is None: return default
    try:
        return int(str(s).strip())
    except (ValueError, TypeError):
        return default


# Números por extenso (EN + PT) 1-10 — usados na detecção de "Three rounds…".
_NUM_PALAVRA = {
    'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
    'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
    'um': 1, 'uma': 1, 'dois': 2, 'duas': 2, 'tres': 3, 'três': 3,
    'quatro': 4, 'cinco': 5, 'seis': 6, 'sete': 7, 'oito': 8, 'nove': 9, 'dez': 10,
}
# Alternância de regex que casa dígito OU palavra (EN/PT).
_NUM_TOKEN_RE = (r'(\d+|one|two|three|four|five|six|seven|eight|nine|ten|'
                 r'um|uma|dois|duas|tr[eê]s|quatro|cinco|seis|sete|oito|nove|dez)')


def _num_ext(tok) -> Optional[int]:
    """'three'/'três'/'3' → int. None se não reconhece."""
    if tok is None:
        return None
    t = str(tok).strip().lower()
    if t.isdigit():
        return int(t)
    return _NUM_PALAVRA.get(t)


# Bloco de rounds ANINHADO: 'then, 2 rounds of:', 'então, 3 rounds de:'. Vira
# banner de seção (não multiplica reps — o buy-in antes do 'then' não entra na
# conta). Diferente de 'N rounds for time' (que é o workout inteiro).
_ROUNDS_BLOCK_RE = re.compile(
    r'^\s*(?:then|ent[ãa]o|depois|after|ap[óo]s)?\s*,?\s*'
    + _NUM_TOKEN_RE + r'\s+rounds?\s+(?:of|de)\b',
    re.I,
)


# Excel diz que NÃO tem rep de chegada (ou que ela não conta/pontua). Quando
# bate, o for_time NÃO ganha a linha de chegada. Cobre várias formas em PT/EN:
#   'chegada não conta/vale/pontua/contabiliza', 'não conta a chegada',
#   'sem chegada', 'não tem/há chegada', 'no finish rep', 'finish doesn't count'.
_CHEGADA_NEGADA_RE = re.compile(
    r'(?:'
    r'chegada[^.\n]{0,40}?n[ãa]o[^.\n]{0,20}?(?:cont|val|pontu|contabiliz)'
    r'|n[ãa]o[^.\n]{0,20}?(?:cont|contabiliz|pontu)[^.\n]{0,25}?chegada'
    r'|(?:sem|n[ãa]o\s+tem|n[ãa]o\s+h[áa]|n[ãa]o\s+possui)\s+(?:a\s+|rep\w*\s+d[eo]\s+)?chegada'
    r'|no\s+finish(?:\s+rep)?\b'
    r'|finish[^.\n]{0,20}?(?:does\s*n.?t|doesn.?t|not)\s*count'
    r'|without\s+(?:the\s+)?finish'
    r')',
    re.I,
)


# AMRAP multi-janela (PWRD Loop): 2+ blocos 'AMRAP N min' separados por 'Rest'.
# Cada janela tem reps prescritas + uma linha 'Max ...' (o que pontua).
_AMRAP_JANELA_RE = re.compile(r'^\s*amrap\s+(\d+)\s*(?:min|minutes?|minutos?)\b', re.I)
_REST_JANELA_RE  = re.compile(r'^\s*(?:rest|descanso|descanse)\b', re.I)
# Linha 'Max. Wall-Ball ...' / 'Max reps of ...' — movimento SEM reps fixas, é
# o que acumula pontuação. Sem número na frente (por isso escapa do _parse_mov_line).
_MAX_MOV_RE = re.compile(r'^\s*max[.:\s]+(?:reps?\s+(?:of|de)\s+)?(.+)$', re.I)


# ── Texto livre de workout ──────────────────────────────────────────────────────
def _parse_mov_line(line: str) -> Optional[tuple[int, str]]:
    """Extrai (reps, nome_upper) de uma linha de movimento.

    Suporta 3 formatos do número inicial:
      `20 Pull-Ups`            → reps=20, nome='PULL-UPS'
      `20-metres DB Lunges`    → reps=20, nome='20-METRES DB LUNGES'  (hífen)
      `900m Ski Erg`           → reps=900, nome='900M SKI ERG'        (unidade colada)
      `3k Treadmill Run`       → reps=3000, nome='3000M TREADMILL RUN' (k → metros)
      `5km Run`                → reps=5000, nome='5000M RUN'

    Rejeita linhas que parecem frase explicativa (`2 atletas nadarão...`)
    pra evitar que virem movimentos.
    """
    s = line.strip()
    # 4 formatos: NUM/NUM resto (gendered), NUM+unit+ESP, NUM-resto, NUM ESP resto
    tem_unidade = False   # NUM colado a unidade (1000m, 20cal) — distância válida ≥1000
    m = re.match(r'^(\d{1,4})/(\d{1,4})\s+(.+)$', s)            # 30/24 cal Row
    if m:
        num_s, num_f, rest = m.group(1), m.group(2), m.group(3).strip()
        nome = f"{num_s}/{num_f} {rest}".upper()
    else:
        m = re.match(r'^(\d{1,4})([a-z]+)\s+(.+)$', s, re.I)    # 900m Ski Erg
        if m:
            num_s, unit, rest = m.group(1), m.group(2), m.group(3).strip()
            # 'k'/'km' = quilômetros → converte pra metros (3k → 3000m) pra reps
            # (acumulado) e display ficarem certos. 'kg'/outras unidades não mexem.
            if unit.lower() in ('k', 'km'):
                num_s = str(int(num_s) * 1000)
                unit = 'm'
            nome = f"{num_s}{unit} {rest}".upper()
            tem_unidade = True
        else:
            m = re.match(r'^(\d{1,4})([-\s])(.+)$', s)
            if not m: return None
            num_s, sep, rest = m.group(1), m.group(2), m.group(3).strip()
            if sep == '-':
                nome = f"{num_s}-{rest}".upper()
            else:
                nome = rest.upper()
    try: num = int(num_s)
    except ValueError: return None
    # ≥1000 sem unidade provavelmente é ano ('2026 ...'); com unidade é distância
    # legítima (1000m Ski Erg / Row / Run).
    if num >= 1000 and not tem_unidade: return None
    # Rejeita frases explicativas (`2 atletas nadarão`, `5 times escolherão`, etc).
    # Antes de testar, descarta descritor `(dois atletas)` / `(N atleta)` entre
    # parens — é metadata operacional, não frase explicativa. Caso real Storm:
    # `200m Run (dois atletas)` é movimento legítimo, não regulamento.
    nome_p_filtro = re.sub(r'\s*\([^)]*atletas?[^)]*\)', '', nome, flags=re.I)
    if _FRASE_NAO_MOVIMENTO_RE.search(nome_p_filtro):
        return None
    return (num, nome)


_FOR_LOAD_IGNORE_RE = re.compile(
    r'^(?:for\s+load|max\s+(?:lift|load)|carga\s+m[áa]xima|time\s+cap|tempo|'
    r'\d+\s+tentativas?|notas?|observa[çc][ãa]o|descanso|rest\b|entre\s+|'
    r'cada\s+atleta|cada\s+tentativa|score|pontua)',
    re.IGNORECASE,
)
_FOR_LOAD_BUYIN_RE = re.compile(
    r'^\s*(?:buy[\s-]?in|aquecimento|warm[\s-]?up)\s*[:\-]?\s*(.*)$',
    re.IGNORECASE,
)
_FOR_LOAD_THEN_RE = re.compile(
    r'^\s*(?:then|ent[ãa]o|depois|after|ap[óo]s|complex)\b[:\s\-,.]*\s*(.*)$',
    re.IGNORECASE,
)


# Rótulo de atleta após a janela de tempo: 'Athlete A', 'Atleta B', '(feminina)'.
_JANELA_ATLETA_RE = re.compile(r'athlete|atleta', re.I)


def _extrair_janelas_for_load(lines: list[str]) -> list[dict]:
    """For Load com janelas de tempo por atleta/tentativa (Muscle Coffee do
    Pwrd by Coffee): cada `(00:00 - 03:00) [Athlete A]` abre uma janela, e as
    linhas seguintes são o complex daquela janela.

    Retorna lista `[{'label','janela','atleta','complex'}]` (A/B/C…), ou `[]`
    quando não há janelas — aí o caller usa o complex único de sempre.
    """
    lines = _truncar_descricao_em_notas(lines)
    grupos: list[dict] = []
    atual: Optional[dict] = None
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        mw = _TIME_WINDOW_RE.search(s)
        if mw:
            resto = s[mw.end():].strip(' -–—:·').strip()
            atleta = resto if resto and _JANELA_ATLETA_RE.search(resto) else resto
            atual = {'janela': mw.group(0).strip('() '), 'atleta': atleta, 'partes': []}
            grupos.append(atual)
            # complex pode vir grudado na mesma linha, após a janela/rótulo
            continue
        if atual is None:
            continue
        if s.startswith(('"', '“', '‘')):
            continue
        if _FOR_LOAD_IGNORE_RE.match(s):
            continue
        atual['partes'].append(s)
    out: list[dict] = []
    for i, g in enumerate(grupos):
        complex_ = _normalizar_complex(' '.join(g['partes'])) if g['partes'] else None
        if not complex_:
            continue
        out.append({
            'label': chr(ord('A') + len(out)),   # A, B, C…
            'janela': g['janela'],
            'atleta': g['atleta'],
            'complex': complex_,
        })
    # Só faz sentido como "blocos" quando há 2+ janelas distintas.
    return out if len(out) >= 2 else []


def _extrair_sequencia_for_load(lines: list[str], nome: str) -> dict:
    """Extrai sequência pro lembrete do árbitro em For Load.

    Retorna `{'buy_in': str|None, 'complex': str|None, 'janelas': list}` —
    strings enxutas + janelas A/B/C opcionais. Só o que importa pro árbitro:
    o que aquece (buy-in), o que vale carga (complex) e, quando há janelas de
    tempo por atleta, cada bloco separado.

    Estratégia (ordem):
      1. Trunca em NOTAS / OBSERVAÇÕES / ─── (regulamento fora do escopo).
      2. Se houver marcador `COMPLEX:` em alguma linha, usa só o que vem
         APÓS como complex; o que vem antes (cal/cal Air Bike, Row, etc)
         vira buy-in.
      3. Senão, busca marcadores 'Buy-in:' / 'Then:' explícitos.
      4. Pula linhas duplicadas tipo 'ATHLETE 2 ... (MESMO PADRÃO)' — todos
         atletas fazem a mesma sequência, descrita só uma vez.
    """
    # 1) Trunca em NOTAS — regulamento fora do escopo da súmula impressa
    lines = _truncar_descricao_em_notas(lines)

    # Janelas de tempo por atleta (A/B/C) — quando presentes, o render mostra
    # cada bloco separado além do complex corrido.
    janelas = _extrair_janelas_for_load(lines)

    # 2) Procura linha com 'COMPLEX:' (1-RM Complex, Squat Complex, etc)
    complex_re = re.compile(r'^(.*?)\b(?:1[-\s]?rep[-\s]?max\s+complex|complex)\s*:\s*(.+)$', re.I)
    skip_re = re.compile(
        r'mesmo\s+padr[ãa]o|same\s+as|mesma\s+sequ[êe]ncia|'
        r'atleta\s*\d+\s*\(|athlete\s*\d+\s*\(',
        re.I,
    )
    for ln in lines:
        s = ln.strip()
        if not s: continue
        if _FOR_LOAD_IGNORE_RE.match(s): continue
        m = complex_re.match(s)
        if not m: continue
        antes = m.group(1).strip()
        depois = m.group(2).strip()
        # buy-in = caloric/distance work antes do COMPLEX (12-cal Air Bike, 200m Row)
        buy_in = _extrair_buyin_caloric(antes)
        # complex = parte após 'COMPLEX:' — normaliza separadores
        complex_ = _normalizar_complex(depois)
        return {'buy_in': buy_in, 'complex': complex_, 'janelas': janelas}

    # 3) Sem COMPLEX explícito: usa marcadores Buy-in / Then
    buy_in_parts: list[str] = []
    complex_parts: list[str] = []
    is_buyin = False
    for ln in lines:
        s = ln.strip()
        if not s: continue
        if s.startswith('"') or s.startswith('“') or s.startswith('‘'): continue
        if _FOR_LOAD_IGNORE_RE.match(s): continue
        if skip_re.search(s): continue   # pula 'ATHLETE 2 (mesmo padrão)'
        m_buyin = _FOR_LOAD_BUYIN_RE.match(s)
        if m_buyin:
            is_buyin = True
            resto = m_buyin.group(1).strip()
            if resto: buy_in_parts.append(resto)
            continue
        m_then = _FOR_LOAD_THEN_RE.match(s)
        if m_then:
            is_buyin = False
            resto = m_then.group(1).strip()
            if resto: complex_parts.append(resto)
            continue
        (buy_in_parts if is_buyin else complex_parts).append(s)
    buy_in = ' '.join(buy_in_parts).strip().upper() or None
    complex_ = ' '.join(complex_parts).strip().upper() or None
    # 4) Fallback: nome do workout (limpo de 'MAX' / 'CARGA MÁXIMA')
    if not complex_ and nome:
        complex_ = re.sub(r'^(?:max\s+|carga\s+m[áa]xima\s+(?:de\s+)?)',
                          '', nome, flags=re.I).strip().upper() or None
    return {'buy_in': buy_in, 'complex': complex_, 'janelas': janelas}


def _extrair_buyin_caloric(texto: str) -> Optional[str]:
    """Pega o trecho do tipo 'N-cal Air Bike', 'N cal Row', 'Nm Run' do início.

    Usado pra capturar buy-in que vem antes de 'COMPLEX:' na mesma linha.
    Strip prefixos tipo 'ATHLETE 1 (0:00-4:00)'.
    """
    s = texto.strip()
    # Remove 'ATHLETE N (window)' do início
    s = re.sub(r'^\s*(?:athlete|atleta)\s*\d+\s*\([^)]*\)\s*', '', s, flags=re.I).strip()
    if not s: return None
    m = re.search(
        r'(\d+[-\s]?(?:cal|calorie|calorias?|m|metres?|metros?|km|mi)\b[^,.;:]*?'
        r'(?:bike|row|ski|run|swim|jump|rope|ergometer|erg)?[^,.;:]*)',
        s, re.I,
    )
    if not m: return s.upper() or None
    return m.group(1).strip().upper()


def _normalizar_complex(texto: str) -> str:
    """Normaliza '1 X 1 Y 1 Z' → '1 X + 1 Y + 1 Z' (separador visual).

    Preserva combos já com '+' / 'e' / 'and' / '&'. Limita tamanho da string.
    """
    s = texto.strip()
    # Se já tem separador, só uppercase
    if re.search(r'[+&]|\be\b|\band\b', s, re.I):
        return s.upper()
    # Inserir ' + ' entre partes 'N <palavras> N <palavras>'
    s = re.sub(r'(\d+\s+[A-Za-zÀ-ú\-]+(?:\s+[A-Za-zÀ-ú\-]+){0,3})\s+(?=\d+\s+[A-Za-zÀ-ú])',
               r'\1 + ', s)
    return s.upper()


def _extrair_nome_workout(lines: list[str]) -> Optional[str]:
    """Extrai o nome do workout da primeira linha, se ela for um título
    (entre aspas ou texto livre não começando com dígito). Retorna None
    se a primeira linha já parece ser conteúdo (movs/diretrizes)."""
    if not lines: return None
    m = re.match(r'^["“‘](.+?)["”’]', lines[0])
    if m: return m.group(1).strip().upper()
    if not re.match(r'^\d', lines[0]):
        return lines[0].strip('"“”').upper()[:40]
    return None


def _parse_for_load(lines: list[str], wkt: Workout, full: str) -> Workout:
    """Branch For Load: detecta tentativas, captura descrição e sequência.
    Returna o wkt populado pra retorno imediato em parse_workout_text."""
    wkt["tipo"] = "for_load"
    m_tent = re.search(r'(\d+)\s*tentativas?', full)
    if m_tent:
        tent = _safe_int(m_tent.group(1))
        if tent is not None: wkt["tentativas"] = tent
    # Texto livre fica em descricao; trunca em separadores tipo NOTAS pra
    # não bagunçar a súmula com regulamento que estoura A4.
    wkt["descricao"] = _truncar_descricao_em_notas(lines)
    wkt["movimentos"] = []
    seq = _extrair_sequencia_for_load(lines, wkt.get("nome", ""))
    wkt["sequencia_movimentos"] = seq
    # Múltiplas janelas (Muscle Coffee): cada janela é um complex INDEPENDENTE,
    # pontuado e SOMADO ('soma das cargas máximas dos N complex'). A régua ganha
    # 1 linha por janela e o total vira soma — não 'melhor de N tentativas'.
    janelas = seq.get("janelas") or []
    if len(janelas) >= 2:
        # Cada complex é pontuado por SI (soma dos máximos). A régua dá 3
        # tentativas POR complex (padrão) — não 1 linha por complex. O 'tentativas'
        # aqui é attempts por complex; o nº de complexes vem de `janelas`.
        wkt["tentativas"] = 3
        wkt["soma_complexes"] = True
        wkt["janela_labels"] = [j.get("atleta") or f"Complex {j['label']}" for j in janelas]
    return wkt


def _detectar_directives(full: str, lines: list[str], wkt: Workout) -> None:
    """Detecta e popula in-place todas as diretrizes do workout no wkt:
    relay (N rounds per athlete), EMOM, tiebreak (por round e geral),
    progressão de reps, goal de For Time, último round MAX. Mutação."""
    # Relay 'N round(s) per athlete' (For Time típico em trios)
    m_relay = (re.search(r'(\d+)\s+rounds?\s+per\s+athletes?', full, re.I)
               or re.search(r'(\d+)\s+rounds?\s+por\s+atleta', full, re.I))
    if m_relay:
        n = _safe_int(m_relay.group(1))
        if n is not None: wkt["rounds_per_atleta"] = n

    # N rounds for time: 'X rounds for time of:' / 'X rounds por tempo:' /
    # 'For time, X rounds of:'. Atleta faz a sequência completa X vezes,
    # score = tempo total. Marca wkt.rounds_fixos pra render mostrar banner
    # + calcular acumulado total (reps × X).
    # Aceita número por extenso ('Three rounds for time of') e 'For time:' em
    # linha separada do 'Four Rounds of' (separador [:,\s]* cobre ':' + \n).
    m_rounds = (
        re.search(_NUM_TOKEN_RE + r'\s+rounds?\s+for\s+time', full, re.I)
        or re.search(_NUM_TOKEN_RE + r'\s+rounds?\s+por\s+tempo', full, re.I)
        or re.search(r'for\s+time[:,\s]*' + _NUM_TOKEN_RE + r'\s+rounds?(?:\s+of\b)?', full, re.I)
        or re.search(r'por\s+tempo[:,\s]*' + _NUM_TOKEN_RE + r'\s+rounds?(?:\s+de\b)?', full, re.I)
        # 'N RFT' (abreviação Rounds For Time)
        or re.search(_NUM_TOKEN_RE + r'\s*rft\b', full, re.I)
        # Linha SÓ com a declaração de rounds ('5 Rounds:', '5 Rounds of:') — sem
        # precisar de 'for time' ao lado. Não pega 'then, N rounds of' (buy-in,
        # começa com 'then') nem 'N rounds per athlete'.
        or re.search(r'^\s*' + _NUM_TOKEN_RE + r'\s+rounds?(?:\s+(?:of|de))?\s*:?\s*$',
                     full, re.I | re.M)
    )
    if m_rounds:
        n = _num_ext(m_rounds.group(1))
        if n is not None and 2 <= n <= 30:   # sanity cap
            wkt["rounds_fixos"] = n
            wkt["tipo"] = "for_time"

    # EMOM (`every X minutes, for Y rounds`) — usa scorecard AMRAP
    m_emom = re.search(r'every\s+(\d+(?::\d+)?)\s*minutes?\s*,?\s*for\s+(\d+)\s+rounds?', full, re.I)
    if m_emom:
        wkt["tipo"] = "amrap"
        wkt["emom_janela"] = m_emom.group(1)
        n = _safe_int(m_emom.group(2))
        if n is not None: wkt["emom_rounds"] = n

    # Tie-break por round (cobrar tempo no final de cada round)
    tb_por_round = (
        re.search(r'(?:tie[\s-]?break|tb|desempate)[:\s-]*'
                  r'(?:tempo|time)?[^.\n]{0,40}'
                  r'(?:final|fim|end)\s+(?:de|of)?\s*(?:the\s+)?(?:cada|each)\s+(?:round|rodada)', full, re.I)
        or re.search(r'(?:tie[\s-]?break|tb|desempate)\s+(?:por|per|each|a\s+cada)\s+(?:round|rodada)', full, re.I))
    if tb_por_round:
        wkt["tiebreak_por_round"] = True
    else:
        # Tiebreak geral (For Time com critério específico — 'tempo ao fim das 21 pull-ups')
        for ln in lines:
            m_tb = re.match(r'\s*(?:tie[\s-]?break|tb|desempate)\s*[:\-]\s*(.+)$', ln, re.I)
            if m_tb:
                wkt["tiebreak"] = m_tb.group(1).strip()
                break
        # Fallback: header sozinho + linha-bullet na sequência (formato Monstar:
        # "Tiebreak\n- Será o tempo no último Pull-Up do Part 3.").
        if not wkt.get("tiebreak"):
            for i, ln in enumerate(lines):
                if re.match(r'^\s*(?:tie[\s-]?break|tb|desempate)\s*$', ln, re.I):
                    # Olha próxima(s) linha(s) — pula brancas e pega bullet/texto
                    for j in range(i + 1, min(i + 4, len(lines))):
                        nxt = lines[j].strip()
                        m_bullet = re.match(r'^[-•*·]\s*(.+)$', nxt)
                        if m_bullet:
                            wkt["tiebreak"] = m_bullet.group(1).strip()
                            break
                    if wkt.get("tiebreak"): break

    # Progressão de reps (*Add N reps each round)
    m_prog = (re.search(r'\*\s*(?:add|acrescent[ae]r?|adicione)\s+(\d+)\s+reps?\s+(?:each|a\s+cada|por)\s+round', full, re.I)
              or re.search(r'\*\s*\+\s*(\d+)\s+reps?\s+(?:each|a\s+cada|por)\s+round', full, re.I)
              or re.search(r'(?:add|acrescent[ae]r?|adicione)\s+(\d+)\s+reps?\s+(?:each|a\s+cada|por)\s+round', full, re.I))
    if m_prog:
        delta = _safe_int(m_prog.group(1))
        if delta is not None: wkt["reps_delta_por_round"] = delta

    # Goal de For Time tipo Simple Dimension/Mind. Para ANTES de '+', 'and',
    # dígitos extras, palavras-chave de chegada — evita capturar combos longos.
    # `[\w\-/.]` inclui ponto pra capturar "Sync. Snatches" inteiro (trios).
    m_goal = re.search(
        r'(?:goal|objetivo|alvo)\s*[:\-]?\s*(\d+)\s+'
        r'((?:[A-Za-zÀ-ú][\w\-/.]*)(?:\s+(?!\+|and\b|e\b|&|\d|finishing|chegada|cross)[A-Za-zÀ-ú][\w\-/.]*){0,3})',
        full, re.I)
    if m_goal:
        n = _safe_int(m_goal.group(1))
        if n is not None: wkt["goal_reps"] = n
        nome_mov = m_goal.group(2).strip().upper()
        nome_mov = re.sub(r'\s*(?:\+|and|e|&)\s.*$', '', nome_mov, flags=re.I).strip()
        wkt["goal_movimento"] = nome_mov

    # Tipo for_time_goal: detecta padrão literal "Goal: N X + finishing rep"
    # (Simple Dimension / Simple Mind). Só promove se já era for_time — não
    # mexe em AMRAP/Express/For Load mesmo que tenham palavra "goal" solta.
    if (wkt.get("tipo") == "for_time"
        and re.search(r'\bgoal\s*[:\-]\s*\d+.*?(?:finishing\s+rep|cross\s+the\s+line)',
                      full, re.I)):
        wkt["tipo"] = "for_time_goal"
        # Carga do goal: primeira linha 'Max <mov> (carga)' que aparecer.
        for line in lines:
            m_max = re.search(
                r'\bmax\b\s+[A-Za-zÀ-ú][\w\s\-/.]+?\s*\(([^)]+)\)', line, re.I)
            if m_max:
                wkt["goal_carga"] = m_max.group(1).strip()
                break

    # Último round vira MAX / AMRAP
    if (re.search(r'(?:last|[úu]ltimo|final)\s+round\s+(?:is\s+)?(?:max|amrap)', full, re.I)
        or re.search(r'(?:last|[úu]ltimo|final)\s+(?:round\s+)?(?:max|amrap)\s+reps?', full, re.I)
        or re.search(r'(?:round|rd)\s*\d+\s*[:=]\s*(?:max|amrap)', full, re.I)):
        wkt["ultimo_round_max"] = True


# Regexes pré-compiladas usadas no loop de movimentos (perf + clareza)
_PARALELO_RE = re.compile(r'^\s*(?:simultaneous(?:ly)?|paralelo|simultaneamente|'
                          r'simultane[oa])\b.*:\s*$', re.I)
_FIM_PARALELO_RE = re.compile(r'^\s*(?:after\s+both|after\s+all|then|ap[óo]s\s+(?:os\s+)?'
                              r'(?:dois|todos|ambos))\b', re.I)
_SKIP_PREFIXES = ('for time', 'por tempo', 'amrap', 'as many reps', 'rest',
                  'atenção', 'atencao', 'obs', 'note', '"', '“')
_MARKER_END_RE = re.compile(r'[*★↑↗](?:\s*\([^)*]*\))?\s*$')
_MARKER_INLINE_RE = re.compile(r'\((?:prog|progressivo|progressive|\+)\)\s*$', re.I)
_DIRECTIVE_PROG_RE = re.compile(
    r'^\s*\*?\s*(?:add|acrescent[ae]r?|adicione|\+)\s+\d+\s+reps?\s+(?:each|a\s+cada|por)\s+round', re.I)
_DIRECTIVE_GOAL_RE = re.compile(r'^\s*(?:goal|objetivo|alvo)\s*[:\-]', re.I)


def _parse_movimentos(lines: list[str], wkt: Workout) -> tuple[list[Movimento], str]:
    """Itera lines extraindo movimentos. Aplica regras de paralelo, marcadores
    progressivos, blocos `then...` e fallback flex-mov pra workouts com Goal.
    Retorna (lista de movimentos, time_cap). Não mutates wkt."""
    movs: list[Movimento] = []
    block = 1
    in_paralelo = False
    time_cap = ""
    has_then = any(re.match(r'^then\.+$', l, re.I) for l in lines)
    has_atleta = any(_ATLETA_HEADER_RE.match(l) for l in lines)
    has_seps = has_then or has_atleta
    has_goal = bool(wkt.get("goal_reps"))
    # Label "ATLETA N" ativo enquanto não chegar `then...` (que reseta).
    # Setado pelo header `Atleta N` no loop abaixo.
    atleta_label: Optional[str] = None

    for line in lines:
        ll = line.lower()
        # Aceita 'Time cap: 16 min', '16 minutes', '12:30 minutes' (mm:ss) e
        # 'Time cap: 16' (sem unidade). mm:ss preserva o formato (12:30 min).
        tc = re.search(r'time\s*cap[:\s]+(\d+(?::\d+)?)\s*(?:min\w*|minutos?)?', line, re.I)
        if tc:
            time_cap = f"{tc.group(1)} min"
            continue
        m_atl = _ATLETA_HEADER_RE.match(line)
        if m_atl:
            n = int(m_atl.group(1))
            atleta_label = f"ATLETA {n}"
            in_paralelo = False
            continue
        if re.match(r'^then[\.\s]*$', line, re.I):
            if movs: movs.append({"separador": "then..."})
            block += 1
            atleta_label = None
            in_paralelo = False
            continue
        if _PARALELO_RE.match(line):
            in_paralelo = True
            continue
        if _FIM_PARALELO_RE.match(line):
            in_paralelo = False   # não consome — 'After both: 21 Pull-Ups' tem mov após
        if any(ll.startswith(p) for p in _SKIP_PREFIXES): continue
        if _DIRECTIVE_GOAL_RE.match(line): continue   # 'Goal:' já capturada

        # Headers informativos ('Part 1 (00:00-06:00)', 'Stage 2', linha só com
        # janela de tempo) — NÃO são movimento, mas merecem aparecer na súmula
        # como banner. Criados como {secao: texto} pra render renderizar.
        s_clean = line.strip()
        # Bloco de rounds aninhado ('then, 2 rounds of:') — banner, preserva o
        # buy-in que veio antes (não vira {rounds_fixos} pra não multiplicar).
        if (m_rb := _ROUNDS_BLOCK_RE.match(s_clean)):
            # buy-in (o que veio antes) roda 1x; o bloco depois roda N rounds.
            # Guardamos N no marcador de seção — o render divide buy-in × bloco.
            n_rb = _num_ext(m_rb.group(1)) or 2
            movs.append({"secao": s_clean.rstrip(':').upper(), "rounds_bloco": n_rb})
            wkt["rounds_bloco"] = n_rb
            continue
        if (_SECTION_HEADER_RE.match(s_clean)
                or (not re.match(r'^\d', s_clean) and _TIME_WINDOW_RE.search(s_clean))):
            movs.append({"secao": s_clean.upper()})
            continue

        # Marca progressivo + remove markers do nome
        s_strip = line.strip()
        is_progressivo = bool(_MARKER_END_RE.search(s_strip)) or bool(_MARKER_INLINE_RE.search(s_strip))
        line_clean = line
        if not line_clean.lstrip().startswith(('*', '★', '↑', '↗')):
            line_clean = re.sub(r'[*★↑↗](?=\s|\(|$)', '', line_clean)
        line_clean = _MARKER_INLINE_RE.sub('', line_clean).rstrip()
        if _DIRECTIVE_PROG_RE.match(line_clean): continue   # '*Add N reps each round'

        parsed = _parse_mov_line(line_clean)
        if parsed:
            reps, nome = parsed
            nome_limpo, carga = _extrair_carga(nome)
            mov: Movimento = {"nome": nome_limpo}
            if reps is not None: mov["reps"] = reps
            if carga: mov["carga"] = carga
            if atleta_label:
                mov["label"] = atleta_label
            elif has_then and not has_atleta and block in BLOCK_LABELS:
                # Compat com workouts que usam só `then...` (sem header Atleta N)
                mov["label"] = BLOCK_LABELS[block]
            if in_paralelo: mov["paralelo"] = True
            if is_progressivo: mov["progressivo"] = True
            movs.append(mov)
        elif has_goal:
            # For Time com Goal: aceita movs sem reps líderes (Snatches 95/65 lb)
            mov_flex = _tentar_flex_mov(line_clean, has_seps, block, in_paralelo)
            if mov_flex:
                if atleta_label:
                    mov_flex["label"] = atleta_label
                elif has_then and not has_atleta and block in BLOCK_LABELS:
                    mov_flex["label"] = BLOCK_LABELS[block]
                movs.append(mov_flex)

    return movs, time_cap


def _tentar_flex_mov(line_clean: str, has_seps: bool,
                     block: int, in_paralelo: bool) -> Optional[Movimento]:
    """Fallback flex-mov pra For Time com Goal: aceita 'Snatches 95/65 lb'
    (sem reps líderes). Filtro estrito pra não pegar notas/regulamento."""
    if re.match(r'^\d', line_clean.strip()): return None   # tem reps — não é flex
    nome_raw = line_clean.strip()
    # Conta palavras IGNORANDO parênteses (carga "(75lb/55lb)", "(2 athletes)")
    # pra não desclassificar movs legítimos tipo "Max Sync. Snatches (75lb) (2 athletes)".
    nome_para_contar = re.sub(r'\s*\([^)]*\)', '', nome_raw).strip()
    palavras = nome_para_contar.split()
    # Limite: 5 palavras normalmente, até 10 se for combo "X + Y" (Simple
    # Dimension trio: "Max Sync. Wall-Ball Shots + Dumbbell Front Squats").
    max_palavras = 10 if '+' in nome_para_contar else 5
    # Restrições: sem pontuação final (frases têm); sem bullet de lista;
    # não pode ser header de seção tipo 'Part 1 (0:00-6:00)' nem ter janela
    # de tempo entre parens (sinal forte de header informativo).
    if not (1 <= len(palavras) <= max_palavras
            and not nome_raw.endswith(('.', ':', '!', '?', ';'))
            and not nome_raw.startswith(('-', '*', '•', '→'))
            and not _SECTION_HEADER_RE.match(nome_raw)
            and not _TIME_WINDOW_RE.search(nome_raw)
            and not _FRASE_NAO_MOVIMENTO_RE.search(nome_raw)
            and not _NOTA_LIKELY_RE.search(nome_raw)):
        return None
    # Strip sufixo "(N athletes)" / "(N athlete to completion)" / "(two dumbbells)"
    # antes da carga — trios encadeiam "(75lb/55lb) (2 athletes)", senão a carga
    # não fica no final. Faz múltiplas passadas pra cobrir sufixos empilhados.
    nome_pre_carga = nome_raw
    for _ in range(3):
        novo = re.sub(
            r'\s*\(\s*(?:\d+\s+athletes?\b[^)]*|two\s+dumbbells?|one\s+dumbbell)\s*\)\s*$',
            '', nome_pre_carga, flags=re.I).strip()
        if novo == nome_pre_carga: break
        nome_pre_carga = novo
    # Normaliza "75lb/55lb" → "75/55 lb" pra _CARGA_END_RE casar o par M/F
    nome_pre_carga = re.sub(
        r'(\d+(?:[.,]\d+)?)\s*(kg|lb|lbs|#|pood)\s*/\s*(\d+(?:[.,]\d+)?)\s*\2',
        r'\1/\3 \2', nome_pre_carga, flags=re.I)
    nome_limpo, carga = _extrair_carga(nome_pre_carga.upper())
    if not nome_limpo or len(nome_limpo) < 3: return None
    mov: Movimento = {"nome": nome_limpo}
    if carga: mov["carga"] = carga
    # Label aplicado pelo caller (_parse_movimentos sabe se é ATLETA N ou
    # Nº BLOCO baseado em has_atleta vs only-then). Ver linhas 547-555.
    if in_paralelo: mov["paralelo"] = True
    return mov


def _aplicar_progressao_reps(wkt: Workout) -> None:
    """Pós-processamento: gera mov.reps_por_round nos movs marcados como
    progressivos. Strict: SÓ aplica nos marcados com '*' explícito —
    diretriz '*Add N reps' sem markers não chuta geral. Mutação in-place."""
    delta = wkt.get("reps_delta_por_round", 0)
    if not delta: return
    movs = wkt.get("movimentos") or []
    if not movs: return
    n_rounds = wkt.get("emom_rounds") or wkt.get("n_rounds") or 5
    ultimo_max = wkt.get("ultimo_round_max", False)
    for m in movs:
        if m.get("chegada") or m.get("separador"): continue
        if not m.get("progressivo"): continue
        base = m.get("reps")
        if not isinstance(base, int): continue
        seq: list = [base + i * delta for i in range(n_rounds)]
        if ultimo_max and seq: seq[-1] = 'MAX'
        m["reps_por_round"] = seq


_COMPOSTO_HEADER_RE = re.compile(
    r'^["“‘](?P<f1>.+?)["”’]\s*\+\s*["“‘](?P<f2>.+?)["”’]\s*$'
)
# Linha que é SÓ um título entre aspas + janela de tempo OU '— Atleta N' opcional
# — cabeçalho de um sub-workout. Cobre o individual ('"Muscle Swim" (00:00-08:00)')
# e a dupla ('"Muscle Swim" — Atleta 1' / '"2k" — Atleta 2').
_COMPOSTO_TITULO_RE = re.compile(
    r'^\s*["“‘](?P<t>.+?)["”’]\s*'
    r'(?:\(\s*\d+:\d+\s*[-–—]\s*\d+:\d+\s*\)\s*)?'          # janela opcional
    r'(?:[—–-]\s*(?:athlete|atleta)\s*\d+\s*)?$',           # '— Atleta N' opcional
    re.I,
)
# NOTAS com qualquer traço decorativo (inclui U+2015 do Pwrd).
_COMPOSTO_NOTAS_RE = re.compile(r'[─―—–]{2,}\s*NOTAS\s*[─―—–]{2,}', re.I)


def _extrair_janela(texto: str, nome: str) -> str:
    """Pega a janela `(0:00-5:00)` que aparece após o nome da fórmula.

    Aceita também `–` (en-dash) entre os tempos. Retorna '' se não achar.
    """
    if not texto or not nome:
        return ''
    nome_esc = re.escape(nome)
    m = re.search(
        rf'["“‘]{nome_esc}["”’]\s*\(\s*(\d+:\d+)\s*[-–]\s*(\d+:\d+)\s*\)',
        texto, re.I)
    return f'{m.group(1)}–{m.group(2)}' if m else ''


def _detectar_composto(lines: list[str]) -> Optional[tuple[str, str, str, str]]:
    """Detecta workout composto pelo header `"X" + "Y"` na 1ª linha.

    Retorna (nome_f1, nome_f2, texto_f1, texto_f2) se for composto, None
    caso contrário. Split: tudo entre o header e a próxima ocorrência de
    `"NOME2"` (linha que abre F2) vira F1; do `"NOME2"` em diante vira F2.
    `─── NOTAS ───` corta o final de F2 — o regulamento fica anexado
    fora dos blocos pra cada F2 conseguir parsear sem ruído.
    """
    if not lines:
        return None
    m = _COMPOSTO_HEADER_RE.match(lines[0])
    if not m:
        # Estratégia 2: dois títulos entre aspas em linhas próprias, cada um
        # abrindo um sub-workout com sua janela ('"Muscle Swim" (00:00-08:00)'
        # … '"3k" (20:00-35:00)'). Vira composto F1/F2 (2 pontuações).
        return _detectar_composto_por_titulos(lines)
    nome_f1 = m.group('f1').strip()
    nome_f2 = m.group('f2').strip()
    # Acha onde F2 começa: linha que abre com `"NOME2"`
    idx_f2 = None
    nome_f2_norm = nome_f2.lower()
    for i, ln in enumerate(lines[1:], start=1):
        m2 = re.match(r'^["“‘](.+?)["”’]', ln)
        if m2 and m2.group(1).strip().lower() == nome_f2_norm:
            idx_f2 = i
            break
    if idx_f2 is None:
        return None  # não acha o início da F2 → trata como workout simples
    # NOTAS marca fim dos blocos
    idx_notas = next((i for i, ln in enumerate(lines)
                      if re.search(r'─{2,}\s*NOTAS\s*─{2,}', ln, re.I)), None)
    fim_f2 = idx_notas if idx_notas is not None else len(lines)
    # F1 começa em 1 (pula header) e vai até início de F2
    texto_f1 = '\n'.join(lines[1:idx_f2])
    texto_f2 = '\n'.join(lines[idx_f2:fim_f2])
    return nome_f1, nome_f2, texto_f1, texto_f2


def _detectar_composto_por_titulos(
    lines: list[str],
) -> Optional[tuple[str, str, str, str]]:
    """Composto sem header `X + Y`: dois títulos entre aspas em linhas próprias,
    cada um abrindo um sub-workout (Muscle Swim + 3k do Pwrd by Coffee).

    Só dispara com EXATAMENTE 2 títulos distintos (3+ é outro formato). Cada
    bloco tem seu 'For time:' e é pontuado separadamente. Retorna a mesma tupla
    de `_detectar_composto` ou None.
    """
    # NOTAS/regulamento pode conter aspas ('"Muscle Swim" e "2k"') — conta
    # títulos só ANTES das notas pra não inflar a contagem.
    idx_notas = next((i for i, ln in enumerate(lines)
                      if _COMPOSTO_NOTAS_RE.search(ln)), None)
    lim = idx_notas if idx_notas is not None else len(lines)
    titulos = [(i, m.group('t').strip())
               for i, ln in enumerate(lines[:lim])
               if (m := _COMPOSTO_TITULO_RE.match(ln)) and m.group('t').strip()]
    if len(titulos) != 2:
        return None
    (i1, nome_f1), (i2, nome_f2) = titulos
    if nome_f1.lower() == nome_f2.lower() or i2 <= i1:
        return None
    fim_f2 = lim
    if fim_f2 <= i2:
        fim_f2 = len(lines)
    texto_f1 = '\n'.join(lines[i1:i2])
    texto_f2 = '\n'.join(lines[i2:fim_f2])
    return nome_f1, nome_f2, texto_f1, texto_f2


def parse_workout_text(text: str, numero: int) -> Workout:
    """Converte texto livre de uma célula/seção num dict de workout.

    Pipeline:
      1. Extrai nome (primeira linha entre aspas, ou texto livre não-numérico)
      2. Detecta tipo (composto / express / for_load / for_time / amrap)
      3. Detecta diretrizes (relay, EMOM, tiebreak, progressão, Goal, MAX)
      4. Loop de movimentos (filtra noise, marca paralelo/progressivo, captura carga)
      5. Adiciona chegada (For Time) e aplica progressão (movs marcados)
    """
    lines = [l.strip() for l in str(text).split('\n') if l.strip()]
    wkt: Workout = {"numero": numero, "nome": f"WKT {numero}", "tipo": "for_time",
                    "modalidade": "individual", "time_cap": "",
                    "movimentos": [], "descricao": []}

    # 0) Composto — `"X" + "Y"` no header → 2 sub-workouts encadeados.
    # Pega o time_cap geral da última diretriz `Time cap:` no texto completo.
    composto = _detectar_composto(lines)
    if composto:
        nome_f1, nome_f2, texto_f1, texto_f2 = composto
        f1 = parse_workout_text(texto_f1, 1)
        f2 = parse_workout_text(texto_f2, 2)
        if not f1.get('nome') or f1['nome'].startswith('WKT '):
            f1['nome'] = nome_f1.upper()
        if not f2.get('nome') or f2['nome'].startswith('WKT '):
            f2['nome'] = nome_f2.upper()
        # Janela `(0:00-5:00)` de cada fórmula → tempo absoluto na timeline
        f1['janela'] = _extrair_janela(texto_f1, nome_f1)
        f2['janela'] = _extrair_janela(texto_f2, nome_f2)
        full = '\n'.join(lines)
        m_cap = re.search(r'time\s*cap\s*[:\-]?\s*([^\n]+?)(?:\.|$)', full, re.I)
        time_cap_total = m_cap.group(1).strip() if m_cap else (f2.get('time_cap') or f1.get('time_cap') or '')
        m_desc = re.search(
            r'descanse?\s+(?:de\s+)?(?:um|uma|dois|duas|tr[êe]s|\d+)\s+(?:minutos?|min|segundos?|s)\b',
            full, re.I)
        descanso = m_desc.group(0) if m_desc else ''
        return {
            'numero': numero,
            'nome': f'{nome_f1.upper()} + {nome_f2.upper()}',
            'tipo': 'composto',
            'modalidade': 'individual',
            'time_cap': time_cap_total,
            'movimentos': [],
            'descricao': [],
            'f1': f1,
            'f2': f2,
            'descanso': descanso,
        }

    # 1) Nome
    nome = _extrair_nome_workout(lines)
    if nome: wkt["nome"] = nome

    # 2) Tipo — Express, AMRAP multi-janela e For Load têm paths dedicados
    if any(re.search(r'express formula', l, re.I) for l in lines):
        return _parse_express(lines, wkt)
    # AMRAP de 2+ janelas ('AMRAP N min' repetido), ex.: PWRD Loop
    if sum(1 for l in lines if _AMRAP_JANELA_RE.match(l)) >= 2:
        return _parse_amrap_multijanela(lines, wkt, '\n'.join(lines))
    full = '\n'.join(lines).lower()
    if ('for load' in full or 'max lift' in full or 'max load' in full
        or re.search(r'\bcarga m[áa]xima\b', full)):
        return _parse_for_load(lines, wkt, full)
    if 'for time' in full or 'por tempo' in full:
        wkt["tipo"] = "for_time"
    elif 'amrap' in full or 'as many reps' in full:
        wkt["tipo"] = "amrap"

    # 3) Diretrizes (relay, EMOM, tiebreak, progressão, goal, MAX)
    _detectar_directives(full, lines, wkt)

    # 4) Movimentos — trunca antes em NOTAS pra não pegar regulamento
    lines_movs = _truncar_descricao_em_notas(lines)
    movs, time_cap = _parse_movimentos(lines_movs, wkt)
    if time_cap:
        wkt["time_cap"] = time_cap
    elif not wkt.get("time_cap"):
        # Fallback: o 'Time cap' pode aparecer no FIM, depois de Note/Score
        # (fora da região de movimentos truncada). É metadata — lê de qualquer
        # posição do texto completo. Negação ('não terá time cap') não casa (sem dígito).
        m_tc = re.search(r'time\s*cap[:\s]+(\d+(?::\d+)?)\s*(?:min\w*|minutos?)?',
                         '\n'.join(lines), re.I)
        if m_tc:
            wkt["time_cap"] = f"{m_tc.group(1)} min"

    # 5) For Time / For Time Goal fecham com chegada — A MENOS que o Excel diga
    #    que a chegada não conta como repetição (a especificidade vem do texto).
    chegada_nao_conta = bool(_CHEGADA_NEGADA_RE.search(full))
    if wkt["tipo"] in ("for_time", "for_time_goal") and movs and not chegada_nao_conta:
        movs.append({"chegada": True})
    wkt["movimentos"] = movs
    _aplicar_progressao_reps(wkt)

    # 6) For Time Goal: marca linhas que começam com "Max <X>" como goal:true.
    # Render usa pra renderizar badge GOAL e omitir checkbox de reps.
    if wkt["tipo"] == "for_time_goal":
        for mov in wkt["movimentos"]:
            nome_up = (mov.get("nome") or "").upper().strip()
            if nome_up.startswith("MAX "):
                mov["goal"] = True
        # 7) Identifica o mov-âncora do tiebreak ("último Pull-Up do Part 3"
        # → última linha PULL-UPS no '3º BLOCO'). Render usa pra colocar caixa
        # de tempo no momento certo. Sem o padrão, tiebreak fica só no rodapé.
        _marcar_mov_tiebreak(wkt)
    return wkt


def _marcar_mov_tiebreak(wkt: Workout) -> None:
    """Marca mov['tiebreak_aqui']=True na linha exata em que o juiz deve
    cobrar o tiebreak. Procura padrão 'último <X> do Part <N>' no texto do
    tiebreak e bate com o último mov com aquele label/nome."""
    tb = (wkt.get('tiebreak') or '').strip()
    if not tb: return
    m = re.search(r'[úu]ltim[oa]\s+(.+?)\s+(?:do|da)\s+part\s+(\d+)', tb, re.I)
    if not m: return
    nome_alvo = m.group(1).strip()
    part_n = _safe_int(m.group(2))
    if not part_n: return
    label_alvo = BLOCK_LABELS.get(part_n)
    if not label_alvo: return
    movs = wkt.get('movimentos') or []
    def _norm(s: str) -> str:
        # Compara só letras: 'Pull-Up' bate 'PULL-UPS' (descartando -, espaço, s final).
        return re.sub(r'[^A-Z]', '', (s or '').upper()).rstrip('S')
    alvo_norm = _norm(nome_alvo)
    if not alvo_norm: return
    alvo_idx = None
    for i, mv in enumerate(movs):
        if mv.get('label') != label_alvo: continue
        nm = _norm(mv.get('nome', ''))
        if alvo_norm in nm or nm in alvo_norm:
            alvo_idx = i   # mantém o ÚLTIMO match (não dá break)
    if alvo_idx is not None:
        movs[alvo_idx]['tiebreak_aqui'] = True


def _parse_express(lines: list[str], wkt: Workout) -> Workout:
    """Extrai fórmulas 1 e 2 de um workout Express."""
    wkt["tipo"] = "express"; wkt["estilo"] = "express"
    f1_lines: list[str] = []
    f2_lines: list[str] = []
    current = None
    f1_janela = f2_janela = ""

    for line in lines:
        m1 = re.search(r'Express Formula 1.{0,5}[([]?([0-9]{2}:[0-9]{2}[^)\]]*)', line, re.I)
        m2 = re.search(r'Express Formula 2.{0,5}[([]?([0-9]{2}:[0-9]{2}[^)\]]*)', line, re.I)
        if re.search(r'Express Formula 1', line, re.I):
            current = 'f1'
            if m1:
                j = m1.group(1).strip().strip(')').replace('-', ' -> ')
                f1_janela = j + '  .  AMRAP'
            continue
        if re.search(r'Express Formula 2', line, re.I):
            current = 'f2'
            if m2:
                j = m2.group(1).strip().strip(')').replace('-', ' -> ')
                f2_janela = j + '  .  FOR TIME'
            continue
        if current == 'f1': f1_lines.append(line)
        elif current == 'f2': f2_lines.append(line)

    def extract_movs(flines: list[str], add_chegada: bool = False) -> tuple[list[Movimento], str]:
        movs: list[Movimento] = []
        tc_val = ""
        for line in flines:
            tc = re.search(r'time\s*cap[:\s]+(\d+)\s*min', line, re.I)
            if tc: tc_val = f"{tc.group(1)} min"; continue
            if re.match(r'^then[\.\s]*$', line, re.I):
                if movs: movs.append({"separador": "then..."}); continue
            p = _parse_mov_line(line)
            if p:
                reps, nome = p
                mov: Movimento = {"nome": nome}
                if reps is not None: mov["reps"] = reps
                movs.append(mov)
        if add_chegada and movs: movs.append({"chegada": True})
        return movs, tc_val

    f1_movs, _     = extract_movs(f1_lines, False)
    f2_movs, tc    = extract_movs(f2_lines, True)
    if tc: wkt["time_cap"] = tc

    if re.search(r'\s+[12]$', wkt["nome"]):
        wkt["nome"] = re.sub(r'\s+[12]$', '', wkt["nome"]).strip()

    wkt["formula1"] = {"janela": f1_janela or "00:00 → 05:00  ·  AMRAP 5 MIN",
                       "descricao": [], "movimentos": f1_movs}
    wkt["formula2"] = {"janela": f2_janela or "06:00 → 12:00  ·  FOR TIME",
                       "descricao": [], "movimentos": f2_movs}
    return wkt


def _extrair_regra_pontuacao(full: str) -> str:
    """Pega a regra de pontuação das NOTAS ('Pontuação\\n- Será o total de ...').
    Retorna a frase principal (o que soma / o que conta), ou ''."""
    m = re.search(r'pontua[çc][ãa]o\s*[:\n]+(.+?)(?:\n\s*\n|\Z)', full, re.I | re.S)
    if not m:
        return ''
    linhas = [re.sub(r'^[\-–—•*\s]+', '', l).strip()
              for l in m.group(1).split('\n') if l.strip()]
    # 1ª linha costuma ser a regra do que soma; ignora a de "não conta".
    for l in linhas:
        if not re.search(r'n[ãa]o\s+cont', l, re.I):
            return l
    return linhas[0] if linhas else ''


def _parse_amrap_multijanela(lines: list[str], wkt: Workout, full: str) -> Workout:
    """AMRAP de múltiplas janelas (PWRD Loop do Pwrd by Coffee): cada bloco
    'AMRAP N min' é uma janela com reps prescritas + (opcional) uma linha 'Max'
    que é o que pontua. Descanso entre janelas. Score = soma das reps 'Max'.

    Estrutura: wkt['janelas'] = [{'titulo','movimentos':[{nome,reps,pontua} |
    {nome,max:True,pontua:True}], 'rest_depois'?}]. Movimento prescrito (reps
    fixas) NÃO pontua quando existe linha Max; sem Max, é AMRAP normal (tudo conta).
    """
    wkt["tipo"] = "amrap"
    lines_movs = _truncar_descricao_em_notas(lines)
    janelas: list[dict] = []
    atual: Optional[dict] = None
    rest_txt = ""
    for line in lines_movs:
        if _AMRAP_JANELA_RE.match(line):
            atual = {"titulo": line.strip().rstrip(':'), "movimentos": []}
            janelas.append(atual)
            continue
        if _REST_JANELA_RE.match(line):
            rest_txt = line.strip()
            if atual is not None:
                atual["rest_depois"] = line.strip()
            continue
        if atual is None:
            continue
        mm = _MAX_MOV_RE.match(line)
        if mm:
            nome = mm.group(1).strip().rstrip('.').upper()
            atual["movimentos"].append({"nome": nome, "max": True, "pontua": True})
            continue
        p = _parse_mov_line(line)
        if p:
            reps, nome = p
            mov: Movimento = {"nome": nome, "pontua": False}
            if reps is not None:
                mov["reps"] = reps
            atual["movimentos"].append(mov)
    # Sem linha Max em nenhuma janela → AMRAP normal: tudo conta pontuação.
    tem_max = any(m.get("max") for j in janelas for m in j["movimentos"])
    if not tem_max:
        for j in janelas:
            for m in j["movimentos"]:
                m["pontua"] = True
    wkt["janelas"] = janelas
    wkt["rest_entre"] = rest_txt
    wkt["score_regra"] = _extrair_regra_pontuacao(full)
    wkt["movimentos"] = []
    return wkt


# ── Excel import ────────────────────────────────────────────────────────────────
def _inferir_modalidade(nome_categoria: str) -> str:
    """Infere a modalidade ('individual'|'dupla'|'trio'|'quarteto'|'time') a
    partir do nome da categoria. Usado pelo renderer pra escolher labels
    ('Nome do Atleta' vs 'Nome do Trio' etc).

    Heurística por palavra-chave em PT-BR e EN. Ordem importa — checa
    palavras mais específicas antes (quarteto antes de time/equipe).
    """
    s = (nome_categoria or '').lower()
    if 'quarteto' in s or 'quartet' in s:    return 'quarteto'
    if 'trio' in s:                          return 'trio'
    if 'dupla' in s or ' pair' in s or 'pairs' in s: return 'dupla'
    if 'time' in s or 'team' in s or 'equipe' in s:  return 'time'
    return 'individual'


def _is_categoria_grid(ws) -> bool:
    """Detecta se a aba tem formato grade (colunas=categorias, linhas=workouts)."""
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if len(rows) < 2: return False
    r1 = [c for c in rows[0] if c is not None]
    r2 = [c for c in rows[1] if c is not None]
    # A 2ª linha precisa ter ALGUMA célula multi-linha (o texto do workout). Não
    # exigimos que seja a 1ª célula: em grades com coluna de dia (col A = 'Sexta'),
    # o dia pode vir sem data/quebra-de-linha e o workout fica na coluna seguinte.
    return (len(r1) >= 2
            and all(isinstance(v, str) for v in r1[:4])
            and any(isinstance(v, str) and '\n' in v for v in r2))


# ── Fallback de parsing (Fase 2: IA como reparador) ─────────────────────────
# Hook opcional injetado pelo app. Assinatura:
#   reparador(raw_text, numero, wkt_regex, problemas) -> Workout | None
# Mantém parsers.py livre de dependência de IA/anthropic — o core segue
# testável sozinho. O app registra o reparador no startup.
_REPARADOR_WORKOUT = None


def registrar_reparador(fn) -> None:
    """Registra o callback de reparo (ex.: ai_parser.reparar_workout_ia).
    Passar None desliga o fallback."""
    global _REPARADOR_WORKOUT
    _REPARADOR_WORKOUT = fn


def parse_workout_text_robusto(text: str, numero: int) -> Workout:
    """parse_workout_text + fallback: se a regex produz um parse que FALHA no
    schema canônico e há um reparador registrado, tenta reparar (IA). Só aceita
    o resultado reparado se ele passar no schema — senão devolve o da regex.
    NUNCA fica pior que a regex sozinha."""
    wkt = parse_workout_text(text, numero)
    if _REPARADOR_WORKOUT is None:
        return wkt
    problemas = validar_workout_schema(wkt, text)
    if not problemas:
        return wkt
    try:
        reparado = _REPARADOR_WORKOUT(text, numero, wkt, problemas)
    except Exception:
        return wkt
    if reparado and not validar_workout_schema(reparado, text):
        return reparado
    return wkt


# ── Schema canônico + validação ─────────────────────────────────────────────
# Tipos canônicos de workout — todo parse deve cair num destes.
TIPOS_WORKOUT = ('for_time', 'for_time_goal', 'amrap', 'express', 'for_load', 'composto')

_TC_NEGADO_RE = re.compile(
    r'n[ãa]o\s+(?:ter[áa]|tem|h[áa])\s+time\s*cap|sem\s+time\s*cap|no\s+time\s*cap', re.I)


def validar_workout_schema(wkt: Workout, raw: str = '') -> list[tuple[str, str]]:
    """Valida invariantes estruturais de um workout parseado (schema canônico).

    Contrato de um workout parseado:
      - `tipo` ∈ TIPOS_WORKOUT
      - tem conteúdo: movimentos | janelas | fórmula (express/composto) | for_load
      - `nome` não é a linha 'Arena:' nem o placeholder 'WKT N'
      - se o texto cru tem linha 'Max'/'Goal', o parse representa a pontuação
        (goal_reps, janela.max ou movimento.goal) — não pode DROPAR o que pontua
      - se o texto tem 'Time cap: N' (não negado), `time_cap` é capturado

    Retorna lista de (codigo, detalhe); vazia = ok. Base do corpus de regressão
    e reutilizável pelo linter de import. NÃO garante correção semântica total
    (formatos exóticos podem passar aqui e ainda precisar de revisão) — cobre as
    falhas estruturais que já mordem em produção.
    """
    probs: list[tuple[str, str]] = []
    tipo = wkt.get('tipo')
    if tipo not in TIPOS_WORKOUT:
        probs.append(('tipo_desconhecido', repr(tipo)))
    tem_conteudo = bool(wkt.get('movimentos') or wkt.get('janelas')
                        or wkt.get('formula1') or wkt.get('f1') or tipo == 'for_load')
    if not tem_conteudo:
        probs.append(('sem_conteudo', 'nenhum movimento/janela/fórmula'))
    nome = wkt.get('nome') or ''
    if nome.lower().startswith('arena') or nome.startswith('WKT '):
        probs.append(('nome_invalido', nome[:40]))
    if raw:
        rl = raw.lower()
        tem_score_raw = bool(re.search(r'(?m)^\s*max\.?\s', rl)) or 'goal:' in rl
        tem_score_par = (bool(wkt.get('goal_reps'))
                         or any(m.get('max') for j in (wkt.get('janelas') or [])
                                for m in j.get('movimentos', []))
                         or any(m.get('goal') for m in (wkt.get('movimentos') or [])))
        if tem_score_raw and not tem_score_par:
            probs.append(('pontuacao_perdida', 'texto tem Max/Goal e o parse não capturou'))
        if ('time cap' in rl and not _TC_NEGADO_RE.search(raw)
                and not wkt.get('time_cap') and tipo not in ('for_load', 'composto')):
            probs.append(('timecap_perdido', 'texto tem Time cap e não foi capturado'))
    return probs


def parse_excel(data: bytes) -> dict[str, Any]:
    """Parser unificado de Excel.

    Sempre retorna shape `evento_multidia`. Os formatos legados (categoria_grid
    e template) são detectados e convertidos por adapters internos pra que o
    resto do sistema trabalhe num modelo único.

    Quando há aba `Equipamento(s)`/`Equipment`, aplica as anilhas + unidade
    globais a todos os workouts For Load do evento.
    """
    if not HAS_EXCEL:
        raise RuntimeError("openpyxl não disponível — instale com: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    # Formato multi-dia: aba `Workouts` + abas `<Dia>` e `<Dia> - Montagem`
    if _is_evento_multidia(wb):
        result = parse_excel_multidia(wb)
    elif _is_layout_grades_e_dias(wb):
        # Grades-por-modalidade: 1+ abas grade (ex: Individuais, Duplas, Times)
        # + abas <Dia> e <Dia> - Montagem (sem aba unificada Workouts)
        result = parse_excel_grades_e_dias(wb)
    else:
        # Formato categoria_grid (modelo legado: 1 aba grade categoria × workout)
        todas_categorias: dict[str, list[Workout]] = {}
        evento_nome = ""
        for sname in wb.sheetnames:
            ws = wb[sname]
            if _is_categoria_grid(ws):
                resultado = _parse_excel_grade(wb, sname)
                todas_categorias.update(resultado.get("por_categoria", {}))
                if not evento_nome:
                    evento_nome = resultado.get("evento_nome", "")
        atletas_por_categoria = _parse_atletas(wb)
        if todas_categorias:
            result = _adaptar_categoria_grid_para_multidia(
                evento_nome, todas_categorias, atletas_por_categoria,
            )
        else:
            # Fallback final: formato template (1 evento, lista plana de workouts)
            template_result = _parse_excel_template(wb)
            if not template_result.get('workouts') and not template_result.get('evento', {}).get('nome'):
                # Nenhum formato reconhecido: erro explícito é melhor que
                # estrutura fantasma "Único / Geral" que confunde a UI.
                return {'tipo': 'erro', 'erro': 'Excel sem dados reconhecíveis — esperava grade categoria×workout, formato multi-dia, ou template Evento+WKT.'}
            result = _adaptar_template_para_multidia(template_result)

    # Aba Equipamento: aplica anilhas + unidade globais a todos os For Load
    equip = _parse_equipamento(wb)
    if equip:
        result['equipamento'] = equip
        result['unidade_default'] = equip['unidade']
        _aplicar_equipamento_a_for_load(result, equip)
    # Enriquece roster com categoria (via faixa de número da aba Inscritos).
    # Usado pra súmulas "pré-evento" — atleta inscrito mas sem bateria/raia ainda.
    inscritos_faixas = _parse_inscritos(wb)
    if inscritos_faixas and result.get('roster'):
        _enriquecer_roster_com_categoria(result['roster'], inscritos_faixas, result.get('dias') or [])
    return result


def _enriquecer_roster_com_categoria(roster: list[dict],
                                      inscritos: dict[str, tuple[int, int]],
                                      dias: list[dict]) -> None:
    """Adiciona campo 'categoria' a cada entry do roster baseado na faixa de
    número da aba Inscritos. Mutação in-place.

    Categoria armazenada é o nome ORIGINAL da categoria nos dias[] (não a
    normalizada usada como chave em inscritos). O match usa duas estratégias
    em ordem:
      1. Match estrito de normalização (preserva distinções tipo
         'Rx Misto (Iniciante)' vs 'Rx Misto (Avançado)')
      2. Cruza com a FAIXA de número da categoria nos dias (se houver) pra
         desambiguar quando inscritos usa forma relaxada.
    Sem match, categoria=''.
    """
    # Índices por normalização (estrita e relaxada) → lista de nomes reais.
    # Lista (não setdefault sozinho) preserva múltiplas candidatas pra detectar
    # ambiguidade depois.
    por_estrita: dict[str, list[str]] = {}
    por_relaxada: dict[str, list[str]] = {}
    for dia in dias:
        for cat in dia.get('categorias', []) or []:
            nome = cat.get('nome', '')
            if not nome: continue
            ke = _normalizar_categoria(nome)
            kr = _normalizar_categoria_relaxada(nome)
            if nome not in por_estrita.setdefault(ke, []):
                por_estrita[ke].append(nome)
            if nome not in por_relaxada.setdefault(kr, []):
                por_relaxada[kr].append(nome)
    for atl in roster:
        try:
            num = int(str(atl.get('numero', '')).strip())
        except (ValueError, AttributeError):
            atl.setdefault('categoria', '')
            continue
        cat_match = ''
        for cat_norm, (n_ini, n_fim) in inscritos.items():
            if not (n_ini <= num <= n_fim): continue
            # Tenta estrita primeiro (preserva 'Rx Misto (Iniciante)' vs '(Avançado)')
            candidatas = por_estrita.get(cat_norm) or por_relaxada.get(cat_norm) or []
            if len(candidatas) == 1:
                cat_match = candidatas[0]
            # Múltiplas candidatas (ambíguo) ou nenhuma → cat_match='' silencioso;
            # usuário resolve manualmente, evita assumir errado.
            break
        atl['categoria'] = cat_match


def _aplicar_equipamento_a_for_load(result: dict[str, Any], equip: dict[str, Any]) -> None:
    """Itera dias→categorias→workouts e injeta anilhas + unidade nos For Load.

    Só seta se o workout ainda não tiver valor explícito — config manual
    posterior (no front) sobrescreve.
    """
    anilhas = equip['anilhas']
    unidade = equip['unidade']
    for dia in result.get('dias', []) or []:
        for cat in dia.get('categorias', []) or []:
            for wkt in cat.get('workouts', []) or []:
                if wkt.get('tipo') != 'for_load': continue
                wkt.setdefault('anilhas', anilhas)
                wkt.setdefault('unidade', unidade)


def _adaptar_categoria_grid_para_multidia(
    evento_nome: str,
    por_categoria: dict[str, list[Workout]],
    atletas_por_categoria: dict[str, list[Atleta]],
) -> dict[str, Any]:
    """Adapter: shape antigo categoria_grid → shape novo evento_multidia (1 dia 'Único')."""
    cats: list[dict[str, Any]] = []
    for workouts in por_categoria.values():
        padronizar_workouts(workouts)
    for cat_nome, workouts in por_categoria.items():
        atletas = atletas_por_categoria.get(cat_nome, [])
        baterias: list[dict[str, Any]] = []
        if atletas:
            baterias.append({
                'numero': '1',
                'codigo_evento': '',
                'horario_aquecimento': '',
                'horario_fila': '',
                'workouts_que_rodam': list(range(1, len(workouts) + 1)),
                'alocacoes': [
                    {
                        'raia':   a.get('raia', '') or str(i + 1),
                        'numero': a.get('numero', ''),
                        'nome':   a.get('nome', ''),
                        'box':    a.get('box', ''),
                    }
                    for i, a in enumerate(atletas)
                ],
            })
        cats.append({'nome': cat_nome, 'workouts': workouts, 'baterias': baterias})

    return {
        'tipo': 'evento_multidia',
        'evento_nome': evento_nome,
        'dias': [{'label': 'Único', 'categorias': cats}],
        'roster': [],
    }


def _adaptar_template_para_multidia(template_result: dict[str, Any]) -> dict[str, Any]:
    """Adapter: shape antigo template (1 evento, lista plana) → evento_multidia."""
    evento = template_result.get('evento', {}) or {}
    workouts = template_result.get('workouts', []) or []
    cat_nome = evento.get('categoria', '') or 'Geral'
    return {
        'tipo': 'evento_multidia',
        'evento_nome': evento.get('nome', ''),
        'dias': [{
            'label': 'Único',
            'categorias': [{
                'nome': cat_nome,
                'workouts': workouts,
                'baterias': [],
            }],
        }],
        'roster': [],
    }


def _parse_atletas(wb) -> dict[str, list[Atleta]]:
    """Lê atletas de abas dedicadas (Atleta(s), Inscritos, Athletes, Participants).

    Antes varria toda aba procurando coluna 'Nome' — frágil, pegava header
    'Nome' da Montagem (que tem N blocos repetidos) e contaminava o roster.
    Agora filtra por nome de aba pra ser explícito sobre intenção do usuário.

    Retorna dict { categoria: [ {nome, box, raia, bateria, numero}, ... ] }.
    """
    # Prefixos/substrings que indicam aba de atletas. Tolera variações como
    # 'Atletas - Individuais', 'Inscritos', 'Athletes (RX)', etc.
    ATLETA_SHEET_KEYWORDS = (
        'atleta', 'atletas', 'inscritos', 'athletes', 'participants', 'participantes',
    )

    CAMPOS = {
        "nome":    ["nome", "atleta", "name", "athlete"],
        "box":     ["box", "afiliacao", "afiliação", "affiliate", "team"],
        "raia":    ["raia", "lane"],
        "bateria": ["bateria", "heat", "bat"],
        "numero":  ["numero", "número", "nº", "no", "number", "id", "inscricao", "inscrição"],
        "categoria": ["categoria", "category", "cat"],
    }

    def encontrar_col(header_row, opcoes):
        for i, v in enumerate(header_row):
            if v and str(v).strip().lower() in opcoes:
                return i
        return None

    def eh_aba_de_atletas(sname: str) -> bool:
        sl = sname.strip().lower()
        return any(kw in sl for kw in ATLETA_SHEET_KEYWORDS)

    resultado: dict[str, list[Atleta]] = {}

    for sname in wb.sheetnames:
        if not eh_aba_de_atletas(sname):
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue

        header_row_idx = None
        header_row = None
        for ri in range(min(3, len(rows))):
            row = [str(v).strip().lower() if v else "" for v in rows[ri]]
            if any(op in row for op in CAMPOS["nome"]):
                header_row_idx = ri
                header_row = rows[ri]
                break
        if header_row is None: continue

        col = {k: encontrar_col(header_row, v) for k, v in CAMPOS.items()}
        if col["nome"] is None: continue

        for row in rows[header_row_idx + 1:]:
            if not row or all(v is None for v in row): continue
            def cell(idx):
                if idx is None: return ""
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ""

            nome = cell(col["nome"])
            if not nome: continue

            atleta: Atleta = {
                "nome":    nome,
                "box":     cell(col["box"]),
                "raia":    cell(col["raia"]),
                "bateria": cell(col["bateria"]),
                "numero":  cell(col["numero"]),
            }

            cat = cell(col["categoria"]) if col["categoria"] is not None else sname
            if not cat or cat.lower() in ("atletas", "inscritos", "participants", "athletes"):
                cat = sname

            if cat not in resultado:
                resultado[cat] = []
            resultado[cat].append(atleta)

    return resultado


_DIAS_SEMANA_NORM = {
    # PT-BR (com e sem acento)
    'sexta': 'sexta', 'sábado': 'sábado', 'sabado': 'sábado', 'domingo': 'domingo',
    'segunda': 'segunda', 'terça': 'terça', 'terca': 'terça',
    'quarta': 'quarta', 'quinta': 'quinta',
    # EN
    'friday': 'sexta', 'saturday': 'sábado', 'sunday': 'domingo',
    'monday': 'segunda', 'tuesday': 'terça', 'wednesday': 'quarta', 'thursday': 'quinta',
    # Curtos comuns
    'sex': 'sexta', 'sáb': 'sábado', 'sab': 'sábado', 'dom': 'domingo',
}


def _extrair_dia_de_celula(col_a: str) -> tuple[str, str]:
    """Lê coluna A de uma linha de workout grade ('Sexta\\n29/05/2026') e
    retorna (dia_normalizado, data). Vazio se não bater num dia conhecido."""
    if not col_a: return ('', '')
    primeira_linha = str(col_a).split('\n')[0].strip().lower()
    if not primeira_linha: return ('', '')
    # Pega só a palavra do dia (ignora data colada tipo 'Sexta 29/05')
    palavra = re.split(r'[\s,/\-]+', primeira_linha)[0]
    dia_norm = _DIAS_SEMANA_NORM.get(palavra, '')
    if not dia_norm: return ('', '')
    # Tenta extrair data (formato BR ou ISO)
    m_data = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2})', str(col_a))
    data = m_data.group(1) if m_data else ''
    return (dia_norm, data)


def _parse_excel_grade(wb, sname: str) -> dict[str, Any]:
    """Parseia formato grade: col=categoria, linha=workout.

    Coluna A (opcional) pode conter o dia + data da linha — workouts ganham
    `_dia_label` e `_dia_data` pra filtragem posterior no shape multi-dia.
    """
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"erro": "Planilha vazia"}

    categorias: list[tuple[int, str]] = []
    for col_idx, val in enumerate(rows[0]):
        if val is not None:
            categorias.append((col_idx, str(val).strip()))

    por_categoria: dict[str, list[Workout]] = {}
    for cat_idx, cat_nome in categorias:
        modalidade = _inferir_modalidade(cat_nome)
        workouts: list[Workout] = []
        for row_num, row in enumerate(rows[1:], 1):
            if cat_idx >= len(row) or row[cat_idx] is None: continue
            cell_text = str(row[cat_idx]).strip()
            if not cell_text: continue
            # Extrai 'Arena: <nome>' antes de parsear pra evitar que vire o nome
            # do workout. parse_workout_text pega a primeira linha como nome —
            # se a 1ª linha for 'Arena: HeleFitness', sem essa extração o nome
            # vira 'ARENA: HELEFITNESS' em vez do nome real (próxima linha).
            arena, texto_limpo = _extrair_arena(cell_text)
            wkt = parse_workout_text_robusto(texto_limpo, row_num)
            if arena:
                wkt['arena'] = arena
            wkt['modalidade'] = modalidade   # inferido do nome da categoria
            # Coluna A: dia + data desta linha. Se presente, marca o wkt
            # pra filtragem por dia no shape multi-dia.
            dia_label, dia_data = _extrair_dia_de_celula(str(row[0]) if row[0] else '')
            if dia_label:
                wkt['_dia_label'] = dia_label
                if dia_data: wkt['_dia_data'] = dia_data
            workouts.append(wkt)
        if workouts:
            por_categoria[cat_nome] = workouts

    evento_nome = sname if sname.lower() not in ('individuais', 'duplas', 'equipamento') else ""

    return {
        "tipo": "categoria_grid",
        "evento_nome": evento_nome,
        "categorias": [c for _, c in categorias if c in por_categoria],
        "por_categoria": por_categoria,
    }


def _parse_excel_template(wb) -> dict[str, Any]:
    """Parseia formato template (Evento + Workouts + WKT1, WKT2...)."""
    config: dict[str, Any] = {"evento": {"nome": "", "categoria": "", "data": ""}, "workouts": []}
    wkt_map: dict[int, Workout] = {}
    for sname in wb.sheetnames:
        sl = sname.strip().lower()
        if sl == "evento":
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]: continue
                k = str(row[0]).strip().lower()
                v = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if k in ("nome", "name", "evento"): config["evento"]["nome"] = v
                elif k in ("categoria", "category"): config["evento"]["categoria"] = v
                elif k in ("data", "date"): config["evento"]["data"] = v
        m = re.match(r'^(?:wkt|workout)\s*[-_]?\s*(\d+)$', sl)
        if not m: continue
        num = int(m.group(1))
        ws = wb[sname]; hdrs = None; movs: list[Movimento] = []
        for row in ws.iter_rows(values_only=True):
            if not any(row): continue
            if hdrs is None: hdrs = [str(c or "").strip().lower() for c in row]; continue
            first = str(row[0] or "").strip().lower()
            if first in ("then...", "then", "então", "---"): movs.append({"separador": "then..."}); continue
            if first in ("chegada", "finish", "arrival"): movs.append({"chegada": True}); continue
            mov: Movimento = {}
            for i, h in enumerate(hdrs):
                if i >= len(row) or row[i] is None: continue
                v = str(row[i]).strip()
                if not v: continue
                if h in ("movimento", "exercise", "movement", "nome", "name"): mov["nome"] = v.upper()
                elif h in ("reps", "rep", "repetições"):
                    try: mov["reps"] = int(float(v))
                    except (ValueError, TypeError): mov["reps"] = v
                elif h in ("label", "bloco", "grupo", "block"): mov["label"] = v
            if "nome" in mov: movs.append(mov)
        wkt: Workout = {"numero": num, "nome": f"WKT {num}", "tipo": "for_time",
                        "modalidade": "individual", "time_cap": "", "movimentos": movs}
        config["workouts"].append(wkt); wkt_map[num] = wkt
    config["workouts"].sort(key=lambda w: w.get("numero", 0))
    return config


# ── PDF import ──────────────────────────────────────────────────────────────────
def parse_pdf(data: bytes) -> dict[str, Any]:
    if not HAS_PDF:
        raise RuntimeError("pdfplumber não disponível — instale com: pip install pdfplumber")
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    sections = re.split(r'\n(?=(?:Workout|WKT)\s+\d+)', full_text, flags=re.I)

    config: dict[str, Any] = {"evento": {"nome": "", "categoria": "", "data": ""}, "workouts": []}

    header_lines = [l.strip() for l in full_text.split('\n')[:8] if l.strip()]
    for line in header_lines:
        if len(line) > 4 and not re.match(r'^(workout|wkt|\d)', line, re.I):
            config["evento"]["nome"] = line
            break

    wkt_num = 0
    for sec in sections:
        sec = sec.strip()
        if not sec: continue
        has_wkt_hdr = re.match(r'^(?:Workout|WKT)\s+(\d+)', sec, re.I)
        has_quoted  = re.search(r'["“].+["”]', sec)
        has_movs    = re.search(r'^\d{1,3}\s+\w', sec, re.M)
        if not (has_wkt_hdr or (has_quoted and has_movs)): continue
        wkt_num += 1
        wkt = parse_workout_text_robusto(sec, wkt_num)
        config["workouts"].append(wkt)

    if not config["workouts"]:
        lines = [l.strip() for l in full_text.split('\n') if l.strip()]
        current = None
        for line in lines:
            tc = re.search(r'time\s*cap[:\s]+(\d+)\s*min', line, re.I)
            p = _parse_mov_line(line)
            m_name = re.match(r'^["“](.+?)["”]', line)
            if m_name and current is None:
                wkt_num += 1
                current = {"numero": wkt_num, "nome": m_name.group(1).upper(),
                           "tipo": "for_time", "modalidade": "individual",
                           "time_cap": "", "movimentos": [], "descricao": []}
                config["workouts"].append(current)
            elif current:
                if tc: current["time_cap"] = f"{tc.group(1)} min"
                elif p:
                    reps, nome = p
                    mov = {"nome": nome}
                    if reps is not None: mov["reps"] = reps
                    current["movimentos"].append(mov)
                elif re.match(r'^then[\.\s]*$', line, re.I):
                    current["movimentos"].append({"separador": "then..."})
        for wkt in config["workouts"]:
            if wkt.get("tipo") == "for_time" and wkt.get("movimentos"):
                wkt["movimentos"].append({"chegada": True})

    for wkt in config["workouts"]:
        if wkt.get("tipo") == "for_time" and wkt.get("movimentos"):
            if not any(m.get("chegada") for m in wkt["movimentos"]):
                wkt["movimentos"].append({"chegada": True})
    return config


# ── Helpers de ordenação e numeração ────────────────────────────────────────────
def _atleta_sort_key(a: Atleta) -> tuple:
    """Chave de ordenação para impressão sequencial: bateria → raia → nome.
    Raia é tratada numericamente quando possível ("10" depois de "2")."""
    bateria  = str(a.get('bateria', '') or '').strip().upper()
    raia_raw = str(a.get('raia', '') or '').strip()
    m = re.match(r'^(\d+)', raia_raw)
    raia_num = int(m.group(1)) if m else 10**9
    nome = str(a.get('nome', '') or '').strip().lower()
    return (bateria, raia_num, raia_raw.lower(), nome)


def assign_workout_numbers(workouts: list[Workout]) -> list[Workout]:
    """Recalcula números de workouts considerando slots.
    Express e Composto ocupam 2 slots (N e N+1). Outros ocupam 1 slot.
    Modifica a lista in-place e retorna ela.
    """
    counter = 1
    for wkt in workouts:
        wkt['numero'] = counter
        if wkt.get('tipo') in ('express', 'composto'):
            wkt['numero_f2'] = counter + 1
            counter += 2
        else:
            wkt.pop('numero_f2', None)
            counter += 1
    return workouts


def assign_workout_numbers_global(dias: list) -> None:
    """Numera workouts em sequência CONTÍNUA por categoria, atravessando dias.

    Ex: Elite Masc com 3 wkts na Sexta + 2 no Sábado vira 1,2,3 e 4,5 — em vez
    de reiniciar 1,2 no Sábado. Express e Composto ocupam 2 slots.
    Mutação in-place nos wkt['numero'] / wkt['numero_f2'] de cada dia.
    """
    counters: dict[str, int] = {}
    for dia in dias or []:
        for cat in (dia.get('categorias') or []):
            nome = (cat.get('nome') or '').strip()
            counter = counters.get(nome, 1)
            for wkt in (cat.get('workouts') or []):
                wkt['numero'] = counter
                if wkt.get('tipo') in ('express', 'composto'):
                    wkt['numero_f2'] = counter + 1
                    counter += 2
                else:
                    wkt.pop('numero_f2', None)
                    counter += 1
            counters[nome] = counter


# ── Excel multi-dia (formato real do evento) ──────────────────────────────────
# Formato esperado:
#   - Aba `Workouts`: grade dia (col A) × categoria (cols B+); cada célula é
#     o texto livre do workout. Linha de header de categorias se repete.
#   - Aba `<Dia>` (ex: `Sexta`, `Sábado`, `Domingo`): cronograma de baterias
#     com colunas Eventos | Categoria | Bateria | Arbitragem | Quantidade |
#     Aquecimento | <em branco> | Fila.
#   - Aba `<Dia> - Montagem`: blocos por bateria com header em 3 linhas
#     (horário, código+categoria, "Raia | Número | Nome | Box") seguido das
#     linhas de raia. Raias com #N/A são vazias.
#   - Aba `Atletas` (opcional): roster informativo de individuais.
#
# Convenção de arena: linha `Arena: <nome>` em qualquer ponto do texto livre
# do workout (na aba `Workouts`). É extraída e mostrada no header da súmula.

_DIA_LABELS_VALIDOS = ("segunda", "terça", "terca", "quarta", "quinta",
                        "sexta", "sábado", "sabado", "domingo")


def _is_evento_multidia(wb) -> bool:
    """Detecta se o arquivo é um evento multi-dia.

    Critérios: existe uma aba chamada `Workouts` E pelo menos uma aba do tipo
    `<Dia> - Montagem` (qualquer dia da semana).
    """
    nomes_lower = [s.lower() for s in wb.sheetnames]
    if 'workouts' not in nomes_lower:
        return False
    return any(' - montagem' in n for n in nomes_lower)


def _extrair_arena(texto: str) -> tuple[str, str]:
    """Extrai a primeira linha `Arena: <nome>` do texto livre do workout.

    Retorna (arena, texto_sem_linha_de_arena). Case-insensitive. Se não houver
    linha de arena, retorna ("", texto_original).
    """
    if not texto:
        return "", texto or ""
    linhas = texto.split('\n')
    arena = ""
    out: list[str] = []
    for linha in linhas:
        if not arena:
            m = re.match(r'^\s*arena\s*:\s*(.+?)\s*$', linha, re.I)
            if m:
                arena = m.group(1).strip()
                continue   # remove a linha do texto
        out.append(linha)
    return arena, '\n'.join(out)


def _parse_workouts_grade_multidia(ws) -> dict[str, dict[str, dict[str, Any]]]:
    """Lê a aba `Workouts` e retorna mapa { dia → { categoria → workout_parsed } }.

    A aba tem linhas de header de categoria que se repetem; a coluna A traz
    o rótulo do dia (Sexta/Sábado/Domingo) — é "sticky", vale até o próximo
    rótulo. Cada célula é texto livre que entra em parse_workout_text.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    resultado: dict[str, dict[str, dict[str, Any]]] = {}
    categorias_atuais: list[str] = []
    dia_atual: str = ""
    contador_workout = 0

    for row in rows:
        if not row or all(c is None for c in row):
            continue
        col_a = str(row[0]).strip() if row[0] is not None else ""
        # Linha de header de categoria: col A vazia, cols B+ com strings de categoria
        cells_b_em_diante = [c for c in row[1:] if c is not None]
        eh_header_categorias = (
            not col_a
            and len(cells_b_em_diante) >= 2
            and all(isinstance(c, str) and '\n' not in c for c in cells_b_em_diante[:3])
        )
        if eh_header_categorias:
            categorias_atuais = [str(c).strip() if c else "" for c in row[1:]]
            continue

        # Linha de dia (rótulo na col A) ou linha de workout (col A vazia, dia sticky)
        if col_a.lower() in _DIA_LABELS_VALIDOS:
            dia_atual = col_a

        if not dia_atual or not categorias_atuais:
            continue

        # Cada célula B+ é o texto de workout daquela categoria
        contador_workout += 1
        if dia_atual not in resultado:
            resultado[dia_atual] = {}
        for idx, cat in enumerate(categorias_atuais):
            if not cat:
                continue
            cell = row[idx + 1] if idx + 1 < len(row) else None
            if cell is None or not str(cell).strip():
                continue
            arena, texto_limpo = _extrair_arena(str(cell))
            wkt = parse_workout_text_robusto(texto_limpo, contador_workout)
            if arena:
                wkt['arena'] = arena
            if cat not in resultado[dia_atual]:
                resultado[dia_atual][cat] = []
            resultado[dia_atual][cat].append(wkt)

    return resultado


def _detectar_blocos_cronograma(header_row: list[str]) -> list[dict[str, int | None]]:
    """Procura TODAS as ocorrências de 'categoria' numa linha de header.

    Cada ocorrência define um bloco (arena). Pra cada bloco, identifica colunas
    auxiliares (eventos, bateria, aquecimento, fila) à esquerda e direita
    delimitadas pela próxima 'categoria' (se houver).
    """
    cat_cols = [i for i, h in enumerate(header_row) if h == 'categoria']
    if not cat_cols:
        return []
    blocos: list[dict[str, int | None]] = []
    for k, col_cat in enumerate(cat_cols):
        # Limites do bloco: do próximo 'categoria' anterior (ou 0) até o próximo (ou fim)
        lo = cat_cols[k - 1] + 1 if k > 0 else 0
        hi = cat_cols[k + 1] if k + 1 < len(cat_cols) else len(header_row)

        def _find(names: tuple[str, ...], ini: int, fim: int) -> int | None:
            for i in range(ini, fim):
                if header_row[i] in names:
                    return i
            return None

        # Estrutura de cada bloco: Eventos | Categoria | Bateria | ... | Fila.
        # 'Eventos' fica À ESQUERDA da categoria; 'Bateria'/'Aquecimento'/'Fila'
        # à DIREITA. Buscar em toda a faixa pegava a Bateria do bloco ANTERIOR
        # (multi-arena) — bug que perdia baterias (ex: Tap Control no 3º bloco).
        blocos.append({
            'eventos':     _find(('eventos',), lo, col_cat),
            'categoria':   col_cat,
            'bateria':     _find(('bateria',), col_cat + 1, hi),
            'aquecimento': _find(('aquecimento',), col_cat + 1, hi),
            'fila':        _find(('fila',), col_cat + 1, hi),
            'lo':          lo,
            'hi':          hi,
        })
    return blocos


def _parse_cronograma_dia(ws) -> list[dict[str, Any]]:
    """Lê uma aba de cronograma (`Sexta`, `Sábado`, `Domingo`).

    Suporta N arenas em colunas paralelas no mesmo dia. Cada arena tem seu
    próprio conjunto de colunas `Eventos | Categoria | Bateria | Aquecimento | Fila`.

    Retorna lista de baterias: cada uma com numero, codigo_evento (ex: '#1',
    '#2 & #3'), categoria, horario_aquecimento, horario_fila.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Encontra a linha de header (aquela que contém "Categoria")
    header_idx = None
    for i, row in enumerate(rows[:5]):
        valores = [str(c).strip().lower() if c else "" for c in row]
        if 'categoria' in valores:
            header_idx = i
            break
    if header_idx is None:
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[header_idx]]
    blocos = _detectar_blocos_cronograma(header)
    if not blocos:
        return []

    # Nome da arena por bloco: linha(s) acima do header têm 'Arena: X / Dia' na
    # faixa de colunas do bloco. Usado pra rotular colisão de bateria (número é
    # por arena, então mesma numeração em arenas diferentes NÃO é colisão).
    arena_por_bloco = [""] * len(blocos)
    for bidx, bloco in enumerate(blocos):
        for ri in range(header_idx):
            for ci in range(bloco['lo'], min(bloco['hi'], len(rows[ri]))):
                cell = rows[ri][ci]
                m = re.search(r'arena\s*:\s*(.+?)\s*(?:/|$)', str(cell), re.I) if cell else None
                if m:
                    arena_por_bloco[bidx] = m.group(1).strip()
                    break
            if arena_por_bloco[bidx]:
                break

    baterias: list[dict[str, Any]] = []
    # Cada bloco tem seu próprio 'codigo_atual' sticky (arenas não compartilham)
    codigo_atual_por_bloco = [""] * len(blocos)

    for row in rows[header_idx + 1:]:
        if not row:
            continue
        for bidx, bloco in enumerate(blocos):
            col_cat = bloco['categoria']
            col_bat = bloco['bateria']
            if col_cat is None or col_bat is None:
                continue
            cat_val = row[col_cat] if col_cat < len(row) else None
            bat_val = row[col_bat] if col_bat < len(row) else None
            if not cat_val or bat_val is None:
                continue

            # codigo sticky por bloco
            col_ev = bloco['eventos']
            if col_ev is not None and col_ev < len(row):
                ev_val = row[col_ev]
                if ev_val:
                    codigo_atual_por_bloco[bidx] = str(ev_val).strip()

            col_aq = bloco['aquecimento']
            col_fl = bloco['fila']
            baterias.append({
                'numero': str(bat_val).strip(),
                'codigo_evento': codigo_atual_por_bloco[bidx],
                'categoria': str(cat_val).strip(),
                'horario_aquecimento': _fmt_horario(row[col_aq]) if col_aq is not None and col_aq < len(row) else "",
                'horario_fila': _fmt_horario(row[col_fl]) if col_fl is not None and col_fl < len(row) else "",
                '_bloco': bidx,
                '_arena_cron': arena_por_bloco[bidx],
            })
    return baterias


def _fmt_horario(v: Any) -> str:
    """Converte célula de horário em string `HH:MM`. Aceita time, datetime ou string."""
    if v is None:
        return ""
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%H:%M')
        except Exception:
            return str(v)
    s = str(v).strip()
    # "18:20:00" → "18:20"
    m = re.match(r'^(\d{1,2}:\d{2})(:\d{2})?$', s)
    if m:
        return m.group(1)
    return s


def _detectar_blocos_montagem(valores_linha: list[str]) -> list[dict[str, int | None]]:
    """Procura TODAS as ocorrências de 'raia' numa linha e identifica colunas
    relacionadas (numero, nome, box) à direita de cada uma.

    Suporta arenas paralelas (múltiplos blocos lado a lado na mesma aba).
    Retorna lista de dicts {'raia', 'numero', 'nome', 'box'} com posições.
    """
    blocos: list[dict[str, int | None]] = []
    n = len(valores_linha)
    for col_raia, v in enumerate(valores_linha):
        if v != 'raia':
            continue
        # 'nome' deve aparecer até 4 colunas à direita de 'raia'
        col_nome = None
        for off in range(1, min(5, n - col_raia)):
            if valores_linha[col_raia + off] == 'nome':
                col_nome = col_raia + off
                break
        if col_nome is None:
            continue
        col_numero = None
        for off in range(1, col_nome - col_raia):
            if valores_linha[col_raia + off] in ('número', 'numero'):
                col_numero = col_raia + off
                break
        col_box = None
        for off in range(1, min(4, n - col_nome)):
            if valores_linha[col_nome + off] == 'box':
                col_box = col_nome + off
                break
        blocos.append({'raia': col_raia, 'numero': col_numero,
                       'nome': col_nome, 'box': col_box})
    return blocos


def _parse_montagem_dia(ws) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    """Lê uma aba `<Dia> - Montagem`. Suporta N arenas paralelas (blocos lado a lado).

    Estrutura repetida por bateria (1+ blocos por linha):
        L1: [horário, nº_bateria, ...]    L1: [horário_arena2, nº_bat2, ...]
        L2: [codigo, categoria, ...]      L2: [codigo_arena2, categoria_arena2, ...]
        L3: ["Raia", "Número", "Nome", "Box"]   L3: ["Raia", ..., "Box"]
        L4..N: dados de raia              L4..N: dados de raia

    Retorna dict mapeando (codigo_evento, categoria, numero_bateria) → lista de
    alocações (raia, numero, nome, box). Raias com nome `#N/A` são puladas.
    """
    rows = list(ws.iter_rows(values_only=True))
    resultado: dict[tuple[str, str, str], list[dict[str, Any]]] = {}

    i = 0
    while i < len(rows):
        row = rows[i]
        if not row:
            i += 1
            continue
        valores = [str(c).strip().lower() if c else "" for c in row]
        blocos = _detectar_blocos_montagem(valores)
        if not blocos:
            i += 1
            continue

        # Pra cada bloco detectado nessa linha, lê metadata 1-2 linhas acima
        # e processa as alocações abaixo. Suporta 2 layouts:
        #   Layout A (Sun Challenge):
        #     L-2: [horário, nº_bateria]
        #     L-1: [#N (código de evento), categoria]
        #   Layout B (Monstar — sem código de evento, com arena+workout no topo):
        #     L-3: [Arena: ...]
        #     L-2: [horário, nome_workout]
        #     L-1: [nº_bateria, categoria]
        # Detecção: se L-1 col[col_raia] começa com '#' → Layout A.
        for bloco in blocos:
            col_raia = bloco['raia']
            col_numero = bloco['numero']
            col_nome = bloco['nome']
            col_box = bloco['box']

            def _cell(r, c):
                return r[c] if (r is not None and c is not None and c < len(r)) else None

            prev = rows[i - 1] if i >= 1 else None
            prev2 = rows[i - 2] if i >= 2 else None
            v_acima_0 = _cell(prev, col_raia)
            v_acima_1 = _cell(prev, col_raia + 1)
            v_acima2_0 = _cell(prev2, col_raia)
            v_acima2_1 = _cell(prev2, col_raia + 1)

            s_acima_0 = str(v_acima_0).strip() if v_acima_0 is not None else ""
            eh_layout_sun = s_acima_0.startswith('#') or re.match(r'^(?:wkt|workout)\b', s_acima_0, re.I)

            if eh_layout_sun:
                # Sun: L-1 = (codigo, categoria), L-2 = (_, bateria)
                codigo    = s_acima_0
                categoria = str(v_acima_1).strip() if v_acima_1 is not None else ""
                numero_bat = str(v_acima2_1).strip() if v_acima2_1 is not None else ""
            else:
                # Monstar: L-1 = (bateria, categoria), L-2 = (_, workout_name)
                # Usa o nome do workout como 'codigo' (sem '#') pra fins de matching.
                numero_bat = s_acima_0
                categoria  = str(v_acima_1).strip() if v_acima_1 is not None else ""
                codigo     = str(v_acima2_1).strip() if v_acima2_1 is not None else ""

            alocacoes: list[dict[str, Any]] = []
            j = i + 1
            while j < len(rows):
                r = rows[j]
                if r is None:
                    break
                raia_v = _cell(r, col_raia)
                # Bloco acaba quando a coluna raia fica vazia nesse bloco
                if raia_v is None:
                    break
                # Ou quando aparece novo header (raia/nome) nessa coluna
                vals_j = [str(c).strip().lower() if c else "" for c in r]
                if col_raia < len(vals_j) and vals_j[col_raia] == 'raia':
                    break
                nome_v = _cell(r, col_nome)
                if nome_v is None:
                    j += 1
                    continue
                nome_str = str(nome_v).strip()
                if not nome_str or nome_str.upper() == '#N/A':
                    j += 1
                    continue
                num_v = _cell(r, col_numero)
                box_v = _cell(r, col_box)
                alocacoes.append({
                    'raia':   str(raia_v).strip(),
                    'numero': str(num_v).strip() if num_v is not None else "",
                    'nome':   nome_str,
                    'box':    str(box_v).strip() if box_v is not None else "",
                })
                j += 1

            if alocacoes:
                resultado[(codigo, categoria, numero_bat)] = alocacoes

        i += 1   # avança 1; loop natural pula linhas já processadas (col raia já não é 'raia')

    return resultado


# Parênteses que rotulam bateria/heat — removidos na normalização estrita.
# Cobre: '(Heat 1)', '(Heat 2-3)', '(Heats)', '(Single Heat)', '(Final Heat)'.
_HEAT_PAREN_RE = re.compile(
    r'\s*\((?:single heat|final heat|heat\s*[\d\-/]*|heats?)\)\s*',
    re.I,
)


def _normalizar_categoria(s: str) -> str:
    """Normaliza nome de categoria pra comparação — remove sufixos de bateria.

    Tira só `(Heat N)`, `(Single Heat)`, `(Final Heat)` etc. — parênteses que
    rotulam bateria no cronograma mas não fazem parte do nome da categoria.
    Preserva descritores livres como `(identico ao amador)` ou `(Iniciante)`
    que diferenciam categorias distintas dentro do mesmo evento.
    """
    if not s:
        return ""
    s = _HEAT_PAREN_RE.sub(' ', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    s = s.replace('begginer', 'beginner')
    return s


def _normalizar_categoria_relaxada(s: str) -> str:
    """Versão relaxada: remove TODOS os parênteses, não só os de heat.

    Usada como fallback no match. Útil quando a grade tem descritor extra
    (ex: 'Master 35-39 (identico ao amador)') que o cronograma não repete.
    Só deve ser aplicada quando não há ambiguidade — duas categorias da grade
    com mesma versão relaxada precisam ser desambiguadas pelo nome cheio.
    """
    if not s:
        return ""
    s = re.sub(r'\s*\([^)]*\)\s*', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    s = s.replace('begginer', 'beginner')
    return s


# Folding de gênero pra chave fuzzy: radical comum por gênero. NÃO cruza entre
# si (masc/fem/mist são disjuntos), então não confunde categorias distintas.
_GENERO_FOLD = {
    'masculino': 'masc', 'masculina': 'masc', 'masculinos': 'masc', 'masculinas': 'masc',
    'feminino': 'fem', 'feminina': 'fem', 'femininos': 'fem', 'femininas': 'fem',
    'misto': 'mist', 'mista': 'mist', 'mistos': 'mist', 'mistas': 'mist', 'mixto': 'mist',
}


def _chave_categoria_fuzzy(s: str) -> str:
    """Chave tolerante a ordem das palavras, concordância de gênero e posição
    do '+'. Casa variações humanas que descrevem a MESMA categoria:

      'Master Masculino 40-44'  ==  'Master 40-44 Masculino'   (ordem)
      'Dupla Rx Masculino'      ==  'Dupla Rx Masculina'       (gênero)
      'Trio Master Misto 110+'  ==  'Trio Master Misto +110'   (sinal +)

    Gênero é folded pra radical comum (masc/fem/mist) — disjuntos, não cruzam.
    Tokens são ordenados pra ignorar a ordem. Só deve ser usada como camada de
    match com guarda anti-ambiguidade (duas cats da grade com mesma chave fuzzy
    não podem usar essa camada).
    """
    s = _normalizar_categoria_relaxada(s)
    if not s:
        return ""
    s = s.replace('+', ' ')            # +110 / 110+ → 110
    s = re.sub(r'\s+', ' ', s).strip()
    toks = [_GENERO_FOLD.get(t, t) for t in s.split()]
    return ' '.join(sorted(toks))


def _split_codigo_evento(codigo: str) -> list[str]:
    """Quebra um código tipo '#2 & #3' em ['#2', '#3']. Códigos simples viram [codigo]."""
    if not codigo:
        return []
    return [p.strip() for p in re.split(r'\s*&\s*', codigo) if p.strip()]


def _workout_numero_de_codigo(codigo: str) -> int | None:
    """Extrai o número do workout. Aceita '#1', '#02', 'WKT 4', 'Workout 04'.

    Exige prefixo explícito ('#', 'WKT' ou 'Workout') pra evitar pegar dígito
    de texto qualquer (ex: 'Bat 12' não deve virar workout 12).
    """
    m = re.match(r'^\s*(?:#|wkt|workout)\s*(\d+)\s*$', codigo, re.I)
    return int(m.group(1)) if m else None


def _workouts_que_rodam_da_bateria(codigo_evento: str, workouts: list[dict]) -> list[int]:
    """Mapeia o codigo_evento de uma bateria pra lista de posições 1-based
    de workouts da categoria. Aceita 4 formas:
      1. Nº direto: '#1', '#2 & #3', 'WKT 4'  → match por número
      2. Nome do workout entre aspas: '"Simple Dimension"'  → match por nome
      3. Misto: '"Simple Dimension" & "Spin"'  → split por & + match por nome
      4. Composto: parte do split bate com nome de F1 ou F2 do composto
         (Storm Dupla: cronograma diz `"Barbells and Jump & Run in the Park"`,
         mas o composto se chama `BARBELLS AND JUMP + RUN IN THE PARK`.
         Split por & retorna `BARBELLS AND JUMP` e `RUN IN THE PARK`, e
         cada um bate com F1.nome ou F2.nome do composto).
    Retorna [] se nada bateu.
    """
    if not codigo_evento or not workouts:
        return []
    # Antes de split, testa match exato do codigo cheio com nome de algum
    # workout (cobre composto onde o nome dele JÁ tem `+` ou `&` literal).
    nome_full = codigo_evento.strip().strip('"“”\'').upper()
    for idx, w in enumerate(workouts, start=1):
        nome_w = (w.get('nome', '') or '').strip().upper()
        if nome_w and nome_w == nome_full:
            return [idx]
    # Split em '&' pra suportar baterias mistas (workout A & workout B)
    partes = _split_codigo_evento(codigo_evento) or [codigo_evento]
    posicoes: list[int] = []
    for p in partes:
        # Forma 1: '#N' / 'WKT N'
        n = _workout_numero_de_codigo(p)
        if n is not None:
            if n not in posicoes: posicoes.append(n)
            continue
        # Forma 2/3: nome do workout (entre aspas ou não) — match case-insensitive
        nome_busca = p.strip().strip('"“”\'').upper()
        if not nome_busca: continue
        for idx, w in enumerate(workouts, start=1):
            nome_w = (w.get('nome', '') or '').strip().upper()
            if nome_w and nome_w == nome_busca:
                if idx not in posicoes: posicoes.append(idx)
                break
            # Forma 4: composto — F1 ou F2 do composto bate com a parte
            if w.get('tipo') == 'composto':
                f1_nome = ((w.get('f1') or {}).get('nome') or '').strip().upper()
                f2_nome = ((w.get('f2') or {}).get('nome') or '').strip().upper()
                if nome_busca == f1_nome or nome_busca == f2_nome:
                    if idx not in posicoes: posicoes.append(idx)
                    break
    return posicoes


def _roster_individuais(wb) -> list[dict[str, str]]:
    """Lê a aba `Atletas` (roster informativo): número, nome, box."""
    if 'Atletas' not in wb.sheetnames:
        return []
    ws = wb['Atletas']
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        numero = str(row[0]).strip() if row[0] is not None else ""
        nome   = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        box    = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not nome:
            continue
        key = (numero, nome.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append({'numero': numero, 'nome': nome, 'box': box})
    return out


def parse_excel_multidia(wb) -> dict[str, Any]:
    """Parser do formato evento multi-dia (Workouts + cronograma + montagem).

    Retorna estrutura aninhada:
        { tipo: 'evento_multidia',
          evento_nome: str,
          dias: [
            { label: 'Sexta',
              categorias: [
                { nome: 'Trio Rx Misto',
                  workouts: [Workout, ...],
                  baterias: [
                    { numero: '1',
                      codigo_evento: '#1',
                      horario_aquecimento: '18:20',
                      horario_fila: '18:45',
                      workouts_que_rodam: [1],   # nº dos workouts (índices em workouts da categoria)
                      alocacoes: [{raia, numero, nome, box}, ...]
                    }
                  ]
                }
              ]
            }
          ],
          roster: [{numero, nome, box}, ...],
        }
    """
    # 1) Workouts: dia → categoria → [workouts]
    if 'Workouts' not in wb.sheetnames:
        return {'tipo': 'erro', 'erro': 'Aba Workouts ausente'}
    workouts_por_dia_cat = _parse_workouts_grade_multidia(wb['Workouts'])
    if not workouts_por_dia_cat:
        return {'tipo': 'erro', 'erro': 'Aba Workouts vazia ou ilegível'}

    # 2) Pra cada dia detectado em Workouts, lê cronograma + montagem
    nomes_lower = {s.lower(): s for s in wb.sheetnames}
    dias_resultado: list[dict[str, Any]] = []
    for dia_label in workouts_por_dia_cat.keys():
        sname = nomes_lower.get(dia_label.lower())
        montagem_sname = nomes_lower.get(f"{dia_label.lower()} - montagem")
        cronograma = _parse_cronograma_dia(wb[sname]) if sname else []
        montagem   = _parse_montagem_dia(wb[montagem_sname]) if montagem_sname else {}

        # Agrupa por categoria do dia
        cats_resultado: list[dict[str, Any]] = []
        for cat_nome, lista_workouts in workouts_por_dia_cat[dia_label].items():
            # Filtra baterias do cronograma cuja categoria casa (ignora sufixo "(Single Heat)" etc.)
            cat_norm = _normalizar_categoria(cat_nome)
            baterias_da_cat = [b for b in cronograma if cat_norm in _normalizar_categoria(b['categoria'])]
            # Para cada bateria, monta as alocações via montagem
            for b in baterias_da_cat:
                # codigo_evento pode estar vazio no cronograma; nesse caso usa a Montagem
                # como fonte primária do código.
                codigos_cronograma = set(_split_codigo_evento(b['codigo_evento']))
                aloc: list[dict[str, Any]] = []
                codigo_montagem = ""
                for chave, alocs in montagem.items():
                    chave_codigo, chave_cat, chave_bat = chave
                    if chave_bat != b['numero']:
                        continue
                    if cat_norm not in _normalizar_categoria(chave_cat):
                        continue
                    # Se cronograma trouxe códigos, exige interseção com os da montagem.
                    # Ambos podem ser compostos (ex: '#2 & #3'), então comparo conjuntos.
                    if codigos_cronograma:
                        codigos_chave = set(_split_codigo_evento(chave_codigo))
                        if not (codigos_cronograma & codigos_chave):
                            continue
                    aloc = alocs
                    codigo_montagem = chave_codigo
                    break
                # codigo final: o que veio do cronograma OU o que a montagem revelou
                codigo_final = b['codigo_evento'] or codigo_montagem
                codigos_finais = _split_codigo_evento(codigo_final) or ([codigo_final] if codigo_final else [])
                workouts_que_rodam = [n for n in (_workout_numero_de_codigo(c) for c in codigos_finais) if n is not None]
                b_full = {
                    **b,
                    'codigo_evento': codigo_final,
                    'workouts_que_rodam': workouts_que_rodam,
                    'alocacoes': aloc,
                }
                cat_existing = next((c for c in cats_resultado if c['nome'] == cat_nome), None)
                if cat_existing is None:
                    cat_existing = {'nome': cat_nome, 'workouts': lista_workouts, 'baterias': []}
                    cats_resultado.append(cat_existing)
                cat_existing['baterias'].append(b_full)

            # Categorias sem bateria no cronograma ainda entram (workouts mostrados, sem alocação)
            if not baterias_da_cat:
                cats_resultado.append({'nome': cat_nome, 'workouts': lista_workouts, 'baterias': []})

        dias_resultado.append({'label': dia_label, 'categorias': cats_resultado})

    # Padroniza nomes de movimentos (PT-BR/EN/case → forma canônica)
    for d in dias_resultado:
        for c in d.get('categorias', []) or []:
            padronizar_workouts(c.get('workouts', []) or [])

    return {
        'tipo': 'evento_multidia',
        'evento_nome': '',  # pode ser preenchido pela UI a partir do nome do arquivo ou config
        'dias': dias_resultado,
        'roster': _roster_individuais(wb),
    }


# ── Layout grades-por-modalidade + dias com Montagem ──────────────────────────
# Caso de uso: planilhas com workouts em abas separadas por modalidade
# (ex: `Individuais` + `Duplas`, ou `Times` + `Solo`, etc.) e cronograma +
# montagem por dia (`<Dia>` + `<Dia> - Montagem`). Sem aba unificada `Workouts`.

def _tem_cronograma(ws) -> bool:
    """Aba parece um cronograma de dia: header com 'Categoria' E 'Bateria'.

    Mesma heurística usada no corpo de `parse_excel_grades_e_dias` pra detectar
    dias sem par `- Montagem`. Multi-arena: basta uma ocorrência de cada.
    """
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri >= 5:
            break
        valores = [str(c).strip().lower() if c else "" for c in row]
        if 'categoria' in valores and 'bateria' in valores:
            return True
    return False


def _is_layout_grades_e_dias(wb) -> bool:
    nomes_lower = [s.lower() for s in wb.sheetnames]
    if 'workouts' in nomes_lower:
        return False  # se tem Workouts, o detector multidia clássico cuida disso
    tem_grade = any(_is_categoria_grid(wb[s]) for s in wb.sheetnames)
    if not tem_grade:
        return False
    # Dias podem vir COM montagem (atletas alocados) ou SÓ com cronograma
    # (planejamento, roster ainda não fechado → o corpo gera súmulas em branco).
    # Não exigir `- Montagem` aqui: o porteiro não pode ser mais rígido que o
    # corpo, que já trata `montagem = {}` quando não há a aba.
    tem_montagem = any(' - montagem' in n for n in nomes_lower)
    tem_dia_cronograma = any(_tem_cronograma(wb[s]) for s in wb.sheetnames)
    return tem_montagem or tem_dia_cronograma


def _split_partes_categoria(s: str) -> list[str]:
    """Split bruto de string de bateria composta em pedaços (sem normalizar).

    Separadores aceitos: ` & ` e `, `. Regrudar pedaços com parens
    desbalanceados — vírgula DENTRO de parens (ex: `Iniciante (8, 9 anos)`)
    não é separador.
    """
    if not s:
        return []
    bruto = re.split(r'\s+&\s+|,\s+', s)
    partes: list[str] = []
    buffer = ''
    for p in bruto:
        candidato = f'{buffer}, {p}' if buffer else p
        if candidato.count('(') != candidato.count(')'):
            buffer = candidato
        else:
            partes.append(candidato)
            buffer = ''
    if buffer:
        partes.append(buffer)
    return [p for p in partes if p.strip()]


def _quebrar_categoria_composta(s: str) -> list[str]:
    """'A (Heat 1) & B (Heat 2)' → ['a', 'b'] (cada parte normalizada).

    Aceita também `,` como separador (`A (Heat 1), B (Heat 2) & C (Heat 3)`)
    — formato que aparece em eventos com 3+ cats compartilhando bateria.

    Diferente de `_normalizar_categoria`, que perde tudo depois do primeiro `(`
    e portanto descarta a segunda categoria de baterias mistas.
    """
    return [_normalizar_categoria(p) for p in _split_partes_categoria(s)]


def _bateria_casa_categoria(
    bateria_categoria: str,
    cat_grade_norm: str,
    cat_grade_relaxada: str | None = None,
    permite_relaxado: bool = False,
    cat_grade_fuzzy: str | None = None,
    permite_fuzzy: bool = False,
) -> bool:
    """Match exato (após normalização e quebra de '&').

    Substring causa falso positivo entre 'Rx Masculino' (Sábado) e 'Dupla Rx
    Masculino' (Domingo) — categorias diferentes que rodam em dias diferentes.

    Fallback relaxado (sem parênteses) só roda quando `permite_relaxado=True`,
    o que o caller deve passar apenas se a categoria não colide com outra.
    """
    partes = _quebrar_categoria_composta(bateria_categoria)
    if cat_grade_norm in partes:
        return True
    if permite_relaxado and cat_grade_relaxada:
        partes_relax = [_normalizar_categoria_relaxada(p)
                        for p in _split_partes_categoria(bateria_categoria)]
        if cat_grade_relaxada in partes_relax:
            return True
    if permite_fuzzy and cat_grade_fuzzy:
        partes_fuzzy = [_chave_categoria_fuzzy(p)
                        for p in _split_partes_categoria(bateria_categoria)]
        if cat_grade_fuzzy in partes_fuzzy:
            return True
    return False


def _propagar_codigos_da_montagem(
    cronograma: list[dict[str, Any]],
    montagem: dict[tuple[str, str, str], list[dict[str, Any]]],
) -> None:
    """Quando o cronograma vem sem códigos de evento (`#1`, `#2 & #3`, etc),
    procura o código correspondente na montagem pelo número da bateria.

    Mutates `cronograma` in-place, preenchendo `codigo_evento`.
    """
    if any(b.get('codigo_evento') for b in cronograma):
        return  # cronograma já tem códigos — não interfere
    cods_por_bat: dict[str, str] = {}
    for (cod, _cat, bat), _ in montagem.items():
        if cod and bat:
            cods_por_bat.setdefault(bat, cod)
    for b in cronograma:
        if not b.get('codigo_evento'):
            b['codigo_evento'] = cods_por_bat.get(b.get('numero', ''), '')


def _roster_de_abas_atletas(wb) -> list[dict[str, str]]:
    """Lê abas com lista de atletas/duplas/trios.

    Aceita 3 padrões de nome:
      - `Atletas` ou `Atleta` (puro)
      - `Atleta - X` / `Atletas - X` (Sun Challenge style: separado por tipo)
      - `Athletes` / `Athlete` (inglês)

    Estrutura esperada: col A = número, col B = nome, col C = box. Sem header.
    """
    out: list[dict[str, str]] = []
    for sname in wb.sheetnames:
        sl = sname.lower().strip()
        eh_atletas = (
            sl in ('atleta', 'atletas', 'athlete', 'athletes')
            or sl.startswith('atleta - ')
            or sl.startswith('atletas - ')
            or sl.startswith('athlete - ')
            or sl.startswith('athletes - ')
        )
        if not eh_atletas:
            continue
        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None for c in row):
                continue
            numero = str(row[0]).strip() if row[0] is not None else ""
            nome   = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            box    = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            if not nome or nome.upper() == '#N/A':
                continue
            out.append({'numero': numero, 'nome': nome, 'box': box})
    # Dedup por (numero, nome lower): linhas duplicadas no Excel (típico em
    # copia/cola) não viram atletas repetidos na súmula combinada de pré-evento.
    # Match por nome também (não só número) preserva colisão legítima entre
    # modalidades — Storm reusa #101 pra `MATHEUS` (Individual) e `GOKU E
    # KURIRIN` (Dupla); ambos ficam.
    seen: set[tuple[str, str]] = set()
    deduped: list[dict[str, str]] = []
    for a in out:
        key = (a['numero'], a['nome'].lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(a)
    return deduped


def _parse_equipamento(wb) -> Optional[dict[str, Any]]:
    """Lê aba `Equipamento(s)` / `Equipment` (se houver) → dict com anilhas + unidade.

    Estrutura esperada: header com `Anilha | Peso | Qtd` (ordem livre). Cada
    linha lista um tipo de anilha disponível no evento. O peso pode vir como
    número puro (assume kg) ou com unidade colada/separada (`25kg`, `25 kg`,
    `45lb`, `45 lb`).

    Retorna `{anilhas: [pesos únicos ordenados desc], unidade: 'kg'|'lb'}`
    ou None se a aba não existe ou está vazia. Detecta unidade global do
    evento: se qualquer célula tem 'lb', evento inteiro é lb (assume coerência).
    """
    sname = next((s for s in wb.sheetnames
                  if s.strip().lower() in ('equipamento', 'equipamentos', 'equipment')),
                 None)
    if not sname:
        return None
    ws = wb[sname]
    pesos: set[float] = set()
    unidade = 'kg'
    col_peso_idx: Optional[int] = None
    # Procura header: a linha que tem 'peso' em alguma célula
    header_row: Optional[int] = None
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri > 10: break
        for ci, c in enumerate(row):
            if c and str(c).strip().lower() == 'peso':
                col_peso_idx = ci
                header_row = ri
                break
        if col_peso_idx is not None: break
    if col_peso_idx is None:
        # Fallback: assume Anilha=A, Peso=B, Qtd=C e tenta a partir da linha 2
        col_peso_idx = 1
        header_row = 0
    # Lê valores da coluna Peso a partir da linha seguinte ao header
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri <= header_row: continue
        if col_peso_idx >= len(row): continue
        cell = row[col_peso_idx]
        if cell is None or cell == '': continue
        s = str(cell).strip().lower()
        if 'lb' in s: unidade = 'lb'
        # Extrai número (aceita '25', '25kg', '25 kg', '2,5', '2.5kg')
        m = re.match(r'^([\d]+(?:[\.,]\d+)?)', s)
        if not m: continue
        try:
            peso = float(m.group(1).replace(',', '.'))
        except ValueError:
            continue
        if peso > 0:
            pesos.add(peso)
    if not pesos:
        # Fallback formato `Categoria | Equipamento | Qtd` (Pwrd by Coffee):
        # o peso vem no NOME do equipamento ('Anilha Color 5 kg'). Varre todas
        # as células e pega só linhas de ANILHA (barra/dumbbell/med ball ficam
        # de fora da régua de anilhas).
        anilha_re = re.compile(r'anilha[^\d]*(\d+(?:[.,]\d+)?)\s*(kg|lbs?)?', re.I)
        anilha_tem_lb = anilha_tem_kg = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if not c:
                    continue
                s = str(c).strip()
                if 'anilha' not in s.lower():
                    continue
                m = anilha_re.search(s)
                if not m:
                    continue
                if m.group(2):
                    if 'lb' in m.group(2).lower():
                        anilha_tem_lb = True
                    else:
                        anilha_tem_kg = True
                try:
                    peso = float(m.group(1).replace(',', '.'))
                except ValueError:
                    continue
                if peso > 0:
                    pesos.add(peso)
        # A unidade da régua vem DAS ANILHAS — não das med balls (que são lb por
        # convenção mesmo em evento kg e contaminavam a coluna acima).
        if pesos:
            if anilha_tem_lb and not anilha_tem_kg:
                unidade = 'lb'
            elif anilha_tem_kg:
                unidade = 'kg'
    if not pesos:
        return None
    # Heurística: se a unidade não veio explícita ('45', '35'), tenta inferir.
    # 45 e 55 são anilhas icônicas em lb (não existem em kg padrão).
    # 1.25 e 2.5 são fracionárias típicas de kg.
    if unidade == 'kg':   # default — só re-avalia se ninguém escreveu 'kg'/'lb'
        tem_lb_typical = any(p in (45, 55) for p in pesos)
        tem_kg_typical = any(p in (1.25, 2.5) for p in pesos)
        if tem_lb_typical and not tem_kg_typical:
            unidade = 'lb'
    # Dumbbells disponíveis (pro linter de carga fora do rol). Nome típico:
    # 'Dumbbell 22,5 kg'. Barra/anilha não entram aqui.
    dumbbells: set[float] = set()
    db_re = re.compile(r'dumbbell[^\d]*(\d+(?:[.,]\d+)?)\s*(kg|lbs?)?', re.I)
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if not c or 'dumbbell' not in str(c).lower():
                continue
            m = db_re.search(str(c))
            if not m:
                continue
            try:
                w = float(m.group(1).replace(',', '.'))
            except ValueError:
                continue
            if w > 0:
                dumbbells.add(w)
    return {
        'anilhas': sorted(pesos, reverse=True),
        'unidade': unidade,
        'dumbbells': sorted(dumbbells, reverse=True),
    }


def _parse_inscritos(wb) -> dict[str, tuple[int, int]]:
    """Lê aba `Inscritos` (se houver) → mapa categoria_normalizada → (n_ini, n_fim).

    Estrutura esperada: header com `Nome` + colunas que contenham `inicial` e
    `final`. Múltiplos blocos (separados por linhas vazias) são suportados —
    típico quando há Individuais e Duplas no mesmo evento. Retorna `{}` se a
    aba não existir ou estiver fora do padrão.

    Usado pra desambiguar alocações de baterias mistas (atletas de duas
    categorias rodando juntos): a faixa de número diz quem pertence a qual.
    """
    return {cat: (ini, fim) for cat, (ini, fim, _) in _parse_inscritos_full(wb).items()}


def _parse_inscritos_full(wb) -> dict[str, tuple[int, int, Optional[bool]]]:
    """Versão completa: retorna `(ini, fim, is_individual)`.

    `is_individual=True/False` se a coluna `Individual` existir no header
    (valor `Sim`/`Não`); `None` se a coluna não está presente. Usado pra
    desambiguar quando duas categorias compartilham a mesma faixa de número
    (Individual `Rx Masculino` e Dupla `Dupla Rx Masculino` usam 101-199).
    """
    sname = next((s for s in wb.sheetnames if s.strip().lower() == 'inscritos'), None)
    if not sname:
        return {}
    ws = wb[sname]

    resultado: dict[str, tuple[int, int, Optional[bool]]] = {}
    col_nome = col_ini = col_fim = None
    col_indiv: Optional[int] = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            col_nome = col_ini = col_fim = None  # quebra de bloco: re-detecta header
            col_indiv = None
            continue
        vals = [str(c).strip().lower() if c else '' for c in row]
        # Header novo: precisa ter coluna 'nome' + uma 'inicial' + uma 'final'
        if 'nome' in vals and any('inicial' in v for v in vals) and any('final' in v for v in vals):
            col_nome = vals.index('nome')
            col_ini  = next(i for i, v in enumerate(vals) if 'inicial' in v)
            col_fim  = next(i for i, v in enumerate(vals) if 'final' in v)
            col_indiv = next((i for i, v in enumerate(vals) if 'individual' in v), None)
            continue
        if col_nome is None:
            continue
        nome = row[col_nome] if col_nome < len(row) else None
        ini  = row[col_ini]  if col_ini  < len(row) else None
        fim  = row[col_fim]  if col_fim  < len(row) else None
        if not nome or ini is None or fim is None:
            continue
        try:
            ini_int, fim_int = int(ini), int(fim)
        except (TypeError, ValueError):
            continue
        if ini_int > fim_int:
            continue
        is_indiv: Optional[bool] = None
        if col_indiv is not None and col_indiv < len(row):
            v = str(row[col_indiv] or '').strip().lower()
            if v in ('sim', 'yes', 's', 'y', 'true', '1'):
                is_indiv = True
            elif v in ('não', 'nao', 'no', 'n', 'false', '0'):
                is_indiv = False
        resultado[_normalizar_categoria(str(nome))] = (ini_int, fim_int, is_indiv)
    return resultado


def _filtrar_alocacoes_por_faixa(
    alocs: list[dict[str, Any]], faixa: tuple[int, int]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Particiona alocações: as que caem em [ini, fim] vs as que ficam fora.

    Retorna (mantidas, descartadas). Alocações sem `numero` numérico vão pra
    descartadas — não há como saber a qual categoria pertencem.
    """
    ini, fim = faixa
    mantidas: list[dict[str, Any]] = []
    descartadas: list[dict[str, Any]] = []
    for a in alocs:
        try:
            n = int(str(a.get('numero', '')).strip())
        except (ValueError, AttributeError):
            descartadas.append(a)
            continue
        if ini <= n <= fim:
            mantidas.append(a)
        else:
            descartadas.append(a)
    return mantidas, descartadas


def _alocacoes_tem_atleta_na_faixa(
    alocs: list[dict[str, Any]], faixa: tuple[int, int],
) -> bool:
    """True se alguma alocação tem número dentro da faixa."""
    if not faixa:
        return False
    ini, fim = faixa
    for a in alocs:
        try:
            n = int(str(a.get('numero', '')).strip())
        except (ValueError, AttributeError):
            continue
        if ini <= n <= fim:
            return True
    return False


def _bateria_tem_atleta_na_faixa(
    bateria_numero: str,
    montagem: dict[tuple[str, str, str], list[dict[str, Any]]],
    faixa: tuple[int, int],
) -> bool:
    """True se a Montagem dessa bateria tem ao menos 1 atleta na faixa.

    Cobre o caso onde o nome textual da categoria diverge entre Inscritos,
    cronograma e Montagem (ex: Inscritos diz `Teen Scaled 14-15 Feminino`
    mas cronograma só diz `14-15 Feminino (Single Heat)`). A faixa de
    número do Inscritos vira a fonte de verdade.
    """
    if not faixa or not bateria_numero:
        return False
    for (_cod, _cat, chave_bat), alocs in montagem.items():
        if chave_bat != bateria_numero:
            continue
        if _alocacoes_tem_atleta_na_faixa(alocs, faixa):
            return True
    return False


def parse_excel_grades_e_dias(wb) -> dict[str, Any]:
    """Parser para layout: grades de workout por modalidade + dias com Montagem.

    Estratégia:
      1. Lê todas as abas grade (Individuais, Duplas, ...) → categoria → workouts.
      2. Detecta pares `<Dia>` + `<Dia> - Montagem`.
      3. Pra cada dia, lê cronograma + montagem. Se cronograma vem sem códigos,
         puxa código da montagem pela bateria.
      4. Pra cada categoria da grade, anexa ao dia onde aparece no cronograma.
      5. Roster lido das abas `Atleta(s) - X`.
    """
    # 1) Grades — junta todas
    grade_por_categoria: dict[str, list[Workout]] = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        if not _is_categoria_grid(ws):
            continue
        r = _parse_excel_grade(wb, sname)
        # Não sobrescreve: categorias com mesmo nome em grades diferentes mantêm
        # a primeira ocorrência (raro — modelagem do usuário deve evitar).
        for cat, wkts in r.get('por_categoria', {}).items():
            grade_por_categoria.setdefault(cat, wkts)
    if not grade_por_categoria:
        return {'tipo': 'erro', 'erro': 'Nenhuma grade categoria×workout detectada'}

    for workouts in grade_por_categoria.values():
        padronizar_workouts(workouts)

    # Faixas de número por categoria — desambigua atletas em baterias mistas.
    # Versão `_full` traz `is_individual` quando a coluna Individual existe;
    # usado pra separar modalidades quando faixas colidem (`Rx Masculino` vs
    # `Dupla Rx Masculino` usam 101-199 cada um).
    inscritos_full = _parse_inscritos_full(wb)
    inscritos_faixas: dict[str, tuple[int, int]] = {
        cat: (ini, fim) for cat, (ini, fim, _) in inscritos_full.items()
    }
    inscritos_modalidade: dict[str, Optional[bool]] = {
        cat: is_indiv for cat, (_, _, is_indiv) in inscritos_full.items()
    }
    # Faixas que COLIDEM entre duas ou mais categorias
    _contagem_faixa: dict[tuple[int, int], int] = {}
    for f in inscritos_faixas.values():
        _contagem_faixa[f] = _contagem_faixa.get(f, 0) + 1
    _faixas_colisao: set[tuple[int, int]] = {
        f for f, n in _contagem_faixa.items() if n > 1
    }

    # Detecta categorias da grade com mesma "relaxada" — match relaxado é
    # ambíguo nesse caso e não pode ser usado como fallback. Ex: se aparecer
    # 'Rx Misto (Iniciante)' e 'Rx Misto (Avançado)' na mesma grade, as duas
    # têm relaxada 'rx misto' → match relaxado proibido pra essas.
    cats_grade_relaxadas = {cat: _normalizar_categoria_relaxada(cat)
                            for cat in grade_por_categoria}
    _contagem_relaxada: dict[str, int] = {}
    for r in cats_grade_relaxadas.values():
        _contagem_relaxada[r] = _contagem_relaxada.get(r, 0) + 1
    cats_ambiguas = {cat for cat, r in cats_grade_relaxadas.items()
                     if _contagem_relaxada[r] > 1}

    # Chave fuzzy (ordem/gênero/±) por categoria da grade + guarda de ambiguidade:
    # duas cats da grade com a mesma chave fuzzy não podem usar essa camada.
    cats_grade_fuzzy = {cat: _chave_categoria_fuzzy(cat)
                        for cat in grade_por_categoria}
    _contagem_fuzzy: dict[str, int] = {}
    for f in cats_grade_fuzzy.values():
        _contagem_fuzzy[f] = _contagem_fuzzy.get(f, 0) + 1
    cats_ambiguas_fuzzy = {cat for cat, f in cats_grade_fuzzy.items()
                           if f and _contagem_fuzzy[f] > 1}

    # 2) Dias detectados — em ordem de preferência:
    #    (a) abas <Dia> que TÊM par <Dia> - Montagem (atletas alocados)
    #    (b) abas <Dia> sozinhas (planejamento; gera súmulas em branco)
    # Aceita dias da semana em PT-BR ou EN, ou qualquer nome de aba que não
    # seja meta (Inscritos, Atletas, Heats, etc).
    nomes_lower = {s.lower(): s for s in wb.sheetnames}
    META_SHEETS = {
        'inscritos', 'categorias', 'atletas', 'atleta', 'athletes', 'athlete',
        'heats', 'time caps', 'timecaps', 'equipamento', 'equipamentos',
    }
    dias_com_montagem: list[str] = []
    for sname in wb.sheetnames:
        sl = sname.lower()
        if sl.endswith(' - montagem'):
            dia_sl = sl[: -len(' - montagem')]
            if dia_sl in nomes_lower:
                dias_com_montagem.append(nomes_lower[dia_sl])
    # Dias sem montagem: qualquer aba que pareça ser um dia mas não tem par
    dias_sem_montagem: list[str] = []
    for sname in wb.sheetnames:
        sl = sname.lower().strip()
        if sl in META_SHEETS: continue
        if sl.endswith(' - montagem'): continue
        if any(p in sl for p in ('workouts', 'inscritos', 'roster')): continue
        if nomes_lower.get(sname.lower()) in dias_com_montagem: continue
        # Heurística: aba é dia se tem cronograma de baterias — header precisa
        # ter tanto 'categoria' QUANTO 'bateria' nas primeiras linhas. Evita
        # falso positivo com abas tipo 'Finalistas' (Categoria + Place + Number).
        ws = wb[sname]
        for ri, r in enumerate(ws.iter_rows(values_only=True)):
            if ri >= 5: break
            valores = [str(c).strip().lower() if c else "" for c in r]
            if 'categoria' in valores and 'bateria' in valores:
                dias_sem_montagem.append(sname)
                break
    dias_detectados = dias_com_montagem + dias_sem_montagem
    if not dias_detectados:
        return {'tipo': 'erro', 'erro': 'Nenhum dia encontrado — esperava aba <Dia> com cronograma (coluna Categoria) ou par <Dia> + <Dia> - Montagem'}

    # 3-4) Pra cada dia, lê e agrupa categorias presentes
    dias_resultado: list[dict[str, Any]] = []
    avisos_import: list[dict[str, str]] = []
    # Rastreia categorias da grade que casaram em ≥1 dia — as que sobram não
    # geram súmula (typo de nome, categoria só na grade, etc). Linter 2.0.
    cats_grade_casadas: set[str] = set()
    for dia_label in dias_detectados:
        # dia_norm pra comparar com wkt._dia_label (sempre normalizado)
        dia_norm_atual = _DIAS_SEMANA_NORM.get(dia_label.strip().lower(), dia_label.strip().lower())
        sname_dia = nomes_lower[dia_label.lower()]
        sname_mont = nomes_lower.get(f"{dia_label.lower()} - montagem")
        cronograma = _parse_cronograma_dia(wb[sname_dia])
        montagem   = _parse_montagem_dia(wb[sname_mont]) if sname_mont else {}
        _propagar_codigos_da_montagem(cronograma, montagem)

        # Linter: colisão de bateria — mesmo número na MESMA arena (bloco) com
        # categorias diferentes (número de bateria é por arena). Pegou a #72 do
        # Pwrd (Trio Interm Masc Heat 3 e Master 45+ ambas 72 na Quadra, 15:11).
        _por_arena_num: dict[tuple, list[dict]] = {}
        for b in cronograma:
            _por_arena_num.setdefault((b.get('_bloco'), b.get('numero')), []).append(b)
        for (_bloco, num), grupo in _por_arena_num.items():
            cats_distintas = {b.get('categoria', '').strip() for b in grupo}
            if len(cats_distintas) > 1:
                arena = grupo[0].get('_arena_cron') or f'arena {(_bloco or 0) + 1}'
                horarios = sorted({b.get('horario_fila', '') for b in grupo if b.get('horario_fila')})
                avisos_import.append({
                    'nivel': 'erro',
                    'msg':   f'Bateria {num} duplicada na arena "{arena}" — '
                             f'{len(cats_distintas)} categorias diferentes com o mesmo número'
                             + (f' (horários {", ".join(horarios)})' if horarios else '')
                             + '. Renumere ou separe.',
                    'onde':  f'{dia_label}/{arena}',
                })

        # Conjunto de categorias (normalizadas) presentes neste dia — coleta as
        # duas formas pra suportar match estrito ou relaxado.
        cats_no_dia_norm: set[str] = set()
        cats_no_dia_relax: set[str] = set()
        cats_no_dia_fuzzy: set[str] = set()
        for b in cronograma:
            cat_str = b.get('categoria', '')
            cats_no_dia_norm.update(_quebrar_categoria_composta(cat_str))
            cats_no_dia_relax.update(
                _normalizar_categoria_relaxada(p)
                for p in _split_partes_categoria(cat_str)
            )
            cats_no_dia_fuzzy.update(
                _chave_categoria_fuzzy(p)
                for p in _split_partes_categoria(cat_str)
            )

        # Modalidades presentes neste dia (Individual=True / Dupla=False),
        # deduzidas das cats casadas por nome no cronograma. Usado pra
        # bloquear match por faixa colidida em modalidade ausente do dia
        # — ex: Sábado só tem Individuais, então faixa 1301-1399 não pode
        # mapear pra `Dupla Iniciante Mista` (mesmo número, modalidade
        # ausente). Vazio = todas as modalidades aceitas (sem coluna
        # `Individual` no Inscritos → comportamento legado).
        mods_no_dia: set[bool] = set()
        for cat_norm, (_ini, _fim, is_indiv) in inscritos_full.items():
            if is_indiv is None:
                continue
            if cat_norm in cats_no_dia_norm or cat_norm in cats_no_dia_relax:
                mods_no_dia.add(is_indiv)

        cats_resultado: list[dict[str, Any]] = []
        # Tracker de chaves da montagem que foram consumidas via cronograma.
        # As que sobrarem são promovidas a baterias extras no fim do loop —
        # cobre Heat órfão (cronograma incompleto mas montagem com alocação).
        chaves_consumidas: set[tuple] = set()
        for cat_grade, workouts in grade_por_categoria.items():
            cat_grade_norm = _normalizar_categoria(cat_grade)
            cat_grade_relax = cats_grade_relaxadas[cat_grade]
            permite_relax = cat_grade not in cats_ambiguas
            cat_grade_fuzzy = cats_grade_fuzzy[cat_grade]
            permite_fuzzy = cat_grade not in cats_ambiguas_fuzzy
            # Faixa de número desta categoria no Inscritos (se houver).
            # Usado pra 3ª camada de match: bateria onde o nome textual não
            # menciona a cat, mas a Montagem tem atletas dela.
            faixa_cat = inscritos_faixas.get(cat_grade_norm) or (
                inscritos_faixas.get(cat_grade_relax) if permite_relax else None
            )
            # Faixa colidida (mesma faixa em duas cats, típico Individuais×Duplas)
            # ainda pode servir pra match por faixa neste dia, desde que:
            #   (a) a(s) outra(s) cat(s) com a mesma faixa NÃO estejam no
            #       cronograma deste dia (sem conflito de nome direto), OU
            #   (b) a modalidade da cat (Individual/Dupla via coluna Inscritos)
            #       esteja presente no dia. Cobre Storm: faixa 1301-1399 é
            #       Teen Intermediario 16-17 Masculino (individual) e Dupla
            #       Iniciante Mista (dupla); no Sábado só Individual aparece,
            #       então a Dupla não pode entrar mesmo sem conflito por nome.
            faixa_cat_unica = faixa_cat
            if faixa_cat in _faixas_colisao:
                conflitos = [c for c, f in inscritos_faixas.items()
                             if f == faixa_cat and c != cat_grade_norm and c != cat_grade_relax]
                conflito_presente = any(
                    c in cats_no_dia_norm or c in cats_no_dia_relax
                    for c in conflitos
                )
                # Modalidade desta cat (via Inscritos). Bloqueia se modalidade
                # não está no dia. Tolerante se coluna `Individual` ausente.
                mod_cat = inscritos_modalidade.get(cat_grade_norm) or (
                    inscritos_modalidade.get(cat_grade_relax) if permite_relax else None
                )
                modalidade_ausente = (
                    mod_cat is not None and mods_no_dia and mod_cat not in mods_no_dia
                )
                if conflito_presente or modalidade_ausente:
                    faixa_cat_unica = None
            # Só anexa categoria se ela aparece em alguma bateria deste dia.
            # Camadas de match, em ordem:
            #   (1) match estrito de normalização (nome)
            #   (2) match relaxado (nome sem parênteses, se não-ambígua)
            #   (3) match por faixa de número (atletas da faixa caem em
            #       alguma bateria da Montagem). Cobre Teens & similares
            #       onde Inscritos usa nome longo (`Teen Scaled 14-15 Fem`)
            #       mas cronograma usa nome curto (`14-15 Feminino`).
            casou_por_nome = (
                cat_grade_norm in cats_no_dia_norm
                or (permite_relax and cat_grade_relax in cats_no_dia_relax)
                or (permite_fuzzy and cat_grade_fuzzy
                    and cat_grade_fuzzy in cats_no_dia_fuzzy)
            )
            casou_por_faixa = faixa_cat_unica is not None and any(
                _bateria_tem_atleta_na_faixa(b.get('numero', ''), montagem, faixa_cat_unica)
                for b in cronograma
            )
            if not (casou_por_nome or casou_por_faixa):
                continue
            cats_grade_casadas.add(cat_grade)

            # Workouts deste DIA específico (filtra wkts da grade pelo _dia_label).
            # Usado tanto pra match de codigo_evento (nome do workout) quanto pra
            # popular cats_resultado.
            workouts_do_dia = [
                w for w in workouts
                if not w.get('_dia_label') or w.get('_dia_label') == dia_norm_atual
            ]

            baterias_da_cat = [
                b for b in cronograma
                if _bateria_casa_categoria(b.get('categoria', ''), cat_grade_norm,
                                           cat_grade_relax, permite_relax,
                                           cat_grade_fuzzy, permite_fuzzy)
                or (faixa_cat_unica is not None
                    and _bateria_tem_atleta_na_faixa(b.get('numero', ''), montagem, faixa_cat_unica))
            ]

            baterias_full: list[dict[str, Any]] = []
            for b in baterias_da_cat:
                codigos_b = set(_split_codigo_evento(b.get('codigo_evento', '')))
                bat_cat_exata = b.get('categoria', '')
                # 1ª passada: match estrito (bat + cat + interseção de códigos).
                # 2ª passada: match relaxado (só bat + cat) — só vale quando há um
                #             único candidato (sem ambiguidade).
                # 3ª passada: match por (codigo + categoria exata), ignorando bateria
                #             — pra eventos multi-arena onde cronograma usa bateria
                #             local por arena e montagem usa bateria global. Só vale
                #             quando há único candidato pra essa categoria+código.
                # Cada candidato carrega a chave completa (cod, cat, bat) pra
                # poder marcar como consumida no tracker.
                candidatos_estrito: list[tuple] = []
                candidatos_relaxado: list[tuple] = []
                candidatos_sem_bat: list[tuple] = []
                for chave, alocs in montagem.items():
                    chave_cod, chave_cat, chave_bat = chave
                    cat_bate = _bateria_casa_categoria(
                        chave_cat, cat_grade_norm, cat_grade_relax, permite_relax,
                        cat_grade_fuzzy, permite_fuzzy,
                    )
                    # 3ª camada: nome textual não bateu mas a Montagem tem
                    # atletas dentro da faixa de número desta categoria.
                    # Só usa faixa única (sem colisão Individual/Dupla).
                    if not cat_bate and faixa_cat_unica is not None:
                        cat_bate = _alocacoes_tem_atleta_na_faixa(alocs, faixa_cat_unica)
                    if not cat_bate:
                        continue
                    cat_exata_bate = (
                        chave_cat.strip().lower() == bat_cat_exata.strip().lower()
                    )
                    if chave_bat == b['numero']:
                        candidatos_relaxado.append((chave, alocs))
                        if codigos_b:
                            codigos_chave = set(_split_codigo_evento(chave_cod))
                            if codigos_b & codigos_chave:
                                candidatos_estrito.append((chave, alocs))
                    elif cat_exata_bate and codigos_b:
                        codigos_chave = set(_split_codigo_evento(chave_cod))
                        if codigos_b & codigos_chave:
                            candidatos_sem_bat.append((chave, alocs))
                escolhido = (
                    candidatos_estrito[0] if candidatos_estrito
                    else (candidatos_relaxado[0] if len(candidatos_relaxado) == 1
                          else (candidatos_sem_bat[0] if len(candidatos_sem_bat) == 1 else None))
                )
                if escolhido:
                    chave_escolhida, aloc = escolhido
                    codigo_montagem = chave_escolhida[0]
                    chaves_consumidas.add(chave_escolhida)
                else:
                    codigo_montagem, aloc = "", []

                # Filtra alocações pela faixa de número da categoria atual
                # quando: (a) bateria mista por nome (`X & Y`) ou (b) bateria
                # casada por faixa (nome do cronograma não menciona a cat).
                # Em ambos os casos a Montagem pode trazer atletas de outras
                # categorias compartilhando o horário; a faixa do Inscritos
                # separa quem é de quem.
                nome_e_misto = len(_quebrar_categoria_composta(b.get('categoria', ''))) > 1
                if aloc and faixa_cat and (nome_e_misto or not casou_por_nome):
                    aloc, descartados = _filtrar_alocacoes_por_faixa(aloc, faixa_cat)
                    # Aviso só pra atletas que NÃO caem em NENHUMA faixa
                    # conhecida — esses são erros reais (typo no Excel ou
                    # atleta fora da numeração). Atletas descartados que
                    # pertencem a outra categoria conhecida estão no lugar
                    # certo, só não é a categoria sendo processada agora.
                    for d in descartados:
                        n_str = str(d.get('numero', '')).strip()
                        try:
                            n_int = int(n_str)
                        except ValueError:
                            continue   # número inválido — outro problema
                        if any(lo <= n_int <= hi for lo, hi in inscritos_faixas.values()):
                            continue   # pertence a outra categoria conhecida
                        nome = (d.get('nome') or '?').strip() or '?'
                        avisos_import.append({
                            'nivel': 'aviso',
                            'msg':   f'Atleta #{n_str} ({nome}) com número fora de toda faixa do Inscritos — não foi atribuído a nenhuma categoria',
                            'onde':  f'{dia_label}/Bat {b.get("numero", "?")}',
                        })

                codigo_final = b.get('codigo_evento') or codigo_montagem
                # workouts_do_dia já está filtrado por dia (após v1.38). Mapeia
                # codigo_evento → posição 1-based no array de workouts da cat.
                # Suporta '#N', 'WKT N' e nome do workout entre aspas
                # ('"Simple Dimension"') — comum no formato Monstar.
                workouts_que_rodam = _workouts_que_rodam_da_bateria(
                    codigo_final, workouts_do_dia)
                baterias_full.append({
                    **b,
                    'codigo_evento': codigo_final,
                    'workouts_que_rodam': workouts_que_rodam,
                    'alocacoes': aloc,
                })

            # Mantém no dia SÓ os workouts que alguma bateria roda — senão a lista
            # da categoria mostra workouts de outros dias (o _dia_label da grade é
            # não-confiável: um mesmo workout roda em dias diferentes por categoria,
            # e o cronograma é a fonte de verdade). Remapeia as posições. Se nenhum
            # roda (dia de planejamento sem código casado), mantém a lista cheia.
            posicoes = sorted({p for b in baterias_full
                               for p in (b.get('workouts_que_rodam') or [])})
            if posicoes and len(posicoes) < len(workouts_do_dia):
                remap = {old: new for new, old in enumerate(posicoes, start=1)}
                workouts_do_dia = [workouts_do_dia[p - 1] for p in posicoes]
                for b in baterias_full:
                    b['workouts_que_rodam'] = [remap[p] for p in b.get('workouts_que_rodam', [])
                                               if p in remap]

            cats_resultado.append({
                'nome':      cat_grade,
                'workouts':  workouts_do_dia,
                'baterias':  baterias_full,
            })

        # Promoção de baterias órfãs: chaves da montagem que ficaram sem match
        # via cronograma (ex: organizador esqueceu de adicionar Heat 2 no
        # cronograma mas registrou as alocações na Montagem). Cada chave órfã
        # vira bateria extra na categoria correspondente da grade.
        for chave, alocs in montagem.items():
            if chave in chaves_consumidas:
                continue
            chave_cod, chave_cat, chave_bat = chave
            partes = _quebrar_categoria_composta(chave_cat)
            # Acha categoria da grade que case com alguma parte dessa chave.
            cat_correspondente = None
            for cat_grade in grade_por_categoria:
                cat_grade_norm = _normalizar_categoria(cat_grade)
                if cat_grade_norm in partes:
                    cat_correspondente = cat_grade
                    break
            if not cat_correspondente:
                continue
            # Pega ou cria a entrada de categoria no resultado
            cat_entry = next((c for c in cats_resultado if c['nome'] == cat_correspondente), None)
            if cat_entry is None:
                cat_entry = {
                    'nome':      cat_correspondente,
                    'workouts':  grade_por_categoria[cat_correspondente],
                    'baterias':  [],
                }
                cats_resultado.append(cat_entry)
            # Filtra alocações se bateria mista + faixa Inscritos disponível
            aloc_final = alocs
            if len(partes) > 1:
                cat_norm = _normalizar_categoria(cat_correspondente)
                faixa = inscritos_faixas.get(cat_norm) or inscritos_faixas.get(
                    _normalizar_categoria_relaxada(cat_correspondente)
                )
                if faixa:
                    aloc_final, _ = _filtrar_alocacoes_por_faixa(alocs, faixa)
            codigos_finais = _split_codigo_evento(chave_cod) or ([chave_cod] if chave_cod else [])
            workouts_que_rodam = [
                n for n in (_workout_numero_de_codigo(c) for c in codigos_finais) if n is not None
            ]
            cat_entry['baterias'].append({
                'numero':              chave_bat,
                'codigo_evento':       chave_cod,
                'categoria':           chave_cat,
                'horario_aquecimento': '',
                'horario_fila':        '',
                'workouts_que_rodam':  workouts_que_rodam,
                'alocacoes':           aloc_final,
            })
            avisos_import.append({
                'nivel': 'aviso',
                'msg':   f'Bateria {chave_bat} ({chave_cat}) tem alocações na Montagem mas não está no cronograma — promovida como bateria extra',
                'onde':  f'{dia_label}/{cat_correspondente}',
            })

        dias_resultado.append({'label': dia_label, 'categorias': cats_resultado})

    # Linter: categoria da grade que não casou com NENHUMA bateria — não gera
    # súmula. Quase sempre é nome divergente entre grade e cronograma/Inscritos.
    for cat_grade in grade_por_categoria:
        if cat_grade not in cats_grade_casadas:
            avisos_import.append({
                'nivel': 'erro',
                'msg':   f'Categoria "{cat_grade}" tem workouts na grade mas não '
                         f'aparece em nenhuma bateria do cronograma — não vai gerar súmula. '
                         f'Confira se o nome bate com o Inscritos/cronograma.',
                'onde':  'grade × cronograma',
            })

    return {
        'tipo':         'evento_multidia',
        'evento_nome':  '',
        'dias':         dias_resultado,
        'roster':       _roster_de_abas_atletas(wb),
        'avisos_import': avisos_import,
    }
